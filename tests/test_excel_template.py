"""Tests for suiteview.utils.excel_template (template → workbook conversion)."""
import os
import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from suiteview.utils.excel_template import (
    is_excel_template,
    workbook_filename,
    copy_as_workbook,
)


def _make_template(path, macro=False):
    """Write a minimal Excel template at *path* and stamp it as a template."""
    wb = Workbook()
    wb.template = True
    wb.save(path)
    return path


@pytest.mark.parametrize(
    "name,expected",
    [
        ("Interest.xltx", True),
        ("Interest.XLTX", True),
        ("Macro.xltm", True),
        ("Legacy.xlt", True),
        ("Interest.xlsx", False),
        ("Macro.xlsm", False),
        ("Legacy.xls", False),
        ("notes.txt", False),
        ("noext", False),
    ],
)
def test_is_excel_template(name, expected):
    assert is_excel_template(name) is expected


@pytest.mark.parametrize(
    "name,expected",
    [
        ("Interest.xltx", "Interest.xlsx"),
        ("Macro.xltm", "Macro.xlsm"),
        ("Legacy.xlt", "Legacy.xls"),
        ("Interest.XLTX", "Interest.xlsx"),
        ("Interest.xlsx", "Interest.xlsx"),
        ("notes.txt", "notes.txt"),
        ("noext", "noext"),
    ],
)
def test_workbook_filename(name, expected):
    assert workbook_filename(name) == expected


def test_copy_as_workbook_converts_xltx(tmp_path):
    src = _make_template(str(tmp_path / "Interest.xltx"))
    dest = str(tmp_path / f"12345 - {workbook_filename('Interest.xltx')}")

    copy_as_workbook(src, dest)

    assert os.path.basename(dest) == "12345 - Interest.xlsx"

    content_types = zipfile.ZipFile(dest).read("[Content_Types].xml").decode("utf-8")
    assert "spreadsheetml.sheet.main+xml" in content_types
    assert "template.main+xml" not in content_types

    # Reopened as a regular workbook, no longer flagged as a template.
    assert load_workbook(dest).template is False


def test_copy_as_workbook_converts_xltm_macro_content_type(tmp_path):
    # openpyxl can't author a real .xltm, so craft a minimal zip whose
    # content type is the macro-template main type and verify the swap.
    src = str(tmp_path / "Macro.xltm")
    macro_template_ct = (
        "application/vnd.ms-excel.template.macroEnabled.main+xml"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        f'<Override PartName="/xl/workbook.xml" ContentType="{macro_template_ct}"/>'
        '</Types>'
    )
    with zipfile.ZipFile(src, "w") as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("xl/workbook.xml", "<workbook/>")

    dest = str(tmp_path / "Macro.xlsm")
    copy_as_workbook(src, dest)

    out = zipfile.ZipFile(dest).read("[Content_Types].xml").decode("utf-8")
    assert "application/vnd.ms-excel.sheet.macroEnabled.main+xml" in out
    assert macro_template_ct not in out


def test_copy_as_workbook_passthrough_non_template(tmp_path):
    src = str(tmp_path / "data.xlsx")
    Workbook().save(src)

    dest = str(tmp_path / "data_copy.xlsx")
    copy_as_workbook(src, dest)

    assert os.path.getsize(dest) == os.path.getsize(src)
    # Still a valid, openable workbook.
    assert load_workbook(dest) is not None
