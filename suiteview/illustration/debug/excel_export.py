"""Export projection results to Excel for debug/validation.

Usage:
    from suiteview.illustration.debug.excel_export import export_projection_to_excel
    export_projection_to_excel(results, "debug_output.xlsx", policy_data=policy)
"""
from __future__ import annotations

from dataclasses import fields
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from suiteview.illustration.models.calc_state import MonthlyState
from suiteview.illustration.models.policy_data import IllustrationPolicyData


# Pipeline-ordered field list — matches the calculation sequence
_PIPELINE_ORDER = [
    # COUNTERS
    "date", "policy_year", "policy_month", "duration", "attained_age",
    "is_anniversary",
    # TRACKING (inputs carried forward — needed before premium calc)
    "premiums_ytd", "premiums_to_date", "withdrawals_to_date", "cost_basis",
    # LOAN CAP (capitalized loan balances — before premium, per RERUN row 16)
    "rg_loan_princ", "rg_loan_accrued",
    "pf_loan_princ", "pf_loan_accrued",
    "vbl_loan_princ", "vbl_loan_accrued",
    # PREMIUM
    "gross_premium", "prem_under_target", "prem_over_target",
    "target_load", "excess_load", "flat_load", "total_premium_load",
    "net_premium", "av_after_premium",
    # DEDUCTION
    "nar_av", "corridor_rate", "standard_db", "gross_db", "corr_amount",
    "discounted_db_cov1", "discounted_db_corr", "discounted_db",
    "nar_cov1", "nar_corr", "nar",
    "coi_rate", "coi_charge_cov1", "coi_charge_corr", "coi_charge",
    "epu_rate",
    "epu_charge", "mfee_charge", "av_charge", "pw_charge", "benefit_charges",
    "rider_charges", "total_deduction",
    "av_after_deduction",
    # INFORCE MONTHLY DEDUCTION CHECK
    "system_coi_charge", "system_expense_charge", "system_other_charge",
    "system_monthly_deduction", "md_check_av_before_deduction",
    "md_check_calculated_deduction", "md_check_deduction_variance",
    "md_check_calculated_av_after_deduction", "md_check_av_variance",
    # INTEREST
    "days_in_month", "annual_interest_rate", "bonus_interest_rate",
    "effective_annual_rate", "monthly_interest_rate",
    "reg_impaired_int", "pref_impaired_int",
    "interest_credited",
    "av_end_of_month",
    # LOAN ACCRUAL (after interest — per RERUN cols 587-592)
    "reg_loan_charge", "pref_loan_charge", "vbl_loan_charge",
    "end_rg_loan_princ", "end_rg_loan_accrued",
    "end_pf_loan_princ", "end_pf_loan_accrued",
    "end_vbl_loan_princ", "end_vbl_loan_accrued",
    "policy_debt",
    # END OF MONTH
    "scr_rate", "surrender_charge", "surrender_value", "ending_sv", "ending_db",
    # SHADOW ACCOUNT (CCV)
    "shadow_bav", "shadow_wd_charges", "shadow_sa",
    "shadow_target_prem", "shadow_prem_under_target", "shadow_prem_over_target",
    "shadow_target_load", "shadow_excess_load", "shadow_prem_load",
    "shadow_net_prem", "shadow_nar_av", "shadow_db",
    "shadow_coi_rate", "shadow_coi", "shadow_dbd_rate", "shadow_nar",
    "shadow_epu_rate", "shadow_epu", "shadow_mfee",
    "shadow_rider_charges", "shadow_md", "shadow_av",
    "shadow_days", "shadow_int_rate", "shadow_eff_rate",
    "shadow_interest", "shadow_eav", "shadow_eav_less_debt",
    # SAFETY NET / LAPSE PROTECTION
    "monthly_mtp", "accumulated_mtp", "accum_mtp_less_prem",
    "snet_active", "shadow_protection", "positive_sv", "av_less_loans",
    # CUMULATIVE
    "cumulative_interest", "cumulative_charges",
    # STATUS
    "lapsed",
]

# Section boundaries — first field in each section
_SECTION_MAP = {
    "date": "COUNTERS",
    "premiums_ytd": "TRACKING",
    "rg_loan_princ": "LOAN CAP",
    "gross_premium": "PREMIUM",
    "nar_av": "DEDUCTION",
    "system_coi_charge": "MD CHECK",
    "days_in_month": "INTEREST",
    "reg_loan_charge": "LOAN ACCRUAL",
    "scr_rate": "END OF MONTH",
    "shadow_bav": "SHADOW (CCV)",
    "monthly_mtp": "SAFETY NET",
    "cumulative_interest": "CUMULATIVE",
    "lapsed": "STATUS",
}

# Section colors (alternating for visual grouping)
_SECTION_COLORS = {
    "COUNTERS":     "2F5496",
    "TRACKING":     "548235",
    "LOAN CAP":     "BF6900",
    "PREMIUM":      "4472C4",
    "DEDUCTION":    "C00000",
    "MD CHECK":     "833C0C",
    "INTEREST":     "7030A0",
    "LOAN ACCRUAL": "BF6900",
    "END OF MONTH": "BF8F00",
    "SHADOW (CCV)": "375623",
    "SAFETY NET":  "4A235A",
    "CUMULATIVE":   "548235",
    "STATUS":       "808080",
}

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_FIELD_FONT = Font(bold=True, size=9)
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)

# Columns to format as currency
_CURRENCY_FIELDS = {
    "gross_premium", "prem_under_target", "prem_over_target",
    "target_load", "excess_load", "flat_load", "total_premium_load",
    "net_premium", "av_after_premium", "nar_av", "standard_db",
    "gross_db", "corr_amount", "discounted_db_cov1", "discounted_db_corr",
    "discounted_db", "nar_cov1", "nar_corr", "nar",
    "total_db", "total_discounted_db", "total_nar",
    "coi_charge_cov1", "coi_charge_corr", "coi_charge", "total_coi_charge",
    "epu_charge", "mfee_charge", "av_charge",
    "pw_charge", "benefit_charges", "rider_charges", "total_deduction", "av_after_deduction",
    "system_coi_charge", "system_expense_charge", "system_other_charge",
    "system_monthly_deduction", "md_check_av_before_deduction",
    "md_check_calculated_deduction", "md_check_deduction_variance",
    "md_check_calculated_av_after_deduction", "md_check_av_variance",
    "reg_impaired_int", "pref_impaired_int", "interest_credited",
    "av_end_of_month", "surrender_charge", "surrender_value", "ending_sv",
    "ending_db", "premiums_ytd", "premiums_to_date", "withdrawals_to_date", "cost_basis",
    "cumulative_interest", "cumulative_charges",
    "rg_loan_princ", "rg_loan_accrued",
    "pf_loan_princ", "pf_loan_accrued",
    "reg_loan_charge", "pref_loan_charge", "vbl_loan_charge",
    "vbl_loan_princ", "vbl_loan_accrued",
    "end_rg_loan_princ", "end_rg_loan_accrued",
    "end_pf_loan_princ", "end_pf_loan_accrued",
    "end_vbl_loan_princ", "end_vbl_loan_accrued",
    "policy_debt",
    "shadow_bav", "shadow_wd_charges", "shadow_sa",
    "shadow_target_prem", "shadow_prem_under_target", "shadow_prem_over_target",
    "shadow_target_load", "shadow_excess_load", "shadow_prem_load",
    "shadow_net_prem", "shadow_nar_av", "shadow_db",
    "shadow_coi", "shadow_nar",
    "shadow_epu", "shadow_mfee",
    "shadow_rider_charges", "shadow_md", "shadow_av",
    "shadow_interest", "shadow_eav", "shadow_eav_less_debt",
    "monthly_mtp", "accumulated_mtp", "accum_mtp_less_prem",
    "av_less_loans",
}

_RATE_FIELDS = {
    "corridor_rate", "coi_rate", "epu_rate", "scr_rate",
    "annual_interest_rate", "bonus_interest_rate",
    "effective_annual_rate", "monthly_interest_rate",
    "shadow_coi_rate", "shadow_dbd_rate", "shadow_epu_rate",
    "shadow_int_rate", "shadow_eff_rate",
}

_INFORCE_FILL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid",
)


def export_projection_to_excel(
    results: List[MonthlyState],
    filepath: str | Path,
    policy_data: Optional[IllustrationPolicyData] = None,
) -> Path:
    """Export projection results to an Excel workbook.

    Args:
        results: List of MonthlyState from engine.project().
        filepath: Output .xlsx path.
        policy_data: Optional policy data to include as a summary sheet.

    Returns:
        Path to the created file.
    """
    filepath = Path(filepath)
    wb = Workbook()

    # ── Projection sheet ──────────────────────────────────────
    ws = wb.active
    ws.title = "Projection"
    projection_order = _build_projection_order(policy_data)

    # Row 1: Section headers (colored per section)
    current_section = ""
    for col_idx, fname in enumerate(projection_order, 1):
        section = _SECTION_MAP.get(fname)
        if section:
            current_section = section
        color = _SECTION_COLORS.get(current_section, "4472C4")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell = ws.cell(row=1, column=col_idx, value=current_section)
        cell.font = _SECTION_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Row 2: Field name headers
    for col_idx, fname in enumerate(projection_order, 1):
        section = _SECTION_MAP.get(fname)
        if section:
            current_section = section
        color = _SECTION_COLORS.get(current_section, "4472C4")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell = ws.cell(row=2, column=col_idx, value=fname)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows (row 3+)
    for row_offset, state in enumerate(results):
        row_idx = row_offset + 3
        is_inforce = row_offset == 0 and state.gross_premium == 0.0
        for col_idx, fname in enumerate(projection_order, 1):
            val = _get_projection_value(state, fname)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if fname in _CURRENCY_FIELDS or _is_benefit_amount_or_charge(fname):
                cell.number_format = '#,##0.00'
            elif fname in _RATE_FIELDS or _is_benefit_rate(fname):
                cell.number_format = '0.0000000'

            if is_inforce:
                cell.fill = _INFORCE_FILL

    # Auto-width
    for col_idx, fname in enumerate(projection_order, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(fname) + 2, 12)

    # Freeze first two header rows + date column
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(projection_order))}{len(results) + 2}"

    # ── Policy summary sheet (optional) ───────────────────────
    if policy_data is not None:
        _write_policy_summary(wb, policy_data)

    wb.save(str(filepath))
    return filepath


def _build_projection_order(
    policy_data: Optional[IllustrationPolicyData],
) -> List[str]:
    order = list(_PIPELINE_ORDER)
    if policy_data is None:
        return order

    collapsed_deduction_fields = [
        "standard_db", "gross_db", "corr_amount", "epu_rate", "epu_charge",
        "discounted_db_cov1", "discounted_db_corr", "discounted_db",
        "nar_cov1", "nar_corr", "nar",
        "coi_rate", "coi_charge_cov1", "coi_charge_corr", "coi_charge",
    ]
    for field in collapsed_deduction_fields:
        if field in order:
            order.remove(field)

    coverage_keys = [f"cov{i}" for i in range(1, max(1, len(policy_data.segments)) + 1)]
    deduction_detail_fields = []
    deduction_detail_fields.extend([f"db_{key}" for key in coverage_keys])
    deduction_detail_fields.append("db_corr")
    deduction_detail_fields.extend([f"discounted_db_{key}" for key in coverage_keys])
    deduction_detail_fields.append("discounted_db_corr")
    deduction_detail_fields.extend([f"nar_{key}" for key in coverage_keys])
    deduction_detail_fields.append("nar_corr")
    deduction_detail_fields.extend([f"coi_rate{i}" for i in range(1, len(coverage_keys) + 1)])
    deduction_detail_fields.append("coi_rate_corr")
    deduction_detail_fields.extend([f"coi_charge_{key}" for key in coverage_keys])
    deduction_detail_fields.append("coi_charge_corr")
    deduction_detail_fields.extend([
        "total_db", "total_discounted_db", "total_nar", "total_coi_charge",
    ])
    for index in range(1, len(coverage_keys) + 1):
        deduction_detail_fields.extend([f"epu_rate{index}", f"epu_charge{index}"])
    deduction_detail_fields.append("epu_charge")

    if "corridor_rate" in order:
        idx = order.index("corridor_rate") + 1
        order[idx:idx] = deduction_detail_fields

    has_benefits = any(
        benefit.is_active and not (benefit.benefit_type or "").startswith("#")
        for benefit in policy_data.benefits
    )
    has_riders = any(rider.is_active for rider in policy_data.riders)

    if not has_benefits and "benefit_charges" in order:
        order.remove("benefit_charges")
    if not has_riders and "rider_charges" in order:
        order.remove("rider_charges")

    benefit_fields = []
    seen = set()
    for benefit in policy_data.benefits:
        if not benefit.is_active:
            continue
        if (benefit.benefit_type or "").startswith("#"):
            continue
        code = (benefit.benefit_type or "") + (benefit.benefit_subtype or "")
        if not code or code in seen:
            continue
        seen.add(code)
        benefit_fields.extend([
            f"benefit_{code}_amount",
            f"benefit_{code}_rate",
            f"benefit_{code}_charge",
        ])

    if benefit_fields and "benefit_charges" in order:
        idx = order.index("benefit_charges") + 1
        order[idx:idx] = benefit_fields

    rider_fields = []
    for rider in policy_data.riders:
        if not rider.is_active:
            continue
        key = rider.export_key
        rider_fields.extend([
            f"rider_{key}_amount",
            f"rider_{key}_rate",
            f"rider_{key}_charge",
        ])

    if rider_fields and "rider_charges" in order:
        idx = order.index("rider_charges") + 1
        order[idx:idx] = rider_fields

    if "scr_rate" in order:
        order.remove("scr_rate")
    if "surrender_charge" in order:
        order.remove("surrender_charge")
    surrender_fields = []
    for index in range(1, len(coverage_keys) + 1):
        surrender_fields.extend([f"scr_rate{index}", f"surrender_charge{index}"])
    surrender_fields.append("surrender_charge")
    if "surrender_value" in order:
        idx = order.index("surrender_value")
        order[idx:idx] = surrender_fields
    return order


def _get_projection_value(state: MonthlyState, field_name: str):
    if field_name.startswith("db_cov"):
        key = field_name[len("db_"):]
        return state.db_by_coverage.get(key, 0.0)
    if field_name == "db_corr":
        return state.corr_amount
    if field_name.startswith("discounted_db_cov"):
        key = field_name[len("discounted_db_"):]
        return state.discounted_db_by_coverage.get(key, 0.0)
    if field_name == "discounted_db_corr":
        return state.discounted_db_corr
    if field_name.startswith("nar_cov"):
        key = field_name[len("nar_"):]
        return state.nar_by_coverage.get(key, 0.0)
    if field_name == "nar_corr":
        return state.nar_corr
    if field_name.startswith("coi_rate"):
        suffix = field_name[len("coi_rate"):]
        if suffix == "_corr":
            return state.coi_rate
        if suffix.isdigit():
            return state.coi_rates_by_coverage.get(f"cov{suffix}", 0.0)
    if field_name.startswith("coi_charge_cov"):
        key = field_name[len("coi_charge_"):]
        return state.coi_charges_by_coverage.get(key, 0.0)
    if field_name == "coi_charge_corr":
        return state.coi_charge_corr
    if field_name.startswith("epu_rate"):
        suffix = field_name[len("epu_rate"):]
        if suffix.isdigit():
            return state.epu_rates_by_coverage.get(f"cov{suffix}", 0.0)
    if field_name.startswith("epu_charge") and field_name != "epu_charge":
        suffix = field_name[len("epu_charge"):]
        if suffix.isdigit():
            return state.epu_charges_by_coverage.get(f"cov{suffix}", 0.0)
    if field_name.startswith("scr_rate") and field_name != "scr_rate":
        suffix = field_name[len("scr_rate"):]
        if suffix.isdigit():
            return state.scr_rates_by_coverage.get(f"cov{suffix}", 0.0)
    if field_name.startswith("surrender_charge") and field_name != "surrender_charge":
        suffix = field_name[len("surrender_charge"):]
        if suffix.isdigit():
            return state.surrender_charges_by_coverage.get(f"cov{suffix}", 0.0)
    if field_name.startswith("benefit_"):
        parts = field_name.split("_")
        if len(parts) >= 3:
            code = parts[1]
            value_type = parts[2]
            if value_type == "amount":
                return state.benefit_amounts.get(code, 0.0)
            if value_type == "rate":
                return state.benefit_rates.get(code, 0.0)
            if value_type == "charge":
                return state.benefit_charge_detail.get(code, 0.0)
    if field_name.startswith("rider_"):
        remainder = field_name[len("rider_"):]
        for suffix, values in (
            ("_amount", state.rider_amounts),
            ("_rate", state.rider_rates),
            ("_charge", state.rider_charge_detail),
        ):
            if remainder.endswith(suffix):
                key = remainder[: -len(suffix)]
                return values.get(key, 0.0)
    return getattr(state, field_name)


def _is_benefit_amount_or_charge(field_name: str) -> bool:
    return _is_deduction_amount(field_name) or (
        field_name.startswith("benefit_") or field_name.startswith("rider_")
    ) and (field_name.endswith("_amount") or field_name.endswith("_charge"))


def _is_benefit_rate(field_name: str) -> bool:
    return _is_deduction_rate(field_name) or (
        field_name.startswith("benefit_") or field_name.startswith("rider_")
    ) and field_name.endswith("_rate")


def _is_deduction_amount(field_name: str) -> bool:
    return (
        field_name.startswith("db_cov")
        or field_name == "db_corr"
        or field_name.startswith("discounted_db_cov")
        or field_name == "discounted_db_corr"
        or field_name.startswith("nar_cov")
        or field_name == "nar_corr"
        or field_name.startswith("coi_charge_cov")
        or field_name == "coi_charge_corr"
        or (field_name.startswith("epu_charge") and field_name != "epu_charge")
        or (field_name.startswith("surrender_charge") and field_name != "surrender_charge")
        or field_name in {"total_db", "total_discounted_db", "total_nar", "total_coi_charge"}
    )


def _is_deduction_rate(field_name: str) -> bool:
    if field_name.startswith("coi_rate"):
        suffix = field_name[len("coi_rate"):]
        return field_name == "coi_rate_corr" or suffix.isdigit()
    if field_name.startswith("epu_rate"):
        return field_name[len("epu_rate"):].isdigit()
    if field_name.startswith("scr_rate"):
        return field_name[len("scr_rate"):].isdigit()
    return False


def _write_policy_summary(wb: Workbook, policy: IllustrationPolicyData) -> None:
    """Write a summary sheet with policy input data."""
    ws = wb.create_sheet("Policy Data")

    ws.cell(row=1, column=1, value="Field").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=2, value="Value").font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL

    # Write policy fields (skip internal/list fields)
    row = 2
    for f in fields(policy):
        if f.name.startswith("_"):
            continue
        val = getattr(policy, f.name)
        if isinstance(val, list):
            continue
        ws.cell(row=row, column=1, value=f.name)
        ws.cell(row=row, column=2, value=str(val) if val is not None else "")
        row += 1

    # Write segments
    if policy.segments:
        row += 1
        ws.cell(row=row, column=1, value="SEGMENTS").font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        row += 1
        for i, seg in enumerate(policy.segments):
            ws.cell(row=row, column=1, value=f"Segment {i+1}").font = Font(bold=True)
            row += 1
            for sf in fields(seg):
                sval = getattr(seg, sf.name)
                ws.cell(row=row, column=1, value=f"  {sf.name}")
                ws.cell(row=row, column=2, value=str(sval) if sval is not None else "")
                row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.freeze_panes = "A2"
