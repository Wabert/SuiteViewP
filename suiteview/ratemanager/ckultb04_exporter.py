"""
CKULTB04 surrender-charge Excel exporters.

Two output shapes, both written with openpyxl's write-only workbook so the
very large (700k+ row) reports stream to disk without holding the whole
workbook in memory:

  * ``export_raw``   — every CKULTB04 column, one row per record.
  * ``export_table`` — a trimmed rate table: State Code, Sex, Rate Class,
    Band, Issue Age, Duration, Rate.
"""

from __future__ import annotations

import os
from typing import Callable, Iterable, Optional

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font

from suiteview.ratemanager.ckultb04_parser import (
    RAW_HEADERS,
    RAW_KEYS,
    iter_records,
)

# "Excel Table" layout: (record key, column header).
TABLE_COLUMNS = [
    ("PLAN_CODE", "Plan Code"),
    ("STATE_CODE", "State Code"),
    ("SEX_CODE", "Sex"),
    ("RATE_CLASS", "Rate Class"),
    ("BAND_CODE", "Band"),
    ("HIGH_ISSUE_AGE", "Issue Age"),
    ("HIGH_DURATION", "Duration"),
    ("CHARGE_PCT", "Rate"),
]

_HEADER_FONT = Font(bold=True)


def _header_cell(ws, value):
    cell = WriteOnlyCell(ws, value=value)
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    return cell


def _split_progress(
    progress_cb: Optional[Callable[[float], None]],
    parse_fraction: float = 0.9,
):
    """Wrap ``progress_cb`` so parsing fills 0..parse_fraction of the bar."""
    if progress_cb is None:
        return None

    def _cb(frac: float) -> None:
        progress_cb(frac * parse_fraction)

    return _cb


def export_raw(
    input_path: str,
    output_path: str,
    progress_cb: Optional[Callable[[float], None]] = None,
    plan_codes: Optional[Iterable[str]] = None,
) -> int:
    """Write every CKULTB04 column to ``output_path``. Returns row count.

    ``plan_codes`` optionally limits output to the given PLAN CODE values;
    when ``None`` (or empty) every plan code is written.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("CKULTB04")

    ws.append([_header_cell(ws, h) for h in RAW_HEADERS])

    allow = set(plan_codes) if plan_codes else None
    count = 0
    for record in iter_records(input_path, progress_cb=_split_progress(progress_cb)):
        if allow is not None and record["PLAN_CODE"] not in allow:
            continue
        ws.append([record[key] for key in RAW_KEYS])
        count += 1

    if progress_cb is not None:
        progress_cb(0.95)
    wb.save(output_path)
    if progress_cb is not None:
        progress_cb(1.0)
    return count


def export_table(
    input_path: str,
    output_path: str,
    progress_cb: Optional[Callable[[float], None]] = None,
    plan_codes: Optional[Iterable[str]] = None,
) -> int:
    """Write the trimmed rate table to ``output_path``. Returns row count.

    ``plan_codes`` optionally limits output to the given PLAN CODE values;
    when ``None`` (or empty) every plan code is written.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("CKULTB04 Table")

    ws.append([_header_cell(ws, label) for _, label in TABLE_COLUMNS])

    allow = set(plan_codes) if plan_codes else None
    keys = [key for key, _ in TABLE_COLUMNS]
    count = 0
    for record in iter_records(input_path, progress_cb=_split_progress(progress_cb)):
        if allow is not None and record["PLAN_CODE"] not in allow:
            continue
        ws.append([record[key] for key in keys])
        count += 1

    if progress_cb is not None:
        progress_cb(0.95)
    wb.save(output_path)
    if progress_cb is not None:
        progress_cb(1.0)
    return count


def default_output_name(input_path: str, suffix: str) -> str:
    """Build ``<stem> - <suffix>.xlsx`` from the input filename."""
    base = os.path.basename(input_path)
    stem = os.path.splitext(base)[0].strip()
    return f"{stem} - {suffix}.xlsx"
