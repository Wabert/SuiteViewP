"""
CKULTB04 surrender-charge DB Format builder.

Turns the CKULTB04 surrender-charge report into UL_Rates-ready tables, one
workbook per run with two sheets:

  * SCR POINTER — one row per (State, Sex, RateClass, Band) combo of a plan
    code, mapping it to an Index(SCR):
        Plancode, IssueVersion, Sex, RateClass, Band, State, Index(SCR)
  * RATE_SCR    — the expanded surrender-charge schedule per index:
        Index(SCR), IssueAge, Duration, Rate

Rules (confirmed with the product owner):
  * Surrender charges are the SAME current and guaranteed, so there is no Scale
    column — one schedule per index.
  * Durations are 1-based (source year 0 becomes Duration 1). Each issue-age
    schedule is filled with 0 out to maturity: Durations 1..(MaturityAge -
    IssueAge). The source "999" terminal marker (rate 0 thereafter) is dropped;
    the trailing years are written explicitly as 0.
  * IssueVersion is always 1.
  * Index(SCR) starts at the user-supplied Starting Index per plan code and
    counts up once per DISTINCT schedule — combos with identical schedules
    share an index (same index assignment as the other DB formats).
"""

from __future__ import annotations

from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font

from suiteview.ratemanager.ckultb04_parser import iter_records

POINTER_HEADERS = [
    "Plancode", "IssueVersion", "Sex", "RateClass", "Band", "State", "Index(SCR)",
]
SCR_HEADERS = ["Index(SCR)", "IssueAge", "Duration", "Rate"]

ISSUE_VERSION = 1
# HIGH_DURATION at/above this is the "0 charge thereafter" terminal marker.
_SENTINEL_DUR = 900

_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")

# combo = (state, sex, rate_class, band)
ComboKey = Tuple[str, str, str, str]
# schedule = {issue_age: {source_duration: rate}}
Schedule = Dict[int, Dict[int, float]]


@dataclass
class CKULTB04DBSpec:
    """Per-plan-code DB Format settings supplied by the user."""
    plan_code: str
    maturity_age: int
    start_index: int


def _collect(
    input_path: str,
    plan_codes: List[str],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Dict[ComboKey, Schedule]]:
    """Gather the explicit surrender-charge rates for the selected plan codes.

    Returns ``{plan_code: {combo: {issue_age: {source_duration: rate}}}}``. The
    terminal 999 marker rows are skipped (durations past the schedule are filled
    with 0 at build time).
    """
    wanted = set(plan_codes)
    out: Dict[str, Dict[ComboKey, Schedule]] = {
        pc: defaultdict(lambda: defaultdict(dict)) for pc in wanted
    }
    for rec in iter_records(input_path, progress_cb=progress_cb):
        pc = rec["PLAN_CODE"]
        if pc not in wanted:
            continue
        dur = rec["HIGH_DURATION"]
        if dur >= _SENTINEL_DUR:
            continue
        combo: ComboKey = (
            rec["STATE_CODE"], rec["SEX_CODE"], rec["RATE_CLASS"], rec["BAND_CODE"],
        )
        out[pc][combo][rec["HIGH_ISSUE_AGE"]][dur] = rec["CHARGE_PCT"]
    return out


def _signature(schedule: Schedule) -> tuple:
    """A hashable signature of a combo's explicit rates (for dedupe)."""
    return tuple(
        (ia, tuple(sorted(schedule[ia].items())))
        for ia in sorted(schedule)
    )


def _schedule_rows(index: int, schedule: Schedule, maturity_age: int) -> List[list]:
    """Expand one combo's schedule to 1-based durations, 0-filled to maturity."""
    rows: List[list] = []
    for ia in sorted(schedule):
        explicit = schedule[ia]
        max_dur = maturity_age - ia          # 1-based Durations 1..max_dur
        if max_dur < 1:
            continue
        for dur in range(1, max_dur + 1):
            rate = explicit.get(dur - 1, 0.0)   # source year 0 -> Duration 1
            rows.append([index, ia, dur, round(rate, 6)])
    return rows


def _header_cell(ws, value):
    cell = WriteOnlyCell(ws, value=value)
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    return cell


def _write_sheet(ws, headers: List[str], rows: List[list]) -> None:
    ws.append([_header_cell(ws, h) for h in headers])
    for row in rows:
        ws.append(row)


def build_ckultb04_db(
    input_path: str,
    specs: List[CKULTB04DBSpec],
    output_path: str,
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Dict[str, int]]:
    """Build the SCR POINTER + RATE_SCR workbook. Returns per-plancode counts."""
    def _parse_cb(frac: float) -> None:
        if progress_cb:
            progress_cb(frac * 0.7)

    data = _collect(input_path, [s.plan_code for s in specs], progress_cb=_parse_cb)

    pointer_rows: List[list] = []
    scr_rows: List[list] = []
    counts: Dict[str, Dict[str, int]] = {}

    total = max(len(specs), 1)
    for si, spec in enumerate(specs):
        combos_data = data.get(spec.plan_code, {})
        combos = sorted(combos_data.keys())

        groups: "OrderedDict[tuple, int]" = OrderedDict()
        combo_index: Dict[ComboKey, int] = {}
        next_index = spec.start_index
        for combo in combos:
            sig = _signature(combos_data[combo])
            idx = groups.get(sig)
            if idx is None:
                idx = next_index
                groups[sig] = idx
                scr_rows.extend(
                    _schedule_rows(idx, combos_data[combo], spec.maturity_age))
                next_index += 1
            combo_index[combo] = idx

        for combo in combos:
            state, sex, rate_class, band = combo
            pointer_rows.append([
                spec.plan_code, ISSUE_VERSION, sex, rate_class, band, state,
                combo_index[combo],
            ])

        counts[spec.plan_code] = {
            "pointer": len(combos),
            "scr_groups": len(groups),
        }
        if progress_cb:
            progress_cb(0.7 + 0.25 * (si + 1) / total)

    wb = Workbook(write_only=True)
    _write_sheet(wb.create_sheet("SCR POINTER"), POINTER_HEADERS, pointer_rows)
    _write_sheet(wb.create_sheet("RATE_SCR"), SCR_HEADERS, scr_rows)
    if progress_cb:
        progress_cb(0.97)
    wb.save(output_path)
    if progress_cb:
        progress_cb(1.0)

    counts["_totals"] = {"pointer": len(pointer_rows), "scr": len(scr_rows)}
    return counts
