"""
Benefit DB Format builder.

Turns the benefit rider rates in an IAF into UL_Rates-ready tables:

  * BENEFIT POINTER — one row per BASE-plancode (Sex, RateClass, Band) combo,
    mapping it to a benefit COI index and a benefit target index.
  * RATE_BENCOI    — Index(BENCOI), Scale, IssueAge, Duration, Rate.
  * RATE_BENTRG    — Index(BENTRG), IssueAge, Rate(MTP), Rate(CTP).

Rules (confirmed with the product owner):
  * The pointer must cover the SAME Sex/Class/Band combos as the base plancode.
    Benefit rates are usually unbanded (band '0'); when a benefit does not vary
    by band, every base band for a given (sex, class) points to the same index.
  * The user supplies one starting BENEFIT INDEX per benefit. BENCOI and BENTRG
    indices are assigned independently — each starts at that index and counts
    up per distinct rate group (combos with identical rates share an index).
  * Renewable benefit → COI varies by attained age (duration d uses the
    ultimate rate at attained age issue+d-1). Non-renewable → the issue-age
    rate is held level across durations.
  * Benefit COI is ultimate-only. Current == guaranteed unless the benefit has
    its own G rates, so RATE_BENCOI carries Scale 0 (guaranteed) and Scale 1
    (current) with identical rates in the common case.
"""

from __future__ import annotations

from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from suiteview.ratemanager.parser import IAFParser, ParseResult

BASE_OPTIONS = {"**", "E*"}

_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")

POINTER_HEADERS = [
    "Plancode", "BenefitType", "Benefit", "IssueVersion",
    "Sex", "RateClass", "Band", "Index(BENCOI)", "Index(BENTRG)",
]
BENCOI_HEADERS = ["Index(BENCOI)", "Scale", "IssueAge", "Duration", "Rate"]
BENTRG_HEADERS = ["Index(BENTRG)", "IssueAge", "Rate(MTP)", "Rate(CTP)"]

ComboKey = Tuple[str, str, str]   # (sex, rate_class, band)


@dataclass
class BenefitDBSpec:
    """Per-benefit reformat settings supplied by the user."""
    code: str
    renewable: bool
    start_index: int


def _norm(opt: str) -> str:
    return (opt or "").strip()


def _base_combos(result: ParseResult) -> List[ComboKey]:
    """Sex/Class/Band combos of the BASE plancode (its current COI rates)."""
    combos = set()
    for r in result.rates:
        if r.rate_type == "C" and _norm(r.plan_option) == "**":
            combos.add((r.gender, r.rate_class, r.band))
    return sorted(combos)


def _pay_age(result: ParseResult) -> int:
    if result.products and result.products[0].pay_age:
        return result.products[0].pay_age
    return 121


def _plancode(result: ParseResult) -> str:
    return result.products[0].plancode.strip() if result.products else ""


def _issue_version(result: ParseResult) -> str:
    """The plan's issue version — '1' when the IAF's V column is blank."""
    if result.products and result.products[0].version.strip():
        return result.products[0].version.strip()
    return "1"


def _benefit_rates_by_combo(
    result: ParseResult, code: str, rate_type: str,
) -> Dict[ComboKey, Dict[int, float]]:
    """``{combo: {age: rate}}`` for one benefit code and rate type.

    Keyed by the rate's age field (attained age for COI ultimate rows, issue
    age for the duration-0 targets). Deduplicated to the most recent
    scale_start when multiple vintages exist.
    """
    best_date: Dict[Tuple[ComboKey, int], str] = {}
    out: Dict[ComboKey, Dict[int, float]] = defaultdict(dict)
    for r in result.rates:
        if r.rate_type != rate_type or _norm(r.plan_option) != code:
            continue
        combo = (r.gender, r.rate_class, r.band)
        key = (combo, r.attained_age)
        prev = best_date.get(key)
        if prev is None or r.scale_start > prev:
            best_date[key] = r.scale_start
            out[combo][r.attained_age] = r.rate
    return out


def _map_key(base_combo: ComboKey, benefit_bands: set) -> Optional[ComboKey]:
    """Map a base (banded) combo to the benefit's rate key.

    If the benefit varies by the combo's band, match it directly; otherwise
    (unbanded benefit) collapse to band '0'.
    """
    s, c, b = base_combo
    if b in benefit_bands:
        return (s, c, b)
    if "0" in benefit_bands:
        return (s, c, "0")
    return None


def _expand_bencoi_rows(
    index: int,
    current: Dict[int, float],
    guaranteed: Dict[int, float],
    renewable: bool,
    max_att_age: int,
) -> List[list]:
    """Fully-select expansion of one benefit COI rate set (Scale 0 + 1)."""
    ages = sorted(current)
    if not ages:
        return []
    ia_min, ia_max = ages[0], ages[-1]
    rows: List[list] = []
    for scale, rates in ((0, guaranteed), (1, current)):
        if not rates:
            continue
        for ia in range(ia_min, ia_max + 1):
            if ia not in rates:
                continue
            max_dur = max_att_age - ia + 1
            for dur in range(1, max_dur + 1):
                att = ia + dur - 1
                rate = rates.get(att) if renewable else rates.get(ia)
                if rate is not None:
                    rows.append([index, scale, ia, dur, round(rate, 6)])
    return rows


def _bentrg_rows(index: int, mtp: Dict[int, float], ctp: Dict[int, float]) -> List[list]:
    rows: List[list] = []
    for ia in sorted(set(mtp) | set(ctp)):
        rows.append([
            index, ia,
            round(mtp[ia], 6) if ia in mtp else "",
            round(ctp[ia], 6) if ia in ctp else "",
        ])
    return rows


def _write_sheet(ws, headers: List[str], rows: List[list]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(len(str(header)) + 2, 12)


def build_benefit_rows(
    result: ParseResult,
    specs: List[BenefitDBSpec],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Tuple[List[list], List[list], List[list], Dict[str, Dict[str, int]]]:
    """Build the benefit DB rows without writing a workbook.

    Returns ``(pointer_rows, bencoi_rows, bentrg_rows, counts)`` — the row
    lists match POINTER_HEADERS / BENCOI_HEADERS / BENTRG_HEADERS. Used by
    both the Benefits DB workbook export and the Rate Workup builder.
    """
    plancode = _plancode(result)
    issue_version = _issue_version(result)
    max_att_age = _pay_age(result) - 1
    base_combos = _base_combos(result)

    pointer_rows: List[list] = []
    bencoi_rows: List[list] = []
    bentrg_rows: List[list] = []
    counts: Dict[str, Dict[str, int]] = {}

    total = max(len(specs), 1)
    for si, spec in enumerate(specs):
        code = spec.code
        cur = _benefit_rates_by_combo(result, code, "C")
        guar = _benefit_rates_by_combo(result, code, "G")
        ctp = _benefit_rates_by_combo(result, code, "T")
        mtp = _benefit_rates_by_combo(result, code, "M")

        coi_bands = {b for (_s, _c, b) in cur}
        trg_bands = {b for (_s, _c, b) in set(ctp) | set(mtp)}

        # ── BENCOI indices: group base combos by identical COI content ──
        bencoi_index: Dict[ComboKey, int] = {}
        coi_groups: "OrderedDict[tuple, int]" = OrderedDict()
        next_coi = spec.start_index
        coi_pointer_rows = 0
        for bc in base_combos:
            key = _map_key(bc, coi_bands)
            cur_rates = cur.get(key) if key else None
            if not cur_rates:
                continue
            guar_rates = guar.get(key) or cur_rates
            sig = (
                tuple(sorted(cur_rates.items())),
                tuple(sorted(guar_rates.items())),
                spec.renewable,
            )
            idx = coi_groups.get(sig)
            if idx is None:
                idx = next_coi
                coi_groups[sig] = idx
                bencoi_rows.extend(_expand_bencoi_rows(
                    idx, cur_rates, guar_rates, spec.renewable, max_att_age))
                next_coi += 1
            bencoi_index[bc] = idx
            coi_pointer_rows += 1

        # ── BENTRG indices: group base combos by identical target content ──
        bentrg_index: Dict[ComboKey, int] = {}
        trg_groups: "OrderedDict[tuple, int]" = OrderedDict()
        next_trg = spec.start_index
        for bc in base_combos:
            key = _map_key(bc, trg_bands)
            c_rates = ctp.get(key, {}) if key else {}
            m_rates = mtp.get(key, {}) if key else {}
            if not c_rates and not m_rates:
                continue
            sig = (tuple(sorted(m_rates.items())), tuple(sorted(c_rates.items())))
            idx = trg_groups.get(sig)
            if idx is None:
                idx = next_trg
                trg_groups[sig] = idx
                bentrg_rows.extend(_bentrg_rows(idx, m_rates, c_rates))
                next_trg += 1
            bentrg_index[bc] = idx

        # ── Pointer rows: one per base combo the benefit applies to ──
        for bc in base_combos:
            ci = bencoi_index.get(bc)
            ti = bentrg_index.get(bc)
            if ci is None and ti is None:
                continue
            pointer_rows.append([
                plancode, code, "", issue_version,
                bc[0], bc[1], bc[2],
                ci if ci is not None else "",
                ti if ti is not None else "",
            ])

        counts[code] = {
            "pointer": coi_pointer_rows or len(bentrg_index),
            "bencoi_groups": len(coi_groups),
            "bentrg_groups": len(trg_groups),
        }
        if progress_cb:
            progress_cb((si + 1) / total)

    return pointer_rows, bencoi_rows, bentrg_rows, counts


def build_benefit_db(
    result: ParseResult,
    specs: List[BenefitDBSpec],
    output_path: str,
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Dict[str, int]]:
    """Build the combined DB Format workbook. Returns per-benefit counts."""
    pointer_rows, bencoi_rows, bentrg_rows, counts = build_benefit_rows(
        result, specs, progress_cb=progress_cb)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet("BENEFIT POINTER"), POINTER_HEADERS, pointer_rows)
    _write_sheet(wb.create_sheet("RATE_BENCOI"), BENCOI_HEADERS, bencoi_rows)
    _write_sheet(wb.create_sheet("RATE_BENTRG"), BENTRG_HEADERS, bentrg_rows)
    wb.save(output_path)

    counts["_totals"] = {
        "pointer": len(pointer_rows),
        "bencoi": len(bencoi_rows),
        "bentrg": len(bentrg_rows),
    }
    return counts


def export_benefit_db(
    input_path: str,
    output_path: str,
    specs: List[BenefitDBSpec],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Dict[str, int]]:
    """Parse ``input_path`` and build the benefit DB workbook for ``specs``."""
    def _parse_cb(frac: float) -> None:
        if progress_cb:
            progress_cb(frac * 0.6)

    result = IAFParser().parse(input_path, progress_cb=_parse_cb)
    if result.error:
        raise ValueError(result.error)

    def _build_cb(frac: float) -> None:
        if progress_cb:
            progress_cb(0.6 + frac * 0.4)

    return build_benefit_db(result, specs, output_path, progress_cb=_build_cb)
