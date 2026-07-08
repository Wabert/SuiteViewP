"""
IAF benefit-rate helpers.

Benefit rates live in the same IAF file as the base plancode and use the same
rate types (``C`` cost-of-insurance, ``T`` commission target, ``M`` minimum
target, optionally ``G`` guaranteed). They are distinguished by the last two
characters of the IDENT key (the parser's ``plan_option`` field):

  * ``**`` — base plancode
  * ``E*`` — base plancode Table-4 targets
  * anything else (e.g. ``21``, ``4M``, ``7G``) — a benefit type+subtype code

Benefit COI rates are ultimate (duration 99): one rate per attained age.
Benefit targets are keyed by issue age (duration 0), split into MTP (``M``)
and CTP (``T``).

This module lists the benefit codes present in a parsed IAF and exports the
"Excel Table" workbook — one COI sheet and one Targets sheet per selected
benefit.
"""

from __future__ import annotations

import re
from collections import OrderedDict
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from suiteview.ratemanager.parser import IAFParser, ParseResult

# plan_option values that are NOT benefits.
BASE_OPTIONS = {"**", "E*"}

_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")

COI_HEADERS = ["Plancode", "BenefitType", "Age", "Sex", "Class", "Band", "Rate"]
TARGET_HEADERS = [
    "Plancode", "BenefitType", "Age", "Sex", "Class", "Band", "RateMTP", "RateCTP",
]


def _norm_option(opt: str) -> str:
    return (opt or "").strip()


def list_benefit_codes(result: ParseResult) -> List[str]:
    """Return the distinct benefit codes (type+subtype) found in the rates."""
    codes: "OrderedDict[str, int]" = OrderedDict()
    for r in result.rates:
        opt = _norm_option(r.plan_option)
        if opt and opt not in BASE_OPTIONS:
            codes[opt] = codes.get(opt, 0) + 1
    return sorted(codes.keys())


def benefit_summary(result: ParseResult) -> List[Tuple[str, int, int]]:
    """Return ``(code, coi_count, target_count)`` for each benefit code."""
    coi: Dict[str, int] = {}
    trg: Dict[str, int] = {}
    for r in result.rates:
        opt = _norm_option(r.plan_option)
        if not opt or opt in BASE_OPTIONS:
            continue
        if r.rate_type == "C":
            coi[opt] = coi.get(opt, 0) + 1
        elif r.rate_type in ("T", "M"):
            trg[opt] = trg.get(opt, 0) + 1
    codes = sorted(set(coi) | set(trg))
    return [(c, coi.get(c, 0), trg.get(c, 0)) for c in codes]


def _plancode_map(result: ParseResult) -> Dict[int, str]:
    return {p.ref: (p.plancode or "").strip() for p in result.products}


def _base_plancode(result: ParseResult) -> str:
    return result.products[0].plancode.strip() if result.products else ""


def _coi_rows(result: ParseResult, code: str, ref2plan: Dict[int, str]) -> List[tuple]:
    """One ultimate COI rate per (sex, class, band, attained age).

    If multiple scale vintages exist, the most recent scale_start wins.
    """
    best: Dict[tuple, Tuple[str, tuple]] = {}
    for r in result.rates:
        if r.rate_type != "C" or _norm_option(r.plan_option) != code:
            continue
        key = (r.gender, r.rate_class, r.band, r.attained_age)
        prev = best.get(key)
        if prev is None or r.scale_start > prev[0]:
            plancode = ref2plan.get(r.product_ref, "")
            best[key] = (r.scale_start, (
                plancode, code, r.attained_age, r.gender, r.rate_class, r.band, r.rate,
            ))
    rows = [v[1] for v in best.values()]
    rows.sort(key=lambda x: (x[3], x[4], x[5], x[2]))
    return rows


def _target_rows(result: ParseResult, code: str, plancode: str) -> List[tuple]:
    """Merge T (CTP) and M (MTP) targets by (sex, class, band, issue age)."""
    ctp: Dict[tuple, float] = {}
    mtp: Dict[tuple, float] = {}
    for r in result.rates:
        if _norm_option(r.plan_option) != code:
            continue
        key = (r.gender, r.rate_class, r.band, r.attained_age)
        if r.rate_type == "T":
            ctp[key] = r.rate
        elif r.rate_type == "M":
            mtp[key] = r.rate
    keys = sorted(set(ctp) | set(mtp), key=lambda k: (k[0], k[1], k[2], k[3]))
    rows = []
    for k in keys:
        rows.append((
            plancode, code, k[3], k[0], k[1], k[2],
            mtp.get(k, 0.0), ctp.get(k, 0.0),
        ))
    return rows


def _sheet_title(code: str, suffix: str) -> str:
    safe = re.sub(r"[:\\/?*\[\]]", "", code)
    return f"{safe} {suffix}"[:31]


def _write_sheet(ws, headers: List[str], rows: List[tuple]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(len(header) + 2, 10)


def build_benefit_table(
    result: ParseResult,
    codes: List[str],
    output_path: str,
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Tuple[int, int]]:
    """Write one workbook with a COI + Targets sheet per selected benefit.

    Returns ``{code: (coi_row_count, target_row_count)}``.
    """
    ref2plan = _plancode_map(result)
    plancode = _base_plancode(result)

    wb = Workbook()
    wb.remove(wb.active)   # start with no sheets; add per benefit

    counts: Dict[str, Tuple[int, int]] = {}
    total = max(len(codes), 1)
    for i, code in enumerate(codes):
        coi_rows = _coi_rows(result, code, ref2plan)
        trg_rows = _target_rows(result, code, plancode)
        counts[code] = (len(coi_rows), len(trg_rows))

        if coi_rows:
            ws = wb.create_sheet(_sheet_title(code, "COI"))
            _write_sheet(ws, COI_HEADERS, coi_rows)
        if trg_rows:
            ws = wb.create_sheet(_sheet_title(code, "Targets"))
            _write_sheet(ws, TARGET_HEADERS, trg_rows)

        if progress_cb:
            progress_cb(0.6 + 0.4 * (i + 1) / total)

    if not wb.sheetnames:
        ws = wb.create_sheet("No Benefits")
        ws.cell(row=1, column=1, value="No benefit COI or target rates found for the selection.")

    wb.save(output_path)
    return counts


def export_benefit_table(
    input_path: str,
    output_path: str,
    codes: List[str],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, Tuple[int, int]]:
    """Parse ``input_path`` and write the benefit Excel Table for ``codes``."""
    def _parse_cb(frac: float) -> None:
        if progress_cb:
            progress_cb(frac * 0.6)

    result = IAFParser().parse(input_path, progress_cb=_parse_cb)
    if result.error:
        raise ValueError(result.error)
    return build_benefit_table(result, codes, output_path, progress_cb=progress_cb)
