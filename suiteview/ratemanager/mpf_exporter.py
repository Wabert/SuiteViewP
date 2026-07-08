"""
MPF Supplemental exporters.

Three outputs from the type-2 (Supplemental) records of a Misc Premium File:

  * ``export_raw``   — flat dump of every age/premium row for the selected
    premium codes (type 2 only).
  * ``export_table`` — one sheet per premium code, rates expanded to
    IssueAge x Duration per the per-code renewal flag.
  * ``build_db``     — one combined workbook: a POINTER sheet + a COI sheet
    (Index, Scale, IssueAge, Duration, Rate). COI only — no targets.

Renewal logic mirrors the IAF benefits tab:
  * Renewable   → premium varies by attained age (duration d uses the value at
    attained age issue+d-1).
  * Non-renewable → the issue-age premium is held level across durations.

Percentage premiums load as decimals (``5.64%`` -> 0.0564); factor premiums
(``1.22``) load unchanged. The Excel Raw dump keeps the original printed token.
"""

from __future__ import annotations

import re
from collections import OrderedDict
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from suiteview.ratemanager.mpf_parser import (
    group_by_combo,
    iter_records,
)

_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")

RAW_HEADERS = [
    "Company", "BenefitType", "Sex", "Class", "Band",
    "PremiumCode", "Sequence", "Age", "Premium",
]
TABLE_HEADERS = [
    "BenefitType", "PremiumCode", "Sex", "Class", "Band",
    "IssueAge", "Duration", "Rate",
]
POINTER_HEADERS = [
    "Company", "BenefitType", "PremiumCode", "Sex", "Class", "Band",
    "Index(BENCOI)",
]
COI_HEADERS = ["Index(BENCOI)", "Scale", "IssueAge", "Duration", "Rate"]

# combo key = (company, benefit, sex, class, band, premcode)
ComboKey = Tuple[str, str, str, str, str, str]


def summarize(
    path: str, progress_cb: Optional[Callable[[float], None]] = None,
) -> List[Tuple[str, str, int, int]]:
    """Return ``(premcode, benefit, combo_count, row_count)`` per premium code."""
    grouped = group_by_combo(iter_records(path, progress_cb=progress_cb))
    pc: "OrderedDict[str, dict]" = OrderedDict()
    for (company, benefit, sex, cls, band, premcode), table in grouped.items():
        e = pc.setdefault(premcode, {"benefits": set(), "combos": 0, "rows": 0})
        e["benefits"].add(benefit)
        e["combos"] += 1
        e["rows"] += len(table)
    out = [
        (premcode, "/".join(sorted(e["benefits"])), e["combos"], e["rows"])
        for premcode, e in pc.items()
    ]
    out.sort(key=lambda r: (r[1], r[0]))
    return out


def _expand(table: Dict[int, tuple], renewable: bool) -> List[Tuple[int, int, float]]:
    """Expand an attained-age table into ``(issue_age, duration, rate)`` rows."""
    ages = sorted(table)
    if not ages:
        return []
    max_age = ages[-1]
    rows: List[Tuple[int, int, float]] = []
    for ia in ages:
        for dur in range(1, max_age - ia + 2):
            att = ia + dur - 1
            if renewable:
                if att in table:
                    rows.append((ia, dur, table[att][0]))
            else:
                rows.append((ia, dur, table[ia][0]))
    return rows


def _safe_sheet(name: str) -> str:
    return re.sub(r"[:\\/?*\[\]]", "", name)[:31] or "Sheet"


def _split_progress(progress_cb, parse_fraction=0.6):
    if progress_cb is None:
        return None

    def _cb(frac: float) -> None:
        progress_cb(frac * parse_fraction)

    return _cb


def export_raw(
    path: str,
    output_path: str,
    premcodes: List[str],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> int:
    """Dump every age/premium row for the selected premium codes."""
    sel = set(premcodes)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("MPF Supplemental")
    ws.append(RAW_HEADERS)

    count = 0
    for rec in iter_records(path, progress_cb=_split_progress(progress_cb)):
        k = rec.key
        if k.premcode not in sel:
            continue
        for age, prem_str, _val, _pct in rec.pairs:
            ws.append([k.company, k.benefit, k.sex, k.rate_class, k.band,
                       k.premcode, k.seq, age, prem_str])
            count += 1
    if progress_cb:
        progress_cb(0.95)
    wb.save(output_path)
    if progress_cb:
        progress_cb(1.0)
    return count


def _combos_by_premcode(grouped) -> "OrderedDict":
    by_pc: "OrderedDict[str, list]" = OrderedDict()
    for ck, table in grouped.items():
        by_pc.setdefault(ck[5], []).append((ck, table))
    return by_pc


def export_table(
    path: str,
    output_path: str,
    specs: List[Tuple[str, bool]],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, int]:
    """One sheet per premium code, rates expanded per the renewal flag.

    ``specs`` = list of ``(premcode, renewable)``.
    """
    grouped = group_by_combo(iter_records(path, progress_cb=_split_progress(progress_cb)))
    ren = {pc: renew for pc, renew in specs}
    by_pc = _combos_by_premcode(grouped)

    wb = Workbook()
    wb.remove(wb.active)
    counts: Dict[str, int] = {}

    codes = [pc for pc, _ in specs]
    total = max(len(codes), 1)
    for i, pc in enumerate(codes):
        items = by_pc.get(pc, [])
        ws = wb.create_sheet(_safe_sheet(pc))
        _write_header(ws, TABLE_HEADERS)
        renewable = ren.get(pc, False)
        n = 0
        for ck, table in items:
            _company, benefit, sex, cls, band, premcode = ck
            for ia, dur, rate in _expand(table, renewable):
                ws.append([benefit, premcode, sex, cls, band, ia, dur, rate])
                n += 1
        _autofit(ws, TABLE_HEADERS)
        counts[pc] = n
        if progress_cb:
            progress_cb(0.6 + 0.4 * (i + 1) / total)

    if not wb.sheetnames:
        wb.create_sheet("No Data")
    wb.save(output_path)
    return counts


def build_db(
    path: str,
    output_path: str,
    specs: List[Tuple[str, bool, int]],
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Dict[str, dict]:
    """Combined POINTER + COI workbook. ``specs`` = ``(premcode, renewable, start_index)``."""
    grouped = group_by_combo(iter_records(path, progress_cb=_split_progress(progress_cb)))
    by_pc = _combos_by_premcode(grouped)

    pointer_rows: List[list] = []
    coi_rows: List[list] = []
    counts: Dict[str, dict] = {}

    total = max(len(specs), 1)
    for si, (pc, renewable, start_index) in enumerate(specs):
        items = sorted(by_pc.get(pc, []), key=lambda it: it[0])
        groups: "OrderedDict[tuple, int]" = OrderedDict()
        next_idx = start_index
        for ck, table in items:
            company, benefit, sex, cls, band, premcode = ck
            sig = (tuple(sorted((a, table[a][0]) for a in table)), renewable)
            idx = groups.get(sig)
            if idx is None:
                idx = next_idx
                groups[sig] = idx
                for scale in (0, 1):
                    for ia, dur, rate in _expand(table, renewable):
                        coi_rows.append([idx, scale, ia, dur, rate])
                next_idx += 1
            pointer_rows.append([company, benefit, premcode, sex, cls, band, idx])
        counts[pc] = {"combos": len(items), "indexes": len(groups)}
        if progress_cb:
            progress_cb(0.6 + 0.4 * (si + 1) / total)

    wb = Workbook()
    wb.remove(wb.active)
    ws_p = wb.create_sheet("POINTER")
    _write_header(ws_p, POINTER_HEADERS)
    for row in pointer_rows:
        ws_p.append(row)
    _autofit(ws_p, POINTER_HEADERS)
    ws_c = wb.create_sheet("RATE_BENCOI")
    _write_header(ws_c, COI_HEADERS)
    for row in coi_rows:
        ws_c.append(row)
    _autofit(ws_c, COI_HEADERS)
    wb.save(output_path)

    counts["_totals"] = {"pointer": len(pointer_rows), "coi": len(coi_rows)}
    return counts


def _write_header(ws, headers: List[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"


def _autofit(ws, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(len(str(header)) + 2, 11)
