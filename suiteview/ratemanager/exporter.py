"""
IAF Rate File Excel Exporter

Builds a structured .xlsx workbook from parsed IAF data.
Ports the VBA CPDReady() logic from ProgressBar.frm.
"""

import os
from collections import OrderedDict
from typing import List, Optional, Callable
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from suiteview.ratemanager.parser import ParseResult, ProductInfo, AdvProductInfo, RateRecord


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean_plan_option(raw: str) -> str:
    """Strip leading/trailing '*' from plan option (matches VBA logic)."""
    result = raw
    if result and result[0] == '*':
        result = result[1:]
    if result and result[-1] == '*':
        result = result[:-1]
    if not result:
        result = ''
    return result


def _format_date_for_tab(date_str: str) -> str:
    """Convert MM/DD/YYYY to MM-DD-YYYY for sheet tab names."""
    return date_str.replace('/', '-')


def _make_sheet_name(rate_type, scale_start, plan_option, is_select):
    """Build the worksheet tab name matching VBA naming convention."""
    opt = _clean_plan_option(plan_option)
    date_part = _format_date_for_tab(scale_start)
    name = f'{rate_type} - {date_part} {opt}'
    if is_select:
        name += ' Select'
    return name[:31]


def _safe_num(val):
    """Try to convert to int or float for Excel cells."""
    if isinstance(val, (int, float)):
        return val
    try:
        v = val.strip().replace(',', '') if isinstance(val, str) else str(val)
        if '.' in v:
            return float(v)
        return int(v)
    except (ValueError, TypeError):
        return val


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name='Segoe UI', bold=True, size=10)
_SECTION_FONT = Font(name='Segoe UI', bold=True, size=11, color='1F4E79')


# ---------------------------------------------------------------------------
# IAFExporter
# ---------------------------------------------------------------------------

class IAFExporter:
    """Builds an Excel workbook from parsed IAF data, matching VBA CPDReady()."""

    @staticmethod
    def export(result: ParseResult, output_path: str,
               progress_cb: Optional[Callable[[float], None]] = None) -> str:
        """
        Build and save the xlsx workbook.

        Args:
            result: A ParseResult from IAFParser.parse().
            output_path: Full path for the output .xlsx file.
            progress_cb: Optional callback receiving progress 0.0 - 1.0.

        Returns:
            The output file path.
        """
        wb = Workbook()
        prod = result.products[0] if result.products else ProductInfo()

        # 1. Product Information tab
        ws_prod = wb.active
        ws_prod.title = "Product Information"
        IAFExporter._build_product_info(ws_prod, prod, result)

        if progress_cb:
            progress_cb(0.1)

        # 2. GROUPS Prem Lmts And Restrs tab
        if result.adv_products:
            ws_adv = wb.create_sheet("GROUPS Prem Lmts And Restrs")
            IAFExporter._build_adv_product(ws_adv, prod, result)

        if progress_cb:
            progress_cb(0.2)

        # 3. Rate tabs
        IAFExporter._build_rate_tabs(wb, result, progress_cb)

        # Save
        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    # Product Information tab
    # ------------------------------------------------------------------

    @staticmethod
    def _build_product_info(ws, prod: ProductInfo, result: ParseResult):
        """Build the Product Information tab matching VBA layout."""
        ws.cell(1, 1, "RATE INTRODUCTION").font = _SECTION_FONT
        ws.cell(2, 1, "Plancode")
        ws.cell(2, 2, prod.plancode)
        ws.cell(3, 1, "Rate Set Version Code")
        ws.cell(3, 2, prod.version)
        ws.cell(4, 1, "Rate Set Effective Date")
        ws.cell(4, 2, prod.eff_date)

        if result.adv_products:
            ws.cell(5, 1, "Product Line Code")
            ws.cell(5, 2, "U")

        ws.cell(7, 1, "ISSUE LIMITS AND RESTRICTIONS").font = _SECTION_FONT

        # Low / High age from all rates
        if result.rates:
            low_age = min(r.attained_age for r in result.rates)
            high_age = max(r.attained_age for r in result.rates)
        else:
            low_age = high_age = 0

        ws.cell(8, 1, "Low Age")
        ws.cell(8, 2, low_age)
        ws.cell(9, 1, "High Age")
        ws.cell(9, 2, high_age)
        ws.cell(10, 1, "Issue Age Use Code")
        ws.cell(10, 2, _safe_num(prod.iar_use))

        # Premium label depends on pay_age_use
        if str(prod.pay_age_use) == "1":
            ws.cell(11, 1, "Premium Cease Age")
        else:
            ws.cell(11, 1, "Premium Years Number")
        ws.cell(11, 2, _safe_num(prod.pay_age))

        ws.cell(12, 1, "Premium Cease Duration or Age Indicator")
        ws.cell(12, 2, _safe_num(prod.pay_age_use))

        # Benefit label depends on me_age_use
        if str(prod.me_age_use) == "1":
            ws.cell(13, 1, "Benefit Period Attained Age")
        else:
            ws.cell(13, 1, "Benefit Period Policy Duration")
        ws.cell(13, 2, _safe_num(prod.me_age))

        ws.cell(14, 1, "Benefit Period Duration or Age Indicator")
        ws.cell(14, 2, _safe_num(prod.me_age_use))
        ws.cell(15, 1, "Value Per Unit Amount")
        ws.cell(15, 2, _safe_num(prod.val_per_unit))
        ws.cell(16, 1, "Deficient Reserves Code")
        ws.cell(16, 2, _safe_num(prod.deficient))
        ws.cell(17, 1, "Special Rate Designation Code")
        ws.cell(17, 2, prod.spec_benefits)

        ws.cell(19, 1, "AGENT PRODUCTION").font = _SECTION_FONT

        pcu = str(prod.prod_cred_use)
        if pcu == "1":
            ws.cell(20, 1, "Production Credit per Unit Amount")
        elif pcu == "2":
            ws.cell(20, 1, "Production Credit Percent")
        elif pcu == "3":
            ws.cell(20, 1, "Production Credit Flat Amount")
        else:
            ws.cell(20, 1, "Production Credit per Unit Amount")
        ws.cell(20, 2, _safe_num(prod.prod_cred_amt))

        ws.cell(21, 1, "Production Credit Use Code")
        ws.cell(21, 2, _safe_num(prod.prod_cred_use))
        ws.cell(22, 1, "Million Dollar Round Table Code")
        ws.cell(22, 2, prod.mdrt)

    # ------------------------------------------------------------------
    # GROUPS Premium Limits and Restraints tab
    # ------------------------------------------------------------------

    @staticmethod
    def _build_adv_product(ws, prod: ProductInfo, result: ParseResult):
        """Build the GROUPS Premium Limits and Restraints tab."""
        ws.cell(1, 1, "PREMUM RESTRICTIONS").font = _SECTION_FONT  # VBA typo

        first_ref = prod.ref if prod.ref else 1
        rule_code = ""
        corr_rule = ""
        for adv in result.adv_products:
            if adv.product_ref == first_ref or adv.product_ref == 0:
                if not rule_code:
                    rule_code = adv.rule_code
                if not corr_rule:
                    corr_rule = adv.corr_rule

        ws.cell(2, 1, "Minimum Initial Premium Rule Code")
        ws.cell(2, 2, rule_code)
        ws.cell(3, 1, "Corridor Rule Code")
        ws.cell(3, 2, corr_rule)
        ws.cell(5, 1, "PREMUM LIMITS").font = _SECTION_FONT  # VBA typo

        # Header row (row 6) — matches VBA fields 0-11 excluding 4 and 8
        headers = [
            "0", "Issue Age", "Initial Minimum", "Initial Maximum",
            "Periodic Premium Minimum", "Periodic Premium Maximum",
            "Required First Year Premium", "Corridor Percentage",
            "Corridor Amount", "MAP Period",
        ]
        for ci, hdr in enumerate(headers, start=1):
            ws.cell(6, ci, hdr).font = _HEADER_FONT

        # Data rows (row 7+) — fields 1-11 excluding 4 and 8, cols 2-10
        row = 7
        for adv in result.adv_products:
            vals = [
                adv.issue_age, adv.init_prem_min, adv.init_prem_max,
                adv.per_prem_min, adv.per_prem_max, adv.fy_prem,
                adv.corr_pct, adv.corr_amt, adv.map_period,
            ]
            for ci, val in enumerate(vals, start=2):
                ws.cell(row, ci, _safe_num(val))
            row += 1

    # ------------------------------------------------------------------
    # Rate tabs (attained-age and select)
    # ------------------------------------------------------------------

    @staticmethod
    def _build_rate_tabs(wb, result: ParseResult,
                         progress_cb: Optional[Callable[[float], None]] = None):
        """Build all rate tabs matching VBA CPDReady() layout."""
        if not result.rates:
            return

        # Organise rates into two buckets
        #   attained: {sheet_name: {group_name: [(age, rate), ...]}}
        #   select:   {sheet_name: {group_name: {(duration, attained_age): rate}}}
        attained_sheets: OrderedDict = OrderedDict()
        select_sheets: OrderedDict = OrderedDict()

        for r in result.rates:
            plan_opt = _clean_plan_option(r.plan_option)
            is_select = r.duration != 0
            date_part = _format_date_for_tab(r.scale_start)

            if is_select:
                sheet_name = f"{r.rate_type} - {date_part} {plan_opt} Select"
            else:
                sheet_name = f"{r.rate_type} - {date_part} {plan_opt}"
            sheet_name = sheet_name[:31]

            group_name = f"{r.gender} {r.rate_class} {r.band}"

            if not is_select:
                if sheet_name not in attained_sheets:
                    attained_sheets[sheet_name] = OrderedDict()
                grp = attained_sheets[sheet_name]
                if group_name not in grp:
                    grp[group_name] = []
                grp[group_name].append((r.attained_age, r.rate))
            else:
                if sheet_name not in select_sheets:
                    select_sheets[sheet_name] = OrderedDict()
                grp = select_sheets[sheet_name]
                if group_name not in grp:
                    grp[group_name] = {}
                grp[group_name][(r.duration, r.attained_age)] = r.rate

        total_sheets = len(attained_sheets) + len(select_sheets)
        done = 0

        # --- Attained-age sheets ---
        for sname, groups in attained_sheets.items():
            ws = wb.create_sheet(sname)
            col = 1
            for gname, age_rates in groups.items():
                ws.cell(1, col, gname).font = _HEADER_FONT
                age_rates.sort(key=lambda x: x[0])
                row = 2
                for age, rate in age_rates:
                    ws.cell(row, col, age)
                    ws.cell(row, col + 1, rate)
                    row += 1
                col += 3   # 3-column spacing

            done += 1
            if progress_cb and total_sheets:
                progress_cb(0.2 + 0.8 * done / total_sheets)

        # --- Select sheets ---
        for sname, groups in select_sheets.items():
            ws = wb.create_sheet(sname)
            group_start_row = 1

            for gname, dur_age_rates in groups.items():
                durations = sorted(set(d for d, _ in dur_age_rates))
                ages = sorted(set(a for _, a in dur_age_rates))

                # Group header + duration column headers
                ws.cell(group_start_row, 1, gname).font = _HEADER_FONT
                dur_col = {}
                for ci, dur in enumerate(durations, start=2):
                    ws.cell(group_start_row, ci, dur)
                    dur_col[dur] = ci

                # Age rows
                age_row = {}
                for ri, age in enumerate(ages, start=group_start_row + 1):
                    ws.cell(ri, 1, age)
                    age_row[age] = ri

                # Fill rates
                for (dur, age), rate in dur_age_rates.items():
                    r_pos = age_row.get(age)
                    c_pos = dur_col.get(dur)
                    if r_pos and c_pos:
                        ws.cell(r_pos, c_pos, rate)

                # VBA uses 122-row block spacing between groups
                group_start_row += 122

            done += 1
            if progress_cb and total_sheets:
                progress_cb(0.2 + 0.8 * done / total_sheets)


# ---------------------------------------------------------------------------
# Filename utilities
# ---------------------------------------------------------------------------

def generate_output_filename(result: ParseResult, input_filepath: str = "") -> str:
    """Generate the output filename from parsed data.

    Format: ``{REGION} - {plancode} - {YYYY-MM-DD}.xlsx``
    """
    if not result.products:
        return "rates_output.xlsx"

    prod = result.products[0]
    plancode = prod.plancode.strip()

    # Parse eff_date MM/DD/YYYY → YYYY-MM-DD
    parts = prod.eff_date.split('/')
    if len(parts) == 3:
        date_str = f"{parts[2]}-{parts[0]}-{parts[1]}"
    else:
        date_str = prod.eff_date

    region = extract_region_from_filename(input_filepath) if input_filepath else ""
    if region:
        return f"{region} - {plancode} - {date_str}.xlsx"
    return f"{plancode} - {date_str}.xlsx"


def extract_region_from_filename(filepath: str) -> str:
    """Extract the region code (e.g. 'CKAS') from an IAF filename.

    Expected pattern: ``...PRINT.REGION MMDDYYYY.txt``
    """
    basename = os.path.basename(filepath)
    upper = basename.upper()
    if '.PRINT.' in upper:
        after = basename[upper.index('.PRINT.') + 7:]
        region = after.split()[0] if after else ""
        return region
    return ""

