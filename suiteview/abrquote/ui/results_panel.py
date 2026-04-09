"""
ABR Quote — Results Panel (Step 3).

Displays full and partial acceleration results, premium impact,
and provides export to Excel functionality.
"""

from __future__ import annotations

import logging
from typing import Optional

from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QGroupBox, QFrame,
    QLineEdit, QFileDialog, QMessageBox,
)

from ..models.abr_data import ABRPolicyData, MedicalAssessment, ABRQuoteResult
from ..models.abr_constants import PLAN_CODE_INFO
from .abr_styles import (
    CRIMSON_DARK, CRIMSON_PRIMARY, CRIMSON_RICH, CRIMSON_SUBTLE,
    SLATE_PRIMARY, SLATE_DARK,
    WHITE, GRAY_DARK, GRAY_MID, GRAY_TEXT,
    GROUP_BOX_STYLE, BUTTON_PRIMARY_STYLE, BUTTON_SLATE_STYLE, BUTTON_NAV_STYLE,
    LABEL_MONEY_STYLE, LABEL_MONEY_LARGE_STYLE, DIVIDER_STYLE,
    INPUT_STYLE,
)
from .calc_viewer import CalcViewerDialog

logger = logging.getLogger(__name__)


class ResultsPanel(QWidget):
    """Step 3 panel — display ABR quote results and export.

    Signals:
        new_quote_requested: User wants to start over.
    """

    new_quote_requested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._result: Optional[ABRQuoteResult] = None
        self._policy: Optional[ABRPolicyData] = None
        self._assessment: Optional[MedicalAssessment] = None
        # Detailed calc data for viewer
        self._mort_detail: list[dict] = []
        self._apv_detail: list[dict] = []
        self._apv_summary: dict = {}
        self._policy_info: str = ""
        self._derived_values: dict = {}
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        # ── Face Amount to Accelerate ───────────────────────────────────
        face_row = QHBoxLayout()
        face_row.setSpacing(8)

        face_lbl = QLabel("Face Amount to Accelerate:")
        face_lbl.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {CRIMSON_DARK};")
        face_row.addWidget(face_lbl)

        self._face_input = QLineEdit()
        self._face_input.setStyleSheet(INPUT_STYLE)
        self._face_input.setFixedWidth(160)
        self._face_input.setEnabled(False)
        face_row.addWidget(self._face_input)

        self._face_change_btn = QPushButton("Change")
        self._face_change_btn.setFixedWidth(90)
        self._face_change_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._face_change_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {CRIMSON_RICH}, stop:1 {CRIMSON_PRIMARY});
                color: {WHITE};
                border: 1px solid {CRIMSON_DARK};
                border-radius: 5px;
                padding: 4px 12px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {CRIMSON_PRIMARY}, stop:1 {CRIMSON_DARK});
            }}
        """)
        self._face_change_btn.clicked.connect(self._on_face_change_accept)
        self._face_change_btn.setVisible(False)
        face_row.addWidget(self._face_change_btn)
        face_row.addStretch()

        layout.addLayout(face_row)

        # ── Full Acceleration ───────────────────────────────────────────
        self.full_group = QGroupBox("Full Acceleration")
        self.full_group.setStyleSheet(GROUP_BOX_STYLE)
        full_grid = QGridLayout(self.full_group)
        full_grid.setContentsMargins(12, 16, 12, 8)
        full_grid.setSpacing(4)
        full_grid.setHorizontalSpacing(8)

        self._full_labels = {}
        for i, (label_text, key) in enumerate([
            ("Eligible Death Benefit:", "eligible_db"),
            ("Actuarial Discount:", "actuarial_discount"),
            ("Administrative Fee:", "admin_fee"),
        ]):
            lbl = QLabel(label_text)
            lbl.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK};")
            full_grid.addWidget(lbl, i, 0, Qt.AlignmentFlag.AlignRight)
            val = QLabel("\u2014")
            val.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK}; font-weight: bold;")
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            full_grid.addWidget(val, i, 1, Qt.AlignmentFlag.AlignLeft)
            self._full_labels[key] = val

        # Loan Repayment row (visible only for UL/IUL/ISWL)
        self._full_loan_lbl = QLabel("Loan Repayment:")
        self._full_loan_lbl.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK};")
        full_grid.addWidget(self._full_loan_lbl, 3, 0, Qt.AlignmentFlag.AlignRight)
        self._full_labels["loan_repayment"] = QLabel("\u2014")
        self._full_labels["loan_repayment"].setStyleSheet(f"font-size: 11px; color: {GRAY_DARK}; font-weight: bold;")
        self._full_labels["loan_repayment"].setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        full_grid.addWidget(self._full_labels["loan_repayment"], 3, 1, Qt.AlignmentFlag.AlignLeft)
        self._full_loan_lbl.setVisible(False)
        self._full_labels["loan_repayment"].setVisible(False)

        divider = QFrame()
        divider.setStyleSheet(DIVIDER_STYLE)
        divider.setFixedHeight(2)
        full_grid.addWidget(divider, 4, 0, 1, 2)

        lbl_ab = QLabel("Accelerated Benefit:")
        lbl_ab.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {CRIMSON_DARK};")
        full_grid.addWidget(lbl_ab, 5, 0, Qt.AlignmentFlag.AlignRight)

        self.full_benefit_label = QLabel("\u2014")
        self.full_benefit_label.setStyleSheet(LABEL_MONEY_LARGE_STYLE)
        full_grid.addWidget(self.full_benefit_label, 5, 1, Qt.AlignmentFlag.AlignLeft)

        lbl_br = QLabel("Benefit Ratio:")
        lbl_br.setStyleSheet(f"font-size: 10px; color: {GRAY_TEXT};")
        full_grid.addWidget(lbl_br, 6, 0, Qt.AlignmentFlag.AlignRight)
        self.full_ratio_label = QLabel("\u2014")
        self.full_ratio_label.setStyleSheet(f"font-size: 10px; color: {GRAY_TEXT}; font-weight: bold;")
        full_grid.addWidget(self.full_ratio_label, 6, 1, Qt.AlignmentFlag.AlignLeft)

        # Vertical separator
        vsep_full = QFrame()
        vsep_full.setFrameShape(QFrame.Shape.VLine)
        vsep_full.setStyleSheet(f"color: {GRAY_MID}; background: {GRAY_MID};")
        full_grid.addWidget(vsep_full, 0, 2, 7, 1)

        # APV component labels — right column
        apv_lbl_style = f"font-size: 11px; color: {GRAY_DARK};"
        apv_val_style = f"font-size: 11px; color: {GRAY_DARK}; font-weight: bold;"
        self._full_apv_labels = {}
        for j, (apv_label, apv_key) in enumerate([
            ("APV_FB:", "apv_fb"),
            ("APV_FP:", "apv_fp"),
            ("APV_FD:", "apv_fd"),
        ]):
            lbl = QLabel(apv_label)
            lbl.setStyleSheet(apv_lbl_style)
            full_grid.addWidget(lbl, j, 3, Qt.AlignmentFlag.AlignRight)
            val = QLabel("\u2014")
            val.setStyleSheet(apv_val_style)
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            full_grid.addWidget(val, j, 4, Qt.AlignmentFlag.AlignLeft)
            self._full_apv_labels[apv_key] = val

        full_grid.setColumnStretch(1, 2)
        full_grid.setColumnStretch(4, 2)
        layout.addWidget(self.full_group)

        # ── Max Partial Acceleration ────────────────────────────────────
        self.partial_group = QGroupBox("Max Partial Acceleration")
        self.partial_group.setStyleSheet(GROUP_BOX_STYLE)
        partial_grid = QGridLayout(self.partial_group)
        partial_grid.setContentsMargins(12, 16, 12, 8)
        partial_grid.setSpacing(4)
        partial_grid.setHorizontalSpacing(8)

        self._partial_labels = {}
        self._partial_static_widgets = []
        for i, (label_text, key) in enumerate([
            ("Eligible Death Benefit:", "eligible_db"),
            ("Actuarial Discount:", "actuarial_discount"),
            ("Administrative Fee:", "admin_fee"),
        ]):
            lbl = QLabel(label_text)
            lbl.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK};")
            partial_grid.addWidget(lbl, i, 0, Qt.AlignmentFlag.AlignRight)
            val = QLabel("\u2014")
            val.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK}; font-weight: bold;")
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            partial_grid.addWidget(val, i, 1, Qt.AlignmentFlag.AlignLeft)
            self._partial_labels[key] = val
            self._partial_static_widgets.append(lbl)

        # Loan Repayment row (visible only for UL/IUL/ISWL)
        self._partial_loan_lbl = QLabel("Loan Repayment:")
        self._partial_loan_lbl.setStyleSheet(f"font-size: 11px; color: {GRAY_DARK};")
        partial_grid.addWidget(self._partial_loan_lbl, 3, 0, Qt.AlignmentFlag.AlignRight)
        self._partial_labels["loan_repayment"] = QLabel("\u2014")
        self._partial_labels["loan_repayment"].setStyleSheet(f"font-size: 11px; color: {GRAY_DARK}; font-weight: bold;")
        self._partial_labels["loan_repayment"].setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        partial_grid.addWidget(self._partial_labels["loan_repayment"], 3, 1, Qt.AlignmentFlag.AlignLeft)
        self._partial_loan_lbl.setVisible(False)
        self._partial_labels["loan_repayment"].setVisible(False)
        self._partial_static_widgets.append(self._partial_loan_lbl)

        divider2 = QFrame()
        divider2.setStyleSheet(DIVIDER_STYLE)
        divider2.setFixedHeight(2)
        partial_grid.addWidget(divider2, 4, 0, 1, 2)
        self._partial_static_widgets.append(divider2)

        lbl_pab = QLabel("Accelerated Benefit:")
        lbl_pab.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {CRIMSON_DARK};")
        partial_grid.addWidget(lbl_pab, 5, 0, Qt.AlignmentFlag.AlignRight)
        self._partial_static_widgets.append(lbl_pab)

        self.partial_benefit_label = QLabel("\u2014")
        self.partial_benefit_label.setStyleSheet(LABEL_MONEY_LARGE_STYLE)
        partial_grid.addWidget(self.partial_benefit_label, 5, 1, Qt.AlignmentFlag.AlignLeft)

        lbl_pbr = QLabel("Benefit Ratio:")
        lbl_pbr.setStyleSheet(f"font-size: 10px; color: {GRAY_TEXT};")
        partial_grid.addWidget(lbl_pbr, 6, 0, Qt.AlignmentFlag.AlignRight)
        self._partial_static_widgets.append(lbl_pbr)
        self.partial_ratio_label = QLabel("\u2014")
        self.partial_ratio_label.setStyleSheet(f"font-size: 10px; color: {GRAY_TEXT}; font-weight: bold;")
        partial_grid.addWidget(self.partial_ratio_label, 6, 1, Qt.AlignmentFlag.AlignLeft)

        # Vertical separator
        vsep_partial = QFrame()
        vsep_partial.setFrameShape(QFrame.Shape.VLine)
        vsep_partial.setStyleSheet(f"color: {GRAY_MID}; background: {GRAY_MID};")
        partial_grid.addWidget(vsep_partial, 0, 2, 7, 1)
        self._partial_static_widgets.append(vsep_partial)

        # APV component labels — right column
        self._partial_apv_labels = {}
        for j, (apv_label, apv_key) in enumerate([
            ("APV_FB:", "apv_fb"),
            ("APV_FP:", "apv_fp"),
            ("APV_FD:", "apv_fd"),
        ]):
            lbl = QLabel(apv_label)
            lbl.setStyleSheet(apv_lbl_style)
            partial_grid.addWidget(lbl, j, 3, Qt.AlignmentFlag.AlignRight)
            val = QLabel("\u2014")
            val.setStyleSheet(apv_val_style)
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            partial_grid.addWidget(val, j, 4, Qt.AlignmentFlag.AlignLeft)
            self._partial_apv_labels[apv_key] = val
            self._partial_static_widgets.append(lbl)

        # "Not allowed" overlay
        self._partial_not_allowed_label = QLabel(
            "MAX PARTIAL NOT ALLOWED\nPOLICY ALREADY AT MINIMUM FACE"
        )
        self._partial_not_allowed_label.setStyleSheet(
            f"font-size: 14px; font-weight: bold; color: {CRIMSON_DARK}; padding: 16px;"
        )
        self._partial_not_allowed_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._partial_not_allowed_label.setVisible(False)
        partial_grid.addWidget(self._partial_not_allowed_label, 0, 0, 7, 5)

        partial_grid.setColumnStretch(1, 2)
        partial_grid.setColumnStretch(4, 2)
        layout.addWidget(self.partial_group)

        # ── Premium Impact ──────────────────────────────────────────────
        self.premium_group = QGroupBox("Premium Impact")
        self.premium_group.setStyleSheet(GROUP_BOX_STYLE)
        premium_layout = QGridLayout(self.premium_group)
        premium_layout.setContentsMargins(12, 16, 12, 8)
        premium_layout.setSpacing(4)

        self._prem_row_labels = []
        for i, label_text in enumerate(["Premium Before:", "After (Full Accel):", "After (Partial):"]):
            lbl = QLabel(label_text)
            lbl.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {CRIMSON_DARK};")
            premium_layout.addWidget(lbl, i, 0, Qt.AlignmentFlag.AlignRight)
            self._prem_row_labels.append(lbl)

        self.premium_before_label = QLabel("\u2014")
        self.premium_before_label.setStyleSheet(LABEL_MONEY_STYLE)
        premium_layout.addWidget(self.premium_before_label, 0, 1)

        self.premium_after_full_label = QLabel("\u2014")
        self.premium_after_full_label.setStyleSheet(LABEL_MONEY_STYLE)
        premium_layout.addWidget(self.premium_after_full_label, 1, 1)

        # After (Partial): read-only label for non-UL, editable input for UL
        self.premium_after_partial_label = QLabel("\u2014")
        self.premium_after_partial_label.setStyleSheet(LABEL_MONEY_STYLE)
        premium_layout.addWidget(self.premium_after_partial_label, 2, 1)

        self.premium_after_partial_input = QLineEdit()
        self.premium_after_partial_input.setPlaceholderText("0.00")
        self.premium_after_partial_input.setStyleSheet(INPUT_STYLE)
        self.premium_after_partial_input.setFixedWidth(120)
        self.premium_after_partial_input.setVisible(False)
        premium_layout.addWidget(self.premium_after_partial_input, 2, 1)

        premium_layout.setColumnStretch(2, 1)
        layout.addWidget(self.premium_group)

        # ── Messages ────────────────────────────────────────────────────
        msg_group = QGroupBox("Messages")
        msg_group.setStyleSheet(GROUP_BOX_STYLE)
        msg_layout = QVBoxLayout(msg_group)
        msg_layout.setContentsMargins(12, 20, 12, 8)
        self.messages_label = QLabel("")
        self.messages_label.setWordWrap(True)
        self.messages_label.setStyleSheet(
            "color: #C62828; font-size: 13px; font-weight: bold; padding: 4px;"
        )
        msg_layout.addWidget(self.messages_label)
        layout.addWidget(msg_group)

        # ── Action buttons ──────────────────────────────────────────────
        btn_row = QHBoxLayout()

        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setStyleSheet(BUTTON_SLATE_STYLE)
        self.export_btn.clicked.connect(self._on_export)
        btn_row.addWidget(self.export_btn)

        self.copy_btn = QPushButton("Copy Summary")
        self.copy_btn.setStyleSheet(BUTTON_PRIMARY_STYLE)
        self.copy_btn.clicked.connect(self._on_copy)
        btn_row.addWidget(self.copy_btn)

        self.view_calc_btn = QPushButton("View Calculated Values")
        self.view_calc_btn.setStyleSheet(BUTTON_PRIMARY_STYLE)
        self.view_calc_btn.setToolTip(
            "Inspect month-by-month mortality derivation and APV computation"
        )
        self.view_calc_btn.clicked.connect(self._on_view_calc)
        self.view_calc_btn.setEnabled(False)
        btn_row.addWidget(self.view_calc_btn)

        btn_row.addStretch()

        self.new_quote_btn = QPushButton("New Quote")
        self.new_quote_btn.setStyleSheet(BUTTON_NAV_STYLE)
        self.new_quote_btn.clicked.connect(self.new_quote_requested.emit)
        btn_row.addWidget(self.new_quote_btn)

        layout.addLayout(btn_row)

    # ── Helpers ──────────────────────────────────────────────────────────

    def _make_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(f"font-size: 12px; font-weight: bold; color: {CRIMSON_DARK};")
        return lbl

    @staticmethod
    def _fmt_money(amount: float) -> str:
        if amount < 0:
            return f"(${abs(amount):,.2f})"
        return f"${amount:,.2f}"

    # ── Public API ──────────────────────────────────────────────────────

    def set_policy(self, policy: ABRPolicyData):
        self._policy = policy
        # UL/IUL/ISWL: rename Premium Impact → Monthly Deduction Impact
        # and switch After (Partial) to an editable input
        is_ul = policy.product_type in ("UL", "IUL", "ISWL")
        self.premium_group.setTitle(
            "Monthly Deduction Impact" if is_ul else "Premium Impact"
        )
        # Rename "Premium Before:" → "Last Monthly Deduction:" for UL
        self._prem_row_labels[0].setText(
            "Last Monthly Deduction:" if is_ul else "Premium Before:"
        )
        self.premium_after_partial_label.setVisible(not is_ul)
        self.premium_after_partial_input.setVisible(is_ul)
        if is_ul:
            self.premium_after_partial_input.clear()

    def set_assessment(self, assessment: MedicalAssessment):
        self._assessment = assessment

    # ── Face Amount Change / Accept ─────────────────────────────────────

    def _on_face_change_accept(self):
        """Toggle between Change and Accept modes for the face amount input."""
        if self._face_change_btn.text() == "Change":
            # Switch to edit mode
            self._face_input.setEnabled(True)
            self._face_input.setFocus()
            self._face_input.selectAll()
            self._face_change_btn.setText("Accept")
        else:
            # Accept mode — validate and recalculate
            self._accept_face_amount()

    def _revert_face_input(self):
        """Reset the face input to the current displayed face and lock it."""
        face = self._policy.face_amount if self._policy else 0.0
        self._face_input.setText(f"${face:,.2f}")
        self._face_input.setEnabled(False)
        self._face_change_btn.setText("Change")

    def _accept_face_amount(self):
        """Validate the entered face amount and recalculate acceleration."""
        if not self._policy or not self._result:
            return

        raw = self._face_input.text().replace("$", "").replace(",", "").strip()
        try:
            custom_face = float(raw)
        except ValueError:
            QMessageBox.warning(
                self, "Invalid Amount",
                "Please enter a valid numeric face amount.",
            )
            self._revert_face_input()
            return

        total_face = self._policy.face_amount
        min_allowed = 10_000.0

        if custom_face > total_face:
            QMessageBox.warning(
                self, "Invalid Amount",
                f"Face amount cannot exceed the total policy face "
                f"of ${total_face:,.2f}.",
            )
            self._revert_face_input()
            return

        if custom_face < min_allowed:
            QMessageBox.warning(
                self, "Invalid Amount",
                f"Face amount cannot be less than the minimum "
                f"of ${min_allowed:,.2f}.",
            )
            self._revert_face_input()
            return

        # Check if amount exceeds max partial eligible DB
        max_partial_eligible = self._result.partial_eligible_db
        if (max_partial_eligible > 0
                and custom_face > max_partial_eligible
                and abs(custom_face - total_face) >= 0.01):
            QMessageBox.warning(
                self, "Invalid Amount",
                f"Accelerating this amount would drop the face below the "
                f"minimum. If you want to partial accelerate you must put "
                f"an amount less than or equal to ${max_partial_eligible:,.2f}.",
            )
            self._revert_face_input()
            return

        # Lock input back down
        self._face_input.setEnabled(False)
        self._face_input.setText(f"${custom_face:,.2f}")
        self._face_change_btn.setText("Change")

        # Recalculate the full acceleration group with the custom face
        is_full = abs(custom_face - total_face) < 0.01

        if is_full:
            self.full_group.setTitle("Full Acceleration")
        else:
            self.full_group.setTitle("Partial Acceleration")

        # Proportional recalculation
        orig_eligible = self._result.full_eligible_db
        orig_discount = self._result.full_actuarial_discount
        admin_fee = self._result.full_admin_fee

        if orig_eligible > 0:
            ratio = custom_face / orig_eligible
            new_discount = round(orig_discount * ratio, 2)
        else:
            ratio = 0.0
            new_discount = 0.0

        self._full_labels["eligible_db"].setText(self._fmt_money(custom_face))
        self._full_labels["actuarial_discount"].setText(self._fmt_money(new_discount))
        self._full_labels["admin_fee"].setText(self._fmt_money(admin_fee))

        # Loan repayment — proportionally scaled
        orig_loan = self._result.full_loan_repayment
        if orig_eligible > 0 and orig_loan > 0:
            new_loan = round(orig_loan * ratio, 2)
            self._full_loan_lbl.setVisible(True)
            self._full_labels["loan_repayment"].setVisible(True)
            self._full_labels["loan_repayment"].setText(self._fmt_money(new_loan))
            new_benefit = round(custom_face - new_discount - admin_fee - new_loan, 2)
        else:
            new_benefit = round(custom_face - new_discount - admin_fee, 2)

        new_ratio = new_benefit / custom_face if custom_face > 0 else 0.0

        if new_benefit < 0:
            self.full_benefit_label.setText(
                f"$0.00  (calc result: {self._fmt_money(new_benefit)})"
            )
        else:
            self.full_benefit_label.setText(self._fmt_money(new_benefit))
        self.full_ratio_label.setText(f"{new_ratio * 100:.2f}%")

        # APV components — proportionally scaled
        if orig_eligible > 0:
            self._full_apv_labels["apv_fb"].setText(
                self._fmt_money(self._result.apv_fb * ratio)
            )
            self._full_apv_labels["apv_fp"].setText(
                self._fmt_money(self._result.apv_fp * ratio)
            )
            self._full_apv_labels["apv_fd"].setText(
                self._fmt_money(self._result.apv_fd * ratio)
            )

    def set_calc_data(
        self,
        mort_detail: list[dict],
        apv_detail: list[dict],
        apv_summary: dict,
        policy_info: str = "",
    ):
        """Store detailed calculation tables for the viewer."""
        self._mort_detail = mort_detail
        self._apv_detail = apv_detail
        self._apv_summary = apv_summary
        self._policy_info = policy_info
        self.view_calc_btn.setEnabled(bool(mort_detail))

    def set_derived_values(self, derived_values: dict):
        self._derived_values = derived_values or {}

    def display_results(self, result: ABRQuoteResult):
        """Populate all result fields."""
        self._result = result

        # Populate face amount input and reset to full state
        if self._policy:
            self._face_input.setText(f"${self._policy.face_amount:,.2f}")
            self._face_change_btn.setVisible(True)
            self._face_change_btn.setText("Change")
            self._face_input.setEnabled(False)
        self.full_group.setTitle("Full Acceleration")

        # Full acceleration
        self._full_labels["eligible_db"].setText(self._fmt_money(result.full_eligible_db))
        self._full_labels["actuarial_discount"].setText(
            self._fmt_money(result.full_actuarial_discount)
        )
        self._full_labels["admin_fee"].setText(self._fmt_money(result.full_admin_fee))

        # Loan Repayment row — visible only for UL/IUL/ISWL with non-zero loan
        has_loan = result.full_loan_repayment > 0
        self._full_loan_lbl.setVisible(has_loan)
        self._full_labels["loan_repayment"].setVisible(has_loan)
        if has_loan:
            self._full_labels["loan_repayment"].setText(
                self._fmt_money(result.full_loan_repayment)
            )

        if result.full_accel_benefit < 0:
            self.full_benefit_label.setText(
                f"$0.00  (calc result: {self._fmt_money(result.full_accel_benefit)})"
            )
        else:
            self.full_benefit_label.setText(self._fmt_money(result.full_accel_benefit))
        self.full_ratio_label.setText(f"{result.full_benefit_ratio * 100:.2f}%")

        # Full APV components
        self._full_apv_labels["apv_fb"].setText(self._fmt_money(result.apv_fb))
        self._full_apv_labels["apv_fp"].setText(self._fmt_money(result.apv_fp))
        self._full_apv_labels["apv_fd"].setText(self._fmt_money(result.apv_fd))

        # Partial acceleration — check if at minimum face
        at_min_face = result.partial_eligible_db <= 0
        has_partial_loan = result.partial_loan_repayment > 0
        self._partial_not_allowed_label.setVisible(at_min_face)
        for w in self._partial_static_widgets:
            w.setVisible(not at_min_face)
        for key, val in self._partial_labels.items():
            if key == "loan_repayment":
                val.setVisible(not at_min_face and has_partial_loan)
            else:
                val.setVisible(not at_min_face)
        self._partial_loan_lbl.setVisible(not at_min_face and has_partial_loan)
        self.partial_benefit_label.setVisible(not at_min_face)
        self.partial_ratio_label.setVisible(not at_min_face)
        for val in self._partial_apv_labels.values():
            val.setVisible(not at_min_face)

        if not at_min_face:
            self._partial_labels["eligible_db"].setText(self._fmt_money(result.partial_eligible_db))
            self._partial_labels["actuarial_discount"].setText(
                self._fmt_money(result.partial_actuarial_discount)
            )
            self._partial_labels["admin_fee"].setText(self._fmt_money(result.partial_admin_fee))
            if has_partial_loan:
                self._partial_labels["loan_repayment"].setText(
                    self._fmt_money(result.partial_loan_repayment)
                )

            if result.partial_accel_benefit < 0:
                self.partial_benefit_label.setText(
                    f"$0.00  (calc result: {self._fmt_money(result.partial_accel_benefit)})"
                )
            else:
                self.partial_benefit_label.setText(self._fmt_money(result.partial_accel_benefit))
            self.partial_ratio_label.setText(f"{result.partial_benefit_ratio * 100:.2f}%")

            # Partial APV components (proportionally scaled)
            if result.full_eligible_db > 0:
                ratio = result.partial_eligible_db / result.full_eligible_db
            else:
                ratio = 0.0
            self._partial_apv_labels["apv_fb"].setText(
                self._fmt_money(result.apv_fb * ratio)
            )
            self._partial_apv_labels["apv_fp"].setText(
                self._fmt_money(result.apv_fp * ratio)
            )
            self._partial_apv_labels["apv_fd"].setText(
                self._fmt_money(result.apv_fd * ratio)
            )

        # Premium impact
        self.premium_before_label.setText(result.premium_before)
        self.premium_after_full_label.setText(f"${result.premium_after_full:,.2f}")
        is_ul = self._policy and self._policy.product_type in ("UL", "IUL", "ISWL")
        if is_ul:
            # UL: After (Partial) is a user input — don't overwrite it
            self.premium_after_partial_label.setVisible(False)
            self.premium_after_partial_input.setVisible(not at_min_face)
        elif at_min_face:
            self.premium_after_partial_label.setText("NOT ALLOWED")
        else:
            self.premium_after_partial_label.setText(result.premium_after_partial)

        # Messages
        if result.messages:
            bullets = "\n\n".join(f"\u2022 {m}" for m in result.messages)
            self.messages_label.setText(bullets)
        else:
            self.messages_label.setText("")

    # ── View Calculated Values ────────────────────────────────────────

    def _on_view_calc(self):
        """Open the detailed calculation viewer window (modeless)."""
        if not self._mort_detail:
            return
        after_partial = self.premium_after_partial_input.text().strip()
        viewer = CalcViewerDialog(
            mortality_rows=self._mort_detail,
            apv_rows=self._apv_detail,
            apv_summary=self._apv_summary,
            policy_info=self._policy_info,
            policy=self._policy,
            assessment=self._assessment,
            result=self._result,
            derived_values=self._derived_values,
            after_partial_override=after_partial,
            parent=None,
        )
        viewer.show()
        # Keep a reference so the window isn't garbage-collected
        self._calc_viewer = viewer

    # ── Export ───────────────────────────────────────────────────────────

    def _on_export(self):
        """Export results to an Excel workbook."""
        if self._result is None:
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export ABR Quote",
            f"ABR_Quote_{self._policy.policy_number if self._policy else 'Unknown'}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not filename:
            return

        try:
            self._write_excel(filename)
            QMessageBox.information(
                self, "Export Complete",
                f"ABR Quote exported to:\n{filename}",
            )
        except Exception as e:
            logger.error(f"Export error: {e}", exc_info=True)
            QMessageBox.warning(
                self, "Export Error",
                f"Could not export:\n{e}",
            )

    def _write_excel(self, filepath: str):
        """Write results to Excel workbook matching the ABR output format."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = openpyxl.Workbook()
        r = self._result
        p = self._policy

        # ── Sheet 1: Output - Full ──────────────────────────────────────
        ws = wb.active
        ws.title = "Output - Full"

        # Header styling
        header_font = Font(bold=True, size=12)
        money_font = Font(bold=True, size=11)
        label_font = Font(size=11)

        plan_info = PLAN_CODE_INFO.get(p.plan_code.upper(), ("", "")) if p else ("", "")

        row = 1
        ws.cell(row=row, column=1, value="ABR QUOTE — FULL ACCELERATION").font = header_font
        row += 2
        ws.cell(row=row, column=1, value="Policy Number:").font = label_font
        ws.cell(row=row, column=2, value=p.policy_number if p else "").font = money_font
        row += 1
        ws.cell(row=row, column=1, value="Insured:").font = label_font
        ws.cell(row=row, column=2, value=p.insured_name if p else "").font = money_font
        row += 1
        ws.cell(row=row, column=1, value="Plan:").font = label_font
        ws.cell(row=row, column=2, value=f"{plan_info[1]} ({plan_info[0]}-Year Level)").font = money_font
        row += 2

        # Section 5: Accelerated Benefit Payment
        ws.cell(row=row, column=1, value="ACCELERATED BENEFIT PAYMENT").font = header_font
        row += 1
        ws.cell(row=row, column=1, value="Total Eligible Death Benefit:").font = label_font
        ws.cell(row=row, column=2, value=r.full_eligible_db).number_format = '$#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="An Actuarial Discount:").font = label_font
        ws.cell(row=row, column=2, value=-r.full_actuarial_discount).number_format = '($#,##0.00)'
        row += 1
        ws.cell(row=row, column=1, value="An Administrative Charge:").font = label_font
        ws.cell(row=row, column=2, value=-r.full_admin_fee).number_format = '($#,##0.00)'
        row += 1
        if r.full_loan_repayment > 0:
            ws.cell(row=row, column=1, value="Loan Repayment:").font = label_font
            ws.cell(row=row, column=2, value=-r.full_loan_repayment).number_format = '($#,##0.00)'
            row += 1
        ws.cell(row=row, column=1, value="Any Policy Debt:").font = label_font
        ws.cell(row=row, column=2, value=0).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value="Total Accelerated Benefit Payment:").font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=r.full_accel_benefit).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        row += 2

        ws.cell(row=row, column=1, value="Premium Before:").font = label_font
        ws.cell(row=row, column=2, value=r.premium_before).font = money_font
        row += 1
        ws.cell(row=row, column=1, value="Premium After:").font = label_font
        ws.cell(row=row, column=2, value=f"${r.premium_after_full:,.2f}").font = money_font
        row += 1
        ws.cell(row=row, column=1, value="Benefit Ratio:").font = label_font
        ws.cell(row=row, column=2, value=r.full_benefit_ratio).number_format = '0.00%'

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        # ── Sheet 2: Output - Max Partial ───────────────────────────────
        ws2 = wb.create_sheet("Output - Max Partial")

        row = 1
        ws2.cell(row=row, column=1, value="ABR QUOTE — MAX PARTIAL ACCELERATION").font = header_font
        row += 2
        ws2.cell(row=row, column=1, value="Policy Number:").font = label_font
        ws2.cell(row=row, column=2, value=p.policy_number if p else "").font = money_font
        row += 2

        ws2.cell(row=row, column=1, value="ACCELERATED BENEFIT PAYMENT").font = header_font
        row += 1
        ws2.cell(row=row, column=1, value="Total Eligible Death Benefit:").font = label_font
        ws2.cell(row=row, column=2, value=r.partial_eligible_db).number_format = '$#,##0.00'
        row += 1
        ws2.cell(row=row, column=1, value="An Actuarial Discount:").font = label_font
        ws2.cell(row=row, column=2, value=-r.partial_actuarial_discount).number_format = '($#,##0.00)'
        row += 1
        ws2.cell(row=row, column=1, value="An Administrative Charge:").font = label_font
        ws2.cell(row=row, column=2, value=-r.partial_admin_fee).number_format = '($#,##0.00)'
        row += 1
        if r.partial_loan_repayment > 0:
            ws2.cell(row=row, column=1, value="Loan Repayment:").font = label_font
            ws2.cell(row=row, column=2, value=-r.partial_loan_repayment).number_format = '($#,##0.00)'
            row += 1

        ws2.cell(row=row, column=1, value="Total Accelerated Benefit Payment:").font = Font(bold=True, size=12)
        ws2.cell(row=row, column=2, value=r.partial_accel_benefit).number_format = '$#,##0.00'
        ws2.cell(row=row, column=2).font = Font(bold=True, size=12)
        row += 2

        ws2.cell(row=row, column=1, value="Premium After (Partial):").font = label_font
        ws2.cell(row=row, column=2, value=r.premium_after_partial).font = money_font
        row += 1
        ws2.cell(row=row, column=1, value="Benefit Ratio:").font = label_font
        ws2.cell(row=row, column=2, value=r.partial_benefit_ratio).number_format = '0.00%'

        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 20

        wb.save(filepath)

    def _on_copy(self):
        """Copy summary text to clipboard."""
        if self._result is None:
            return

        r = self._result
        p = self._policy

        lines = [
            "ABR QUOTE SUMMARY",
            f"Policy: {p.policy_number if p else 'N/A'}",
            f"Insured: {p.insured_name if p else 'N/A'}",
            "",
            "FULL ACCELERATION:",
            f"  Eligible DB:        {self._fmt_money(r.full_eligible_db)}",
            f"  Actuarial Discount: {self._fmt_money(r.full_actuarial_discount)}",
            f"  Admin Fee:          {self._fmt_money(r.full_admin_fee)}",
        ]
        if r.full_loan_repayment > 0:
            lines.append(f"  Loan Repayment:     {self._fmt_money(r.full_loan_repayment)}")
        lines += [
            f"  Accel. Benefit:     {self._fmt_money(r.full_accel_benefit)}",
            f"  Benefit Ratio:      {r.full_benefit_ratio * 100:.2f}%",
            f"  APV_FB:             {self._fmt_money(r.apv_fb)}",
            f"  APV_FP:             {self._fmt_money(r.apv_fp)}",
            f"  APV_FD:             {self._fmt_money(r.apv_fd)}",
            "",
            "MAX PARTIAL ACCELERATION:",
            f"  Eligible DB:        {self._fmt_money(r.partial_eligible_db)}",
            f"  Actuarial Discount: {self._fmt_money(r.partial_actuarial_discount)}",
            f"  Admin Fee:          {self._fmt_money(r.partial_admin_fee)}",
        ]
        if r.partial_loan_repayment > 0:
            lines.append(f"  Loan Repayment:     {self._fmt_money(r.partial_loan_repayment)}")
        lines += [
            f"  Accel. Benefit:     {self._fmt_money(r.partial_accel_benefit)}",
            f"  Benefit Ratio:      {r.partial_benefit_ratio * 100:.2f}%",
            "",
            f"Premium Before:  {r.premium_before}",
            f"Premium After (Full):    ${r.premium_after_full:,.2f}",
            f"Premium After (Partial): {r.premium_after_partial}",
        ]

        from PyQt6.QtWidgets import QApplication
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(lines))
        self.messages_label.setText("Summary copied to clipboard.")
