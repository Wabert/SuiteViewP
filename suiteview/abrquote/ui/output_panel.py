"""
Output Panel — File management for ABR Quote policies.

Provides a file explorer interface for the specific policy folder and
relevant tools, similar to PolView's Policy Support tab but simplified.
"""

import logging
import os
import shutil
from typing import Optional, List, Tuple

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QGroupBox, QGridLayout, QMessageBox, QListWidget, QListWidgetItem,
    QAbstractItemView, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal, QPoint, QMimeData
from PyQt6.QtGui import QColor, QDrag, QPixmap, QPainter, QFont

from suiteview.ui.widgets.mini_explorer import (
    MiniExplorer, DropTargetSubfolderList, DraggableToolsList,
    DoubleClickablePathLabel, TightItemDelegate, _icon_for_ext
)
from .abr_styles import (
    CRIMSON_DARK, CRIMSON_PRIMARY, CRIMSON_RICH, CRIMSON_LIGHT, CRIMSON_SUBTLE, CRIMSON_BG,
    SLATE_PRIMARY, SLATE_TEXT, SLATE_DARK, SLATE_LIGHT,
    WHITE, GRAY_DARK, GRAY_MID,
    GROUP_BOX_STYLE, BUTTON_PRIMARY_STYLE,
)
from ..models.abr_data import ABRPolicyData, ABRQuoteResult, MedicalAssessment
from ..models.abr_constants import PLAN_CODE_INFO, MODAL_LABELS

logger = logging.getLogger(__name__)

# ── Custom styles for MiniExplorer components (matching ABR theme) ────────────────
ABR_EXPLORER_BOX_STYLE = f"""
    QGroupBox {{
        font-weight: bold;
        border: 2px solid {CRIMSON_PRIMARY};
        border-radius: 8px;
        margin-top: 24px;
        background-color: {WHITE};
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        left: 12px;
        padding: 2px 8px;
        color: {SLATE_TEXT};
        background: {CRIMSON_PRIMARY};
        border-radius: 4px;
    }}
"""

ABR_NAV_BTN_STYLE = f"""
    QPushButton {{
        background: {CRIMSON_SUBTLE};
        color: {CRIMSON_DARK};
        border: 1px solid {CRIMSON_PRIMARY};
        border-radius: 3px;
        padding: 1px 5px;
        font-size: 11px;
        font-weight: bold;
        min-height: 16px;
        max-height: 18px;
    }}
    QPushButton:hover {{
        background: {CRIMSON_PRIMARY};
        color: {WHITE};
    }}
    QPushButton:disabled {{
        background: {GRAY_MID};
        color: {GRAY_DARK};
        border-color: {GRAY_MID};
    }}
"""

# Makes path labels "light red" (using CRIMSON_BG/CRIMSON_SUBTLE blend)
ABR_PATH_BAR_STYLE = f"""
    font-size: 10px;
    color: {CRIMSON_DARK};
    background: #F9E0E3;
    border: 1px solid {CRIMSON_PRIMARY};
    border-radius: 3px;
    padding: 1px 4px;
"""

ABR_LIST_STYLE = f"""
    QListWidget {{
        font-size: 11px;
        border: none;
        background-color: {WHITE};
        outline: none;
        padding: 0px;
        margin: 0px;
    }}
    QListWidget::item {{
        padding: 0px 4px;
        margin: 0px;
        border: none;
        min-height: 0px;
        color: {GRAY_DARK};
    }}
    QListWidget::item:hover {{
        background-color: {CRIMSON_SUBTLE};
    }}
    QListWidget::item:selected {{
        background-color: {SLATE_LIGHT};
        color: {CRIMSON_DARK};
        font-weight: bold;
    }}
"""

POLICY_INFO_FRAME_STYLE_ABR = f"""
    QGroupBox {{
        font-weight: bold;
        border: 2px solid {CRIMSON_PRIMARY};
        border-radius: 8px;
        margin-top: 10px;
        background-color: {WHITE};
        font-size: 12px;
        color: {CRIMSON_DARK};
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        subcontrol-position: top left;
        left: 12px;
        padding: 2px 12px;
        background-color: {CRIMSON_PRIMARY};
        color: {SLATE_TEXT};
        border-radius: 4px;
    }}
"""


# ── Custom List Classes for Drag Styling ─────────────────────────────

# Duplicate MIME constant from mini_explorer to ensure compatibility
_MIME_TOOL_FILE = "application/x-suiteview-toolfile"

class CrimsonDraggableToolsList(DraggableToolsList):
    """Overrides drag image to use Crimson/Slate theme."""
    def startDrag(self, supportedActions):
        item = self.currentItem()
        if not item:
            return
        # Access role via Qt.ItemDataRole.UserRole (which MiniExplorer.PATH_ROLE maps to)
        path = item.data(Qt.ItemDataRole.UserRole) or ""
        if not path or not os.path.isfile(path):
            return
        
        mime = QMimeData()
        mime.setData(_MIME_TOOL_FILE, path.encode())
        drag = QDrag(self)
        drag.setMimeData(mime)
        
        # Create themed drag pixmap
        pix = QPixmap(200, 20)
        pix.fill(QColor(CRIMSON_SUBTLE)) # Pinkish background
        
        painter = QPainter(pix)
        painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        painter.setPen(QColor(CRIMSON_DARK)) # Crimson text
        
        # Draw border
        painter.setPen(QColor(CRIMSON_PRIMARY))
        painter.drawRect(0, 0, 199, 19)
        
        painter.setPen(QColor(CRIMSON_DARK))
        # Draw text centered vertically
        rect = pix.rect().adjusted(4, 0, -4, 0)
        painter.drawText(rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, 
                         os.path.basename(path))
        painter.end()
        
        drag.setPixmap(pix)
        drag.setHotSpot(QPoint(10, 10))
        drag.exec(Qt.DropAction.CopyAction)

class CrimsonDropTargetList(DropTargetSubfolderList):
    """Overrides drag image to use Crimson/Slate theme."""
    def startDrag(self, supportedActions):
        item = self.currentItem()
        if not item:
            return
        path = item.data(Qt.ItemDataRole.UserRole) or ""
        if not path or not os.path.isfile(path):
            return
            
        mime = QMimeData()
        mime.setData(_MIME_TOOL_FILE, path.encode())
        drag = QDrag(self)
        drag.setMimeData(mime)
        
        pix = QPixmap(200, 20)
        pix.fill(QColor(CRIMSON_SUBTLE))
        
        painter = QPainter(pix)
        painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        painter.setPen(QColor(CRIMSON_PRIMARY)) # Border
        painter.drawRect(0, 0, 199, 19)
        
        painter.setPen(QColor(CRIMSON_DARK)) # Text
        rect = pix.rect().adjusted(4, 0, -4, 0)
        painter.drawText(rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, 
                         os.path.basename(path))
        painter.end()
        
        drag.setPixmap(pix)
        drag.setHotSpot(QPoint(10, 10))
        drag.exec(Qt.DropAction.CopyAction)


class RecommendedFilesList(QListWidget):
    """Draggable list for recommended form files — drag items into Policy Subfolders."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if not item:
            return
        path = item.data(Qt.ItemDataRole.UserRole) or ""
        if not path or not os.path.isfile(path):
            return

        mime = QMimeData()
        mime.setData(_MIME_TOOL_FILE, path.encode())
        drag = QDrag(self)
        drag.setMimeData(mime)

        # Themed drag pixmap
        pix = QPixmap(220, 20)
        pix.fill(QColor(CRIMSON_SUBTLE))

        painter = QPainter(pix)
        painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        painter.setPen(QColor(CRIMSON_PRIMARY))
        painter.drawRect(0, 0, 219, 19)
        painter.setPen(QColor(CRIMSON_DARK))
        rect = pix.rect().adjusted(4, 0, -4, 0)
        painter.drawText(rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                         os.path.basename(path))
        painter.end()

        drag.setPixmap(pix)
        drag.setHotSpot(QPoint(10, 10))
        drag.exec(Qt.DropAction.CopyAction)


class OutputPanel(QWidget):
    """
    Output/File Management Panel.
    
    Layout:
      Left column:
        - Policy Subfolders (compact, drop target)
        - Recommended Files (state-specific forms from DB)
      Right column:
        - Resources (general ABR folder browser)
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._policy: Optional[ABRPolicyData] = None
        self._result: Optional[ABRQuoteResult] = None
        self._assessment: Optional[MedicalAssessment] = None
        self._derived_values: dict = {}
        self._mort_detail: list[dict] = []
        self._apv_detail: list[dict] = []
        self._apv_summary: dict = {}
        self._policy_folder_path = ""
        self._tools_root_path = self._get_tools_root_path()
        
        self._setup_ui()
        
        # Initialize
        self._update_ui_state()

    def _get_process_control_dir(self) -> str:
        username = os.environ.get("USERNAME", os.environ.get("USER", "unknown"))
        return os.path.join(
            "C:\\Users", username,
            "OneDrive - American National Insurance Company",
            "Life Product - Process_Control",
        )

    def _get_tools_root_path(self) -> str:
        # Base path: Process_Control/Task/Accelerated Death Benefit (ABR11 & ABR14)
        return os.path.join(
            self._get_process_control_dir(),
            "Task",
            "Accelerated Death Benefit (ABR11 & ABR14)"
        )

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # ── Top info bar ──────────────────────────────────────────────────
        info_frame = QGroupBox("Policy Support")
        info_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE_ABR)
        ig = QGridLayout(info_frame)
        ig.setContentsMargins(12, 24, 12, 8)
        ig.setHorizontalSpacing(10)
        ig.setVerticalSpacing(4)

        lbl_s = f"font-size: 11px; font-weight: bold; color: {CRIMSON_DARK}; background: transparent; border: none;"
        val_s = f"font-size: 11px; color: {GRAY_DARK}; background: transparent; border: none;"
        path_s = f"""
            font-size: 11px; color: {CRIMSON_DARK};
            background: #F9E0E3; border: 1px solid {CRIMSON_PRIMARY};
            border-radius: 3px; padding: 2px 6px;
        """

        # Row 0: Policy Folder label
        ig.addWidget(self._mk_lbl("Policy Folder:", lbl_s), 0, 0)
        self._policy_folder_label = QLabel("No policy loaded")
        self._policy_folder_label.setStyleSheet(val_s)
        ig.addWidget(self._policy_folder_label, 0, 1, 1, 2)  # span 2 cols

        # Row 1: Library Path — full width (no label)
        self._library_path_label = DoubleClickablePathLabel("")
        self._library_path_label.setStyleSheet(path_s)
        ig.addWidget(self._library_path_label, 1, 0, 1, 3)  # span all 3 cols

        # Row 2: Create Folder button — left justified
        self._create_folder_btn = QPushButton("📁 Create")
        self._create_folder_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {SLATE_TEXT}, stop:1 {SLATE_PRIMARY});
                color: {WHITE};
                border: 1px solid {SLATE_DARK};
                border-radius: 3px;
                padding: 2px 8px;
                font-size: 10px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {SLATE_PRIMARY}, stop:1 {SLATE_DARK});
                color: {WHITE};
            }}
        """)
        self._create_folder_btn.clicked.connect(self._on_create_policy_folder)
        self._create_folder_btn.setVisible(False)
        ig.addWidget(self._create_folder_btn, 2, 0, 1, 1)  # left-justified

        # Hidden status label (used internally for copy feedback)
        self._status_label = QLabel("")
        self._status_label.setStyleSheet(f"font-size: 10px; color: {GRAY_DARK}; background: transparent; border: none;")
        self._status_label.setVisible(False)

        ig.setColumnStretch(1, 1)
        layout.addWidget(info_frame)

        # ── Two-column Explorer ───────────────────────────────────────────
        cols = QWidget()
        cl = QHBoxLayout(cols)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(10)

        # ── Left column: Policy Subfolders + Recommended Files ──────────
        left_col = QWidget()
        left_layout = QVBoxLayout(left_col)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)

        # 1. Policy Subfolders (Drop Target) — compact
        self._subfolder_explorer = MiniExplorer(
            title="Policy Subfolders",
            list_widget_class=CrimsonDropTargetList,
            root_path="" # Will be set when policy loads
        )
        self._apply_abr_style(self._subfolder_explorer)

        # Print Detail button — small green button below subfolder group
        self._print_detail_btn = QPushButton("📋 Print Detail")
        self._print_detail_btn.setFixedHeight(22)
        self._print_detail_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #4CAF50, stop:1 #388E3C);
                color: {WHITE};
                border: 1px solid #2E7D32;
                border-radius: 3px;
                padding: 1px 10px;
                font-size: 10px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #66BB6A, stop:1 #43A047);
            }}
            QPushButton:disabled {{
                background: #C8E6C9;
                color: #A5D6A7;
                border-color: #C8E6C9;
            }}
        """)
        self._print_detail_btn.setEnabled(False)
        self._print_detail_btn.clicked.connect(self._on_print_detail)
        self._subfolder_explorer.list_widget.file_dropped.connect(self._on_file_dropped)
        
        left_layout.addWidget(self._subfolder_explorer, 3)  # stretch=3 (bigger)

        # Place button in a right-aligned row below the subfolder explorer
        detail_row = QHBoxLayout()
        detail_row.addStretch()
        detail_row.addWidget(self._print_detail_btn)
        left_layout.addLayout(detail_row)

        self._copy_up_btn = QPushButton("Copy ↑")
        self._copy_up_btn.setStyleSheet(f"""
            QPushButton {{
                background: {CRIMSON_SUBTLE};
                color: {CRIMSON_DARK};
                border: 1px solid {CRIMSON_PRIMARY};
                border-radius: 4px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {CRIMSON_PRIMARY};
                color: {WHITE};
            }}
        """)
        self._copy_up_btn.clicked.connect(self._on_copy_recommended)
        
        btn_layout_up = QHBoxLayout()
        btn_layout_up.addStretch(1)
        btn_layout_up.addWidget(self._copy_up_btn)
        btn_layout_up.addStretch(1)
        left_layout.addLayout(btn_layout_up)

        # 2. Recommended Forms — state-based forms from DB
        self._recommended_group = QGroupBox("Recommended Forms")
        self._recommended_group.setStyleSheet(ABR_EXPLORER_BOX_STYLE)
        rec_layout = QVBoxLayout(self._recommended_group)
        rec_layout.setContentsMargins(4, 20, 4, 4)
        rec_layout.setSpacing(2)

        # State label at top
        self._rec_state_label = QLabel("No state loaded")
        self._rec_state_label.setStyleSheet(f"""
            font-size: 10px; color: {CRIMSON_DARK}; font-weight: bold;
            background: #F9E0E3; border: 1px solid {CRIMSON_PRIMARY};
            border-radius: 3px; padding: 2px 6px;
        """)
        rec_layout.addWidget(self._rec_state_label)

        # Draggable list of recommended files
        self._recommended_list = RecommendedFilesList()
        self._recommended_list.setStyleSheet(ABR_LIST_STYLE)
        self._recommended_list.setItemDelegate(TightItemDelegate(self._recommended_list))
        self._recommended_list.setUniformItemSizes(True)
        self._recommended_list.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection
        )
        # Double-click to open the file
        self._recommended_list.itemDoubleClicked.connect(self._on_recommended_dblclick)
        rec_layout.addWidget(self._recommended_list, 1)

        # Hint label
        hint = QLabel("Drag files ↑ to Policy Subfolders")
        hint.setStyleSheet(f"font-size: 9px; color: {GRAY_MID}; font-style: italic; background: transparent; border: none;")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        rec_layout.addWidget(hint)

        left_layout.addWidget(self._recommended_group, 2)  # stretch=2

        cl.addWidget(left_col, 1)

        self._copy_left_btn = QPushButton("← Copy")
        self._copy_left_btn.setStyleSheet(f"""
            QPushButton {{
                background: {CRIMSON_SUBTLE};
                color: {CRIMSON_DARK};
                border: 1px solid {CRIMSON_PRIMARY};
                border-radius: 4px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {CRIMSON_PRIMARY};
                color: {WHITE};
            }}
        """)
        self._copy_left_btn.clicked.connect(self._on_copy_resource)
        
        btn_layout_left = QVBoxLayout()
        btn_layout_left.addStretch(1)
        btn_layout_left.addWidget(self._copy_left_btn)
        btn_layout_left.addStretch(1)
        
        cl.addLayout(btn_layout_left)

        # ── Right column: Resources (formerly Available Tools) ─────────
        self._tools_explorer = MiniExplorer(
            title="Resources",
            list_widget_class=CrimsonDraggableToolsList,
            root_path=self._tools_root_path
        )
        self._apply_abr_style(self._tools_explorer)
        if hasattr(self._tools_explorer, 'list_widget') and self._tools_explorer.list_widget:
            self._tools_explorer.list_widget.setSelectionMode(
                QAbstractItemView.SelectionMode.ExtendedSelection
            )
        
        cl.addWidget(self._tools_explorer, 1)

        layout.addWidget(cols, 1)

    def _apply_abr_style(self, explorer: MiniExplorer):
        """Inject ABR stylesheets into MiniExplorer components."""
        # Find the group box (container)
        gb = explorer.findChild(QGroupBox)
        if gb:
            gb.setStyleSheet(ABR_EXPLORER_BOX_STYLE)

        # Find nav buttons (using private members since we know the class structure)
        if hasattr(explorer, '_home_btn'):
            explorer._home_btn.setStyleSheet(ABR_NAV_BTN_STYLE)
        if hasattr(explorer, '_up_btn'):
            explorer._up_btn.setStyleSheet(ABR_NAV_BTN_STYLE)
        
        # Path label
        if hasattr(explorer, '_path_label'):
            explorer._path_label.setStyleSheet(ABR_PATH_BAR_STYLE)
            
        # List widget
        if hasattr(explorer, 'list_widget'):
            explorer.list_widget.setStyleSheet(ABR_LIST_STYLE)

    @staticmethod
    def _mk_lbl(text: str, style: str) -> QLabel:
        l = QLabel(text)
        l.setStyleSheet(style)
        return l

    # ── Policy Loading ──────────────────────────────────────────────────

    def set_policy(self, policy: Optional[ABRPolicyData]):
        self._policy = policy
        self._result = None
        self._assessment = None
        self._derived_values = {}
        self._mort_detail = []
        self._apv_detail = []
        self._apv_summary = {}
        self._print_detail_btn.setEnabled(False)
        self._update_ui_state()

    def set_result(self, result: ABRQuoteResult):
        self._result = result
        self._print_detail_btn.setEnabled(bool(self._policy and self._result))

    def set_assessment(self, assessment: MedicalAssessment):
        self._assessment = assessment

    def set_calc_data(self, mort_detail: list, apv_detail: list, apv_summary: dict):
        self._mort_detail = mort_detail
        self._apv_detail = apv_detail
        self._apv_summary = apv_summary

    def set_derived_values(self, derived_values: dict):
        self._derived_values = derived_values or {}

    def _update_ui_state(self):
        if not self._policy:
            self._policy_folder_label.setText("No policy loaded")
            self._library_path_label.setText("")
            self._status_label.setText("")
            self._create_folder_btn.setVisible(False)
            self._subfolder_explorer.set_root("")
            self._clear_recommended()
            return

        # Path: ...\Policies\PolicyNumber
        pn = self._policy.policy_number
        policies_dir = os.path.join(self._tools_root_path, "Policies")
        self._policy_folder_path = os.path.join(policies_dir, pn)
        
        if os.path.exists(self._policy_folder_path):
            self._policy_folder_label.setText(pn)
            self._policy_folder_label.setStyleSheet(
                f"font-size: 11px; color: {CRIMSON_DARK}; font-weight: bold; background: transparent; border: none;"
            )
            # Show full path (folder exists, double-click works)
            self._library_path_label.setText(self._policy_folder_path)
            self._create_folder_btn.setVisible(False)
            self._subfolder_explorer.set_root(self._policy_folder_path)
        else:
            self._policy_folder_label.setText(f"{pn} — No Folder Found")
            self._policy_folder_label.setStyleSheet(
                f"font-size: 11px; color: #C00000; font-weight: bold; background: transparent; border: none;"
            )
            # Show parent Policies dir (which exists) so double-click opens something useful
            self._library_path_label.setText(policies_dir if os.path.isdir(policies_dir) else self._tools_root_path)
            self._create_folder_btn.setVisible(True)
            self._subfolder_explorer.set_root("")

        # Load recommended files based on state
        self._load_recommended_files()

    # ── Recommended Files ───────────────────────────────────────────────

    def _clear_recommended(self):
        """Clear the recommended files list."""
        self._recommended_list.clear()
        self._rec_state_label.setText("No state loaded")

    def _load_recommended_files(self):
        """Query the state_variations table for the policy's issue_state and populate the list."""
        self._recommended_list.clear()

        if not self._policy or not self._policy.issue_state:
            self._rec_state_label.setText("No state loaded")
            return

        state_abbr = self._policy.issue_state.upper().strip()
        self._rec_state_label.setText(f"State: {state_abbr}")

        # Query the ABR database for form files associated with this state
        try:
            from ..models.abr_database import get_abr_database
            db = get_abr_database()
            conn = db.connect()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT state_name, election_form, "
                "       disclosure_form_critical, disclosure_form_chronic,"
                "       disclosure_form_terminal "
                "FROM [SV_ABR_STATE_VARIATIONS] "
                "WHERE state_abbr = ?",
                (state_abbr,)
            )
            row = cursor.fetchone()
            cursor.close()

            if not row:
                self._rec_state_label.setText(f"State: {state_abbr} (not in DB)")
                item = QListWidgetItem("  No forms found for this state")
                item.setForeground(QColor(GRAY_MID))
                self._recommended_list.addItem(item)
                return

            state_name = row[0] or state_abbr
            self._rec_state_label.setText(f"State: {state_abbr} — {state_name}")

            # Collect form file names (skip empty/None)
            form_fields = [
                ("Election Form", row[1]),
                ("Disclosure (Critical)", row[2]),
                ("Disclosure (Chronic)", row[3]),
                ("Disclosure (Terminal)", row[4]),
            ]

            # Known directories where form files may live
            forms_base = os.path.join(self._tools_root_path, "Forms")
            search_dirs = [
                os.path.join(forms_base, "ABR Election Forms"),
                os.path.join(forms_base, "ABR Disclosure Forms", "ABR14"),
                os.path.join(forms_base, "ABR Disclosure Forms", "ABR11"),
                os.path.join(forms_base, "ABR Disclosure Forms"),
                forms_base,
            ]

            seen_files = set()  # avoid duplicates
            for label, filename in form_fields:
                if not filename or not filename.strip():
                    continue
                filename = filename.strip()
                if filename in seen_files:
                    continue
                seen_files.add(filename)

                # Try to find the actual file on disk
                full_path = self._find_form_file(filename, search_dirs)

                ext = os.path.splitext(filename)[1].lower()
                icon = _icon_for_ext(ext)

                if full_path:
                    item = QListWidgetItem(f"{icon}  {filename}")
                    item.setData(Qt.ItemDataRole.UserRole, full_path)
                    item.setToolTip(f"{label}\n{full_path}")
                else:
                    item = QListWidgetItem(f"⚠️  {filename}")
                    item.setForeground(QColor(GRAY_MID))
                    item.setToolTip(f"{label}\nFile not found on disk")
                
                self._recommended_list.addItem(item)

            if self._recommended_list.count() == 0:
                item = QListWidgetItem("  No forms configured for this state")
                item.setForeground(QColor(GRAY_MID))
                self._recommended_list.addItem(item)

        except Exception as e:
            logger.error(f"Error loading recommended files: {e}", exc_info=True)
            self._rec_state_label.setText(f"State: {state_abbr} (error)")
            item = QListWidgetItem(f"  Error: {e}")
            item.setForeground(QColor("#C00000"))
            self._recommended_list.addItem(item)

    def _find_form_file(self, filename: str, search_dirs: List[str]) -> Optional[str]:
        """Search for a form file in the known directories. Returns full path or None."""
        for d in search_dirs:
            if not os.path.isdir(d):
                continue
            candidate = os.path.join(d, filename)
            if os.path.isfile(candidate):
                return candidate
        return None

    def _on_recommended_dblclick(self, item: QListWidgetItem):
        """Open the recommended file when double-clicked."""
        path = item.data(Qt.ItemDataRole.UserRole) or ""
        if path and os.path.isfile(path):
            os.startfile(path)

    # ── Actions ─────────────────────────────────────────────────────────

    def _on_create_policy_folder(self):
        if not self._policy_folder_path:
            return
        try:
            os.makedirs(self._policy_folder_path, exist_ok=True)
            self._update_ui_state()
            QMessageBox.information(self, "Success", "Policy folder created.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to create policy folder:\n{e}")

    def _do_copy_files(self, source_paths: List[str]):
        if not source_paths:
            return
            
        dest_dir = self._subfolder_explorer.current_path()
        if not dest_dir or not os.path.isdir(dest_dir):
            if self._policy_folder_path and os.path.isdir(self._policy_folder_path):
                 dest_dir = self._policy_folder_path
            else:
                QMessageBox.information(
                    self, "No Destination",
                    "Please create/open the policy folder first."
                )
                return

        copied_count = 0
        errors = []
        for source_path in source_paths:
            if not source_path or not os.path.isfile(source_path):
                continue
                
            filename = os.path.basename(source_path)
            if self._policy:
                dest_filename = f"{self._policy.policy_number} - {filename}"
            else:
                dest_filename = filename

            dest_path = os.path.join(dest_dir, dest_filename)
            if os.path.exists(dest_path):
                errors.append(f"'{dest_filename}' already exists.")
                continue

            try:
                shutil.copy2(source_path, dest_path)
                copied_count += 1
            except Exception as e:
                errors.append(f"Failed to copy '{filename}': {e}")
                
        if copied_count > 0:
            self._status_label.setText(f"✓ Copied {copied_count} file(s)")
            self._subfolder_explorer.refresh()
            
        if errors:
            QMessageBox.warning(self, "Copy Issues", "\n".join(errors))

    def _on_copy_recommended(self):
        items = self._recommended_list.selectedItems()
        paths = [i.data(Qt.ItemDataRole.UserRole) for i in items if i.data(Qt.ItemDataRole.UserRole)]
        self._do_copy_files(paths)

    def _on_copy_resource(self):
        if not hasattr(self._tools_explorer, 'list_widget') or not self._tools_explorer.list_widget:
            return
        items = self._tools_explorer.list_widget.selectedItems()
        paths = [i.data(Qt.ItemDataRole.UserRole) for i in items if i.data(Qt.ItemDataRole.UserRole)]
        self._do_copy_files(paths)

    def _on_file_dropped(self, source_path: str):
        """Handle file copy from Tools/Recommended to Policy folder."""
        self._do_copy_files([source_path])

    # ── Print Detail ────────────────────────────────────────────────────

    def _on_print_detail(self):
        """Generate a detail workbook and save it to the policy folder."""
        if not self._policy or not self._result:
            return

        from datetime import datetime

        now = datetime.now()
        timestamp = now.strftime("%m-%d-%Y %H%M%S")
        pn = self._policy.policy_number
        filename = f"{pn} - details - {timestamp}.xlsx"

        # Determine destination folder
        dest_dir = self._subfolder_explorer.current_path()
        if not dest_dir or not os.path.isdir(dest_dir):
            if self._policy_folder_path and os.path.isdir(self._policy_folder_path):
                dest_dir = self._policy_folder_path
            else:
                QMessageBox.information(
                    self, "No Destination",
                    "Please create the policy folder first.",
                )
                return

        filepath = os.path.join(dest_dir, filename)

        try:
            self._write_detail_workbook(filepath)
            self._subfolder_explorer.refresh()
            QMessageBox.information(
                self, "Print Detail",
                f"Detail workbook saved:\n{filename}",
            )
        except Exception as e:
            logger.error(f"Print Detail error: {e}", exc_info=True)
            QMessageBox.warning(
                self, "Print Detail Error",
                f"Could not create workbook:\n{e}",
            )

    def _write_detail_workbook(self, filepath: str):
        """Write a 5-sheet detail workbook with all quote data."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        p = self._policy
        r = self._result
        a = self._assessment

        header_font = Font(bold=True, size=12)
        label_font = Font(bold=True, size=11)
        value_font = Font(size=11)
        section_fill = PatternFill("solid", fgColor="8B0000")
        section_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="D3D3D3")
        header_col_font = Font(bold=True, size=10)
        data_font = Font(size=10)
        thin_border = Border(
            bottom=Side(style="thin", color="999999")
        )

        def _write_header(ws, title, row=1):
            ws.cell(row=row, column=1, value=title).font = header_font
            return row + 2

        def _write_section(ws, title, row):
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = section_font
            cell.fill = section_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            return row + 1

        def _write_field(ws, row, label, value, col=1):
            ws.cell(row=row, column=col, value=label).font = label_font
            ws.cell(row=row, column=col + 1, value=value).font = value_font
            return row + 1

        # ── Sheet 1: Policy Info ────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Policy Info"
        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 35

        row = _write_header(ws1, "ABR Quote — Policy Information")

        row = _write_section(ws1, "Policy Details", row)
        row = _write_field(ws1, row, "Policy Number:", p.policy_number)
        row = _write_field(ws1, row, "Insured:", p.insured_name or "—")

        plan_info = PLAN_CODE_INFO.get(p.plan_code.upper(), None) if p.plan_code else None
        plan_desc = f"{plan_info[1]} ({plan_info[0]}-Year Level)" if plan_info else "—"
        row = _write_field(ws1, row, "Plancode:", p.plan_code or "—")
        row = _write_field(ws1, row, "Plan Description:", plan_desc)

        sex_display = {"M": "Male", "F": "Female", "U": "Unisex"}.get(p.sex, p.sex or "—")
        row = _write_field(ws1, row, "Sex:", sex_display)
        row = _write_field(ws1, row, "Rate Sex:", p.rate_sex or "—")
        row = _write_field(ws1, row, "Issue Age:", p.issue_age)
        row = _write_field(ws1, row, "Attained Age:", p.attained_age)
        row = _write_field(ws1, row, "Rate Class:", p.rate_class or "—")
        row = _write_field(ws1, row, "Face Amount:", f"${p.face_amount:,.2f}" if p.face_amount else "—")
        row = _write_field(ws1, row, "Min Face:", f"${p.min_face_amount:,.0f}")
        row = _write_field(ws1, row, "Issue State:", p.issue_state or "—")
        if p.issue_date:
            row = _write_field(ws1, row, "Issue Date:", p.issue_date.strftime("%m/%d/%Y"))
        else:
            row = _write_field(ws1, row, "Issue Date:", "—")
        row = _write_field(ws1, row, "Policy Year:", p.policy_year)
        row = _write_field(ws1, row, "Month of Year:", p.policy_month)
        row = _write_field(ws1, row, "Base Plancode:", p.base_plancode or "—")
        row = _write_field(ws1, row, "Billing Mode:", MODAL_LABELS.get(p.billing_mode, str(p.billing_mode)))
        row = _write_field(ws1, row, "Modal Premium:", f"${p.modal_premium:,.2f}" if p.modal_premium else "—")
        row = _write_field(ws1, row, "Table Rating:", p.table_rating)
        row = _write_field(ws1, row, "Annual Flat Extra:", f"${p.flat_extra:.2f}" if p.flat_extra > 0 else "None")
        if p.flat_cease_date:
            row = _write_field(ws1, row, "Flat Cease Date:", p.flat_cease_date.strftime("%m/%d/%Y"))
        else:
            row = _write_field(ws1, row, "Flat Cease Date:", "—")

        row += 1
        row = _write_section(ws1, "Quote Parameters", row)
        row = _write_field(ws1, row, "Quote Date:", r.quote_date.strftime("%m/%d/%Y") if r.quote_date else "—")
        row = _write_field(ws1, row, "ABR Interest Rate:", f"{r.abr_interest_rate * 100:.2f}%")
        row = _write_field(ws1, row, "Per Diem (Daily):", f"${r.per_diem_daily:,.2f}")
        row = _write_field(ws1, row, "Per Diem (Annual):", f"${r.per_diem_annual:,.2f}")

        row += 1
        row = _write_section(ws1, "Riders / Coverages", row)
        if p.riders:
            for rider in p.riders:
                rider_desc = f"{rider.plancode} ({rider.rider_type})"
                if rider.benefit_type:
                    rider_desc += f" — BNF {rider.benefit_type}{rider.benefit_subtype or ''}"
                row = _write_field(ws1, row, rider_desc, f"${rider.fallback_premium:,.2f}/yr")
        else:
            row = _write_field(ws1, row, "No riders.", "")

        # ── Sheet 2: Assessment ─────────────────────────────────────────
        ws2 = wb.create_sheet("Assessment")
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 35

        row = _write_header(ws2, "ABR Quote — Assessment")

        row = _write_section(ws2, "Rider Configuration", row)
        row = _write_field(ws2, row, "Rider Type:", a.rider_type if a else "—")

        row += 1
        row = _write_section(ws2, "Assessment Inputs", row)
        if a:
            if a.use_five_year:
                row = _write_field(ws2, row, "5-Year Survival Rate:", f"{a.five_year_survival}")
                row = _write_field(ws2, row, "  Return to Normal:", "Yes" if a.use_return_5yr else "No")
            if a.use_ten_year:
                row = _write_field(ws2, row, "10-Year Survival Rate:", f"{a.ten_year_survival}")
                row = _write_field(ws2, row, "  Return to Normal:", "Yes" if a.use_return_10yr else "No")
            if a.use_le:
                row = _write_field(ws2, row, "Life Expectancy:", f"{a.life_expectancy_years} years")
            if a.use_table:
                row = _write_field(ws2, row, "Table (rating):", f"{a.direct_table_rating}")
                row = _write_field(ws2, row, "  Start/Stop Year:", f"{a.table_start_year} — {a.table_stop_year}")
            if a.use_flat:
                row = _write_field(ws2, row, "Flat ($/1000):", f"${a.direct_flat_extra:.2f}")
                row = _write_field(ws2, row, "  Start/Stop Year:", f"{a.flat_start_year} — {a.flat_stop_year}")
            if a.use_table_2:
                row = _write_field(ws2, row, "Table 2 (rating):", f"{a.direct_table_rating_2}")
                row = _write_field(ws2, row, "  Start/Stop Year:", f"{a.table_2_start_year} — {a.table_2_stop_year}")
            if a.use_flat_2:
                row = _write_field(ws2, row, "Flat 2 ($/1000):", f"${a.direct_flat_extra_2:.2f}")
                row = _write_field(ws2, row, "  Start/Stop Year:", f"{a.flat_2_start_year} — {a.flat_2_stop_year}")
            row = _write_field(ws2, row, "In Lieu Of:", "Yes" if a.in_lieu_of else "No (In Addition To)")
        else:
            row = _write_field(ws2, row, "No assessment data.", "")

        row += 1
        row = _write_section(ws2, "Derived Substandard Values", row)
        dv = self._derived_values
        if dv:
            # Widen columns for this section — 4 columns used
            ws2.column_dimensions['C'].width = 30
            ws2.column_dimensions['D'].width = 35

            # Sub-headers
            ws2.cell(row=row, column=1, value="Current (Unmodified)").font = Font(bold=True, size=11, underline="single")
            ws2.cell(row=row, column=3, value="Modified (Substandard Applied)").font = Font(bold=True, size=11, underline="single")
            row += 1

            field_pairs = [
                ("5-Year Survival:", "std_survival_5yr", "5-Year Survival:", "mod_survival_5yr"),
                ("10-Year Survival:", "std_survival_10yr", "10-Year Survival:", "mod_survival_10yr"),
                ("Life Expectancy:", "std_le", "Life Expectancy:", "mod_le"),
                ("Table Rating:", "std_table_rating", "Table Ratings:", "table_rating"),
                ("Flat Extra:", "std_flat_extra", "Flat Extras:", "flat_extra"),
            ]
            for std_label, std_key, mod_label, mod_key in field_pairs:
                ws2.cell(row=row, column=1, value=std_label).font = label_font
                ws2.cell(row=row, column=2, value=dv.get(std_key, "—")).font = value_font
                ws2.cell(row=row, column=3, value=mod_label).font = label_font
                ws2.cell(row=row, column=4, value=dv.get(mod_key, "—")).font = value_font
                row += 1
        elif a:
            row = _write_field(ws2, row, "Derived Table Rating:", f"{a.derived_table_rating:.4f}")
            if a.use_five_year and a.use_ten_year:
                row = _write_field(ws2, row, "  5yr Table Rating:", f"{a.derived_table_rating_5yr:.4f}")
                row = _write_field(ws2, row, "  10yr Table Rating:", f"{a.derived_table_rating_10yr:.4f}")
            row = _write_field(ws2, row, "Life Expectancy (rounded):", f"{a.life_expectancy_rounded}")

        row += 1
        row = _write_section(ws2, "Results Summary", row)
        row = _write_field(ws2, row, "Full Accel Benefit:", f"${r.full_accel_benefit:,.2f}")
        row = _write_field(ws2, row, "Full Benefit Ratio:", f"{r.full_benefit_ratio * 100:.2f}%")
        if r.partial_eligible_db > 0:
            row = _write_field(ws2, row, "Partial Accel Benefit:", f"${r.partial_accel_benefit:,.2f}")
            row = _write_field(ws2, row, "Partial Benefit Ratio:", f"{r.partial_benefit_ratio * 100:.2f}%")
        else:
            row = _write_field(ws2, row, "Partial Acceleration:", "NOT ALLOWED — At Minimum Face")
        row = _write_field(ws2, row, "Premium Before:", r.premium_before)
        row = _write_field(ws2, row, "After (Full Accel):", f"${r.premium_after_full:,.2f}")
        if r.partial_eligible_db > 0:
            row = _write_field(ws2, row, "After (Partial):", r.premium_after_partial)
        else:
            row = _write_field(ws2, row, "After (Partial):", "NOT ALLOWED")
        row = _write_field(ws2, row, "APV_FB:", f"${r.apv_fb:,.2f}")
        row = _write_field(ws2, row, "APV_FP:", f"${r.apv_fp:,.2f}")
        row = _write_field(ws2, row, "APV_FD:", f"${r.apv_fd:,.2f}")

        # ── Sheet 3: Mortality Derivation ───────────────────────────────
        ws3 = wb.create_sheet("Mortality Derivation")

        mort_headers = [
            "Quote Month", "Policy Year", "Mo in Yr", "Att Age",
            "qx VBT (annual)", "qx × Mult (annual)", "qx Improved (annual)",
            "Table Rating", "qx + Table (annual)",
            "Flat Extra ($/1000)", "qx + Flat (annual)", "qx Capped (annual)",
            "qx Monthly", "px Monthly", "Cum Survival",
        ]
        for c, h in enumerate(mort_headers, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = header_col_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for i, mrow in enumerate(self._mort_detail, 2):
            tbl_val = mrow.get("table_rating_applied", 0.0)
            flat_val = mrow.get("flat_extra_applied", 0.0)
            values = [
                mrow["quote_month"], mrow["duration_year"],
                mrow["month_in_year"], mrow["attained_age"],
                mrow["qx_vbt"], mrow["qx_multiplied"],
                mrow["qx_improved"],
                tbl_val if tbl_val > 0 else "",
                mrow["qx_table_rated"],
                flat_val if flat_val > 0 else "",
                mrow["qx_flat_extra"], mrow["qx_capped"],
                mrow["qx_monthly"], mrow["px_monthly"],
                mrow["cum_survival"],
            ]
            for c, v in enumerate(values, 1):
                cell = ws3.cell(row=i, column=c, value=v)
                cell.font = data_font
                if isinstance(v, float):
                    cell.number_format = '0.00000000'

        # Auto-width for mortality columns
        for c in range(1, len(mort_headers) + 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14

        # ── Sheet 4: Life Expectancy ────────────────────────────────────
        ws4 = wb.create_sheet("Life Expectancy")

        le_headers = [
            "Quote Month", "Policy Year", "Att Age",
            "qx Monthly", "px Monthly", "tPx (cum surv)",
            "Sum tPx (months)", "Curtate LE (years)",
        ]
        for c, h in enumerate(le_headers, 1):
            cell = ws4.cell(row=1, column=c, value=h)
            cell.font = header_col_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Compute LE development from mortality detail
        tp_x = 1.0
        sum_tpx = 0.0
        le_data_row = 2

        for mrow in self._mort_detail:
            qx_m = mrow["qx_monthly"]
            px_m = 1.0 - qx_m
            tp_x *= px_m
            sum_tpx += tp_x
            curtate_years = sum_tpx / 12.0

            values = [
                mrow["quote_month"], mrow["duration_year"],
                mrow["attained_age"],
                qx_m, px_m, tp_x,
                sum_tpx, curtate_years,
            ]
            for c, v in enumerate(values, 1):
                cell = ws4.cell(row=le_data_row, column=c, value=v)
                cell.font = data_font
                if isinstance(v, float):
                    cell.number_format = '0.000000'
            le_data_row += 1

        # Summary rows
        curtate_le = sum_tpx / 12.0 if self._mort_detail else 0.0
        complete_le = curtate_le + 0.5

        le_data_row += 1  # blank separator
        summary_font_xl = Font(bold=True, size=11, color="8B0000")
        for label, val in [
            ("Sum tPx (months):", sum_tpx),
            ("Curtate LE (years):", curtate_le),
            ("Complete LE (+ 0.5):", complete_le),
        ]:
            ws4.cell(row=le_data_row, column=6, value=label).font = summary_font_xl
            cell = ws4.cell(row=le_data_row, column=8, value=val)
            cell.font = summary_font_xl
            cell.number_format = '0.0000'
            le_data_row += 1

        for c in range(1, len(le_headers) + 1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16

        # ── Sheet 5: APV — Present Value ────────────────────────────────
        ws5 = wb.create_sheet("APV - Present Value")

        apv_headers = [
            "Month", "t", "qx Monthly", "px Monthly", "tpx (cum surv)",
            "v^(t+1) (benefit)", "v^t (premium)", "PVDB(t) (this mo)",
            "PVDB Cum", "Prem Rate (per $1K)", "PVFP(t) (this mo)",
            "PVFP Cum", "tpx End",
        ]
        for c, h in enumerate(apv_headers, 1):
            cell = ws5.cell(row=1, column=c, value=h)
            cell.font = header_col_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for i, arow in enumerate(self._apv_detail, 2):
            values = [
                arow["month"], arow["t"],
                arow["qx_monthly"], arow["px_monthly"], arow["tp_x"],
                arow["v_benefit"], arow["v_premium"],
                arow["pvdb_t"], arow["pvdb_cum"],
                arow["prem_rate"] if arow["prem_rate"] > 0 else "",
                arow["pvfp_t"], arow["pvfp_cum"],
                arow["tp_x_end"],
            ]
            for c, v in enumerate(values, 1):
                cell = ws5.cell(row=i, column=c, value=v)
                cell.font = data_font
                if isinstance(v, float):
                    cell.number_format = '0.000000'

        # APV summary rows
        if self._apv_summary:
            s = self._apv_summary
            apv_sum_row = len(self._apv_detail) + 3
            for label, val, fmt in [
                ("PVFB (raw sum):", s.get("pvfb_raw", 0), '0.000000'),
                ("Cont Mort Adj:", s.get("cont_mort_adj", 0), '0.0000000000'),
                ("PVFB (adj × 1000):", s.get("pvfb_adjusted", 0), '#,##0.00'),
                ("PVFP:", s.get("pvfp", 0), '#,##0.00'),
                ("Actuarial Discount:", s.get("actuarial_discount", 0), '#,##0.00'),
            ]:
                ws5.cell(row=apv_sum_row, column=8, value=label).font = summary_font_xl
                cell = ws5.cell(row=apv_sum_row, column=9, value=val)
                cell.font = summary_font_xl
                cell.number_format = fmt
                apv_sum_row += 1

        for c in range(1, len(apv_headers) + 1):
            ws5.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16

        wb.save(filepath)
