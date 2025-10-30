"""Schema Discovery - Discovers and caches database metadata"""

import logging
from typing import List, Dict, Any, Optional
from sqlalchemy import MetaData, Table, inspect, text
from sqlalchemy.engine import Engine

from suiteview.core.connection_manager import get_connection_manager

logger = logging.getLogger(__name__)


class SchemaDiscovery:
    """Discovers and caches database metadata using SQLAlchemy reflection"""

    def __init__(self):
        self.conn_manager = get_connection_manager()
        # Add simple in-memory cache for columns
        self._columns_cache = {}  # Key: (connection_id, table_name, schema_name)

    def get_tables(self, connection_id: int) -> List[Dict]:
        """
        Get all tables for a connection using SQLAlchemy reflection

        Args:
            connection_id: ID of database connection

        Returns:
            List of table dictionaries with name and schema
        """
        try:
            # Get connection info to check type
            connection = self.conn_manager.repo.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")

            conn_type = connection['connection_type']

            # Handle file-based connections differently
            if conn_type == 'EXCEL':
                return self._get_excel_sheets(connection)
            elif conn_type == 'CSV':
                return self._get_csv_table(connection)
            elif conn_type == 'ACCESS':
                return self._get_access_tables(connection)
            elif conn_type == 'FIXED_WIDTH':
                return self._get_fixed_width_table(connection)

            # Handle SQL Server with direct query (avoid pyodbc reflection issues)
            if conn_type == 'SQL_SERVER':
                return self._get_sql_server_tables(connection_id)
            
            # Handle DB2 with special queries (DataDirect Shadow Client requires LIMIT/WITH workarounds)
            if conn_type == 'DB2':
                return self._get_db2_tables(connection_id)

            # Handle database connections with SQLAlchemy
            engine = self.conn_manager.get_engine(connection_id)
            inspector = inspect(engine)

            tables = []

            # Get all schemas (databases)
            try:
                schemas = inspector.get_schema_names()
            except:
                # Some databases don't support schemas, use default
                schemas = [None]

            for schema in schemas:
                # Skip system schemas
                if schema in ['information_schema', 'mysql', 'performance_schema', 'sys']:
                    continue

                try:
                    table_names = inspector.get_table_names(schema=schema)

                    for table_name in table_names:
                        tables.append({
                            'table_name': table_name,
                            'schema_name': schema,
                            'full_name': f"{schema}.{table_name}" if schema else table_name,
                            'type': 'TABLE'
                        })
                except Exception as e:
                    logger.warning(f"Could not get tables for schema {schema}: {e}")

            logger.info(f"Discovered {len(tables)} tables for connection {connection_id}")
            return sorted(tables, key=lambda x: x['full_name'])

        except Exception as e:
            logger.error(f"Failed to discover tables: {e}")
            raise

    def _get_excel_sheets(self, connection: Dict) -> List[Dict]:
        """Get sheets from an Excel file"""
        import openpyxl
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Excel file not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = []

            for sheet_name in wb.sheetnames:
                sheets.append({
                    'table_name': sheet_name,
                    'schema_name': None,
                    'full_name': sheet_name,
                    'type': 'SHEET'
                })

            wb.close()
            logger.info(f"Discovered {len(sheets)} sheets in Excel file")
            return sheets

        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise

    def _get_sql_server_tables(self, connection_id: int) -> List[Dict]:
        """
        Get tables from SQL Server using direct SQL query
        This avoids pyodbc/SQLAlchemy reflection issues with ODBC Driver 17
        """
        try:
            engine = self.conn_manager.get_engine(connection_id)
            
            # Query to get user tables (not system tables)
            query = text("""
                SELECT 
                    SCHEMA_NAME(schema_id) as schema_name,
                    name as table_name,
                    'TABLE' as table_type
                FROM sys.tables
                WHERE type = 'U'  -- User tables only
                ORDER BY schema_name, name
            """)
            
            tables = []
            with engine.connect() as conn:
                result = conn.execute(query)
                for row in result:
                    schema_name = row[0]
                    table_name = row[1]
                    tables.append({
                        'table_name': table_name,
                        'schema_name': schema_name,
                        'full_name': f"{schema_name}.{table_name}",
                        'type': 'TABLE'
                    })
            
            logger.info(f"Discovered {len(tables)} tables in SQL Server database")
            return tables
            
        except Exception as e:
            logger.error(f"Failed to discover SQL Server tables: {e}")
            raise

    def _get_sql_server_columns(self, connection_id: int, table_name: str, 
                                schema_name: str = None) -> List[Dict]:
        """
        Get columns from SQL Server table using direct SQL query
        This avoids pyodbc/SQLAlchemy reflection issues with ODBC Driver 17
        """
        try:
            engine = self.conn_manager.get_engine(connection_id)
            
            # Default schema is dbo
            if not schema_name:
                schema_name = 'dbo'
            
            # Query to get column information
            query = text("""
                SELECT 
                    c.name as column_name,
                    t.name as data_type,
                    c.max_length,
                    c.precision,
                    c.scale,
                    c.is_nullable,
                    CASE WHEN pk.column_name IS NOT NULL THEN 1 ELSE 0 END as is_primary_key
                FROM sys.columns c
                INNER JOIN sys.types t ON c.user_type_id = t.user_type_id
                LEFT JOIN (
                    SELECT ku.COLUMN_NAME
                    FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS AS tc
                    INNER JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE AS ku
                        ON tc.CONSTRAINT_TYPE = 'PRIMARY KEY' 
                        AND tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
                        AND ku.TABLE_SCHEMA = :schema_name
                        AND ku.TABLE_NAME = :table_name
                ) pk ON c.name = pk.COLUMN_NAME
                WHERE c.object_id = OBJECT_ID(:full_table_name)
                ORDER BY c.column_id
            """)
            
            columns = []
            full_table_name = f"{schema_name}.{table_name}"
            
            with engine.connect() as conn:
                result = conn.execute(query, {
                    'schema_name': schema_name,
                    'table_name': table_name,
                    'full_table_name': full_table_name
                })
                
                for row in result:
                    col_name = row[0]
                    data_type = row[1]
                    max_length = row[2]
                    precision = row[3]
                    scale = row[4]
                    is_nullable = row[5]
                    is_primary_key = bool(row[6])
                    
                    # Format data type with length/precision
                    if data_type in ['varchar', 'char', 'nvarchar', 'nchar']:
                        if max_length == -1:
                            type_str = f"{data_type}(MAX)"
                        else:
                            actual_length = max_length // 2 if data_type.startswith('n') else max_length
                            type_str = f"{data_type}({actual_length})"
                    elif data_type in ['decimal', 'numeric']:
                        type_str = f"{data_type}({precision},{scale})"
                    else:
                        type_str = data_type
                    
                    columns.append({
                        'column_name': col_name,
                        'data_type': type_str,
                        'is_nullable': is_nullable,
                        'is_primary_key': is_primary_key,
                        'default': None,
                        'max_length': max_length if max_length > 0 else None
                    })
            
            logger.debug(f"Discovered {len(columns)} columns for {schema_name}.{table_name}")
            return columns
            
        except Exception as e:
            logger.error(f"Failed to discover SQL Server columns for {table_name}: {e}")
            raise

    def _get_sql_server_unique_values(self, connection_id: int, table_name: str,
                                     column_name: str, schema_name: str = None,
                                     limit: int = 1000) -> List[Any]:
        """
        Get unique values from SQL Server column using direct SQL query
        Uses TOP instead of LIMIT for SQL Server compatibility
        """
        try:
            engine = self.conn_manager.get_engine(connection_id)
            
            # Default schema is dbo
            if not schema_name:
                schema_name = 'dbo'
            
            # Build qualified table name
            qualified_table = f"[{schema_name}].[{table_name}]"
            
            # SQL Server uses TOP instead of LIMIT
            # Use parameterized query for column name safety
            query = text(f"""
                SELECT DISTINCT TOP {limit} [{column_name}]
                FROM {qualified_table}
                WHERE [{column_name}] IS NOT NULL
                ORDER BY [{column_name}]
            """)
            
            unique_values = []
            with engine.connect() as conn:
                result = conn.execute(query)
                for row in result:
                    unique_values.append(row[0])
            
            logger.debug(f"Found {len(unique_values)} unique values for {schema_name}.{table_name}.{column_name}")
            return unique_values
            
        except Exception as e:
            logger.error(f"Failed to get SQL Server unique values for {table_name}.{column_name}: {e}")
            return []

    def _get_sql_server_preview(self, connection_id: int, table_name: str,
                                schema_name: str = None, limit: int = 100) -> tuple:
        """
        Get preview data from SQL Server table
        Uses TOP instead of LIMIT for SQL Server compatibility
        """
        try:
            engine = self.conn_manager.get_engine(connection_id)
            
            # Default schema is dbo
            if not schema_name:
                schema_name = 'dbo'
            
            # Build qualified table name
            qualified_table = f"[{schema_name}].[{table_name}]"
            
            # SQL Server uses TOP instead of LIMIT
            query = text(f"SELECT TOP {limit} * FROM {qualified_table}")
            
            with engine.connect() as conn:
                result = conn.execute(query)
                columns = list(result.keys())
                data = [tuple(row) for row in result.fetchall()]
            
            logger.info(f"Retrieved {len(data)} rows for preview of {schema_name}.{table_name}")
            return (columns, data)
            
        except Exception as e:
            logger.error(f"Failed to get SQL Server preview for {table_name}: {e}")
            raise

    def _get_db2_tables(self, connection_id: int) -> List[Dict]:
        """
        Get tables from DB2 using direct pyodbc connection with workarounds
        DataDirect Shadow Client 7.3 requires special handling:
        - Must use WITH dummy clause for system catalog queries
        - Must include LIMIT clause to prevent crashes
        - Cannot use SQLAlchemy engine (forces MSSQL dialect incompatibility)
        """
        try:
            import pyodbc
            
            # Get connection details
            connection = self.conn_manager.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")
            
            # Build pyodbc connection string from DSN
            dsn = connection.get('connection_string', '').replace('DSN=', '')
            if not dsn:
                raise ValueError("DB2 connection requires DSN")
            
            conn_str = f"DSN={dsn}"
            
            # Connect with pyodbc directly
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            
            # Query to get user tables with DB2-specific workarounds
            # Using WITH DUMMY clause and LIMIT to avoid driver crashes
            query = """
                WITH DUMMY AS (SELECT 1 AS COL1 FROM SYSIBM.SYSDUMMY1)
                SELECT NAME, CREATOR, TYPE
                FROM SYSIBM.SYSTABLES
                WHERE TYPE = 'T'
                ORDER BY CREATOR, NAME
                LIMIT 10000000
            """
            
            cursor.execute(query)
            
            tables = []
            for row in cursor.fetchall():
                table_name = row[0].strip() if row[0] else row[0]
                schema_name = row[1].strip() if row[1] else row[1]
                
                # Skip system tables
                if schema_name and (schema_name.startswith('SYS') or schema_name.startswith('Q')):
                    continue
                if table_name and table_name.startswith('SYS'):
                    continue
                
                tables.append({
                    'table_name': table_name,
                    'schema_name': schema_name,
                    'full_name': f"{schema_name}.{table_name}" if schema_name else table_name,
                    'type': 'TABLE'
                })
            
            conn.close()
            
            logger.info(f"Discovered {len(tables)} tables in DB2 database")
            return tables
            
        except Exception as e:
            logger.error(f"Failed to discover DB2 tables: {e}")
            raise

    def _get_db2_columns(self, connection_id: int, table_name: str, 
                        schema_name: str = None) -> List[Dict]:
        """
        Get columns from DB2 table using direct pyodbc connection with workarounds
        """
        try:
            import pyodbc
            
            if not schema_name:
                raise ValueError("Schema name is required for DB2 tables")
            
            # Get connection details
            connection = self.conn_manager.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")
            
            # Build pyodbc connection string from DSN
            dsn = connection.get('connection_string', '').replace('DSN=', '')
            if not dsn:
                raise ValueError("DB2 connection requires DSN")
            
            conn_str = f"DSN={dsn}"
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            
            # Query to get column information
            # DataDirect Shadow driver doesn't support parameter binding or complex queries well
            # Keep it simple and get primary key info separately if needed
            query = f"""
                SELECT 
                    NAME,
                    COLTYPE,
                    LENGTH,
                    SCALE,
                    NULLS
                FROM SYSIBM.SYSCOLUMNS
                WHERE TBNAME = '{table_name}'
                    AND TBCREATOR = '{schema_name}'
                ORDER BY COLNO
                LIMIT 10000000
            """
            
            cursor.execute(query)
            
            columns = []
            for row in cursor.fetchall():
                col_name = row[0].strip() if row[0] else row[0]
                type_string = row[1].strip() if row[1] else ''
                length = row[2]
                scale = row[3]
                nullable = row[4]
                
                # DB2 COLTYPE returns full type names like 'CHAR', 'VARCHAR', 'INTEGER', etc.
                # Add length/precision info for applicable types
                if type_string in ['VARCHAR', 'CHAR', 'GRAPHIC', 'VARGRAPHIC']:
                    if length:
                        readable_type = f"{type_string}({length})"
                    else:
                        readable_type = type_string
                elif type_string in ['DECIMAL', 'NUMERIC']:
                    if length and scale is not None:
                        readable_type = f"{type_string}({length},{scale})"
                    else:
                        readable_type = type_string
                else:
                    readable_type = type_string
                
                columns.append({
                    'column_name': col_name,
                    'data_type': readable_type,
                    'is_nullable': nullable == 'Y' if nullable else True,
                    'is_primary_key': False,  # Would require separate query
                    'default': None,
                    'max_length': length if length else None
                })
            
            conn.close()
            
            logger.info(f"Discovered {len(columns)} columns in DB2 table {schema_name}.{table_name}")
            return columns
            
        except Exception as e:
            logger.error(f"Failed to discover DB2 columns: {e}")
            raise

    def _get_db2_unique_values(self, connection_id: int, table_name: str,
                              column_name: str, schema_name: str = None,
                              limit: int = 1000) -> List[Any]:
        """
        Get unique values from DB2 column with LIMIT clause
        LIMIT is required to prevent DataDirect driver crashes
        """
        try:
            import pyodbc
            
            if not schema_name:
                raise ValueError("Schema name is required for DB2 tables")
            
            # Get connection details
            connection = self.conn_manager.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")
            
            # Build pyodbc connection string from DSN
            dsn = connection.get('connection_string', '').replace('DSN=', '')
            if not dsn:
                raise ValueError("DB2 connection requires DSN")
            
            conn_str = f"DSN={dsn}"
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            
            # Build qualified table name
            qualified_table = f'{schema_name}.{table_name}'
            
            # DB2 query with LIMIT to prevent crashes
            query = f"""
                SELECT DISTINCT {column_name}
                FROM {qualified_table}
                WHERE {column_name} IS NOT NULL
                ORDER BY {column_name}
                LIMIT {limit}
            """
            
            cursor.execute(query)
            unique_values = [row[0] for row in cursor.fetchall()]
            
            conn.close()
            
            logger.info(f"Found {len(unique_values)} unique values for {qualified_table}.{column_name}")
            return unique_values
            
        except Exception as e:
            logger.error(f"Failed to get DB2 unique values: {e}")
            return []

    def _get_db2_preview(self, connection_id: int, table_name: str,
                        schema_name: str = None, limit: int = 100) -> tuple:
        """
        Get preview data from DB2 table with LIMIT clause
        """
        try:
            import pyodbc
            
            if not schema_name:
                raise ValueError("Schema name is required for DB2 tables")
            
            # Get connection details
            connection = self.conn_manager.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")
            
            # Build pyodbc connection string from DSN
            dsn = connection.get('connection_string', '').replace('DSN=', '')
            if not dsn:
                raise ValueError("DB2 connection requires DSN")
            
            conn_str = f"DSN={dsn}"
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            
            # Build qualified table name - NO QUOTES (works with DB2)
            qualified_table = f'{schema_name}.{table_name}'
            
            # DB2 query with LIMIT
            query = f"SELECT * FROM {qualified_table} LIMIT {limit}"
            
            cursor.execute(query)
            
            # Get column names from cursor.description
            columns = [desc[0] for desc in cursor.description]
            
            # Fetch data
            data = [tuple(row) for row in cursor.fetchall()]
            
            conn.close()
            
            logger.info(f"Retrieved {len(data)} rows for preview of {qualified_table}")
            return (columns, data)
            
        except Exception as e:
            logger.error(f"Failed to get DB2 preview for {table_name}: {e}")
            raise

    def _get_csv_table(self, connection: Dict) -> List[Dict]:
        """Get table from a CSV file (single table)"""
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"CSV file not found: {file_path}")

        filename = os.path.basename(file_path)
        table_name = os.path.splitext(filename)[0]

        return [{
            'table_name': table_name,
            'schema_name': None,
            'full_name': table_name,
            'type': 'CSV'
        }]

    def _get_access_tables(self, connection: Dict) -> List[Dict]:
        """Get tables from MS Access database using ODBC"""
        try:
            import pyodbc
            import os

            file_path = connection.get('connection_string', '')
            if not file_path or not os.path.exists(file_path):
                raise ValueError(f"Access file not found: {file_path}")

            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={file_path};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            tables = []
            for table_info in cursor.tables(tableType='TABLE'):
                # Skip system tables
                if not table_info.table_name.startswith('MSys'):
                    tables.append({
                        'table_name': table_info.table_name,
                        'schema_name': None,
                        'full_name': table_info.table_name,
                        'type': 'TABLE'
                    })

            conn.close()
            logger.info(f"Discovered {len(tables)} tables in Access database")
            return tables

        except Exception as e:
            logger.error(f"Failed to read Access database: {e}")
            raise

    def _get_fixed_width_table(self, connection: Dict) -> List[Dict]:
        """Get table from a fixed width file (single table)"""
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Fixed width file not found: {file_path}")

        filename = os.path.basename(file_path)
        table_name = os.path.splitext(filename)[0]

        return [{
            'table_name': table_name,
            'schema_name': None,
            'full_name': table_name,
            'type': 'FIXED_WIDTH'
        }]

    def get_columns(self, connection_id: int, table_name: str,
                   schema_name: str = None) -> List[Dict]:
        """
        Get columns for a table

        Args:
            connection_id: ID of database connection
            table_name: Name of the table
            schema_name: Schema/database name (optional)

        Returns:
            List of column dictionaries
        """
        # Check cache first
        cache_key = (connection_id, table_name, schema_name)
        if cache_key in self._columns_cache:
            logger.debug(f"Using cached columns for {schema_name}.{table_name}")
            return self._columns_cache[cache_key]
        
        try:
            # Get connection info to check type
            connection = self.conn_manager.repo.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")

            conn_type = connection['connection_type']

            # Handle file-based connections differently
            if conn_type == 'EXCEL':
                columns = self._get_excel_columns(connection, table_name)
            elif conn_type == 'CSV':
                columns = self._get_csv_columns(connection)
            elif conn_type == 'ACCESS':
                columns = self._get_access_columns(connection, table_name)
            elif conn_type == 'FIXED_WIDTH':
                columns = self._get_fixed_width_columns(connection)
            # Handle SQL Server with direct query (avoid pyodbc reflection issues)
            elif conn_type == 'SQL_SERVER':
                columns = self._get_sql_server_columns(connection_id, table_name, schema_name)
            # Handle DB2 with direct query (DataDirect driver requires special handling)
            elif conn_type == 'DB2':
                columns = self._get_db2_columns(connection_id, table_name, schema_name)
            else:
                # Handle database connections with SQLAlchemy
                engine = self.conn_manager.get_engine(connection_id)
                inspector = inspect(engine)

                # Get column information
                columns_info = inspector.get_columns(table_name, schema=schema_name)

                # Get primary keys
                pk_constraint = inspector.get_pk_constraint(table_name, schema=schema_name)
                primary_keys = pk_constraint.get('constrained_columns', []) if pk_constraint else []
                
                columns = []
                for col_info in columns_info:
                    col_name = col_info['name']
                    col_type = str(col_info['type'])

                    columns.append({
                        'column_name': col_name,
                        'data_type': col_type,
                        'is_nullable': col_info.get('nullable', True),
                        'is_primary_key': col_name in primary_keys,
                        'default': col_info.get('default'),
                        'max_length': None  # Could parse from type string if needed
                    })

            logger.debug(f"Discovered {len(columns)} columns for {table_name}")
            
            # Cache the results
            self._columns_cache[cache_key] = columns
            
            return columns

        except Exception as e:
            logger.error(f"Failed to discover columns for {table_name}: {e}")
            raise

    def _get_excel_columns(self, connection: Dict, sheet_name: str) -> List[Dict]:
        """Get columns from an Excel sheet"""
        import openpyxl
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Excel file not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            ws = wb[sheet_name]

            # Get header row (assume first row contains headers)
            columns = []
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

            if header_row:
                for idx, cell_value in enumerate(header_row, start=1):
                    col_name = str(cell_value) if cell_value is not None else f"Column{idx}"

                    # Try to infer data type from first few non-null values
                    data_type = "TEXT"
                    sample_values = []
                    for row in ws.iter_rows(min_row=2, max_row=10, min_col=idx, max_col=idx, values_only=True):
                        if row[0] is not None:
                            sample_values.append(row[0])

                    if sample_values:
                        sample = sample_values[0]
                        if isinstance(sample, (int, float)):
                            data_type = "NUMERIC"
                        elif isinstance(sample, bool):
                            data_type = "BOOLEAN"
                        # openpyxl converts dates to datetime objects
                        elif hasattr(sample, 'strftime'):
                            data_type = "DATE/TIME"

                    columns.append({
                        'column_name': col_name,
                        'data_type': data_type,
                        'is_nullable': True,
                        'is_primary_key': False,
                        'default': None,
                        'max_length': None
                    })
            else:
                # No header row found
                raise ValueError("Excel sheet appears to be empty")

            wb.close()
            logger.info(f"Discovered {len(columns)} columns in Excel sheet '{sheet_name}'")
            return columns

        except Exception as e:
            logger.error(f"Failed to read Excel columns: {e}")
            raise

    def _get_csv_columns(self, connection: Dict) -> List[Dict]:
        """Get columns from a CSV file"""
        import csv
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"CSV file not found: {file_path}")

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header_row = next(reader, None)

                if not header_row:
                    raise ValueError("CSV file appears to be empty")

                columns = []
                for col_name in header_row:
                    columns.append({
                        'column_name': col_name.strip(),
                        'data_type': "TEXT",
                        'is_nullable': True,
                        'is_primary_key': False,
                        'default': None,
                        'max_length': None
                    })

                logger.info(f"Discovered {len(columns)} columns in CSV file")
                return columns

        except Exception as e:
            logger.error(f"Failed to read CSV columns: {e}")
            raise

    def _get_access_columns(self, connection: Dict, table_name: str) -> List[Dict]:
        """Get columns from an Access table"""
        try:
            import pyodbc
            import os

            file_path = connection.get('connection_string', '')
            if not file_path or not os.path.exists(file_path):
                raise ValueError(f"Access file not found: {file_path}")

            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={file_path};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Query the table to get column info instead of using cursor.columns()
            # which has issues with some Access databases
            query = f"SELECT TOP 1 * FROM [{table_name}]"
            cursor.execute(query)
            
            columns = []
            for col_desc in cursor.description:
                # col_desc is tuple: (name, type_code, display_size, internal_size, precision, scale, null_ok)
                col_name = col_desc[0]
                type_code = col_desc[1]
                
                # Map type codes to type names
                type_map = {
                    str: 'TEXT',
                    int: 'INTEGER',
                    float: 'FLOAT',
                    bool: 'BOOLEAN',
                }
                type_name = type_map.get(type_code, 'TEXT')
                
                columns.append({
                    'column_name': col_name,
                    'data_type': type_name,
                    'is_nullable': bool(col_desc[6]) if len(col_desc) > 6 else True,
                    'is_primary_key': False,  # Would need additional query
                    'default': None,
                    'max_length': col_desc[3] if len(col_desc) > 3 else None
                })

            conn.close()
            logger.info(f"Discovered {len(columns)} columns in Access table '{table_name}'")
            return columns

        except Exception as e:
            logger.error(f"Failed to read Access columns: {e}")
            raise

    def _get_fixed_width_columns(self, connection: Dict) -> List[Dict]:
        """Get columns from fixed width file definition"""
        # TODO: Implement based on stored field definitions
        # For now, return a placeholder
        return [{
            'column_name': 'Data',
            'data_type': 'TEXT',
            'is_nullable': True,
            'is_primary_key': False,
            'default': None,
            'max_length': None
        }]

    def get_unique_values(self, connection_id: int, table_name: str,
                         column_name: str, schema_name: str = None,
                         limit: int = 1000) -> List[Any]:
        """
        Get unique values for a column

        Args:
            connection_id: ID of database connection
            table_name: Name of the table
            column_name: Name of the column
            schema_name: Schema/database name (optional)
            limit: Maximum number of unique values to return

        Returns:
            List of unique values (sorted)
        """
        try:
            # Get connection info to check type
            connection = self.conn_manager.repo.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")

            conn_type = connection['connection_type']

            # Handle file-based connections differently
            if conn_type == 'EXCEL':
                return self._get_excel_unique_values(connection, table_name, column_name, limit)
            elif conn_type == 'ACCESS':
                return self._get_access_unique_values(connection, table_name, column_name, limit)
            elif conn_type == 'CSV':
                return self._get_csv_unique_values(connection, table_name, column_name, limit)
            elif conn_type == 'SQL_SERVER':
                return self._get_sql_server_unique_values(connection_id, table_name, column_name, schema_name, limit)
            elif conn_type == 'DB2':
                return self._get_db2_unique_values(connection_id, table_name, column_name, schema_name, limit)
            else:
                # SQL-based connections
                engine = self.conn_manager.get_engine(connection_id)

                # Build qualified table name
                qualified_table = f"{schema_name}.{table_name}" if schema_name else table_name

                # Build query to get unique values
                query = text(f"""
                    SELECT DISTINCT {column_name}
                    FROM {qualified_table}
                    WHERE {column_name} IS NOT NULL
                    ORDER BY {column_name}
                    LIMIT {limit}
                """)

                with engine.connect() as conn:
                    result = conn.execute(query)
                    unique_values = [row[0] for row in result]

                logger.debug(f"Found {len(unique_values)} unique values for {table_name}.{column_name}")
                return unique_values

        except Exception as e:
            logger.error(f"Failed to get unique values for {table_name}.{column_name}: {e}")
            # Return empty list instead of raising - allows graceful degradation
            return []

    def _get_excel_unique_values(self, connection: Dict, sheet_name: str,
                                column_name: str, limit: int = 1000) -> List[Any]:
        """Get unique values from an Excel column"""
        import openpyxl
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]

        # Find column index by header name
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        col_index = None
        for idx, header in enumerate(header_row, start=1):
            if str(header) == column_name:
                col_index = idx
                break

        if col_index is None:
            raise ValueError(f"Column '{column_name}' not found in sheet")

        # Collect unique values
        unique_values = set()
        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index, values_only=True):
            value = row[0]
            if value is not None:
                unique_values.add(value)
                if len(unique_values) >= limit:
                    break

        wb.close()

        # Sort and return
        sorted_values = sorted(list(unique_values), key=lambda x: (x is None, x))
        return sorted_values[:limit]

    def _get_access_unique_values(self, connection: Dict, table_name: str,
                                 column_name: str, limit: int = 1000) -> List[Any]:
        """Get unique values from an Access table column"""
        import pyodbc
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Access file not found: {file_path}")

        # Use pyodbc to query the Access database
        conn_str = (
            r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
            f'DBQ={file_path};'
        )

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Query unique values
        query = f"""
            SELECT DISTINCT [{column_name}]
            FROM [{table_name}]
            WHERE [{column_name}] IS NOT NULL
            ORDER BY [{column_name}]
        """

        cursor.execute(query)
        unique_values = [row[0] for row in cursor.fetchmany(limit)]

        cursor.close()
        conn.close()

        return unique_values

    def _get_csv_unique_values(self, connection: Dict, table_name: str,
                              column_name: str, limit: int = 1000) -> List[Any]:
        """Get unique values from a CSV column"""
        import pandas as pd
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"CSV file not found: {file_path}")

        # Read CSV
        df = pd.read_csv(file_path)

        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in CSV")

        # Get unique values
        unique_values = df[column_name].dropna().unique()
        unique_values = sorted(unique_values)[:limit]

        return list(unique_values)

    def get_preview_data(self, connection_id: int, table_name: str,
                        schema_name: str = None, limit: int = 100) -> tuple:
        """
        Get preview data from a table (first N rows)

        Args:
            connection_id: ID of database connection
            table_name: Name of the table
            schema_name: Schema/database name (optional)
            limit: Maximum number of rows to return (default 100)

        Returns:
            Tuple of (columns, data) where columns is list of column names
            and data is list of row tuples
        """
        try:
            # Get connection info to check type
            connection = self.conn_manager.repo.get_connection(connection_id)
            if not connection:
                raise ValueError(f"Connection {connection_id} not found")

            conn_type = connection['connection_type']

            # Handle file-based connections differently
            if conn_type == 'EXCEL':
                return self._get_excel_preview(connection, table_name, limit)
            elif conn_type == 'ACCESS':
                return self._get_access_preview(connection, table_name, limit)
            elif conn_type == 'CSV':
                return self._get_csv_preview(connection, table_name, limit)
            elif conn_type == 'SQL_SERVER':
                return self._get_sql_server_preview(connection_id, table_name, schema_name, limit)
            elif conn_type == 'DB2':
                return self._get_db2_preview(connection_id, table_name, schema_name, limit)
            else:
                # SQL-based connections
                engine = self.conn_manager.get_engine(connection_id)

                # Build qualified table name
                qualified_table = f"{schema_name}.{table_name}" if schema_name else table_name

                # Build preview query
                query = text(f"SELECT * FROM {qualified_table} LIMIT {limit}")

                with engine.connect() as conn:
                    result = conn.execute(query)
                    columns = list(result.keys())
                    data = [tuple(row) for row in result.fetchall()]

                logger.info(f"Retrieved {len(data)} rows for preview of {table_name}")
                return (columns, data)

        except Exception as e:
            logger.error(f"Failed to get preview data for {table_name}: {e}")
            raise

    def _get_excel_preview(self, connection: Dict, sheet_name: str, limit: int = 100) -> tuple:
        """Get preview data from an Excel sheet"""
        import openpyxl
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]

        # Get column headers
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        columns = [str(h) if h is not None else f"Column{i}" for i, h in enumerate(header_row, 1)]

        # Get data rows
        data = []
        for row in ws.iter_rows(min_row=2, max_row=limit+1, values_only=True):
            data.append(row)

        wb.close()

        return (columns, data)

    def _get_access_preview(self, connection: Dict, table_name: str, limit: int = 100) -> tuple:
        """Get preview data from an Access table"""
        import pyodbc
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"Access file not found: {file_path}")

        conn_str = (
            r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
            f'DBQ={file_path};'
        )

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Query first N rows
        query = f"SELECT TOP {limit} * FROM [{table_name}]"
        cursor.execute(query)

        # Get column names
        columns = [column[0] for column in cursor.description]

        # Get data
        data = [tuple(row) for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return (columns, data)

    def _get_csv_preview(self, connection: Dict, table_name: str, limit: int = 100) -> tuple:
        """Get preview data from a CSV file"""
        import pandas as pd
        import os

        file_path = connection.get('connection_string', '')
        if not file_path or not os.path.exists(file_path):
            raise ValueError(f"CSV file not found: {file_path}")

        # Read CSV with limit
        df = pd.read_csv(file_path, nrows=limit)

        columns = list(df.columns)
        data = [tuple(row) for row in df.itertuples(index=False, name=None)]

        return (columns, data)

    def get_row_count(self, connection_id: int, table_name: str,
                     schema_name: str = None) -> int:
        """
        Get row count for a table

        Args:
            connection_id: ID of database connection
            table_name: Name of the table
            schema_name: Schema/database name (optional)

        Returns:
            Number of rows in the table
        """
        try:
            engine = self.conn_manager.get_engine(connection_id)

            # Build qualified table name
            qualified_table = f"{schema_name}.{table_name}" if schema_name else table_name

            # Build count query
            query = text(f"SELECT COUNT(*) FROM {qualified_table}")

            with engine.connect() as conn:
                result = conn.execute(query)
                count = result.scalar()

            logger.debug(f"Row count for {table_name}: {count}")
            return count or 0

        except Exception as e:
            logger.error(f"Failed to get row count for {table_name}: {e}")
            return 0

    def refresh_metadata(self, connection_id: int):
        """
        Refresh cached metadata for a connection

        Args:
            connection_id: ID of database connection
        """
        # For now, just clear the engine cache so it reconnects
        # In the future, could update cached metadata in database
        self.conn_manager._engines.pop(connection_id, None)
        logger.info(f"Refreshed metadata for connection {connection_id}")


# Singleton instance
_schema_discovery: Optional[SchemaDiscovery] = None


def get_schema_discovery() -> SchemaDiscovery:
    """Get or create singleton schema discovery instance"""
    global _schema_discovery
    if _schema_discovery is None:
        _schema_discovery = SchemaDiscovery()
    return _schema_discovery
