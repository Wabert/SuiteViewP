"""
Enhanced File Explorer V4 - Power User Edition
Features:
- Multi-tab support
- Breadcrumb navigation
- Mainframe upload integration
- Enhanced preview (Office files)
- Batch operations
- Print Directory to Excel
"""

import os
import sys
import shutil
import subprocess
import json
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (QTreeView, QVBoxLayout, QHBoxLayout, QWidget, 
                              QHeaderView, QToolBar, QMessageBox, QInputDialog, 
                              QTextEdit, QPushButton, QSplitter, QLabel, QMenu,
                              QTabWidget, QDialog, QDialogButtonBox, QCheckBox,
                              QComboBox, QLineEdit, QProgressDialog, QFrame)
from PyQt6.QtGui import QIcon, QAction, QStandardItemModel, QStandardItem
from PyQt6.QtCore import Qt, QModelIndex, QThread, pyqtSignal

import logging

logger = logging.getLogger(__name__)


class DirectoryExportThread(QThread):
    """Background thread for exporting directory to Excel"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(str, bool)
    
    def __init__(self, root_path, include_subdirs, output_path):
        super().__init__()
        self.root_path = root_path
        self.include_subdirs = include_subdirs
        self.output_path = output_path
        
    def run(self):
        """Export directory structure to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Directory Contents"
            
            # Headers
            headers = ['Level', 'Type', 'Name', 'Full Path', 'Size', 'Modified', 'Extension']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions['A'].width = 8   # Level
            ws.column_dimensions['B'].width = 10  # Type
            ws.column_dimensions['C'].width = 40  # Name
            ws.column_dimensions['D'].width = 60  # Full Path
            ws.column_dimensions['E'].width = 12  # Size
            ws.column_dimensions['F'].width = 18  # Modified
            ws.column_dimensions['G'].width = 10  # Extension
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            row_num = 2
            total_files = 0
            total_size = 0
            
            # Walk directory
            root = Path(self.root_path)
            
            if self.include_subdirs:
                items = []
                for dirpath, dirnames, filenames in os.walk(root):
                    current = Path(dirpath)
                    level = len(current.relative_to(root).parts)
                    
                    # Add directories
                    for dirname in sorted(dirnames):
                        items.append((current / dirname, level, True))
                    
                    # Add files
                    for filename in sorted(filenames):
                        items.append((current / filename, level, False))
                        
                    self.progress.emit(len(items), f"Scanning... {len(items)} items found")
            else:
                # Just immediate contents
                items = []
                for item in sorted(root.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                    items.append((item, 0, item.is_dir()))
            
            # Add items to Excel
            for idx, (item_path, level, is_dir) in enumerate(items):
                try:
                    stats = item_path.stat()
                    
                    # Type
                    item_type = "📁 Folder" if is_dir else "📄 File"
                    
                    # Name with indentation
                    indent = "  " * level
                    name = f"{indent}{item_path.name}"
                    
                    # Size
                    if is_dir:
                        size_str = ""
                    else:
                        size_bytes = stats.st_size
                        total_size += size_bytes
                        total_files += 1
                        
                        if size_bytes < 1024:
                            size_str = f"{size_bytes} B"
                        elif size_bytes < 1024 * 1024:
                            size_str = f"{size_bytes / 1024:.1f} KB"
                        elif size_bytes < 1024 * 1024 * 1024:
                            size_str = f"{size_bytes / (1024 * 1024):.1f} MB"
                        else:
                            size_str = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
                    
                    # Modified date
                    modified = datetime.fromtimestamp(stats.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Extension
                    extension = item_path.suffix.upper() if not is_dir and item_path.suffix else ""
                    
                    # Add row
                    ws.append([
                        level,
                        item_type,
                        name,
                        str(item_path),
                        size_str,
                        modified,
                        extension
                    ])
                    
                    # Style folder rows
                    if is_dir:
                        for col in range(1, 8):
                            cell = ws.cell(row=row_num, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    row_num += 1
                    
                    if idx % 10 == 0:
                        self.progress.emit(idx + 1, f"Processing... {idx + 1}/{len(items)}")
                        
                except (PermissionError, OSError):
                    continue
            
            # Add summary section
            row_num += 2
            ws.cell(row=row_num, column=1, value="Summary:")
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Files:")
            ws.cell(row=row_num, column=2, value=total_files)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Size:")
            if total_size < 1024 * 1024 * 1024:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024):.2f} MB")
            else:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024 * 1024):.2f} GB")
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Export Date:")
            ws.cell(row=row_num, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Source Path:")
            ws.cell(row=row_num, column=2, value=str(root))
            
            # Save workbook
            wb.save(self.output_path)
            
            self.finished.emit(self.output_path, True)
            
        except Exception as e:
            logger.error(f"Failed to export directory: {e}")
            self.finished.emit(str(e), False)


class PrintDirectoryDialog(QDialog):
    """Dialog for Print Directory options"""
    
    def __init__(self, current_path, parent=None):
        super().__init__(parent)
        self.current_path = current_path
        self.setWindowTitle("Print Directory to Excel")
        self.setModal(True)
        self.resize(500, 250)
        
        layout = QVBoxLayout(self)
        
        # Current path
        path_label = QLabel(f"<b>Directory:</b> {current_path}")
        path_label.setWordWrap(True)
        layout.addWidget(path_label)
        
        layout.addSpacing(10)
        
        # Options
        options_label = QLabel("<b>Options:</b>")
        layout.addWidget(options_label)
        
        # Include subdirectories checkbox
        self.include_subdirs_cb = QCheckBox("Include all subdirectories and files (recursive)")
        self.include_subdirs_cb.setChecked(True)
        layout.addWidget(self.include_subdirs_cb)
        
        info_label = QLabel("Note: Large directories may take a few moments to process.")
        info_label.setStyleSheet("color: gray; font-size: 9pt;")
        layout.addWidget(info_label)
        
        layout.addSpacing(10)
        
        # Output file name
        output_label = QLabel("<b>Output File Name:</b>")
        layout.addWidget(output_label)
        
        self.output_name = QLineEdit()
        default_name = f"Directory_{Path(current_path).name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.output_name.setText(default_name)
        layout.addWidget(self.output_name)
        
        layout.addStretch()
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_options(self):
        """Return selected options"""
        return {
            'include_subdirs': self.include_subdirs_cb.isChecked(),
            'output_name': self.output_name.text()
        }


class FileExplorerTab(QWidget):
    """Single file explorer tab"""
    
    # Signals
    path_changed = pyqtSignal(str)  # Emitted when path changes
    
    def __init__(self, initial_path=None):
        super().__init__()
        self.current_path = initial_path or str(Path.home())
        self.clipboard = {"path": None, "operation": None}
        self.selected_items = []
        
        self.init_ui()
        
        # Load initial path
        if initial_path:
            self.navigate_to(initial_path)
    
    def init_ui(self):
        """Initialize tab UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        
        # Breadcrumb navigation
        self.breadcrumb_widget = self.create_breadcrumb_bar()
        layout.addWidget(self.breadcrumb_widget)
        
        # Tree view
        self.tree_view = QTreeView()
        self.tree_view.setAnimated(True)
        self.tree_view.setIndentation(20)
        self.tree_view.setHeaderHidden(False)
        self.tree_view.setSelectionMode(QTreeView.SelectionMode.ExtendedSelection)  # Multi-select
        
        # Create model
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified'])
        
        self.tree_view.setModel(self.model)
        
        # Configure header
        header_view = self.tree_view.header()
        header_view.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        
        # Set column widths
        self.tree_view.setColumnWidth(0, 300)
        self.tree_view.setColumnWidth(1, 100)
        self.tree_view.setColumnWidth(2, 120)
        self.tree_view.setColumnWidth(3, 150)
        
        # Connect signals
        self.tree_view.expanded.connect(self.on_item_expanded)
        self.tree_view.clicked.connect(self.on_item_clicked)
        self.tree_view.doubleClicked.connect(self.on_item_double_clicked)
        
        # Context menu
        self.tree_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree_view.customContextMenuRequested.connect(self.show_context_menu)
        
        layout.addWidget(self.tree_view)
    
    def create_breadcrumb_bar(self):
        """Create breadcrumb navigation bar"""
        widget = QFrame()
        widget.setFrameShape(QFrame.Shape.StyledPanel)
        widget.setStyleSheet("""
            QFrame {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
        """)
        
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Home button
        home_btn = QPushButton("🏠")
        home_btn.setToolTip("Go to Home")
        home_btn.setFixedSize(30, 25)
        home_btn.clicked.connect(lambda: self.navigate_to(str(Path.home())))
        layout.addWidget(home_btn)
        
        # Breadcrumb path
        self.breadcrumb_label = QLabel()
        self.breadcrumb_label.setTextFormat(Qt.TextFormat.RichText)
        self.breadcrumb_label.setWordWrap(True)
        layout.addWidget(self.breadcrumb_label, 1)
        
        return widget
    
    def update_breadcrumb(self, path):
        """Update breadcrumb navigation"""
        path_obj = Path(path)
        parts = [path_obj.drive or '/'] + list(path_obj.relative_to(path_obj.anchor).parts)
        
        # Create clickable breadcrumb
        breadcrumb_html = ""
        current_path = path_obj.anchor
        
        for i, part in enumerate(parts):
            if i > 0:
                current_path = Path(current_path) / part
            
            # Make each part clickable (we'll use a simplified version for now)
            if i < len(parts) - 1:
                breadcrumb_html += f"<b>{part}</b> › "
            else:
                breadcrumb_html += f"<b style='color: #366092;'>{part}</b>"
        
        self.breadcrumb_label.setText(breadcrumb_html)
        self.current_path = str(path)
        self.path_changed.emit(str(path))
    
    def navigate_to(self, path):
        """Navigate to a specific path"""
        path_obj = Path(path)
        if not path_obj.exists():
            QMessageBox.warning(self, "Path Not Found", f"Path does not exist:\n{path}")
            return
        
        # Update breadcrumb
        self.update_breadcrumb(path)
        
        # Load directory contents
        self.load_directory(path_obj)
    
    def load_directory(self, dir_path):
        """Load directory contents into tree"""
        self.model.clear()
        self.model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified'])
        
        try:
            dir_path = Path(dir_path)
            items = sorted(dir_path.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower()))
            
            for item in items:
                try:
                    if item.is_dir():
                        row_items = self.create_folder_item(item)
                    else:
                        row_items = self.create_file_item(item)
                    
                    self.model.appendRow(row_items)
                except (PermissionError, OSError):
                    continue
                    
        except (PermissionError, OSError) as e:
            QMessageBox.warning(self, "Access Denied", f"Cannot access folder:\n{str(e)}")
    
    def create_folder_item(self, path, icon="📁"):
        """Create folder item (implementation continues...)"""
        # ... (rest of the implementation will follow)
        pass
    
    def create_file_item(self, path):
        """Create file item (implementation continues...)"""
        # ... (rest of the implementation will follow)
        pass
    
    def on_item_expanded(self, index):
        """Handle item expansion"""
        pass
    
    def on_item_clicked(self, index):
        """Handle item click"""
        pass
    
    def on_item_double_clicked(self, index):
        """Handle double click - navigate into folders or open files"""
        item = self.model.itemFromIndex(self.model.index(index.row(), 0, index.parent()))
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        
        path_obj = Path(path)
        
        if path_obj.is_dir():
            # Navigate into folder
            self.navigate_to(str(path_obj))
        elif path_obj.is_file():
            # Open file
            try:
                if os.name == 'nt':
                    os.startfile(str(path_obj))
                elif sys.platform == 'darwin':
                    subprocess.run(['open', str(path_obj)])
                else:
                    subprocess.run(['xdg-open', str(path_obj)])
            except Exception as e:
                logger.error(f"Failed to open file: {e}")
    
    def show_context_menu(self, position):
        """Show context menu"""
        # ... (implementation continues)
        pass


# Main FileExplorerV4 class will continue in next part...
