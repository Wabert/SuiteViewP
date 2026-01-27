"""
Enhanced File Explorer with Custom Model - OneDrive at Top Level
Shows Quick Access shortcuts (OneDrive) alongside system drives
"""

import os
import sys
import shutil
import subprocess
import json
import time
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (QTreeView, QVBoxLayout, QHBoxLayout, QWidget, 
                              QHeaderView, QToolBar, QMessageBox, QInputDialog, 
                              QTextEdit, QPushButton, QSplitter, QLabel, QMenu,
                              QDialog, QDialogButtonBox, QCheckBox, QLineEdit,
                              QProgressDialog, QFileDialog, QFrame, QSizePolicy,
                              QFileIconProvider, QToolButton, QStyle, QStyledItemDelegate,
                              QApplication, QComboBox)
from PyQt6.QtGui import QIcon, QAction, QStandardItemModel, QStandardItem, QDragEnterEvent, QDropEvent, QDragMoveEvent, QDrag
from PyQt6.QtCore import Qt, QModelIndex, QThread, pyqtSignal, QSortFilterProxyModel, QEvent, QFileInfo, QSize, QUrl, QMimeData, QRegularExpression

import logging

logger = logging.getLogger(__name__)


class FileSortProxyModel(QSortFilterProxyModel):
    """Custom sort proxy that uses UserRole+1 data for proper sorting"""
    
    def lessThan(self, left, right):
        """Compare items using custom sort data"""
        # Get the sort data (UserRole + 1) from both items
        left_data = self.sourceModel().data(left, Qt.ItemDataRole.UserRole + 1)
        right_data = self.sourceModel().data(right, Qt.ItemDataRole.UserRole + 1)
        
        # If both have sort data, use it
        if left_data is not None and right_data is not None:
            # Handle numeric comparison
            if isinstance(left_data, (int, float)) and isinstance(right_data, (int, float)):
                return left_data < right_data
            # Handle string comparison (includes folder/file prefix)
            return str(left_data) < str(right_data)
        
        # Fallback to display text
        left_text = self.sourceModel().data(left, Qt.ItemDataRole.DisplayRole)
        right_text = self.sourceModel().data(right, Qt.ItemDataRole.DisplayRole)
        
        if left_text is None:
            left_text = ""
        if right_text is None:
            right_text = ""
            
        return str(left_text).lower() < str(right_text).lower()


class NoFocusDelegate(QStyledItemDelegate):
    """Custom delegate that removes the focus rectangle from items"""
    
    def paint(self, painter, option, index):
        # Remove focus indicator - this removes the dotted focus rectangle
        option.state = option.state & ~QStyle.StateFlag.State_HasFocus
        super().paint(painter, option, index)


class DropTreeView(QTreeView):
    """Custom QTreeView that accepts file drops and supports dragging files out"""
    
    # Signal emitted when files are dropped (list of paths, destination folder)
    files_dropped = pyqtSignal(list, str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QTreeView.DragDropMode.DragDrop)
        self.setDropIndicatorShown(True)
        self.setDefaultDropAction(Qt.DropAction.CopyAction)
        self._file_explorer = None  # Reference to FileExplorerCore for getting folder paths
        self._current_folder = None  # Current folder being displayed
    
    def set_file_explorer(self, explorer):
        """Set reference to the file explorer for accessing current folder"""
        self._file_explorer = explorer
    
    def set_current_folder(self, folder_path):
        """Set the current folder path"""
        self._current_folder = folder_path
    
    def mouseDoubleClickEvent(self, event):
        """Override to prevent edit mode on double-click, but allow signal emission"""
        # Get the index at the click position
        index = self.indexAt(event.pos())
        if index.isValid():
            # Emit the doubleClicked signal manually
            self.doubleClicked.emit(index)
            # Accept the event to prevent further processing
            event.accept()
            return
        # If no valid index, call parent handler
        super().mouseDoubleClickEvent(event)
    
    def startDrag(self, supportedActions):
        """Start drag operation with selected files"""
        indexes = self.selectedIndexes()
        if not indexes:
            return
        
        # Get unique file paths from selected rows
        paths = []
        seen_rows = set()
        for index in indexes:
            row = index.row()
            if row not in seen_rows:
                seen_rows.add(row)
                # Get path from column 0
                col0_index = index.sibling(row, 0)
                path = self.model().data(col0_index, Qt.ItemDataRole.UserRole)
                if path and os.path.exists(path):
                    paths.append(path)
        
        if not paths:
            return
        
        # Create mime data with file URLs
        mime_data = QMimeData()
        urls = [QUrl.fromLocalFile(path) for path in paths]
        mime_data.setUrls(urls)
        
        # Create and execute drag
        drag = QDrag(self)
        drag.setMimeData(mime_data)
        drag.exec(Qt.DropAction.CopyAction)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        """Handle drag enter - accept if it contains file URLs"""
        if event.mimeData().hasUrls():
            # Check if any URLs are local files
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    event.acceptProposedAction()
                    return
        event.ignore()
    
    def dragMoveEvent(self, event: QDragMoveEvent):
        """Handle drag move - highlight target folder if hovering over one"""
        if event.mimeData().hasUrls():
            # Check if hovering over a folder item
            index = self.indexAt(event.position().toPoint())
            if index.isValid():
                # Get the path from the model
                path = self.model().data(index, Qt.ItemDataRole.UserRole)
                if not path:
                    # Try column 0
                    col0_index = index.sibling(index.row(), 0)
                    path = self.model().data(col0_index, Qt.ItemDataRole.UserRole)
                
                if path and os.path.isdir(path):
                    # Hovering over a folder - will drop into it
                    event.acceptProposedAction()
                    return
            
            # Not over a folder item - will drop into current folder
            if self._current_folder and os.path.isdir(self._current_folder):
                event.acceptProposedAction()
                return
        
        event.ignore()
    
    def dropEvent(self, event: QDropEvent):
        """Handle drop - copy files to target folder, but ignore same-view drops entirely"""
        if not event.mimeData().hasUrls():
            event.ignore()
            return
        
        # Get dropped file paths
        dropped_files = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                file_path = url.toLocalFile()
                if os.path.exists(file_path):
                    dropped_files.append(file_path)
        
        if not dropped_files:
            event.ignore()
            return
        
        # SAFETY: If dropping within the same current folder view, ignore completely
        # This prevents accidental copies/moves when dragging and dropping in details view
        if self._current_folder:
            current_folder_resolved = str(Path(self._current_folder).resolve()).lower()
            
            for dropped_path in dropped_files:
                dropped_parent = str(Path(dropped_path).resolve().parent).lower()
                
                # If the dropped item came from the current folder, ignore the drop entirely
                if dropped_parent == current_folder_resolved:
                    event.ignore()
                    return
        
        # Determine destination folder
        dest_folder = None
        index = self.indexAt(event.position().toPoint())
        
        if index.isValid():
            # Get the path from the model
            path = self.model().data(index, Qt.ItemDataRole.UserRole)
            if not path:
                # Try column 0
                col0_index = index.sibling(index.row(), 0)
                path = self.model().data(col0_index, Qt.ItemDataRole.UserRole)
            
            if path:
                if os.path.isdir(path):
                    dest_folder = path
                else:
                    # Dropped on a file - use its parent folder
                    dest_folder = str(Path(path).parent)
        
        # If no folder from drop target, use current folder
        if not dest_folder:
            dest_folder = self._current_folder
        
        if dest_folder and os.path.isdir(dest_folder):
            # Additional safety checks
            dest_folder_resolved = str(Path(dest_folder).resolve()).lower()
            
            for dropped_path in dropped_files:
                dropped_resolved = str(Path(dropped_path).resolve()).lower()
                dropped_parent = str(Path(dropped_path).resolve().parent).lower()
                
                # Case 1: Dropping item onto itself
                if dropped_resolved == dest_folder_resolved:
                    event.ignore()
                    return
                
                # Case 2: Item is already in the destination folder
                if dropped_parent == dest_folder_resolved:
                    event.ignore()
                    return
                
                # Case 3: Trying to drop a folder into itself (would cause recursion)
                if os.path.isdir(dropped_path):
                    if dest_folder_resolved.startswith(dropped_resolved + os.sep):
                        event.ignore()
                        return
            
            # Safe to proceed - emit signal with dropped files and destination
            self.files_dropped.emit(dropped_files, dest_folder)
            event.acceptProposedAction()
        else:
            event.ignore()


class DropFolderTreeView(QTreeView):
    """Custom QTreeView for folder navigation that accepts file drops"""
    
    # Signal emitted when files are dropped (list of paths, destination folder)
    files_dropped = pyqtSignal(list, str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragDropMode(QTreeView.DragDropMode.DropOnly)
        self.setDropIndicatorShown(True)
        self._file_explorer = None
    
    def set_file_explorer(self, explorer):
        """Set reference to the file explorer for handling drops"""
        self._file_explorer = explorer
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        """Handle drag enter - accept if it contains file URLs"""
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    event.acceptProposedAction()
                    return
        event.ignore()
    
    def dragMoveEvent(self, event: QDragMoveEvent):
        """Handle drag move - only accept if over a folder"""
        if event.mimeData().hasUrls():
            index = self.indexAt(event.position().toPoint())
            if index.isValid():
                # Get the path - try UserRole first
                path = self.model().data(index, Qt.ItemDataRole.UserRole)
                if path and os.path.isdir(path):
                    event.acceptProposedAction()
                    return
        event.ignore()
    
    def dropEvent(self, event: QDropEvent):
        """Handle drop - copy files to target folder, but ignore same-folder drops"""
        if not event.mimeData().hasUrls():
            event.ignore()
            return
        
        # Get dropped file paths
        dropped_files = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                file_path = url.toLocalFile()
                if os.path.exists(file_path):
                    dropped_files.append(file_path)
        
        if not dropped_files:
            event.ignore()
            return
        
        # Get destination folder from drop target
        index = self.indexAt(event.position().toPoint())
        if not index.isValid():
            event.ignore()
            return
        
        path = self.model().data(index, Qt.ItemDataRole.UserRole)
        if not path or not os.path.isdir(path):
            event.ignore()
            return
        
        dest_folder = path
        
        # SAFETY CHECK: Prevent same-folder drops to avoid accidental moves/copies
        dest_folder_resolved = str(Path(dest_folder).resolve()).lower()
        
        for dropped_path in dropped_files:
            dropped_resolved = str(Path(dropped_path).resolve()).lower()
            dropped_parent = str(Path(dropped_path).resolve().parent).lower()
            
            # Case 1: Dropping item onto itself
            if dropped_resolved == dest_folder_resolved:
                event.ignore()
                return
            
            # Case 2: Item is already in the destination folder
            if dropped_parent == dest_folder_resolved:
                event.ignore()
                return
            
            # Case 3: Trying to drop a folder into itself (would cause recursion)
            if os.path.isdir(dropped_path):
                if dest_folder_resolved.startswith(dropped_resolved + os.sep):
                    event.ignore()
                    return
        
        # Safe to proceed - emit signal with dropped files and destination
        self.files_dropped.emit(dropped_files, dest_folder)
        event.acceptProposedAction()


class DepthScanWorker(QThread):
    """Background thread for scanning folders at specified depth"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(list)
    
    def __init__(self, root_path, depth_level):
        super().__init__()
        self.root_path = root_path
        self.depth_level = depth_level
        self._cancelled = False
        
    def cancel(self):
        """Request cancellation of the scan"""
        self._cancelled = True
        
    def run(self):
        """Scan folders up to specified depth and return results"""
        results = []
        
        try:
            root = Path(self.root_path)
            
            # Scan folders recursively up to depth_level
            self._scan_folder(root, "", 0, results)
            
        except Exception as e:
            logger.error(f"Error during depth scan: {e}")
        
        # Emit results
        self.finished.emit(results)
    
    def _scan_folder(self, folder_path: Path, relative_path: str, current_depth: int, results: list):
        """Recursively scan folder up to specified depth"""
        # Check if cancelled
        if self._cancelled:
            return
        
        # If depth_level is -1 (Max), scan everything; otherwise check depth limit
        if self.depth_level != -1 and current_depth >= self.depth_level:
            return
        
        try:
            # Get all items in this folder
            with os.scandir(str(folder_path)) as entries:
                folders = []
                files = []
                
                for entry in entries:
                    try:
                        if entry.name.startswith('.'):
                            continue
                        
                        is_dir = entry.is_dir()
                        item_path = Path(entry.path)
                        
                        # Build display name with pipe delimiters
                        if relative_path:
                            display_name = f"{relative_path} | {entry.name}"
                        else:
                            display_name = entry.name
                        
                        # Get file stats (if not network drive)
                        size = 0
                        modified = ""
                        accessed = ""
                        
                        try:
                            stat_info = entry.stat()
                            size = stat_info.st_size if not is_dir else 0
                            modified = datetime.fromtimestamp(stat_info.st_mtime).strftime("%Y-%m-%d %H:%M")
                            accessed = datetime.fromtimestamp(stat_info.st_atime).strftime("%Y-%m-%d %H:%M")
                        except:
                            pass
                        
                        # Add to results
                        item_data = {
                            'path': str(item_path),
                            'display_name': display_name,
                            'is_dir': is_dir,
                            'depth': current_depth + 1,
                            'size': size,
                            'modified': modified,
                            'accessed': accessed
                        }
                        
                        if is_dir:
                            folders.append((item_path, display_name))
                        
                        results.append(item_data)
                        
                        # Emit progress every 50 items
                        if len(results) % 50 == 0:
                            self.progress.emit(len(results), f"Scanning depth {current_depth + 1}...")
                        
                    except (PermissionError, OSError):
                        continue
                
                # Recursively scan subfolders if we haven't reached max depth (or if Max mode)
                if self.depth_level == -1 or current_depth + 1 < self.depth_level:
                    for subfolder_path, subfolder_display in folders:
                        self._scan_folder(subfolder_path, subfolder_display, current_depth + 1, results)
                        
        except (PermissionError, OSError):
            pass


class DirectoryExportThread(QThread):
    """Background thread for exporting directory to Excel"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(str, bool)
    
    def __init__(self, root_path, include_subdirs, output_path):
        super().__init__()
        self.root_path = root_path
        self.include_subdirs = include_subdirs
        self.output_path = output_path
        
    def run(self):
        """Export directory structure to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Directory Contents"
            
            # Headers
            headers = ['Level', 'Type', 'Name', 'Full Path', 'Size', 'Modified', 'Extension']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions['A'].width = 8   # Level
            ws.column_dimensions['B'].width = 10  # Type
            ws.column_dimensions['C'].width = 40  # Name
            ws.column_dimensions['D'].width = 60  # Full Path
            ws.column_dimensions['E'].width = 12  # Size
            ws.column_dimensions['F'].width = 18  # Modified
            ws.column_dimensions['G'].width = 10  # Extension
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            row_num = 2
            total_files = 0
            total_size = 0
            
            # Walk directory
            root = Path(self.root_path)
            
            if self.include_subdirs:
                items = []
                for dirpath, dirnames, filenames in os.walk(root):
                    current = Path(dirpath)
                    level = len(current.relative_to(root).parts)
                    
                    # Add directories
                    for dirname in sorted(dirnames):
                        items.append((current / dirname, level, True))
                    
                    # Add files
                    for filename in sorted(filenames):
                        items.append((current / filename, level, False))
                        
                    self.progress.emit(len(items), f"Scanning... {len(items)} items found")
            else:
                # Just immediate contents
                items = []
                for item in sorted(root.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                    items.append((item, 0, item.is_dir()))
            
            # Add items to Excel
            for idx, (item_path, level, is_dir) in enumerate(items):
                try:
                    stats = item_path.stat()
                    
                    # Type
                    item_type = "Folder" if is_dir else "File"
                    
                    # Name with indentation
                    indent = "  " * level
                    name = f"{indent}{item_path.name}"
                    
                    # Size
                    if is_dir:
                        size_str = ""
                    else:
                        size_bytes = stats.st_size
                        total_size += size_bytes
                        total_files += 1
                        
                        if size_bytes < 1024:
                            size_str = f"{size_bytes} B"
                        elif size_bytes < 1024 * 1024:
                            size_str = f"{size_bytes / 1024:.1f} KB"
                        elif size_bytes < 1024 * 1024 * 1024:
                            size_str = f"{size_bytes / (1024 * 1024):.1f} MB"
                        else:
                            size_str = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
                    
                    # Modified date
                    modified = datetime.fromtimestamp(stats.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Extension
                    extension = item_path.suffix.upper() if not is_dir and item_path.suffix else ""
                    
                    # Add row
                    ws.append([
                        level,
                        item_type,
                        name,
                        str(item_path),
                        size_str,
                        modified,
                        extension
                    ])
                    
                    # Style folder rows
                    if is_dir:
                        for col in range(1, 8):
                            cell = ws.cell(row=row_num, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    row_num += 1
                    
                    if idx % 10 == 0:
                        self.progress.emit(idx + 1, f"Processing... {idx + 1}/{len(items)}")
                        
                except (PermissionError, OSError):
                    continue
            
            # Add summary section
            row_num += 2
            ws.cell(row=row_num, column=1, value="Summary:")
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Files:")
            ws.cell(row=row_num, column=2, value=total_files)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Size:")
            if total_size < 1024 * 1024 * 1024:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024):.2f} MB")
            else:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024 * 1024):.2f} GB")
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Export Date:")
            ws.cell(row=row_num, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Source Path:")
            ws.cell(row=row_num, column=2, value=str(root))
            
            # Save workbook
            wb.save(self.output_path)
            
            self.finished.emit(self.output_path, True)
            
        except Exception as e:
            logger.error(f"Failed to export directory: {e}")
            self.finished.emit(str(e), False)


class PrintDirectoryDialog(QDialog):
    """Dialog for Print Directory options"""
    
    def __init__(self, current_path, parent=None):
        super().__init__(parent)
        self.current_path = current_path
        self.setWindowTitle("Print Directory to Excel")
        self.setModal(True)
        self.resize(550, 220)
        
        layout = QVBoxLayout(self)
        
        # Current path
        path_label = QLabel(f"<b>Directory:</b> {current_path}")
        path_label.setWordWrap(True)
        path_label.setStyleSheet("padding: 10px; background-color: #f0f0f0; border-radius: 5px;")
        layout.addWidget(path_label)
        
        layout.addSpacing(15)
        
        # Options
        options_label = QLabel("<b>Export Options:</b>")
        layout.addWidget(options_label)
        
        # Include subdirectories checkbox
        self.include_subdirs_cb = QCheckBox("Include all subdirectories and files (recursive)")
        # Default OFF: recursive exports can be extremely large/slow on network paths
        self.include_subdirs_cb.setChecked(False)
        layout.addWidget(self.include_subdirs_cb)
        
        info_label = QLabel("📝 The directory listing will open directly in Excel as an unsaved workbook.")
        info_label.setStyleSheet("color: #666; font-size: 9pt; padding-left: 20px;")
        layout.addWidget(info_label)
        
        layout.addSpacing(10)
        
        note_label = QLabel("💡 Use 'Save As' in Excel if you want to keep the file.")
        note_label.setStyleSheet("color: #0066cc; font-size: 9pt; padding-left: 20px;")
        layout.addWidget(note_label)
        
        layout.addStretch()
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_options(self):
        """Return selected options"""
        return {
            'include_subdirs': self.include_subdirs_cb.isChecked()
        }


class FileExplorerCore(QWidget):
    """
    Core File Explorer with custom model showing OneDrive at top level
    Features:
    - OneDrive shortcuts at root level (like Windows Explorer)
    - System drives (C: D: etc.)
    - Lazy loading of directory contents
    - File operations (cut, copy, paste, rename, delete)
    - File preview pane with mainframe upload button
    """
    
    # Icon cache by extension - loaded once, reused for all files
    _icon_cache = {}
    _folder_icon = None
    
    # Extension to icon type mapping for fast lookups
    ICON_TYPES = {
        # Documents
        '.xlsx': 'excel', '.xls': 'excel', '.xlsm': 'excel', '.xlsb': 'excel',
        '.docx': 'word', '.doc': 'word', '.docm': 'word',
        '.pptx': 'powerpoint', '.ppt': 'powerpoint', '.pptm': 'powerpoint',
        '.pdf': 'pdf',
        '.txt': 'text', '.log': 'text', '.md': 'text', '.csv': 'text',
        # Data
        '.accdb': 'database', '.mdb': 'database', '.db': 'database', '.sqlite': 'database',
        '.laccdb': 'database',
        # Code
        '.py': 'code', '.js': 'code', '.java': 'code', '.cpp': 'code', '.c': 'code',
        '.html': 'code', '.css': 'code', '.json': 'code', '.xml': 'code',
        # Images
        '.jpg': 'image', '.jpeg': 'image', '.png': 'image', '.gif': 'image',
        '.bmp': 'image', '.ico': 'image', '.svg': 'image',
        # Archives
        '.zip': 'archive', '.rar': 'archive', '.7z': 'archive', '.tar': 'archive', '.gz': 'archive',
        # Executables
        '.exe': 'exe', '.msi': 'exe', '.bat': 'exe', '.cmd': 'exe', '.ps1': 'exe',
        # Shortcuts
        '.lnk': 'shortcut', '.url': 'shortcut',
    }
    
    def __init__(self):
        super().__init__()
        
        # Initialize icon provider for Windows system icons (used sparingly)
        self.icon_provider = QFileIconProvider()
        
        # Pre-cache common icons on first instance
        self._init_icon_cache()
        
        self.current_file_path = None
        self.current_file_content = None
        self.current_details_folder = None  # Track current folder in details view
        self.clipboard = {"paths": [], "operation": None}
        
        # Depth search feature
        self.depth_search_enabled = False
        self.depth_search_cache = {}  # Cache: {folder_path: {depth_level: [items]}}
        self.depth_search_folder = None  # Folder where depth search was initiated
        self.depth_search_locked = False  # True when depth search is active and locked
        self.depth_search_active_results = None  # Currently displayed depth results
        
        # Folder-specific search terms
        self.folder_search_terms = {}  # Cache: {folder_path: search_text}
        
        # Load custom quick links via centralized bookmark manager
        # Bar ID 1 = sidebar (by convention)
        from suiteview.ui.widgets.bookmark_data_manager import get_bookmark_manager
        self._bookmark_manager = get_bookmark_manager()
        self.custom_quick_links = self._bookmark_manager.get_bar_data(1)
        
        # Load hidden OneDrive paths
        self.hidden_onedrive_file = Path.home() / ".suiteview" / "hidden_onedrive.json"
        self.hidden_onedrive_paths = self.load_hidden_onedrive()
        
        # Column width settings file
        self.column_widths_file = Path.home() / ".suiteview" / "column_widths.json"
        self.column_widths = self.load_column_widths()
        
        # Panel widths persistence
        self.panel_widths_file = Path.home() / '.suiteview' / 'file_explorer_panel_widths.json'
        self.panel_widths = self.load_panel_widths()
        
        self.init_ui()
    
    def _init_icon_cache(self):
        """Initialize icon cache with common file type icons"""
        if FileExplorerCore._folder_icon is not None:
            return  # Already initialized
        
        # Get folder icon once
        FileExplorerCore._folder_icon = self.icon_provider.icon(QFileIconProvider.IconType.Folder)
        
        # Get standard icons from style for common types
        style = self.style()
        
        # Cache file icon as default
        FileExplorerCore._icon_cache['_default'] = self.icon_provider.icon(QFileIconProvider.IconType.File)
    
    def get_emoji_icon_for_path(self, path):
        """Get emoji icon based on file type - for Quick Links panel"""
        path = Path(path) if isinstance(path, str) else path
        
        if path.is_dir():
            # Check for special folder names
            folder_name = path.name.lower()
            if "onedrive" in folder_name:
                return "☁️"
            elif "desktop" in folder_name:
                return "🖥️"
            elif "documents" in folder_name:
                return "📄"
            elif "downloads" in folder_name:
                return "⬇️"
            elif "pictures" in folder_name or "photos" in folder_name:
                return "🖼️"
            elif "music" in folder_name:
                return "🎵"
            elif "videos" in folder_name:
                return "🎬"
            return "📁"
        
        # It's a file - get icon based on extension
        suffix = path.suffix.lower()
        
        # Map extensions to emoji icons
        emoji_map = {
            # Excel
            '.xlsx': '📊', '.xls': '📊', '.xlsm': '📊', '.xlsb': '📊', '.csv': '📊',
            # Word
            '.docx': '📝', '.doc': '📝', '.docm': '📝', '.rtf': '📝',
            # PowerPoint
            '.pptx': '📽️', '.ppt': '📽️', '.pptm': '📽️',
            # PDF
            '.pdf': '📕',
            # Database / Access
            '.accdb': '🗃️', '.mdb': '🗃️', '.db': '🗃️', '.sqlite': '🗃️', '.laccdb': '🗃️',
            # Text
            '.txt': '📄', '.log': '📄', '.md': '📄',
            # Code
            '.py': '🐍', '.js': '📜', '.java': '☕', '.cpp': '⚙️', '.c': '⚙️',
            '.html': '🌐', '.css': '🎨', '.json': '📋', '.xml': '📋',
            # Images
            '.jpg': '🖼️', '.jpeg': '🖼️', '.png': '🖼️', '.gif': '🖼️',
            '.bmp': '🖼️', '.ico': '🖼️', '.svg': '🖼️',
            # Archives
            '.zip': '📦', '.rar': '📦', '.7z': '📦', '.tar': '📦', '.gz': '📦',
            # Executables
            '.exe': '⚙️', '.msi': '⚙️', '.bat': '⚙️', '.cmd': '⚙️', '.ps1': '⚙️',
            # Shortcuts
            '.lnk': '🔗', '.url': '🔗',
        }
        
        return emoji_map.get(suffix, '📄')
    
    def _get_cached_icon(self, path: Path, is_directory: bool = False):
        """Get icon from cache or create it - fast path for common extensions"""
        if is_directory:
            return FileExplorerCore._folder_icon
        
        suffix = path.suffix.lower()
        
        # Check cache first
        if suffix in FileExplorerCore._icon_cache:
            return FileExplorerCore._icon_cache[suffix]
        
        # For local files, get the actual icon and cache it
        path_str = str(path)
        is_network = path_str.startswith('\\\\')
        
        if not is_network:
            # Local file - get real icon and cache by extension
            try:
                # If file exists, get its icon directly
                if path.exists():
                    file_info = QFileInfo(path_str)
                    icon = self.icon_provider.icon(file_info)
                    FileExplorerCore._icon_cache[suffix] = icon
                    return icon
                else:
                    # File doesn't exist - try to find another file with same extension
                    # to get the icon from Windows shell
                    import tempfile
                    temp_file = Path(tempfile.gettempdir()) / f"_icon_temp{suffix}"
                    try:
                        temp_file.touch()
                        file_info = QFileInfo(str(temp_file))
                        icon = self.icon_provider.icon(file_info)
                        FileExplorerCore._icon_cache[suffix] = icon
                        temp_file.unlink()
                        return icon
                    except:
                        if temp_file.exists():
                            temp_file.unlink()
            except:
                pass
        
        # Network file or error - use default file icon
        return FileExplorerCore._icon_cache.get('_default', FileExplorerCore._folder_icon)
        
    def init_ui(self):
        """Initialize the UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Create toolbar
        toolbar = self.create_toolbar()
        main_layout.addWidget(toolbar)
        
        # Create bookmark bar (browser-style)
        from suiteview.ui.dialogs.shortcuts_dialog import BookmarkBar
        self.bookmark_bar = BookmarkBar(self)
        self.bookmark_bar.navigate_to_path.connect(self.navigate_to_bookmark_folder)
        main_layout.addWidget(self.bookmark_bar)
        
        # Create splitter for tree (left) and details (right)
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.main_splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #A0B8D8;
                width: 2px;
            }
            QSplitter::handle:hover {
                background-color: #7DA3CC;
            }
        """)
        
        # Create tree panel (left side - folder navigation only)
        tree_panel = self.create_tree_panel()
        self.main_splitter.addWidget(tree_panel)
        
        # Create details panel (right side - folder contents with details)
        details_panel = self.create_details_panel()
        self.main_splitter.addWidget(details_panel)
        
        # Set initial sizes from saved values or defaults (30% tree, 70% details)
        saved_left = self.panel_widths.get('left_panel', 300)
        saved_middle = self.panel_widths.get('middle_panel', 700)
        self.main_splitter.setSizes([saved_left, saved_middle])
        
        # Connect splitter moved signal to save panel widths
        self.main_splitter.splitterMoved.connect(self.on_splitter_moved)
        
        main_layout.addWidget(self.main_splitter)
        
    def create_toolbar(self):
        """Create toolbar with file operations"""
        self.toolbar = QToolBar()
        self.toolbar.setMovable(False)
        self._apply_compact_toolbar_style(self.toolbar)
        
        # Add Bookmark (star icon) - first item on toolbar
        add_bookmark_action = QAction("⭐ Add Bookmark", self)
        add_bookmark_action.setToolTip("Add bookmark (Ctrl+D)")
        add_bookmark_action.triggered.connect(self._add_bookmark)
        self.toolbar.addAction(add_bookmark_action)
        
        # Print Directory to Excel
        print_dir_action = QAction("Print Directory", self)
        print_dir_action.setToolTip("Export directory structure to Excel")
        print_dir_action.triggered.connect(self.print_directory_to_excel)
        self.toolbar.addAction(print_dir_action)
        
        # Batch Rename
        batch_rename_action = QAction("Batch Rename", self)
        batch_rename_action.setToolTip("Rename multiple selected files")
        batch_rename_action.triggered.connect(self.batch_rename_files)
        self.toolbar.addAction(batch_rename_action)
        
        # Add spacer to push "Open in Explorer" to the right
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        spacer.setStyleSheet("background: transparent;")
        self.toolbar.addWidget(spacer)
        
        # Open in Explorer (moved to far right)
        explorer_action = QAction("📂 Open in Explorer", self)
        explorer_action.triggered.connect(self.open_in_explorer)
        self.toolbar.addAction(explorer_action)
        
        return self.toolbar

    def _add_bookmark(self):
        """Show add bookmark dialog (called from toolbar button)"""
        if hasattr(self, 'bookmark_bar'):
            self.bookmark_bar.add_bookmark()

    def _apply_compact_toolbar_style(self, toolbar: QToolBar, locked: bool = False) -> None:
        """Apply toolbar styling with optional orange background when locked."""

        toolbar.setObjectName("fileExplorerToolbar")
        toolbar.setIconSize(QSize(16, 16))
        toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextOnly)
        
        bg_color = "#FFB366" if locked else "#E3EDFF"  # Orange when locked, blue otherwise
        
        toolbar.setStyleSheet(
            f"""
            QToolBar#fileExplorerToolbar {{
                padding: 0px 6px;
                spacing: 6px;
                min-height: 26px;
                background: {bg_color};
                border: none;
            }}
            QToolBar#fileExplorerToolbar QToolButton {{
                padding: 3px 10px;
                border: 1px solid #4A6FA5;
                border-bottom: 2px solid #3A5A8A;
                border-radius: 4px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #FFFFFF,
                                            stop:0.45 #F0F5FF,
                                            stop:1 #D0E3FF);
                color: #0A1E5E;
                font-weight: 600;
                font-size: 10px;
            }}
            QToolBar#fileExplorerToolbar QToolButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #FFFFFF,
                                            stop:0.35 #E3EDFF,
                                            stop:1 #B8D0F0);
                border-color: #2563EB;
            }}
            QToolBar#fileExplorerToolbar QToolButton:pressed {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #B8D0F0,
                                            stop:1 #E3EDFF);
                border: 1px solid #3A5A8A;
                border-top: 2px solid #3A5A8A;
            }}
            """
        )

    def _apply_depth_search_locked_style(self, locked: bool = False) -> None:
        """Apply locked/unlocked style for depth search.
        
        Base implementation does nothing. FileExplorerTab overrides this
        to change the breadcrumb bar color when depth search is locked.
        """
        pass

    def _create_nav_button(self, icon: QIcon, tooltip: str, handler) -> QToolButton:
        """Build a breadcrumb-nav button with a clear icon."""

        button = QToolButton()
        button.setAutoRaise(True)
        button.setIcon(icon)
        button.setIconSize(QSize(16, 16))
        button.setToolTip(tooltip)
        button.setCursor(Qt.CursorShape.PointingHandCursor)
        button.setStyleSheet(
            """
            QToolButton {
                border: 1px solid #2563EB;
                border-radius: 4px;
                background-color: #E0ECFF;
                padding: 2px;
            }
            QToolButton:hover {
                background-color: #C9DAFF;
            }
            """
        )
        button.clicked.connect(handler)
        return button
        
    def create_tree_panel(self):
        """Create the tree view with custom model (folders only, name column only)"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header
        header = QLabel("📁 Folders")
        header.setStyleSheet(
            """
            QLabel {
                padding: 4px 8px;
                font-weight: 600;
                font-size: 10pt;
                background-color: #C0D4F0;
                color: #1A3A6E;
                border: none;
                border-bottom: 1px solid #A0B8D8;
            }
            """
        )
        layout.addWidget(header)
        
        # Create tree view with drop support
        self.tree_view = DropFolderTreeView()
        self.tree_view.set_file_explorer(self)
        self.tree_view.files_dropped.connect(self.handle_dropped_files)
        self.tree_view.setAnimated(True)
        self.tree_view.setIndentation(20)
        self.tree_view.setHeaderHidden(True)  # Hide header for tree view
        self.tree_view.setSelectionMode(QTreeView.SelectionMode.SingleSelection)
        self.tree_view.setStyleSheet(
            """
            QTreeView {
                outline: none;
                border: none;
                background-color: transparent;
            }
            QTreeView::item {
                padding: 2px 6px;
                margin: 0px;
                border: none;
                border-radius: 0px;
                background-color: transparent;
                min-height: 20px;
            }
            QTreeView::item:hover {
                background-color: #C8DCF0;
                border: none;
            }
            QTreeView::item:selected {
                background-color: #B0C8E8;
                color: #0A1E5E;
                border: none;
            }
            QTreeView::item:selected:!active {
                background-color: #B0C8E8;
                color: #0A1E5E;
                border: none;
            }
            QTreeView::item:focus {
                border: none;
                outline: none;
            }
            QTreeView::branch {
                background-color: transparent;
                border-image: none;
                image: none;
            }
            QTreeView::branch:selected {
                background-color: #B0C8E8;
            }
            QTreeView::branch:has-children:!has-siblings:closed,
            QTreeView::branch:closed:has-children:has-siblings {
                border-image: none;
                image: none;
                background: qradialgradient(cx:0.5, cy:0.5, radius:0.3, fx:0.5, fy:0.5, stop:0 #0078d4, stop:0.7 #0078d4, stop:0.71 transparent);
            }
            QTreeView::branch:has-children:!has-siblings:open,
            QTreeView::branch:open:has-children:has-siblings {
                border-image: none;
                image: none;
                background: qradialgradient(cx:0.5, cy:0.5, radius:0.3, fx:0.5, fy:0.5, stop:0 #0078d4, stop:0.7 #0078d4, stop:0.71 transparent);
            }
            """
        )
        
        # Apply no-focus delegate to remove focus rectangle
        self.tree_view.setItemDelegate(NoFocusDelegate(self.tree_view))
        
        # Create custom model
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(['Name'])  # Only name column for tree
        self.model.setColumnCount(1)  # Explicitly set to 1 column only
        
        # Populate model with Quick Access and drives
        self.populate_tree_model()
        
        self.tree_view.setModel(self.model)
        
        # Ensure only the Name column is visible - hide any extra columns
        for col in range(1, 10):  # Hide columns 1-9 if they somehow exist
            self.tree_view.setColumnHidden(col, True)
        
        # Set the tree to only resize the first column
        self.tree_view.header().setStretchLastSection(False)
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        
        # Connect signals
        self.tree_view.expanded.connect(self.on_item_expanded)
        self.tree_view.clicked.connect(self.on_tree_item_clicked)
        
        # Context menu
        self.tree_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree_view.customContextMenuRequested.connect(self.show_tree_context_menu)
        
        layout.addWidget(self.tree_view)
        
        # Add footer
        footer = QLabel("")
        footer.setStyleSheet("""
            QLabel {
                background-color: #E0E0E0;
                padding: 2px 8px;
                font-size: 9pt;
                color: #555555;
                border: none;
                border-top: 1px solid #A0B8D8;
            }
        """)
        footer.setFixedHeight(20)
        layout.addWidget(footer)
        self.tree_footer = footer
        
        return widget
        
    def populate_tree_model(self):
        """Populate tree with OneDrive and system drives only (no quick links - those go in right panel)"""
        self.model.clear()
        self.model.setHorizontalHeaderLabels(['Name'])
        self.model.setColumnCount(1)  # Explicitly set to 1 column only
        
        # Add OneDrive folders (excluding hidden ones)
        onedrive_paths = self.get_onedrive_paths()
        for od_path in onedrive_paths:
            if od_path.exists() and str(od_path).lower() not in self.hidden_onedrive_paths:
                item = self.create_tree_folder_item(od_path, icon="⭐")
                self.model.appendRow(item)
        
        # Add separator (visual only) if we have OneDrive paths
        if onedrive_paths:
            separator = QStandardItem("─" * 30)
            separator.setEnabled(False)
            self.model.appendRow(separator)
        
        # Add system drives
        drives = self.get_system_drives()
        for drive in drives:
            item = self.create_tree_folder_item(Path(drive), icon="💾")
            self.model.appendRow(item)
        
        # Re-hide any extra columns and ensure proper column sizing
        for col in range(1, 10):
            self.tree_view.setColumnHidden(col, True)
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
    
    def populate_quick_links_model(self, model):
        """Populate a model with quick links only (for the right panel)"""
        model.clear()
        model.setHorizontalHeaderLabels(['Name'])
        model.setColumnCount(1)
        
        # Add custom quick links with system icons (same as details panel)
        for link_path in self.custom_quick_links:
            path = Path(link_path)
            if path.exists():
                item = self.create_quick_link_item(path)
                # Mark as custom link
                item.setData("__QUICK_LINK__", Qt.ItemDataRole.UserRole + 1)
                model.appendRow(item)
    
    def create_quick_link_item(self, path):
        """Create a tree item for a quick link (file or folder) with system icon"""
        path = Path(path)
        
        item = QStandardItem(path.name)
        item.setData(str(path), Qt.ItemDataRole.UserRole)  # Store path
        item.setEditable(False)
        item.setToolTip(str(path))  # Show full path on hover
        
        # Use system icon (same as details panel)
        is_dir = path.is_dir()
        icon = self._get_cached_icon(path, is_dir)
        item.setIcon(icon)
        
        # If it's a directory, add a placeholder for expansion
        if is_dir:
            placeholder = QStandardItem("Loading...")
            placeholder.setEnabled(False)
            item.appendRow(placeholder)
        
        return item
            
    def get_onedrive_paths(self):
        """Get all OneDrive paths (deduplicated, prefer business OneDrive)"""
        onedrive_paths = []
        seen_paths = set()
        has_business_onedrive = False
        
        # Check environment variables - prioritize business OneDrive
        for env_var in ['OneDriveCommercial', 'OneDrive', 'OneDriveConsumer']:
            path = os.environ.get(env_var)
            if path and os.path.exists(path):
                path_obj = Path(path).resolve()  # Resolve to absolute path
                path_str = str(path_obj).lower()  # Normalize for comparison
                if path_str not in seen_paths:
                    seen_paths.add(path_str)
                    onedrive_paths.append(path_obj)
                    # Check if this is a business OneDrive (contains company name)
                    if ' - ' in path_obj.name:
                        has_business_onedrive = True
        
        # Check common locations - only if not already found via environment variables
        home = Path.home()
        
        # First check for business OneDrive with company name pattern
        business_onedrive_found = False
        for item in home.iterdir():
            if item.is_dir() and item.name.startswith("OneDrive - "):
                path_resolved = item.resolve()
                path_str = str(path_resolved).lower()
                if path_str not in seen_paths:
                    seen_paths.add(path_str)
                    onedrive_paths.append(path_resolved)
                    business_onedrive_found = True
                    has_business_onedrive = True
        
        # Only add generic "OneDrive" if no business OneDrive was found
        if not has_business_onedrive:
            generic_onedrive = home / "OneDrive"
            if generic_onedrive.exists():
                path_resolved = generic_onedrive.resolve()
                path_str = str(path_resolved).lower()
                if path_str not in seen_paths:
                    seen_paths.add(path_str)
                    onedrive_paths.append(path_resolved)
        
        return onedrive_paths
        
    def get_system_drives(self):
        """Get all system drives"""
        drives = []
        
        if os.name == 'nt':  # Windows
            import string
            from ctypes import windll
            
            bitmask = windll.kernel32.GetLogicalDrives()
            for letter in string.ascii_uppercase:
                if bitmask & 1:
                    drive_path = f"{letter}:\\"
                    if os.path.exists(drive_path):
                        drives.append(drive_path)
                bitmask >>= 1
        else:  # Unix-like
            drives.append("/")
        
        return drives
        
    def create_tree_folder_item(self, path, icon="📁"):
        """Create a single-column folder item for the tree view"""
        path = Path(path)
        
        # Use more descriptive icons for special folders
        folder_name = path.name.lower() if path.name else ""
        
        # Special folder icons
        if not icon or icon == "📁":  # Only override if default folder icon
            if "desktop" in folder_name:
                icon = "🖥️"
            elif "documents" in folder_name:
                icon = "📄"
            elif "downloads" in folder_name:
                icon = "⬇️"
            elif "pictures" in folder_name or "photos" in folder_name:
                icon = "🖼️"
            elif "music" in folder_name:
                icon = "🎵"
            elif "videos" in folder_name:
                icon = "🎬"
            elif "onedrive" in folder_name:
                icon = "☁️"
            elif folder_name in ["program files", "program files (x86)"]:
                icon = "⚙️"
            elif folder_name == "windows":
                icon = "🪟"
            elif folder_name == "users":
                icon = "👥"
            elif ".git" in folder_name:
                icon = "🔀"
            elif "project" in folder_name or "code" in folder_name:
                icon = "💻"
            else:
                icon = "📁"  # Default folder
        
        # Name column with icon
        name_item = QStandardItem(f"{icon} {path.name if path.name else str(path)}")
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(False)
        
        # Add placeholder child to make it expandable
        name_item.appendRow(QStandardItem("Loading..."))
        
        return name_item
    
    def create_folder_item(self, path, icon=None):
        """Create a row of items for a folder"""
        path = Path(path)
        
        # Get cached folder icon (fast - no network round-trips)
        system_icon = self._get_cached_icon(path, is_directory=True)
        
        # Name column with system icon
        name_item = QStandardItem(system_icon, path.name if path.name else str(path))
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(True)  # Allow editing for F2 rename
        # Store sort data: 0 = folder (sorts first), name in lowercase for case-insensitive sort
        name_item.setData(f"0_{(path.name if path.name else str(path)).lower()}", Qt.ItemDataRole.UserRole + 1)
        
        # Size column (empty for folders)
        size_item = QStandardItem("")
        size_item.setEditable(False)
        size_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)  # Right-align
        size_item.setData(0, Qt.ItemDataRole.UserRole + 1)  # Sort value for empty size
        
        # Type column
        type_item = QStandardItem("Folder")
        type_item.setEditable(False)
        
        # Date modified - skip on network drives for speed
        try:
            # Check if network path
            if str(path).startswith('\\\\'):
                date_str = ""  # Skip date on network drives
                mtime = 0
            else:
                mtime = path.stat().st_mtime
                date_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        except:
            date_str = ""
            mtime = 0
        date_item = QStandardItem(date_str)
        date_item.setEditable(False)
        # Store timestamp for proper sorting
        date_item.setData(mtime, Qt.ItemDataRole.UserRole + 1)
        
        # Date accessed
        try:
            if str(path).startswith('\\\\'):
                adate_str = ""  # Skip date on network drives
                atime = 0
            else:
                atime = path.stat().st_atime
                adate_str = datetime.fromtimestamp(atime).strftime("%Y-%m-%d %H:%M")
        except:
            adate_str = ""
            atime = 0
        adate_item = QStandardItem(adate_str)
        adate_item.setEditable(False)
        adate_item.setData(atime, Qt.ItemDataRole.UserRole + 1)
        
        # Add placeholder child to make it expandable
        name_item.appendRow([
            QStandardItem("Loading..."),
            QStandardItem(""),
            QStandardItem(""),
            QStandardItem(""),
            QStandardItem("")
        ])
        
        return [name_item, size_item, type_item, date_item, adate_item]
        
    def create_file_item(self, path):
        """Create a row of items for a file"""
        path = Path(path)
        
        # Get cached icon by extension (fast - avoids network round-trips)
        icon = self._get_cached_icon(path, is_directory=False)
        
        # Get suffix for type column
        suffix = path.suffix.lower()
        
        # Name column with system icon
        name_item = QStandardItem(icon, path.name)
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(True)  # Allow editing for F2 rename
        # Store sort data: 1 = file (sorts after folders), name in lowercase for case-insensitive sort
        name_item.setData(f"1_{path.name.lower()}", Qt.ItemDataRole.UserRole + 1)
        
        # Size column
        try:
            # Skip size calculation on network drives for speed
            if str(path).startswith('\\\\'):
                size_str = ""  # Skip size on network drives
                size_bytes = 0
            else:
                size = path.stat().st_size
                size_bytes = size
                if size < 1024:
                    size_str = f"{size} B"
                elif size < 1024 * 1024:
                    size_str = f"{size / 1024:.1f} KB"
                elif size < 1024 * 1024 * 1024:
                    size_str = f"{size / (1024 * 1024):.1f} MB"
                else:
                    size_str = f"{size / (1024 * 1024 * 1024):.2f} GB"
        except:
            size_str = ""
            size_bytes = 0
        size_item = QStandardItem(size_str)
        size_item.setEditable(False)
        size_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)  # Right-align
        # Store numeric size for proper sorting
        size_item.setData(size_bytes, Qt.ItemDataRole.UserRole + 1)
        
        # Type column
        type_item = QStandardItem(suffix.upper()[1:] if suffix else "File")
        type_item.setEditable(False)
        
        # Date modified
        try:
            # Skip date on network drives for speed
            if str(path).startswith('\\\\'):
                date_str = ""
                mtime = 0
            else:
                mtime = path.stat().st_mtime
                date_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        except:
            date_str = ""
            mtime = 0
        date_item = QStandardItem(date_str)
        date_item.setEditable(False)
        # Store timestamp for proper sorting
        date_item.setData(mtime, Qt.ItemDataRole.UserRole + 1)
        
        # Date accessed
        try:
            if str(path).startswith('\\\\'):
                adate_str = ""  # Skip date on network drives
                atime = 0
            else:
                atime = path.stat().st_atime
                adate_str = datetime.fromtimestamp(atime).strftime("%Y-%m-%d %H:%M")
        except:
            adate_str = ""
            atime = 0
        adate_item = QStandardItem(adate_str)
        adate_item.setEditable(False)
        adate_item.setData(atime, Qt.ItemDataRole.UserRole + 1)
        
        return [name_item, size_item, type_item, date_item, adate_item]
        
    def on_item_expanded(self, index):
        """Load directory contents when expanded"""
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        # Check if this item has a placeholder child
        if item.rowCount() == 1 and item.child(0, 0).text() == "Loading...":
            # Remove placeholder
            item.removeRow(0)
            
            # Load actual contents (folders only for tree)
            path = item.data(Qt.ItemDataRole.UserRole)
            if path:
                self.load_tree_directory_contents(item, Path(path))
                
    def load_tree_directory_contents(self, parent_item, dir_path):
        """Load folder contents into the tree (folders only)"""
        try:
            dir_path = Path(dir_path)
            
            # Get only folders
            folders = [item for item in sorted(dir_path.iterdir(), key=lambda x: x.name.lower()) 
                      if item.is_dir() and not item.name.startswith('.')]
            
            for folder_path in folders:
                try:
                    item = self.create_tree_folder_item(folder_path)
                    parent_item.appendRow(item)
                except (PermissionError, OSError):
                    # Skip items we can't access
                    continue
                    
        except (PermissionError, OSError) as e:
            # Show error message
            error_item = QStandardItem(f"❌ Access denied")
            error_item.setEnabled(False)
            parent_item.appendRow(error_item)
            
    def create_details_panel(self):
        """Create the details view panel (right side) for folder contents"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header row: search box (filter) + current folder name
        header_widget = QWidget()
        header_widget.setStyleSheet(
            """
            QWidget {
                background-color: #C0D4F0;
                border: none;
                border-bottom: 1px solid #A0B8D8;
            }
            """
        )
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(8, 4, 8, 4)
        header_layout.setSpacing(8)

        self.details_search = QLineEdit()
        self.details_search.setPlaceholderText("Search...")
        self.details_search.setClearButtonEnabled(True)
        self.details_search.setMaximumWidth(240)
        self.details_search.setMaximumHeight(24)
        self.details_search.setStyleSheet(
            """
            QLineEdit {
                padding: 3px 8px;
                border: 1px solid #A0B8D8;
                border-radius: 3px;
                background: white;
                color: #1A3A6E;
            }
            """
        )
        self.details_search.textChanged.connect(self.on_details_search_changed)
        header_layout.addWidget(self.details_search)
        
        # Depth Level combo box (always visible, default 1)
        self.depth_level_combo = QComboBox()
        self.depth_level_combo.addItems(["1", "2", "3", "4", "5", "6", "7", "8", "Max"])
        self.depth_level_combo.setCurrentText("1")
        self.depth_level_combo.setMaximumWidth(60)
        self.depth_level_combo.setMaximumHeight(24)
        self.depth_level_combo.setToolTip("Depth level for subfolder search")
        self.depth_level_combo.setStyleSheet("""
            QComboBox {
                padding: 2px 4px;
                border: 1px solid #A0B8D8;
                border-radius: 3px;
                background: white;
                color: #1A3A6E;
                font-size: 9pt;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #1A3A6E;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                border: 2px solid #2563EB;
                background-color: white;
                selection-background-color: #C9DAFF;
                selection-color: #1A3A6E;
            }
        """)
        self.depth_level_combo.currentTextChanged.connect(self.on_depth_level_changed)
        header_layout.addWidget(self.depth_level_combo)
        
        # On/Off toggle button for depth search
        self.depth_toggle_btn = QPushButton("Off")
        self.depth_toggle_btn.setCheckable(False)
        self.depth_toggle_btn.setMaximumWidth(50)
        self.depth_toggle_btn.setMaximumHeight(24)
        self.depth_toggle_btn.setToolTip("Depth search is off")
        self.depth_toggle_btn.setStyleSheet("""
            QPushButton {
                background-color: #E0ECFF;
                border: 1px solid #2563EB;
                border-radius: 3px;
                padding: 2px 6px;
                font-size: 9pt;
                color: #1A3A6E;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #C9DAFF;
            }
        """)
        self.depth_toggle_btn.clicked.connect(self.toggle_depth_search)
        header_layout.addWidget(self.depth_toggle_btn)

        self.details_header = QLabel("Contents")
        self.details_header.setStyleSheet(
            """
            QLabel {
                font-weight: 600;
                font-size: 10pt;
                color: #1A3A6E;
                background: transparent;
            }
            """
        )
        header_layout.addWidget(self.details_header)
        header_layout.addStretch()

        layout.addWidget(header_widget)
        
        # Create details tree view with drop support
        self.details_view = DropTreeView()
        self.details_view.set_file_explorer(self)
        self.details_view.files_dropped.connect(self.handle_dropped_files)
        self.details_view.setAnimated(False)
        self.details_view.setRootIsDecorated(False)  # No expand arrows
        self.details_view.setIndentation(0)
        self.details_view.setHeaderHidden(False)
        self.details_view.setSelectionMode(QTreeView.SelectionMode.ExtendedSelection)  # Multi-select
        self.details_view.setSortingEnabled(True)
        self.details_view.setEditTriggers(QTreeView.EditTrigger.NoEditTriggers)  # Disable double-click editing
        self.details_view.setDragEnabled(True)  # Enable dragging files out
        self.details_view.setDragDropMode(QTreeView.DragDropMode.DragDrop)
        
        # Style the details view to look like tree view (no borders, no padding)
        self.details_view.setStyleSheet(
            """
            QTreeView {
                outline: none;
                border: none;
                background-color: transparent;
            }
            QTreeView::item {
                padding: 2px 6px;
                margin: 0px;
                border: none;
                border-radius: 0px;
                background-color: transparent;
                min-height: 20px;
                outline: none;
            }
            QTreeView::item:hover {
                background-color: #C8DCF0;
                border: none;
                outline: none;
            }
            QTreeView::item:selected {
                background-color: #B0C8E8;
                color: #0A1E5E;
                border: none;
                outline: none;
            }
            QTreeView::item:selected:!active {
                background-color: #B0C8E8;
                color: #0A1E5E;
                border: none;
                outline: none;
            }
            QTreeView::item:focus {
                border: none;
                outline: none;
            }
            QTreeView:focus {
                border: none;
                outline: none;
            }
            """
        )
        
        # Disable focus frame/rectangle
        self.details_view.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.details_view.setFrameShape(QFrame.Shape.NoFrame)
        
        # Use NoFocusDelegate to remove focus rectangle from items
        self.details_view.setItemDelegate(NoFocusDelegate(self.details_view))
        
        # Install event filter for F2 key
        self.details_view.installEventFilter(self)
        
        # Create details model
        self.details_model = QStandardItemModel()
        self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified', 'Date Accessed'])
        
        # Connect to handle renames
        self.details_model.itemChanged.connect(self.on_item_renamed)
        
        # Create sort/filter proxy model (sorting + search filtering)
        self.details_sort_proxy = FileSortProxyModel()
        self.details_sort_proxy.setSourceModel(self.details_model)
        self.details_sort_proxy.setFilterKeyColumn(0)  # Name column
        self.details_sort_proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        
        # Set the proxy model on the view
        self.details_view.setModel(self.details_sort_proxy)
        
        # Configure header
        header_view = self.details_view.header()
        header_view.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        header_view.setMinimumSectionSize(60)
        header_view.setDefaultSectionSize(100)
        header_view.setFixedHeight(22)
        header_view.setStyleSheet(
            """
            QHeaderView {
                background-color: #E0E0E0;
            }
            QHeaderView::section {
                background-color: #E0E0E0;
                padding: 1px 6px;
                font-size: 11px;
                font-weight: 600;
                color: #333333;
                border: none;
                border-right: 1px solid #C0C0C0;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            """
        )
        
        # Set default sort: Name column, ascending
        self.details_view.sortByColumn(0, Qt.SortOrder.AscendingOrder)
        
        # Apply saved column widths or use defaults
        default_widths = [350, 100, 120, 150, 150]  # Name, Size, Type, Date Modified, Date Accessed
        for col in range(5):
            width = self.column_widths.get(f'col_{col}', default_widths[col])
            self.details_view.setColumnWidth(col, width)
        
        # Connect column resize signal to save widths
        header_view.sectionResized.connect(self.on_column_resized)
        
        # Connect signals
        self.details_view.doubleClicked.connect(self.on_details_item_double_clicked)
        
        # Context menu
        self.details_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.details_view.customContextMenuRequested.connect(self.show_details_context_menu)
        
        layout.addWidget(self.details_view)
        
        # Add footer with item count
        footer = QLabel("")
        footer.setStyleSheet("""
            QLabel {
                background-color: #E0E0E0;
                padding: 2px 8px;
                font-size: 9pt;
                color: #555555;
                border: none;
                border-top: 1px solid #A0B8D8;
            }
        """)
        footer.setFixedHeight(20)
        layout.addWidget(footer)
        self.details_footer = footer
        
        # Connect model change signal to update footer count
        self.details_sort_proxy.rowsInserted.connect(self.update_details_footer)
        self.details_sort_proxy.rowsRemoved.connect(self.update_details_footer)
        self.details_sort_proxy.modelReset.connect(self.update_details_footer)
        
        return widget

    def on_details_search_changed(self, text: str) -> None:
        """Filter the details view contents in real time."""
        if not hasattr(self, 'details_sort_proxy'):
            return

        query = (text or "").strip()
        
        # Update search box border to indicate active filter
        if query:
            # Red border when filter is active
            self.details_search.setStyleSheet(
                """
                QLineEdit {
                    padding: 3px 8px;
                    border: 3px solid #DC143C;
                    border-radius: 3px;
                    background: white;
                    color: #1A3A6E;
                }
                """
            )
        else:
            # Normal border when no filter
            self.details_search.setStyleSheet(
                """
                QLineEdit {
                    padding: 3px 8px;
                    border: 1px solid #A0B8D8;
                    border-radius: 3px;
                    background: white;
                    color: #1A3A6E;
                }
                """
            )
        
        # Save the search term for the current folder
        if hasattr(self, 'current_details_folder') and self.current_details_folder:
            if query:
                self.folder_search_terms[self.current_details_folder] = query
            elif self.current_details_folder in self.folder_search_terms:
                # Remove empty search terms to keep dict clean
                del self.folder_search_terms[self.current_details_folder]
        
        # Standard search using proxy filter (works for both normal and depth results)
        if not query:
            self.details_sort_proxy.setFilterRegularExpression(QRegularExpression())
            return

        # Treat user input as a literal substring match (case-insensitive)
        escaped = QRegularExpression.escape(query)
        regex = QRegularExpression(escaped, QRegularExpression.PatternOption.CaseInsensitiveOption)
        self.details_sort_proxy.setFilterRegularExpression(regex)
    
    def toggle_depth_search(self):
        """Toggle depth search on/off based on button state"""
        # Guard against callback during widget deletion
        try:
            if not self.isVisible() or not hasattr(self, 'depth_search_enabled'):
                return
        except RuntimeError:
            # Widget is being deleted
            return
        
        if self.depth_search_enabled:
            # Currently ON, turn it OFF
            self.depth_search_enabled = False
            self.depth_search_locked = False
            self.depth_search_active_results = None
            self.depth_toggle_btn.setText("Off")
            self.depth_toggle_btn.setToolTip("Depth search is off")
            self.depth_toggle_btn.setStyleSheet("""
                QPushButton {
                    background-color: #E0ECFF;
                    border: 1px solid #2563EB;
                    border-radius: 3px;
                    padding: 2px 6px;
                    font-size: 9pt;
                    color: #1A3A6E;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #C9DAFF;
                }
            """)
            
            # Re-enable tree panel
            if hasattr(self, 'tree_view'):
                self.tree_view.setEnabled(True)
            
            # Restore normal breadcrumb bar style
            self._apply_depth_search_locked_style(locked=False)
            
            # Clear cache and search folder
            self.depth_search_cache.clear()
            self.depth_search_folder = None
            
            # Reset combo to 1
            self.depth_level_combo.setCurrentText("1")
            
            # Clear search and restore normal view
            self.details_search.clear()
            if self.current_details_folder:
                self.load_folder_contents_in_details(Path(self.current_details_folder))
        else:
            # Currently OFF, turn it ON
            depth_level = self.depth_level_combo.currentText()
            
            if depth_level == "1":
                # Depth 1 doesn't need async scan, just notify user
                QMessageBox.information(self, "Depth Search", 
                    "Depth level 1 shows only current folder items.\n\n"
                    "Set depth to 2 or higher for subfolder search.")
                return
            
            # Start depth scan
            if not self.current_details_folder:
                return
            
            self.depth_search_enabled = True
            self.perform_depth_scan_and_populate(depth_level)
    
    def on_depth_level_changed(self, level_text):
        """Handle depth level change - clear cache"""
        # Guard against callback during widget deletion
        try:
            if not self.isVisible() or not hasattr(self, 'depth_search_cache'):
                return
        except RuntimeError:
            # Widget is being deleted
            return
        
        self.depth_search_cache.clear()
        
        # If depth search is active and user changes level, turn it off
        if self.depth_search_enabled:
            self.depth_search_enabled = False
            self.depth_search_locked = False
            self.depth_search_active_results = None
            self.depth_toggle_btn.setText("Off")
            self.depth_toggle_btn.setToolTip("Depth search is off")
            self.depth_toggle_btn.setStyleSheet("""
                QPushButton {
                    background-color: #E0ECFF;
                    border: 1px solid #2563EB;
                    border-radius: 3px;
                    padding: 2px 6px;
                    font-size: 9pt;
                    color: #1A3A6E;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #C9DAFF;
                }
            """)
            self.depth_search_folder = None
            
            # Re-enable tree panel
            if hasattr(self, 'tree_view'):
                self.tree_view.setEnabled(True)
            
            # Restore normal toolbar color
            self._apply_compact_toolbar_style(self.toolbar, locked=False)
            
            # Restore normal view
            if self.current_details_folder:
                self.load_folder_contents_in_details(Path(self.current_details_folder))
    
    def perform_depth_scan_and_populate(self, depth_level):
        """Perform depth scan and populate all results (no search filter)"""
        if not self.current_details_folder:
            return
        
        # Convert "Max" to -1 for unlimited depth
        if depth_level == "Max":
            depth_level_int = -1
        else:
            depth_level_int = int(depth_level)
        
        search_folder = self.current_details_folder
        
        # Check if we have cached results for this folder and depth
        cache_key = search_folder
        if (cache_key in self.depth_search_cache and 
            depth_level_int in self.depth_search_cache[cache_key]):
            # Use cached results
            self._populate_depth_results(self.depth_search_cache[cache_key][depth_level_int])
            self.depth_toggle_btn.setText("On")
            self.depth_toggle_btn.setToolTip("Depth search is on")
            self.depth_toggle_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FFB366;
                    border: 1px solid #FF8C00;
                    border-radius: 3px;
                    padding: 2px 6px;
                    font-size: 9pt;
                    color: #1A3A6E;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #FFC080;
                }
            """)
            return
        
        # Start async scan
        self.depth_search_folder = search_folder
        
        # Create and start worker thread
        self.depth_scan_worker = DepthScanWorker(search_folder, depth_level_int)
        self.depth_scan_worker.finished.connect(self._on_depth_scan_complete)
        self.depth_scan_worker.progress.connect(self._on_depth_scan_progress)
        
        # Create progress dialog with 3 second delay and actual progress bar
        self.depth_progress_dialog = QProgressDialog("Scanning subfolders...", "Stop", 0, 100, self)
        self.depth_progress_dialog.setWindowTitle("Loading Depth Search")
        self.depth_progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self.depth_progress_dialog.setMinimumDuration(3000)  # Show after 3 seconds
        self.depth_progress_dialog.setAutoClose(False)
        self.depth_progress_dialog.setAutoReset(False)
        self.depth_progress_dialog.setValue(0)  # Start at 0
        self.depth_progress_dialog.canceled.connect(self._on_depth_scan_cancelled)
        
        # Store start time for progress calculation
        self.depth_scan_start_count = 0
        
        # Start worker
        self.depth_scan_worker.start()
    
    def _on_depth_scan_progress(self, count: int, message: str):
        """Update progress dialog during depth scan"""
        if hasattr(self, 'depth_progress_dialog') and self.depth_progress_dialog:
            # Calculate a rough progress percentage (we don't know total, so use logarithmic scale)
            # This gives a sense of progress even without knowing the total
            if count > 0:
                # Use logarithmic scale for smoother progress: log(count+1) / log(10000+1) * 100
                # Caps at ~100% around 10000 items
                import math
                progress = min(99, int((math.log(count + 1) / math.log(10000 + 1)) * 100))
                self.depth_progress_dialog.setValue(progress)
            
            self.depth_progress_dialog.setLabelText(f"{message}\n{count} items found")
    
    def _on_depth_scan_cancelled(self):
        """Handle cancellation of depth scan"""
        if hasattr(self, 'depth_scan_worker') and self.depth_scan_worker:
            # Request worker to stop
            self.depth_scan_worker.cancel()
            
            # Wait a bit for worker to finish current iteration
            self.depth_scan_worker.wait(1000)  # Wait up to 1 second
            
            # The worker will emit finished signal with partial results
            # Just close the progress dialog
            if hasattr(self, 'depth_progress_dialog') and self.depth_progress_dialog:
                self.depth_progress_dialog.close()
    
    def _on_depth_scan_complete(self, results: list):
        """Handle completion of depth scan"""
        if not hasattr(self, 'depth_progress_dialog'):
            return
        
        # Check if scan was cancelled (partial results)
        was_cancelled = hasattr(self, 'depth_scan_worker') and self.depth_scan_worker._cancelled
        
        # Cache the results (even if partial from cancellation)
        search_folder = self.depth_search_folder
        depth_level_text = self.depth_level_combo.currentText()
        
        # Convert to cache key
        if depth_level_text == "Max":
            depth_level_key = -1
        else:
            depth_level_key = int(depth_level_text)
        
        # Only cache if not cancelled (partial results shouldn't be cached)
        if not was_cancelled:
            if search_folder not in self.depth_search_cache:
                self.depth_search_cache[search_folder] = {}
            self.depth_search_cache[search_folder][depth_level_key] = results
        
        # Close progress dialog
        if self.depth_progress_dialog:
            self.depth_progress_dialog.close()
        
        # If no results, don't proceed
        if not results:
            self.depth_search_enabled = False
            self.depth_toggle_btn.setText("Off")
            self.depth_toggle_btn.setToolTip("Depth search is off")
            self.depth_toggle_btn.setStyleSheet("""
                QPushButton {
                    background-color: #E0ECFF;
                    border: 1px solid #2563EB;
                    border-radius: 3px;
                    padding: 2px 6px;
                    font-size: 9pt;
                    color: #1A3A6E;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #C9DAFF;
                }
            """)
            return
        
        # Check if we're still in the search folder
        currently_in_search_folder = (self.current_details_folder == search_folder)
        
        if currently_in_search_folder:
            # Populate all results (partial or complete)
            self._populate_depth_results(results)
            
            # Enable lock mode
            self.depth_search_locked = True
            self.depth_search_active_results = results
            
            # Disable tree panel navigation
            if hasattr(self, 'tree_view'):
                self.tree_view.setEnabled(False)
            
            # Change breadcrumb bar to red (if available in FileExplorerTab)
            self._apply_depth_search_locked_style(locked=True)
            
            # Change button to "On" with red background
            self.depth_toggle_btn.setText("On")
            self.depth_toggle_btn.setToolTip("Depth search is on (click to turn off)")
            self.depth_toggle_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FF6B6B;
                    border: 1px solid #CC4444;
                    border-radius: 3px;
                    padding: 2px 6px;
                    font-size: 9pt;
                    color: #FFFFFF;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #FF8888;
                }
            """)
        else:
            # Show dialog with "Go to search folder" button
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Depth Search Complete")
            msg_box.setText(f"Found {len(results)} items at depth {depth_level_text}")
            msg_box.setInformativeText(f"Search folder: {search_folder}")
            
            # Add "Go to search folder" button
            go_button = msg_box.addButton("Go to search folder", QMessageBox.ButtonRole.AcceptRole)
            close_button = msg_box.addButton("Close", QMessageBox.ButtonRole.RejectRole)
            
            msg_box.exec()
            
            # Check which button was clicked
            if msg_box.clickedButton() == go_button:
                # Navigate to search folder and populate results
                self.load_folder_contents_in_details(Path(search_folder))
                self._populate_depth_results(results)
                
                # Enable lock mode
                self.depth_search_locked = True
                self.depth_search_active_results = results
                
                # Disable tree panel navigation
                if hasattr(self, 'tree_view'):
                    self.tree_view.setEnabled(False)
                
                # Change breadcrumb bar to red (if available in FileExplorerTab)
                self._apply_depth_search_locked_style(locked=True)
                
                # Change button to "On" with red background
                self.depth_toggle_btn.setText("On")
                self.depth_toggle_btn.setToolTip("Depth search is on (click to turn off)")
                self.depth_toggle_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #FF6B6B;
                        border: 1px solid #CC4444;
                        border-radius: 3px;
                        padding: 2px 6px;
                        font-size: 9pt;
                        color: #FFFFFF;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #FF8888;
                    }
                """)
            else:
                # User closed dialog, turn off depth search
                self.depth_search_enabled = False
                self.depth_toggle_btn.setText("Off")
                self.depth_toggle_btn.setToolTip("Depth search is off")
                self.depth_toggle_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #E0ECFF;
                        border: 1px solid #2563EB;
                        border-radius: 3px;
                        padding: 2px 6px;
                        font-size: 9pt;
                        color: #1A3A6E;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #C9DAFF;
                    }
                """)
    
    def _populate_depth_results(self, depth_items: list):
        """Populate view with all depth search results (no filtering)"""
        if not depth_items:
            return
        
        # Clear the model
        self.details_model.clear()
        self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified', 'Date Accessed'])
        
        # Restore column widths
        header_view = self.details_view.header()
        try:
            header_view.sectionResized.disconnect(self.on_column_resized)
        except:
            pass
        
        default_widths = [350, 100, 120, 150, 150]
        for col in range(5):
            width = self.column_widths.get(f'col_{col}', default_widths[col])
            self.details_view.setColumnWidth(col, width)
        
        header_view.sectionResized.connect(self.on_column_resized)
        
        # Sort by depth level then alphabetically
        sorted_items = sorted(depth_items, key=lambda x: (x['depth'], x['display_name'].lower()))
        
        # Add all items to view
        for item in sorted_items:
            row_items = self._create_depth_search_item(item)
            self.details_model.appendRow(row_items)
        
        # Update footer
        self.update_details_footer()
    
    def _create_depth_search_item(self, item_data: dict):
        """Create a row item for depth search results"""
        display_name = item_data['display_name']
        full_path = item_data['path']
        is_dir = item_data['is_dir']
        depth = item_data['depth']
        size = item_data.get('size', 0)
        modified = item_data.get('modified', '')
        accessed = item_data.get('accessed', '')
        
        # Create name item with icon
        name_item = QStandardItem(display_name)
        path_obj = Path(full_path)
        icon = self._get_cached_icon(path_obj, is_dir)
        name_item.setIcon(icon)
        name_item.setData(full_path, Qt.ItemDataRole.UserRole)
        
        # Set sort data (depth prefix for proper sorting)
        sort_prefix = f"{depth}_{'0' if is_dir else '1'}_"
        name_item.setData(sort_prefix + display_name.lower(), Qt.ItemDataRole.UserRole + 1)
        
        # Size item
        size_item = QStandardItem()
        if not is_dir and size > 0:
            if size < 1024:
                size_str = f"{size} B"
            elif size < 1024 * 1024:
                size_str = f"{size / 1024:.1f} KB"
            elif size < 1024 * 1024 * 1024:
                size_str = f"{size / (1024 * 1024):.1f} MB"
            else:
                size_str = f"{size / (1024 * 1024 * 1024):.2f} GB"
            size_item.setText(size_str)
            size_item.setData(size, Qt.ItemDataRole.UserRole + 1)
        else:
            size_item.setData(0, Qt.ItemDataRole.UserRole + 1)
        
        # Type item
        type_item = QStandardItem("Folder" if is_dir else path_obj.suffix.upper().lstrip('.'))
        type_item.setData(full_path, Qt.ItemDataRole.UserRole)
        
        # Modified item
        modified_item = QStandardItem(modified)
        modified_item.setData(full_path, Qt.ItemDataRole.UserRole)
        
        # Accessed item
        accessed_item = QStandardItem(accessed)
        accessed_item.setData(full_path, Qt.ItemDataRole.UserRole)
        
        return [name_item, size_item, type_item, modified_item, accessed_item]
    
    def on_tree_item_clicked(self, index):
        """Handle click on tree item - load folder contents in details view"""
        # Block navigation if depth search is locked
        if self.depth_search_locked:
            QMessageBox.warning(self, "Navigation Locked", 
                "Folder navigation is locked while depth search is active.\n\n"
                "Turn off depth search to navigate to other folders.")
            return
        
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if path:
            self.load_folder_contents_in_details(Path(path))
    
    def eventFilter(self, obj, event):
        """Handle keyboard shortcuts in details view (F2 for rename, Ctrl+V for paste, Delete)"""
        if obj == self.details_view and event.type() == event.Type.KeyPress:
            modifiers = event.modifiers()
            key = event.key()
            
            if key == Qt.Key.Key_F2:
                # Get selected item
                indexes = self.details_view.selectedIndexes()
                if indexes:
                    # Get the name column (column 0) of the first selected row
                    name_index = self.details_view.model().index(indexes[0].row(), 0)
                    self.details_view.edit(name_index)
                    return True
            elif key == Qt.Key.Key_Delete:
                # Delete key - delete selected file(s)
                self.delete_file()
                return True
            elif key == Qt.Key.Key_V and (modifiers & Qt.KeyboardModifier.ControlModifier):
                # Ctrl+V - Paste
                self.paste_file()
                return True
            elif key == Qt.Key.Key_C and (modifiers & Qt.KeyboardModifier.ControlModifier):
                # Ctrl+C - Copy
                self.copy_file()
                return True
            elif key == Qt.Key.Key_X and (modifiers & Qt.KeyboardModifier.ControlModifier):
                # Ctrl+X - Cut
                self.cut_file()
                return True
        
        return super().eventFilter(obj, event)
    
    def on_item_renamed(self, item):
        """Handle item rename after F2 edit"""
        # Disconnect temporarily to avoid recursive calls
        self.details_model.itemChanged.disconnect(self.on_item_renamed)
        
        try:
            # Get the old path from UserRole
            old_path = item.data(Qt.ItemDataRole.UserRole)
            if not old_path:
                return
            
            old_path_obj = Path(old_path)
            
            # Extract new name from the item text (remove icon emoji)
            new_text = item.text()
            # Remove emoji icons (they're at the start)
            import re
            new_name = re.sub(r'^[\U0001F300-\U0001F9FF]\s*', '', new_text).strip()
            
            # Validate new name
            if not new_name or new_name == old_path_obj.name:
                # No change or empty name, revert
                icon = item.text().split()[0] if ' ' in item.text() else ""
                item.setText(f"{icon} {old_path_obj.name}" if icon else old_path_obj.name)
                return
            
            # Check for invalid characters
            invalid_chars = r'<>:"/\|?*'
            if any(c in new_name for c in invalid_chars):
                QMessageBox.warning(
                    self, 
                    "Invalid Name", 
                    f"The name cannot contain any of the following characters:\n{invalid_chars}"
                )
                # Revert name
                icon = item.text().split()[0] if ' ' in item.text() else ""
                item.setText(f"{icon} {old_path_obj.name}" if icon else old_path_obj.name)
                return
            
            # Build new path
            new_path_obj = old_path_obj.parent / new_name
            
            # Check if target already exists
            if new_path_obj.exists():
                QMessageBox.warning(
                    self,
                    "Name Conflict",
                    f"A file or folder with the name '{new_name}' already exists."
                )
                # Revert name
                icon = item.text().split()[0] if ' ' in item.text() else ""
                item.setText(f"{icon} {old_path_obj.name}" if icon else old_path_obj.name)
                return
            
            # Perform the rename
            try:
                old_path_obj.rename(new_path_obj)
                
                # Update the item's UserRole data with new path
                item.setData(str(new_path_obj), Qt.ItemDataRole.UserRole)
                
                # Update sort data
                is_folder = new_path_obj.is_dir()
                prefix = "0_" if is_folder else "1_"
                item.setData(f"{prefix}{new_name.lower()}", Qt.ItemDataRole.UserRole + 1)
                
                logger.info(f"Renamed: {old_path_obj} -> {new_path_obj}")
                
            except Exception as e:
                logger.error(f"Failed to rename: {e}")
                QMessageBox.critical(
                    self,
                    "Rename Failed",
                    f"Failed to rename '{old_path_obj.name}':\n{str(e)}"
                )
                # Revert name
                icon = item.text().split()[0] if ' ' in item.text() else ""
                item.setText(f"{icon} {old_path_obj.name}" if icon else old_path_obj.name)
                
        finally:
            # Reconnect signal
            self.details_model.itemChanged.connect(self.on_item_renamed)
    
    def load_folder_contents_in_details(self, dir_path):
        """Load folder contents into the details view - optimized for network drives"""
        # Check if we're returning to the depth search folder while locked
        if (self.depth_search_locked and 
            self.depth_search_folder and 
            str(dir_path) == self.depth_search_folder and
            self.depth_search_active_results):
            # Restore depth search results instead of loading normal contents
            self.current_details_folder = str(dir_path)
            self.details_view.set_current_folder(str(dir_path))
            self.details_header.setText(f"📂 {dir_path.name or str(dir_path)}")
            
            # Restore folder-specific search term for depth search folder too
            if hasattr(self, 'details_search') and hasattr(self, 'folder_search_terms'):
                folder_key = str(dir_path)
                saved_search = self.folder_search_terms.get(folder_key, "")
                
                # Temporarily block signals to prevent triggering on_details_search_changed
                self.details_search.blockSignals(True)
                self.details_search.setText(saved_search)
                self.details_search.blockSignals(False)
                
                # Update border styling based on whether there's a filter
                if saved_search:
                    # Red border when filter is active
                    self.details_search.setStyleSheet(
                        """
                        QLineEdit {
                            padding: 3px 8px;
                            border: 3px solid #DC143C;
                            border-radius: 3px;
                            background: white;
                            color: #1A3A6E;
                        }
                        """
                    )
                    escaped = QRegularExpression.escape(saved_search)
                    regex = QRegularExpression(escaped, QRegularExpression.PatternOption.CaseInsensitiveOption)
                    self.details_sort_proxy.setFilterRegularExpression(regex)
                else:
                    # Normal border when no filter
                    self.details_search.setStyleSheet(
                        """
                        QLineEdit {
                            padding: 3px 8px;
                            border: 1px solid #A0B8D8;
                            border-radius: 3px;
                            background: white;
                            color: #1A3A6E;
                        }
                        """
                    )
                    self.details_sort_proxy.setFilterRegularExpression(QRegularExpression())
            
            self._populate_depth_results(self.depth_search_active_results)
            return
        
        # Start timing and show loading indicator
        import time
        start_time = time.perf_counter()
        
        if hasattr(self, 'details_footer'):
            self.details_footer.setText("...loading")
            # Force UI update to show loading message
            QApplication.processEvents()
        
        try:
            dir_path = Path(dir_path)
            
            # Store current folder for breadcrumb and other operations
            self.current_details_folder = str(dir_path)
            
            # Update the details view's current folder for drag/drop
            self.details_view.set_current_folder(str(dir_path))
            
            # Update header
            self.details_header.setText(f"📂 {dir_path.name or str(dir_path)}")
            
            # Restore folder-specific search term (or clear if first visit)
            if hasattr(self, 'details_search') and hasattr(self, 'folder_search_terms'):
                folder_key = str(dir_path)
                saved_search = self.folder_search_terms.get(folder_key, "")
                
                # Temporarily block signals to prevent triggering on_details_search_changed
                self.details_search.blockSignals(True)
                self.details_search.setText(saved_search)
                self.details_search.blockSignals(False)
                
                # Update border styling and apply filter based on whether there's a saved search
                if saved_search:
                    # Red border when filter is active
                    self.details_search.setStyleSheet(
                        """
                        QLineEdit {
                            padding: 3px 8px;
                            border: 3px solid #DC143C;
                            border-radius: 3px;
                            background: white;
                            color: #1A3A6E;
                        }
                        """
                    )
                    escaped = QRegularExpression.escape(saved_search)
                    regex = QRegularExpression(escaped, QRegularExpression.PatternOption.CaseInsensitiveOption)
                    self.details_sort_proxy.setFilterRegularExpression(regex)
                else:
                    # Normal border when no filter
                    self.details_search.setStyleSheet(
                        """
                        QLineEdit {
                            padding: 3px 8px;
                            border: 1px solid #A0B8D8;
                            border-radius: 3px;
                            background: white;
                            color: #1A3A6E;
                        }
                        """
                    )
                    self.details_sort_proxy.setFilterRegularExpression(QRegularExpression())
            
            # Clear details model BUT preserve column widths
            self.details_model.clear()
            self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified', 'Date Accessed'])
            
            # Temporarily disconnect resize signal to prevent saving while we restore widths
            header_view = self.details_view.header()
            try:
                header_view.sectionResized.disconnect(self.on_column_resized)
            except:
                pass  # Might not be connected yet
            
            # Restore column widths after clearing
            default_widths = [350, 100, 120, 150, 150]
            for col in range(5):
                width = self.column_widths.get(f'col_{col}', default_widths[col])
                self.details_view.setColumnWidth(col, width)
            
            # Reconnect resize signal
            header_view.sectionResized.connect(self.on_column_resized)
            
            # Check if it's a network path
            is_network = str(dir_path).startswith('\\\\')
            
            # Get all items in directory - use os.scandir() which is MUCH faster
            # os.scandir() caches stat info in DirEntry objects (150x faster on network!)
            items_with_type = []
            try:
                with os.scandir(str(dir_path)) as entries:
                    for entry in entries:
                        try:
                            if entry.name.startswith('.'):
                                continue
                            # DirEntry.is_dir() uses cached stat - very fast!
                            is_directory = entry.is_dir()
                            # Convert to Path for compatibility with existing code
                            item_path = Path(entry.path)
                            items_with_type.append((item_path, is_directory))
                        except (PermissionError, OSError):
                            continue
                
                # Sort by name only (case-insensitive), folders first
                items_with_type.sort(key=lambda x: (not x[1], x[0].name.lower()))
            except Exception:
                items_with_type = []
            
            # Add items to view - use cached is_dir result
            for item_path, is_directory in items_with_type:
                try:
                    if is_directory:
                        row_items = self.create_folder_item(item_path)
                    else:
                        row_items = self.create_file_item(item_path)
                    
                    self.details_model.appendRow(row_items)
                except (PermissionError, OSError):
                    # Skip items we can't access
                    continue
            
            # Re-apply current sort order (default: Name ascending)
            header = self.details_view.header()
            sort_column = header.sortIndicatorSection()
            sort_order = header.sortIndicatorOrder()
            self.details_sort_proxy.sort(sort_column, sort_order)
            
            # Calculate elapsed time and update footer with timing
            elapsed_ms = int((time.perf_counter() - start_time) * 1000)
            self.update_details_footer(timing_ms=elapsed_ms)
                    
        except (PermissionError, OSError) as e:
            self.details_model.clear()
            self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified', 'Date Accessed'])
            
            # Temporarily disconnect resize signal
            header_view = self.details_view.header()
            try:
                header_view.sectionResized.disconnect(self.on_column_resized)
            except:
                pass
            
            # Restore column widths even on error
            default_widths = [350, 100, 120, 150, 150]
            for col in range(5):
                width = self.column_widths.get(f'col_{col}', default_widths[col])
                self.details_view.setColumnWidth(col, width)
            
            # Reconnect resize signal
            header_view.sectionResized.connect(self.on_column_resized)
            
            error_item = QStandardItem(f"❌ Access denied: {str(e)}")
            error_item.setEnabled(False)
            self.details_model.appendRow([error_item, QStandardItem(""), QStandardItem(""), QStandardItem(""), QStandardItem("")])
            
            # Update footer even on error
            elapsed_ms = int((time.perf_counter() - start_time) * 1000)
            self.update_details_footer(timing_ms=elapsed_ms)
    
    def on_details_item_double_clicked(self, index):
        """Handle double click in details view"""
        # DEBUG: Log what we're clicking
        print(f"\n=== DOUBLE CLICK DEBUG ===")
        print(f"Proxy Index: row={index.row()}, col={index.column()}")
        print(f"Display text at clicked index: {self.details_sort_proxy.data(index, Qt.ItemDataRole.DisplayRole)}")
        
        # Get the data directly from the proxy model at the clicked index
        # The proxy model handles all the sorting/filtering, so we should query it directly
        path = self.details_sort_proxy.data(index, Qt.ItemDataRole.UserRole)
        
        # If this column doesn't have the path data, get it from column 0 of the same row
        if not path:
            col0_index = index.sibling(index.row(), 0)
            col0_text = self.details_sort_proxy.data(col0_index, Qt.ItemDataRole.DisplayRole)
            print(f"Column 0 text for this row: {col0_text}")
            path = self.details_sort_proxy.data(col0_index, Qt.ItemDataRole.UserRole)
        
        print(f"Path retrieved: {path}")
        print(f"=========================\n")
        
        if not path:
            return
        
        path_obj = Path(path)
        
        # Handle .lnk shortcut files - resolve target and navigate if it's a folder
        if path_obj.suffix.lower() == '.lnk' and path_obj.is_file():
            target_path = self._resolve_shortcut(str(path_obj))
            if target_path:
                target_obj = Path(target_path)
                if target_obj.exists() and target_obj.is_dir():
                    # Navigate to the folder target within File Nav
                    # If depth search is active, turn it off first
                    if self.depth_search_enabled:
                        self.depth_search_enabled = False
                        self.depth_search_locked = False
                        self.depth_search_active_results = None
                        self.depth_toggle_btn.setText("Off")
                        self.depth_toggle_btn.setToolTip("Depth search is off")
                        self.depth_toggle_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #E0ECFF;
                                border: 1px solid #2563EB;
                                border-radius: 3px;
                                padding: 2px 6px;
                                font-size: 9pt;
                                color: #1A3A6E;
                                font-weight: bold;
                            }
                            QPushButton:hover {
                                background-color: #C9DAFF;
                            }
                        """)
                        self.depth_search_folder = None
                        # Re-enable tree panel
                        if hasattr(self, 'tree_view'):
                            self.tree_view.setEnabled(True)
                        # Restore normal toolbar color
                        self._apply_compact_toolbar_style(self.toolbar, locked=False)
                    
                    self.load_folder_contents_in_details(target_obj)
                    return
                elif target_obj.exists() and target_obj.is_file():
                    # Shortcut points to a file, open it
                    try:
                        if os.name == 'nt':
                            os.startfile(str(target_obj))
                        elif sys.platform == 'darwin':
                            subprocess.run(['open', str(target_obj)])
                        else:
                            subprocess.run(['xdg-open', str(target_obj)])
                    except Exception as e:
                        logger.error(f"Failed to open file: {e}")
                        QMessageBox.warning(self, "Cannot Open File", f"Failed to open {target_obj.name}\n\nError: {str(e)}")
                    return
            # If we can't resolve or target doesn't exist, fall through to open the .lnk file itself
        
        if path_obj.is_dir():
            # Navigate into folder - update tree selection and load in details
            self.load_folder_contents_in_details(path_obj)
            # TODO: Could also expand/select this folder in the tree view
        elif path_obj.is_file():
            # Open file with default application
            try:
                if os.name == 'nt':
                    os.startfile(str(path_obj))
                elif sys.platform == 'darwin':
                    subprocess.run(['open', str(path_obj)])
                else:
                    subprocess.run(['xdg-open', str(path_obj)])
            except Exception as e:
                logger.error(f"Failed to open file: {e}")
                QMessageBox.warning(self, "Cannot Open File", f"Failed to open {path_obj.name}\n\nError: {str(e)}")
    
    def _resolve_shortcut(self, lnk_path):
        """Resolve a Windows .lnk shortcut file to get its target path"""
        try:
            import win32com.client
            shell = win32com.client.Dispatch("WScript.Shell")
            shortcut = shell.CreateShortCut(lnk_path)
            return shortcut.Targetpath
        except ImportError:
            # win32com not available, try alternative method
            try:
                # Use PowerShell as fallback
                result = subprocess.run(
                    ['powershell', '-Command', 
                     f"(New-Object -ComObject WScript.Shell).CreateShortcut('{lnk_path}').TargetPath"],
                    capture_output=True, text=True, timeout=5
                )
                if result.returncode == 0 and result.stdout.strip():
                    return result.stdout.strip()
            except Exception as e:
                logger.error(f"Failed to resolve shortcut via PowerShell: {e}")
        except Exception as e:
            logger.error(f"Failed to resolve shortcut: {e}")
        return None

    def show_tree_context_menu(self, position):
        """Show context menu for tree view"""
        index = self.tree_view.indexAt(position)
        if not index.isValid():
            return
        
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        
        # Check if this is a custom quick link (only top-level items in Quick Links)
        is_custom_link = item.data(Qt.ItemDataRole.UserRole + 1) == "__CUSTOM_LINK__"
        
        # Also verify it's a top-level item (no parent except model root)
        is_top_level = item.parent() is None
        
        menu = QMenu()
        
        # Open in Explorer
        open_explorer_action = menu.addAction("📂 Open in File Explorer")
        open_explorer_action.triggered.connect(lambda: self.open_path_in_explorer(path))
        
        # Remove from Sidebar Bookmarks (only for top-level custom links)
        if is_custom_link and is_top_level:
            menu.addSeparator()
            remove_action = menu.addAction("📌 Remove from Sidebar Bookmarks")
            remove_action.triggered.connect(lambda: self.remove_quick_link_by_path(path))
        
        # Show menu
        menu.exec(self.tree_view.viewport().mapToGlobal(position))
    
    def show_details_context_menu(self, position):
        """Show context menu for details view"""
        index = self.details_view.indexAt(position)
        
        menu = QMenu()
        
        # Add "New Folder" option at the top (always available when in a folder)
        if self.current_details_folder:
            new_folder_action = menu.addAction("📁 New Folder")
            new_folder_action.triggered.connect(self.create_new_folder)
            menu.addSeparator()
        
        if index.isValid():
            # Get the data directly from the proxy model at the clicked index
            path = self.details_sort_proxy.data(index, Qt.ItemDataRole.UserRole)
            
            # If this column doesn't have the path data, get it from column 0 of the same row
            if not path:
                col0_index = index.sibling(index.row(), 0)
                path = self.details_sort_proxy.data(col0_index, Qt.ItemDataRole.UserRole)
            
            if path:
                    path_obj = Path(path)
                    
                    # Copy Full Path
                    copy_path_action = menu.addAction("📄 Copy Full Path")
                    copy_path_action.triggered.connect(lambda: self.copy_full_path_to_clipboard(path))
                    
                    # Open folder location (navigate to parent folder in File Nav)
                    open_folder_action = menu.addAction("📂 Open Folder Location")
                    open_folder_action.triggered.connect(lambda: self.open_folder_location_in_file_nav(path))
                    
                    menu.addSeparator()
                    
                    # Cut, Copy operations
                    cut_action = menu.addAction("✂️ Cut")
                    cut_action.triggered.connect(self.cut_file)
                    
                    copy_action = menu.addAction("📋 Copy")
                    copy_action.triggered.connect(self.copy_file)
                    
                    # Rename, Delete
                    rename_action = menu.addAction("✏️ Rename")
                    rename_action.triggered.connect(self.rename_file)
                    
                    delete_action = menu.addAction("🗑️ Delete")
                    delete_action.triggered.connect(self.delete_file)
        
        # Paste (always available if clipboard has content - internal or Windows)
        if self.has_clipboard_content():
            if index.isValid():
                menu.addSeparator()
            paste_action = menu.addAction("📌 Paste")
            paste_action.triggered.connect(self.paste_file)
        elif self.current_details_folder:
            # Show disabled paste option when in a folder but no clipboard content
            if index.isValid():
                menu.addSeparator()
            paste_action = menu.addAction("📌 Paste")
            paste_action.setEnabled(False)
        
        # Show menu
        menu.exec(self.details_view.viewport().mapToGlobal(position))
    
    def open_file(self, path):
        """Open file with default application"""
        try:
            path_obj = Path(path)
            if os.name == 'nt':
                os.startfile(str(path_obj))
            elif sys.platform == 'darwin':
                subprocess.run(['open', str(path_obj)])
            else:
                subprocess.run(['xdg-open', str(path_obj)])
        except Exception as e:
            logger.error(f"Failed to open file: {e}")
            QMessageBox.warning(self, "Error", f"Failed to open file:\n{str(e)}")
    
    def open_path_in_explorer(self, path):
        """Open path in system file explorer"""
        try:
            path_obj = Path(path)
            if os.name == 'nt':
                subprocess.run(['explorer', '/select,', str(path_obj)])
            elif sys.platform == 'darwin':
                subprocess.run(['open', '-R', str(path_obj)])
            else:
                subprocess.run(['xdg-open', str(path_obj.parent)])
        except Exception as e:
            logger.error(f"Failed to open in explorer: {e}")
    
    def get_current_details_folder(self):
        """Get the current folder being displayed in details view"""
        # This would need to track the current folder - for now return empty
        # You could store this as self.current_details_folder when loading
        return ""
        
    def show_file_preview(self, file_path):
        """Show enhanced file preview with support for Excel, CSV, images, etc."""
        self.current_file_path = file_path
        path = Path(file_path)
        suffix = path.suffix.lower()
        
        # Clear preview
        self.preview_text.clear()
        self.upload_button.setEnabled(True)
        
        try:
            # Check file size first
            file_size = path.stat().st_size
            if file_size > 10 * 1024 * 1024:  # 10 MB
                self.preview_text.setText(
                    f"📄 {path.name}\n"
                    f"Size: {file_size / (1024 * 1024):.2f} MB\n\n"
                    f"⚠️ File too large for preview\n"
                    f"Double-click to open in default application"
                )
                return
            
            # Excel files (.xlsx, .xls, .xlsm)
            if suffix in ['.xlsx', '.xls', '.xlsm']:
                self.preview_excel_file(path)
            
            # CSV files
            elif suffix == '.csv':
                self.preview_csv_file(path)
            
            # Text files
            elif suffix in ['.txt', '.log', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', '.sql']:
                self.preview_text_file(path)
            
            # Image files
            elif suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                self.preview_image_file(path)
            
            # PDF files
            elif suffix == '.pdf':
                self.preview_text.setText(
                    f"📕 PDF Document: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in PDF viewer"
                )
            
            # Word documents
            elif suffix in ['.docx', '.doc']:
                self.preview_text.setText(
                    f"📝 Word Document: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in Microsoft Word"
                )
            
            # PowerPoint
            elif suffix in ['.pptx', '.ppt']:
                self.preview_text.setText(
                    f"📽️ PowerPoint: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in PowerPoint"
                )
            
            # Access databases
            elif suffix in ['.accdb', '.mdb']:
                self.preview_text.setText(
                    f"🗃️ Access Database: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in Microsoft Access"
                )
            
            # Unknown/Binary
            else:
                self.preview_text.setText(
                    f"📃 {path.name}\n"
                    f"Type: {suffix.upper()[1:] if suffix else 'Unknown'}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"Preview not available for this file type\n"
                    f"Double-click to open"
                )
                
        except Exception as e:
            logger.error(f"Preview error: {e}")
            self.preview_text.setText(
                f"❌ Error previewing file:\n{str(e)}\n\n"
                f"File: {path.name}"
            )
    
    def preview_excel_file(self, path):
        """Preview Excel file showing first few rows"""
        try:
            import pandas as pd
            
            # Read first sheet using openpyxl engine
            df = pd.read_excel(path, engine='openpyxl', nrows=100)
            
            # Build preview text
            preview = f"📊 Excel File: {path.name}\n"
            preview += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
            preview += "=" * 60 + "\n\n"
            
            # Column names
            preview += "Columns:\n"
            for i, col in enumerate(df.columns, 1):
                preview += f"  {i}. {col}\n"
            
            preview += "\n" + "=" * 60 + "\n"
            preview += "First 10 Rows:\n"
            preview += "=" * 60 + "\n\n"
            
            # Show first 10 rows with formatting
            preview += df.head(10).to_string(index=True, max_colwidth=40)
            
            if len(df) > 10:
                preview += f"\n\n... and {len(df) - 10} more rows"
            
            preview += "\n\n📌 Double-click to open in Excel"
            
            self.preview_text.setText(preview)
            self.current_file_content = preview
            
        except ImportError as e:
            self.preview_text.setText(
                f"📊 Excel File: {path.name}\n\n"
                f"⚠️ Could not preview Excel file:\n"
                f"Missing optional dependency 'xlrd'. Install xlrd >= 2.0.1 for xls Excel support\n"
                f"or use openpyxl for xlsx files.\n\n"
                f"Run: pip install openpyxl\n\n"
                f"Double-click to open in Microsoft Excel"
            )
        except Exception as e:
            self.preview_text.setText(
                f"📊 Excel File: {path.name}\n\n"
                f"⚠️ Could not preview Excel file:\n{str(e)}\n\n"
                f"Double-click to open in Microsoft Excel"
            )
    
    def preview_csv_file(self, path):
        """Preview CSV file"""
        try:
            import pandas as pd
            
            # Try to read CSV
            df = pd.read_csv(path, nrows=100)
            
            preview = f"📑 CSV File: {path.name}\n"
            preview += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
            preview += "=" * 60 + "\n\n"
            
            # Column names
            preview += "Columns:\n"
            for i, col in enumerate(df.columns, 1):
                preview += f"  {i}. {col}\n"
            
            preview += "\n" + "=" * 60 + "\n"
            preview += "First 10 Rows:\n"
            preview += "=" * 60 + "\n\n"
            
            preview += df.head(10).to_string(index=False, max_colwidth=40)
            
            if len(df) > 10:
                preview += f"\n\n... and {len(df) - 10} more rows"
            
            self.preview_text.setText(preview)
            self.current_file_content = preview
            
        except Exception as e:
            # Fallback to text preview
            self.preview_text_file(path)
    
    def preview_text_file(self, path):
        """Preview text files"""
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)  # Read first 50KB
                
            preview = f"📄 Text File: {path.name}\n"
            preview += f"Size: {path.stat().st_size / 1024:.1f} KB\n"
            preview += "=" * 60 + "\n\n"
            preview += content
            
            if path.stat().st_size > 50000:
                preview += "\n\n... (file truncated for preview)"
            
            self.preview_text.setText(preview)
            self.current_file_content = content
            
        except Exception as e:
            self.preview_text.setText(
                f"Cannot read file:\n{str(e)}\n\n"
                f"File: {path.name}"
            )
    
    def preview_image_file(self, path):
        """Preview image file info"""
        try:
            from PIL import Image
            
            img = Image.open(path)
            
            preview = f"🖼️ Image: {path.name}\n"
            preview += f"Size: {path.stat().st_size / 1024:.1f} KB\n"
            preview += f"Dimensions: {img.width} × {img.height} pixels\n"
            preview += f"Format: {img.format}\n"
            preview += f"Mode: {img.mode}\n\n"
            preview += "📌 Double-click to view image"
            
            self.preview_text.setText(preview)
            
        except:
            # If PIL not available, show basic info
            self.preview_text.setText(
                f"🖼️ Image: {path.name}\n"
                f"Size: {path.stat().st_size / 1024:.1f} KB\n\n"
                f"📌 Double-click to view image"
            )
            
    def show_upload_menu(self):
        """Show mainframe upload menu"""
        if not self.current_file_path:
            return
        
        QMessageBox.information(
            self,
            "Upload to Mainframe",
            f"Upload functionality will be integrated here.\n\n"
            f"File: {Path(self.current_file_path).name}"
        )
        
    # File operations
    def get_selected_path(self):
        """Get currently selected file/folder path from details view (single selection)"""
        indexes = self.details_view.selectedIndexes()
        if not indexes:
            return None
        
        # Get the data directly from the proxy model
        path = self.details_sort_proxy.data(indexes[0], Qt.ItemDataRole.UserRole)
        
        # If this column doesn't have the path data, get it from column 0 of the same row
        if not path:
            col0_index = indexes[0].sibling(indexes[0].row(), 0)
            path = self.details_sort_proxy.data(col0_index, Qt.ItemDataRole.UserRole)
        
        return path
    
    def get_selected_paths(self):
        """Get all selected file/folder paths from details view (multi-selection)"""
        selected_rows = {}
        paths = []
        
        # Get unique rows (since selecting a row selects all columns)
        for index in self.details_view.selectedIndexes():
            row = index.row()
            
            if row not in selected_rows:
                selected_rows[row] = True
                # Get column 0 index for this row and query the proxy model
                col0_index = index.sibling(row, 0)
                path = self.details_sort_proxy.data(col0_index, Qt.ItemDataRole.UserRole)
                if path:
                    paths.append(path)
        
        return paths
        
    def cut_file(self):
        """Cut selected file(s)/folder(s)"""
        paths = self.get_selected_paths()
        if paths:
            self.clipboard = {"paths": paths, "operation": "cut"}
            # Also set system clipboard so files can be pasted in Windows Explorer
            self._set_system_clipboard(paths)
            logger.info(f"Cut: {len(paths)} item(s)")
            
    def copy_file(self):
        """Copy selected file(s)/folder(s)"""
        paths = self.get_selected_paths()
        if paths:
            self.clipboard = {"paths": paths, "operation": "copy"}
            # Also set system clipboard so files can be pasted in Windows Explorer
            self._set_system_clipboard(paths)
            logger.info(f"Copy: {len(paths)} item(s)")
    
    def copy_full_path_to_clipboard(self, path):
        """Copy the full path of a file or folder to clipboard as text"""
        clipboard = QApplication.clipboard()
        clipboard.setText(str(path))
        logger.info(f"Copied path to clipboard: {path}")
    
    def open_folder_location_in_file_nav(self, path):
        """Navigate to the parent folder of the given path in File Nav"""
        path_obj = Path(path)
        parent_folder = path_obj.parent if path_obj.is_file() else path_obj
        
        # Check if this is a FileExplorerTab (has navigate_to_path for breadcrumb/history)
        # or just the base FileExplorerCore
        if hasattr(self, 'navigate_to_path'):
            # Use navigate_to_path to update breadcrumb and history
            self.navigate_to_path(str(parent_folder), add_to_history=True)
        else:
            # Fall back to direct loading (base FileExplorerCore)
            self.load_folder_contents_in_details(parent_folder)
        
        # If this is a file, select it in the details view after loading
        if path_obj.is_file():
            # Small delay to ensure the model is populated
            from PyQt6.QtCore import QTimer
            QTimer.singleShot(100, lambda: self.select_file_in_details(path))
        
        logger.info(f"Navigated to folder location: {parent_folder}")
    
    def select_file_in_details(self, file_path):
        """Select a specific file in the details view"""
        try:
            file_name = Path(file_path).name
            # Search through the model to find and select the file
            for row in range(self.details_model.rowCount()):
                item = self.details_model.item(row, 0)
                if item and item.text() == file_name:
                    # Get the index in the proxy model
                    source_index = self.details_model.indexFromItem(item)
                    proxy_index = self.details_sort_proxy.mapFromSource(source_index)
                    
                    # Select and scroll to the item
                    self.details_view.setCurrentIndex(proxy_index)
                    self.details_view.scrollTo(proxy_index)
                    break
        except Exception as e:
            logger.error(f"Failed to select file in details: {e}")
    
    def _set_system_clipboard(self, paths):
        """Set file paths to the system clipboard for use with Windows Explorer"""
        clipboard = QApplication.clipboard()
        mime_data = QMimeData()
        
        # Convert paths to QUrl list
        urls = [QUrl.fromLocalFile(path) for path in paths]
        mime_data.setUrls(urls)
        
        clipboard.setMimeData(mime_data)
            
    def get_clipboard_files(self):
        """Get list of file paths from Windows clipboard"""
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
        
        files = []
        if mime_data.hasUrls():
            for url in mime_data.urls():
                if url.isLocalFile():
                    file_path = url.toLocalFile()
                    if os.path.exists(file_path):
                        files.append(file_path)
        
        return files
    
    def has_clipboard_content(self):
        """Check if there's pasteable content in clipboard (internal or Windows)"""
        # Check internal clipboard
        if self.clipboard.get("paths"):
            return True
        # Check Windows clipboard for files
        return len(self.get_clipboard_files()) > 0
    
    def paste_file(self):
        """Paste cut/copied file/folder from internal clipboard or Windows clipboard"""
        # Determine destination folder
        dest_path = self.get_selected_path()
        
        # If selected item is a file, use its parent folder
        if dest_path and os.path.isfile(dest_path):
            dest_path = str(Path(dest_path).parent)
        
        # If no selection or invalid, use current details folder
        if not dest_path or not os.path.isdir(dest_path):
            if hasattr(self, 'current_details_folder') and self.current_details_folder:
                dest_path = self.current_details_folder
            else:
                QMessageBox.warning(self, "Paste", "No destination folder available")
                return
        
        # Check internal clipboard first
        if self.clipboard.get("paths"):
            success_count = 0
            error_count = 0
            
            for source_path in self.clipboard["paths"]:
                source = Path(source_path)
                dest = Path(dest_path) / source.name
                
                try:
                    # Handle existing file/folder
                    if dest.exists():
                        dest = self._get_unique_dest_path(dest)
                    
                    if self.clipboard["operation"] == "cut":
                        shutil.move(str(source), str(dest))
                    else:
                        if source.is_dir():
                            shutil.copytree(str(source), str(dest))
                        else:
                            shutil.copy2(str(source), str(dest))
                    success_count += 1
                    logger.info(f"Pasted {source.name} to {dest_path}")
                except Exception as e:
                    error_count += 1
                    logger.error(f"Failed to paste {source.name}: {e}")
            
            # Clear clipboard after cut operation
            if self.clipboard["operation"] == "cut":
                self.clipboard = {"paths": [], "operation": None}
            
            # Refresh the details view
            if hasattr(self, 'current_details_folder') and self.current_details_folder:
                self.load_folder_contents_in_details(Path(self.current_details_folder))
            
            if error_count > 0:
                QMessageBox.warning(
                    self, "Paste Results",
                    f"Pasted {success_count} item(s).\n{error_count} item(s) failed."
                )
            return
        
        # Check Windows clipboard for files
        clipboard_files = self.get_clipboard_files()
        if clipboard_files:
            success_count = 0
            error_count = 0
            
            for file_path in clipboard_files:
                source = Path(file_path)
                dest = Path(dest_path) / source.name
                
                try:
                    # Handle existing file/folder
                    if dest.exists():
                        dest = self._get_unique_dest_path(dest)
                    
                    if source.is_dir():
                        shutil.copytree(str(source), str(dest))
                    else:
                        shutil.copy2(str(source), str(dest))
                    success_count += 1
                    logger.info(f"Pasted {source.name} to {dest_path}")
                except Exception as e:
                    error_count += 1
                    logger.error(f"Failed to paste {source.name}: {e}")
            
            # Refresh the details view
            if hasattr(self, 'current_details_folder') and self.current_details_folder:
                self.load_folder_contents_in_details(Path(self.current_details_folder))
            
            if error_count > 0:
                QMessageBox.warning(
                    self, "Paste Results", 
                    f"Pasted {success_count} file(s).\n{error_count} file(s) failed."
                )
            return
        
        # No content to paste
        QMessageBox.information(self, "Paste", "No files in clipboard to paste")
    
    def _get_unique_dest_path(self, dest: Path) -> Path:
        """Get a unique destination path by adding (1), (2), etc. if file exists"""
        if not dest.exists():
            return dest
        
        base = dest.stem
        ext = dest.suffix
        parent = dest.parent
        counter = 1
        
        while True:
            new_name = f"{base} ({counter}){ext}"
            new_dest = parent / new_name
            if not new_dest.exists():
                return new_dest
            counter += 1
    
    def handle_dropped_files(self, file_paths: list, dest_folder: str):
        """Handle files dropped from external sources (Windows Explorer, desktop, etc.)"""
        if not file_paths or not dest_folder:
            return
        
        # Check if all files are from the same folder as the destination
        # If so, ignore the operation (don't create copies)
        all_from_same_folder = True
        for file_path in file_paths:
            source = Path(file_path)
            source_parent = str(source.parent)
            if source_parent != dest_folder:
                all_from_same_folder = False
                break
        
        if all_from_same_folder:
            # All files are being dropped in their own folder - ignore operation
            logger.info(f"Ignored drop operation - files dropped in same folder")
            return
        
        # Check if any folders are being moved and ask for confirmation
        folders_to_move = []
        files_to_copy = []
        for file_path in file_paths:
            source = Path(file_path)
            if source.is_dir():
                folders_to_move.append(source)
            else:
                files_to_copy.append(source)
        
        # If folders are being moved, ask for confirmation
        if folders_to_move:
            dest_name = Path(dest_folder).name
            if len(folders_to_move) == 1:
                folder_name = folders_to_move[0].name
                confirm_msg = f"Move folder '{folder_name}' into '{dest_name}'?"
            else:
                folder_names = ", ".join([f.name for f in folders_to_move[:3]])
                if len(folders_to_move) > 3:
                    folder_names += f", ... ({len(folders_to_move)} folders total)"
                confirm_msg = f"Move folders {folder_names} into '{dest_name}'?"
            
            reply = QMessageBox.question(
                self, "Confirm Folder Move",
                confirm_msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                logger.info("User cancelled folder move operation")
                return
        
        success_count = 0
        error_count = 0
        
        for file_path in file_paths:
            source = Path(file_path)
            
            # Skip if source and destination are the same
            if str(source.parent) == dest_folder:
                logger.info(f"Skipping {source.name} - already in destination folder")
                continue
            
            dest = Path(dest_folder) / source.name
            
            try:
                # Handle existing file/folder
                if dest.exists():
                    dest = self._get_unique_dest_path(dest)
                
                if source.is_dir():
                    shutil.copytree(str(source), str(dest))
                else:
                    shutil.copy2(str(source), str(dest))
                success_count += 1
                logger.info(f"Copied {source.name} to {dest_folder}")
            except Exception as e:
                error_count += 1
                logger.error(f"Failed to copy {source.name}: {e}")
        
        # Refresh the details view
        if hasattr(self, 'current_details_folder') and self.current_details_folder:
            self.load_folder_contents_in_details(Path(self.current_details_folder))
        
        # Show result message for multiple files or errors
        if error_count > 0:
            QMessageBox.warning(
                self, "Copy Results", 
                f"Copied {success_count} file(s).\n{error_count} file(s) failed."
            )
        elif success_count > 1:
            logger.info(f"Successfully copied {success_count} files to {dest_folder}")
            
    def rename_file(self):
        """Rename selected file/folder"""
        path = self.get_selected_path()
        if not path:
            return
        
        old_path = Path(path)
        new_name, ok = QInputDialog.getText(
            self, "Rename", f"Rename '{old_path.name}' to:",
            text=old_path.name
        )
        
        if ok and new_name:
            try:
                new_path = old_path.parent / new_name
                old_path.rename(new_path)
                self.refresh_tree()
                QMessageBox.information(self, "Success", f"Renamed to {new_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Rename failed: {e}")
                
    def delete_file(self):
        """Delete selected file(s)/folder(s)"""
        paths = self.get_selected_paths()
        if not paths:
            return
        
        # Build confirmation message
        if len(paths) == 1:
            message = f"Delete '{Path(paths[0]).name}'?"
        else:
            message = f"Delete {len(paths)} selected items?"
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            deleted_count = 0
            error_count = 0
            
            for path in paths:
                try:
                    path_obj = Path(path)
                    if path_obj.is_dir():
                        shutil.rmtree(path)
                    else:
                        path_obj.unlink()
                    deleted_count += 1
                    logger.info(f"Deleted: {path}")
                except Exception as e:
                    error_count += 1
                    logger.error(f"Failed to delete {path}: {e}")
            
            # Refresh the details view to show files are gone
            if hasattr(self, 'current_details_folder') and self.current_details_folder:
                self.load_folder_contents_in_details(Path(self.current_details_folder))
            
            if error_count > 0:
                QMessageBox.warning(
                    self, "Delete Results",
                    f"Deleted {deleted_count} item(s).\n{error_count} item(s) failed."
                )
                
    def refresh_tree(self):
        """Refresh the tree view"""
        self.populate_tree_model()
    
    def refresh_details_view(self):
        """Refresh the current folder contents in the details view"""
        if self.current_details_folder:
            self.load_folder_contents_in_details(Path(self.current_details_folder))
        else:
            logger.warning("No current folder to refresh")
    
    def create_new_folder(self):
        """Create a new folder in the current directory"""
        if not self.current_details_folder:
            QMessageBox.warning(self, "Error", "No folder selected")
            return
        
        # Prompt for folder name
        folder_name, ok = QInputDialog.getText(
            self, "New Folder", 
            "Enter folder name:",
            QLineEdit.EchoMode.Normal,
            "New Folder"
        )
        
        if ok and folder_name:
            try:
                new_folder_path = Path(self.current_details_folder) / folder_name
                new_folder_path.mkdir(parents=False, exist_ok=False)
                logger.info(f"Created folder: {new_folder_path}")
                
                # Refresh the details view to show the new folder
                self.load_folder_contents_in_details(Path(self.current_details_folder))
                
                QMessageBox.information(self, "Success", f"Folder '{folder_name}' created successfully")
            except FileExistsError:
                QMessageBox.warning(self, "Error", f"Folder '{folder_name}' already exists")
            except Exception as e:
                logger.error(f"Failed to create folder: {e}")
                QMessageBox.warning(self, "Error", f"Failed to create folder:\n{str(e)}")
        
    def open_in_explorer(self):
        """Open selected path in Windows Explorer"""
        path = self.get_selected_path()
        if not path:
            return
        
        try:
            path_obj = Path(path)
            if path_obj.is_file():
                path = str(path_obj.parent)
            
            if os.name == 'nt':
                subprocess.run(['explorer', path])
            else:
                subprocess.run(['xdg-open', path])
        except Exception as e:
            logger.error(f"Failed to open explorer: {e}")
            
    def show_context_menu(self, position):
        """Show context menu with quick link options"""
        menu = QMenu()
        
        # Check if item is a custom quick link
        indexes = self.tree_view.selectedIndexes()
        is_custom_link = False
        selected_path = None
        
        if indexes:
            item = self.model.itemFromIndex(self.model.index(indexes[0].row(), 0, indexes[0].parent()))
            if item:
                if item.data(Qt.ItemDataRole.UserRole + 1) == "__CUSTOM_LINK__":
                    is_custom_link = True
                selected_path = item.data(Qt.ItemDataRole.UserRole)
        
        # Remove from Sidebar Bookmarks (only for custom links)
        if is_custom_link:
            remove_from_quick = menu.addAction("❌ Remove from Sidebar Bookmarks")
        else:
            remove_from_quick = None
        
        explorer_action = menu.addAction("📂 Open in Explorer")
        
        menu.addSeparator()
        
        # File operations at the bottom
        cut_action = menu.addAction("✂️ Cut")
        copy_action = menu.addAction("📋 Copy")
        rename_action = menu.addAction("✏️ Rename")
        delete_action = menu.addAction("🗑️ Delete")
        
        menu.addSeparator()
        paste_action = menu.addAction("📌 Paste")
        
        action = menu.exec(self.tree_view.viewport().mapToGlobal(position))
        
        if action == cut_action:
            self.cut_file()
        elif action == copy_action:
            self.copy_file()
        elif action == paste_action:
            self.paste_file()
        elif action == rename_action:
            self.rename_file()
        elif action == delete_action:
            self.delete_file()
        elif action == remove_from_quick:
            self.remove_from_quick_access()
        elif action == explorer_action:
            self.open_in_explorer()
    
    def upload_to_mainframe(self):
        """Upload selected file to mainframe"""
        path = self.get_selected_path()
        if not path:
            return
        
        path_obj = Path(path)
        if not path_obj.is_file():
            QMessageBox.warning(self, "Invalid Selection", "Please select a file to upload")
            return
        
        try:
            from suiteview.ui.dialogs.mainframe_upload_dialog import MainframeUploadDialog
            
            dialog = MainframeUploadDialog(str(path_obj), self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                dialog.perform_upload()
                
        except Exception as e:
            logger.error(f"Failed to show upload dialog: {e}")
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to open upload dialog:\n{str(e)}"
            )
    
    def batch_rename_files(self):
        """Batch rename multiple selected files"""
        paths = self.get_selected_paths()
        
        if not paths:
            QMessageBox.information(
                self,
                "No Selection",
                "Please select one or more files to rename.\n\n"
                "💡 Tip: Hold Ctrl and click to select multiple files"
            )
            return
        
        # Filter out folders (only rename files in batch)
        file_paths = [p for p in paths if Path(p).is_file()]
        
        if not file_paths:
            QMessageBox.warning(
                self,
                "No Files Selected",
                "Please select files (not folders) to batch rename."
            )
            return
        
        try:
            from suiteview.ui.dialogs.batch_rename_dialog import BatchRenameDialog
            
            dialog = BatchRenameDialog(file_paths, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                if dialog.perform_rename():
                    self.refresh_tree()
                    
        except Exception as e:
            logger.error(f"Failed to show batch rename dialog: {e}")
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to open batch rename dialog:\n{str(e)}"
            )
    
    def load_quick_links(self):
        """Load sidebar quick links from unified bookmarks.json file
        
        Reads from bars.sidebar in the unified format:
        {
            'bars': {
                'sidebar': {
                    'categories': {'Category Name': [{'name': '...', 'path': '...', 'type': 'file|folder'}, ...]},
                    'items': [...],
                    'category_colors': {...}
                }
            },
            'version': 2
        }
        """
        try:
            if self.bookmarks_file.exists():
                with open(self.bookmarks_file, 'r') as f:
                    data = json.load(f)
                    
                    # Read from unified format: bars.sidebar
                    if 'bars' in data and 'sidebar' in data['bars']:
                        sidebar_data = data['bars']['sidebar']
                        # Ensure structure has required keys
                        if 'categories' not in sidebar_data:
                            sidebar_data['categories'] = {}
                        if 'items' not in sidebar_data:
                            sidebar_data['items'] = []
                        if 'category_colors' not in sidebar_data:
                            sidebar_data['category_colors'] = {}
                        return sidebar_data
                    
        except Exception as e:
            logger.error(f"Failed to load quick links: {e}")
        
        # Default structure
        return {
            'categories': {},
            'items': [],
            'category_colors': {}
        }
    
    def save_quick_links(self):
        """Save sidebar quick links via centralized bookmark manager"""
        try:
            self._bookmark_manager.save()
        except Exception as e:
            logger.error(f"Failed to save quick links: {e}")
    
    def get_quick_links_paths(self):
        """Get flat list of paths from quick links for compatibility"""
        paths = []
        for item in self.custom_quick_links.get('items', []):
            if item.get('type') == 'bookmark':
                path = item.get('data', {}).get('path')
                if path:
                    paths.append(path)
        return paths
    
    def is_path_in_quick_links(self, path):
        """Check if a path is already in quick links (at top level or in any category)"""
        path_str = str(Path(path).resolve())
        
        # Check top-level items
        for item in self.custom_quick_links.get('items', []):
            if item.get('type') == 'bookmark':
                if item.get('data', {}).get('path') == path_str:
                    return True
        
        # Check items in categories
        for cat_name, cat_items in self.custom_quick_links.get('categories', {}).items():
            for item in cat_items:
                if item.get('path') == path_str:
                    return True
        
        return False
    
    def add_bookmark_to_quick_links(self, path, insert_at=None):
        """Add a bookmark path to Quick Links items"""
        path_str = str(Path(path).resolve())
        path_obj = Path(path_str)
        
        # Check if already exists
        if self.is_path_in_quick_links(path_str):
            return False
        
        new_item = {
            'type': 'bookmark',
            'data': {
                'name': path_obj.name,
                'path': path_str,
                'type': 'folder' if path_obj.is_dir() else 'file'
            }
        }
        
        if insert_at is not None and 0 <= insert_at <= len(self.custom_quick_links.get('items', [])):
            self.custom_quick_links['items'].insert(insert_at, new_item)
        else:
            self.custom_quick_links['items'].append(new_item)
        
        self.save_quick_links()
        return True
    
    def remove_bookmark_from_quick_links(self, path):
        """Remove a bookmark from Quick Links by path"""
        path_str = str(Path(path).resolve())
        
        # Check top-level items
        items = self.custom_quick_links.get('items', [])
        for i, item in enumerate(items):
            if item.get('type') == 'bookmark':
                if item.get('data', {}).get('path') == path_str:
                    items.pop(i)
                    self.save_quick_links()
                    return True
        
        return False
    
    def add_category_to_quick_links(self, category_name, category_items=None, insert_at=None):
        """Add a category to Quick Links"""
        # Check if category already exists
        if category_name in self.custom_quick_links.get('categories', {}):
            return False
        
        # Add to categories dict
        self.custom_quick_links['categories'][category_name] = category_items or []
        
        # Add to items list
        new_item = {'type': 'category', 'name': category_name}
        
        if insert_at is not None and 0 <= insert_at <= len(self.custom_quick_links.get('items', [])):
            self.custom_quick_links['items'].insert(insert_at, new_item)
        else:
            self.custom_quick_links['items'].append(new_item)
        
        self.save_quick_links()
        return True
    
    def remove_category_from_quick_links(self, category_name):
        """Remove a category from Quick Links"""
        # Remove from categories dict
        if category_name in self.custom_quick_links.get('categories', {}):
            del self.custom_quick_links['categories'][category_name]
        
        # Remove color (it's been transferred elsewhere or deleted)
        category_colors = self.custom_quick_links.get('category_colors', {})
        if category_name in category_colors:
            del category_colors[category_name]
        
        # Remove from items list
        items = self.custom_quick_links.get('items', [])
        for i, item in enumerate(items):
            if item.get('type') == 'category' and item.get('name') == category_name:
                items.pop(i)
                break
        
        self.save_quick_links()
        return True
    
    def rename_category_in_quick_links(self, old_name, new_name):
        """Rename a category in Quick Links"""
        if old_name == new_name:
            return False
        
        # Check if new name already exists
        if new_name in self.custom_quick_links.get('categories', {}):
            return False
        
        # Rename in categories dict
        if old_name in self.custom_quick_links.get('categories', {}):
            self.custom_quick_links['categories'][new_name] = self.custom_quick_links['categories'].pop(old_name)
            
            # Update category field in all bookmarks within the category
            for bookmark in self.custom_quick_links['categories'][new_name]:
                bookmark['category'] = new_name
        
        # Update in items list
        items = self.custom_quick_links.get('items', [])
        for item in items:
            if item.get('type') == 'category' and item.get('name') == old_name:
                item['name'] = new_name
                break
        
        self.save_quick_links()
        return True
    
    def load_hidden_onedrive(self):
        """Load hidden OneDrive paths from JSON file"""
        try:
            if self.hidden_onedrive_file.exists():
                with open(self.hidden_onedrive_file, 'r') as f:
                    paths = json.load(f)
                    # Normalize paths to lowercase for comparison
                    return set(p.lower() for p in paths)
        except Exception as e:
            logger.error(f"Failed to load hidden OneDrive paths: {e}")
        return set()
    
    def save_hidden_onedrive(self):
        """Save hidden OneDrive paths to JSON file"""
        try:
            # Ensure directory exists
            self.hidden_onedrive_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.hidden_onedrive_file, 'w') as f:
                json.dump(list(self.hidden_onedrive_paths), f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save hidden OneDrive paths: {e}")
    
    def hide_onedrive_path(self, path):
        """Hide a OneDrive path from Quick Links"""
        path_str = str(path).lower()
        if path_str not in self.hidden_onedrive_paths:
            self.hidden_onedrive_paths.add(path_str)
            self.save_hidden_onedrive()
            # Refresh tree to remove it
            self.load_tree()
            QMessageBox.information(self, "OneDrive Hidden", f"Hidden from Quick Links:\n{Path(path).name}\n\nTo unhide, delete:\n{self.hidden_onedrive_file}")
    
    def load_column_widths(self):
        """Load column widths from JSON file"""
        try:
            if self.column_widths_file.exists():
                with open(self.column_widths_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load column widths: {e}")
        return {}
    
    def save_column_widths(self):
        """Save column widths to JSON file"""
        try:
            # Ensure directory exists
            self.column_widths_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.column_widths_file, 'w') as f:
                json.dump(self.column_widths, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save column widths: {e}")
    
    def on_column_resized(self, logical_index, old_size, new_size):
        """Handle column resize event - save new width"""
        self.column_widths[f'col_{logical_index}'] = new_size
        self.save_column_widths()
    
    def load_panel_widths(self):
        """Load panel widths from JSON file"""
        try:
            if self.panel_widths_file.exists():
                with open(self.panel_widths_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load panel widths: {e}")
        return {}
    
    def save_panel_widths(self):
        """Save panel widths to JSON file"""
        try:
            # Ensure directory exists
            self.panel_widths_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.panel_widths_file, 'w') as f:
                json.dump(self.panel_widths, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save panel widths: {e}")
    
    def on_splitter_moved(self, pos, index):
        """Handle splitter moved event - save panel widths"""
        if hasattr(self, 'main_splitter'):
            sizes = self.main_splitter.sizes()
            if len(sizes) >= 2:
                self.panel_widths['left_panel'] = sizes[0]
                self.panel_widths['middle_panel'] = sizes[1]
                if len(sizes) >= 3:
                    self.panel_widths['right_panel'] = sizes[2]
                self.save_panel_widths()
    
    def update_details_footer(self, timing_ms=None):
        """Update the details footer with item count and optional timing"""
        if hasattr(self, 'details_footer') and hasattr(self, 'details_sort_proxy'):
            count = self.details_sort_proxy.rowCount()
            
            # Left side: item count
            if count == 0:
                left_text = ""
            elif count == 1:
                left_text = "1 item"
            else:
                left_text = f"{count} items"
            
            # Right side: timing info
            if timing_ms is not None:
                right_text = f"Loaded in {timing_ms}ms"
            else:
                right_text = ""
            
            # Combine with spacing
            if left_text and right_text:
                full_text = f"{left_text}" + " " * 20 + right_text
            else:
                full_text = left_text + right_text
            
            self.details_footer.setText(full_text)
    
    def add_to_quick_access(self):
        """Add selected item to Quick Access"""
        path = self.get_selected_path()
        if not path:
            return
        
        # Use the new helper method
        if self.add_bookmark_to_quick_links(path):
            # Refresh Quick Links panel if it exists
            if hasattr(self, 'refresh_quick_links'):
                self.refresh_quick_links()
    
    def remove_from_quick_access(self):
        """Remove selected item from Quick Access"""
        path = self.get_selected_path()
        if not path:
            return
        
        if self.is_path_in_quick_links(path):
            if self.remove_bookmark_from_quick_links(path):
                # Refresh Quick Links panel if it exists
                if hasattr(self, 'refresh_quick_links'):
                    self.refresh_quick_links()
                QMessageBox.information(self, "Bookmarks", f"Removed from Bookmarks:\n{Path(path).name}")
        else:
            QMessageBox.warning(self, "Bookmarks", "This item is not in Bookmarks")
    
    def remove_quick_link_by_path(self, path):
        """Remove a specific path from Quick Links (used by context menu)"""
        from suiteview.ui.dialogs.shortcuts_dialog import show_compact_confirm
        
        if self.is_path_in_quick_links(path):
            # Show confirmation dialog
            if show_compact_confirm(self, "Remove Bookmark", f"Remove '{Path(path).name}'?"):
                self.remove_bookmark_from_quick_links(path)
                
                # Refresh Quick Links panel if it exists
                if hasattr(self, 'refresh_quick_links'):
                    self.refresh_quick_links()
    
    def add_to_quick_links(self, path, insert_at=None):
        """Add a path to Quick Links"""
        # Use the new helper method
        if self.add_bookmark_to_quick_links(path, insert_at=insert_at):
            # Refresh Quick Links panel if it exists
            if hasattr(self, 'refresh_quick_links'):
                self.refresh_quick_links()
    
    def open_bookmarks_dialog(self):
        """Open/close the Bookmarks panel dialog (toggle)"""
        from suiteview.ui.dialogs.shortcuts_dialog import BookmarksDialog
        from PyQt6.QtCore import QPoint
        
        # Check if dialog is already open - if so, close it
        if hasattr(self, '_bookmarks_dialog') and self._bookmarks_dialog is not None:
            try:
                if self._bookmarks_dialog.isVisible():
                    self._bookmarks_dialog.close()
                    self._bookmarks_dialog = None
                    return
            except RuntimeError:
                # Dialog was deleted
                self._bookmarks_dialog = None
        
        # Calculate position right under the Bookmarks button
        if hasattr(self, 'toolbar'):
            toolbar_geo = self.toolbar.geometry()
            # Get toolbar's global position
            toolbar_global = self.toolbar.mapToGlobal(QPoint(0, 0))
            # Position dialog at start of toolbar, below it
            button_pos = QPoint(toolbar_global.x(), toolbar_global.y() + toolbar_geo.height())
            self._bookmarks_dialog = BookmarksDialog(self, button_pos)
        else:
            self._bookmarks_dialog = BookmarksDialog(self)
            # Fallback positioning
            if self.parent():
                parent_geo = self.parent().geometry()
                self._bookmarks_dialog.move(parent_geo.left() + 10, parent_geo.top() + 50)
        
        # Connect folder navigation signal
        self._bookmarks_dialog.navigate_to_path.connect(self.navigate_to_bookmark_folder)
        
        # Clear reference when dialog closes
        self._bookmarks_dialog.finished.connect(lambda: setattr(self, '_bookmarks_dialog', None))
        
        # Use show() instead of exec() so it's non-blocking and can be toggled
        self._bookmarks_dialog.show()
    
    def navigate_to_bookmark_folder(self, folder_path):
        """Navigate to a folder from bookmark click"""
        try:
            path = Path(folder_path)
            if path.exists() and path.is_dir():
                # Load folder in details view
                self.load_folder_contents_in_details(path)
        except Exception as e:
            logger.error(f"Failed to navigate to bookmark folder: {e}")
    
    def add_to_bookmarks(self, path, name=None):
        """Add a path to Bookmarks via the bookmark bar"""
        # Use the bookmark bar's add functionality
        if hasattr(self, 'bookmark_bar'):
            # Pre-populate the dialog if it's a specific path
            from suiteview.ui.dialogs.shortcuts_dialog import AddBookmarkDialog
            
            categories = list(self.bookmark_bar.bookmarks_data['categories'].keys())
            
            dialog = AddBookmarkDialog(categories, self)
            
            # Pre-fill with the path
            if name:
                dialog.name_input.setText(name)
            else:
                dialog.name_input.setText(Path(path).name)
            dialog.path_input.setText(str(path))
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                bookmark = dialog.get_bookmark_data()
                if bookmark['name'] and bookmark['path']:
                    category = bookmark['category']
                    target_bar_id = bookmark.get('target_bar_id')
                    
                    # Remove the category and target_bar_id from bookmark data for storage
                    del bookmark['category']
                    if 'target_bar_id' in bookmark:
                        del bookmark['target_bar_id']
                    
                    if category == "__BAR__" and target_bar_id is not None:
                        # Add directly to a specific bookmark bar
                        from suiteview.ui.widgets.bookmark_data_manager import get_bookmark_manager
                        from suiteview.ui.widgets.bookmark_widgets import BookmarkContainerRegistry
                        
                        manager = get_bookmark_manager()
                        bar_data = manager.get_bar_data(target_bar_id)
                        
                        # Add to the bar's items
                        if 'items' not in bar_data:
                            bar_data['items'] = []
                        bar_data['items'].append({
                            'type': 'bookmark',
                            'data': bookmark
                        })
                        
                        # Save and refresh the target bar
                        manager.save()
                        target_container = BookmarkContainerRegistry.get(target_bar_id)
                        if target_container:
                            target_container.refresh()
                    else:
                        # Add to category (in the current bar, bar 0)
                        if category not in self.bookmark_bar.bookmarks_data['categories']:
                            self.bookmark_bar.bookmarks_data['categories'][category] = []
                            # Add category to bar_items if it's new
                            self.bookmark_bar.bookmarks_data['bar_items'].append({
                                'type': 'category',
                                'name': category
                            })
                        self.bookmark_bar.bookmarks_data['categories'][category].append(bookmark)
                        
                        self.bookmark_bar.save_bookmarks()
                        self.bookmark_bar.refresh_bookmarks()
    
    def print_directory_to_excel(self):
        """Export current directory structure directly to Excel (no file saved)"""
        # Get current directory from selected item or use current details folder
        path = None
        
        # Try to use current details folder first
        if hasattr(self, 'current_details_folder') and self.current_details_folder:
            path = self.current_details_folder
        else:
            # Get from selection
            path = self.get_selected_path()
        
        # If no selection or selected is a file, try to find a folder
        if not path or (path and not Path(path).is_dir()):
            # Default to user's home directory or OneDrive
            onedrive_paths = self.get_onedrive_paths()
            if onedrive_paths:
                path = str(onedrive_paths[0])
            else:
                path = str(Path.home())
        
        # Show dialog
        dialog = PrintDirectoryDialog(path, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            options = dialog.get_options()

            # Create progress dialog (indeterminate; recursive scope isn't known up-front)
            progress = QProgressDialog("Collecting directory information...", "Cancel", 0, 0, self)
            progress.setWindowTitle("Print Directory")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.setAutoClose(False)
            progress.setAutoReset(False)
            
            # Collect data in current thread (faster for small directories)
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                
                root = Path(path)
                data = []
                files_scanned = 0
                folders_scanned = 0
                last_ui_update = 0.0
                t0 = time.monotonic()
                canceled = False
                
                # Headers
                headers = ['Level', 'Type', 'Name', 'Full Path', 'Size', 'Modified', 'Extension']
                data.append(headers)
                
                def _maybe_update_progress(status: str) -> None:
                    nonlocal last_ui_update
                    now = time.monotonic()
                    # Throttle UI updates to keep scanning fast
                    if (now - last_ui_update) < 0.15:
                        return
                    last_ui_update = now
                    progress.setLabelText(status)
                    QApplication.processEvents()
                
                if options['include_subdirs']:
                    # Use os.walk for speed; compute a real-time estimate using the pending directory count.
                    progress.setLabelText("Scanning folders...")

                    for dirpath, dirnames, filenames in os.walk(root):
                        current = Path(dirpath)
                        level = len(current.relative_to(root).parts)

                        if progress.wasCanceled():
                            canceled = True
                            # Prevent descending into more folders, and stop walking.
                            dirnames[:] = []
                            filenames[:] = []
                            break
                        
                        # Add directories
                        for dirname in sorted(dirnames):
                            if progress.wasCanceled():
                                canceled = True
                                break
                            folder_path = current / dirname
                            try:
                                stat = folder_path.stat()
                                data.append([
                                    level,
                                    'Folder',
                                    dirname,
                                    str(folder_path),
                                    '',
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                                    ''
                                ])
                                folders_scanned += 1
                            except:
                                pass

                        if canceled:
                            dirnames[:] = []
                            filenames[:] = []
                            break
                        
                        # Add files
                        for filename in sorted(filenames):
                            if progress.wasCanceled():
                                canceled = True
                                break
                            file_path = current / filename
                            try:
                                stat = file_path.stat()
                                size = stat.st_size
                                size_str = f"{size:,} B" if size < 1024 else f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.1f} MB"
                                data.append([
                                    level,
                                    'File',
                                    filename,
                                    str(file_path),
                                    size_str,
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                                    file_path.suffix
                                ])
                                files_scanned += 1
                            except:
                                pass

                        if canceled:
                            dirnames[:] = []
                            filenames[:] = []
                            break

                        status = f"Scanning... {folders_scanned:,} folders, {files_scanned:,} files"
                        _maybe_update_progress(status)

                        if canceled:
                            break
                else:
                    # Just immediate contents
                    try:
                        items = list(root.iterdir())
                    except Exception:
                        items = []
                    items = sorted(items, key=lambda x: (not x.is_dir(), x.name.lower()))

                    for item in items:
                        if progress.wasCanceled():
                            canceled = True
                            break
                        try:
                            stat = item.stat()
                            if item.is_dir():
                                data.append([
                                    0, 'Folder', item.name, str(item), '',
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"), ''
                                ])
                                folders_scanned += 1
                            else:
                                size = stat.st_size
                                size_str = f"{size:,} B" if size < 1024 else f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.1f} MB"
                                data.append([
                                    0, 'File', item.name, str(item), size_str,
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"), item.suffix
                                ])
                                files_scanned += 1
                        except:
                            pass

                        status = f"Scanning... {folders_scanned:,} folders, {files_scanned:,} files"
                        _maybe_update_progress(status)
                
                # Scanning finished (possibly canceled)
                if canceled:
                    progress.setLabelText("Canceled — opening Excel with partial results...")
                else:
                    progress.setLabelText("Opening Excel...")
                QApplication.processEvents()
                
                # Open Excel using COM automation (Windows only)
                if os.name == 'nt':
                    try:
                        from win32com.client import dynamic
                        
                        # Create Excel instance using dynamic dispatch to avoid gen_py cache issues
                        excel = dynamic.Dispatch('Excel.Application')
                        excel.Visible = True
                        
                        # Add a new workbook
                        wb = excel.Workbooks.Add()
                        ws = wb.Worksheets(1)
                        ws.Name = "Directory Contents"
                        
                        # Write all data at once (much faster than cell-by-cell)
                        progress.setLabelText("Populating Excel...")
                        if len(data) > 0:
                            # Calculate the range needed
                            num_rows = len(data)
                            num_cols = len(data[0])
                            
                            # Define the range (e.g., "A1:G100")
                            end_col_letter = chr(64 + num_cols)  # A=65, so 64+1=A, 64+7=G, etc.
                            range_address = f"A1:{end_col_letter}{num_rows}"
                            
                            # Write entire data array at once
                            ws.Range(range_address).Value = data
                            
                            QApplication.processEvents()
                        
                        # Format headers
                        header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, len(headers)))
                        header_range.Font.Bold = True
                        header_range.Font.Color = 0xFFFFFF  # White
                        header_range.Interior.Color = 0x926636  # Dark blue
                        header_range.HorizontalAlignment = -4108  # xlCenter
                        
                        progress.setLabelText("Formatting...")
                        QApplication.processEvents()
                        
                        # Auto-fit columns
                        ws.Columns.AutoFit()
                        
                        # Freeze top row
                        ws.Range("A2").Select()
                        excel.ActiveWindow.FreezePanes = True
                        
                        # Select cell A1
                        ws.Range("A1").Select()
                        
                        progress.close()
                        
                        # Don't show message box - just let Excel open
                        # Release COM objects to prevent app freezing
                        ws = None
                        wb = None
                        excel = None
                        
                    except ImportError:
                        progress.close()
                        QMessageBox.critical(
                            self,
                            "Excel COM Error",
                            "Could not access Excel via COM automation.\n\n"
                            "Make sure Microsoft Excel is installed and pywin32 is available."
                        )
                    except Exception as e:
                        progress.close()
                        QMessageBox.critical(
                            self,
                            "Excel Error",
                            f"Failed to open Excel:\n\n{str(e)}"
                        )
                else:
                    progress.close()
                    QMessageBox.warning(
                        self,
                        "Not Supported",
                        "Direct Excel opening is only supported on Windows.\n\n"
                        "This feature requires Microsoft Excel and COM automation."
                    )
                    
            except Exception as e:
                progress.close()
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to collect directory information:\n\n{str(e)}"
                )
