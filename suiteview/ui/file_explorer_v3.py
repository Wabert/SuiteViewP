"""
Enhanced File Explorer with Custom Model - OneDrive at Top Level
Shows Quick Access shortcuts (OneDrive) alongside system drives
"""

import os
import sys
import shutil
import subprocess
import json
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (QTreeView, QVBoxLayout, QHBoxLayout, QWidget, 
                              QHeaderView, QToolBar, QMessageBox, QInputDialog, 
                              QTextEdit, QPushButton, QSplitter, QLabel, QMenu,
                              QDialog, QDialogButtonBox, QCheckBox, QLineEdit,
                              QProgressDialog, QFileDialog, QFrame)
from PyQt6.QtGui import QIcon, QAction, QStandardItemModel, QStandardItem
from PyQt6.QtCore import Qt, QModelIndex, QThread, pyqtSignal

import logging

logger = logging.getLogger(__name__)


class DirectoryExportThread(QThread):
    """Background thread for exporting directory to Excel"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(str, bool)
    
    def __init__(self, root_path, include_subdirs, output_path):
        super().__init__()
        self.root_path = root_path
        self.include_subdirs = include_subdirs
        self.output_path = output_path
        
    def run(self):
        """Export directory structure to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Directory Contents"
            
            # Headers
            headers = ['Level', 'Type', 'Name', 'Full Path', 'Size', 'Modified', 'Extension']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions['A'].width = 8   # Level
            ws.column_dimensions['B'].width = 10  # Type
            ws.column_dimensions['C'].width = 40  # Name
            ws.column_dimensions['D'].width = 60  # Full Path
            ws.column_dimensions['E'].width = 12  # Size
            ws.column_dimensions['F'].width = 18  # Modified
            ws.column_dimensions['G'].width = 10  # Extension
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            row_num = 2
            total_files = 0
            total_size = 0
            
            # Walk directory
            root = Path(self.root_path)
            
            if self.include_subdirs:
                items = []
                for dirpath, dirnames, filenames in os.walk(root):
                    current = Path(dirpath)
                    level = len(current.relative_to(root).parts)
                    
                    # Add directories
                    for dirname in sorted(dirnames):
                        items.append((current / dirname, level, True))
                    
                    # Add files
                    for filename in sorted(filenames):
                        items.append((current / filename, level, False))
                        
                    self.progress.emit(len(items), f"Scanning... {len(items)} items found")
            else:
                # Just immediate contents
                items = []
                for item in sorted(root.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                    items.append((item, 0, item.is_dir()))
            
            # Add items to Excel
            for idx, (item_path, level, is_dir) in enumerate(items):
                try:
                    stats = item_path.stat()
                    
                    # Type
                    item_type = "Folder" if is_dir else "File"
                    
                    # Name with indentation
                    indent = "  " * level
                    name = f"{indent}{item_path.name}"
                    
                    # Size
                    if is_dir:
                        size_str = ""
                    else:
                        size_bytes = stats.st_size
                        total_size += size_bytes
                        total_files += 1
                        
                        if size_bytes < 1024:
                            size_str = f"{size_bytes} B"
                        elif size_bytes < 1024 * 1024:
                            size_str = f"{size_bytes / 1024:.1f} KB"
                        elif size_bytes < 1024 * 1024 * 1024:
                            size_str = f"{size_bytes / (1024 * 1024):.1f} MB"
                        else:
                            size_str = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
                    
                    # Modified date
                    modified = datetime.fromtimestamp(stats.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Extension
                    extension = item_path.suffix.upper() if not is_dir and item_path.suffix else ""
                    
                    # Add row
                    ws.append([
                        level,
                        item_type,
                        name,
                        str(item_path),
                        size_str,
                        modified,
                        extension
                    ])
                    
                    # Style folder rows
                    if is_dir:
                        for col in range(1, 8):
                            cell = ws.cell(row=row_num, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    row_num += 1
                    
                    if idx % 10 == 0:
                        self.progress.emit(idx + 1, f"Processing... {idx + 1}/{len(items)}")
                        
                except (PermissionError, OSError):
                    continue
            
            # Add summary section
            row_num += 2
            ws.cell(row=row_num, column=1, value="Summary:")
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Files:")
            ws.cell(row=row_num, column=2, value=total_files)
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total Size:")
            if total_size < 1024 * 1024 * 1024:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024):.2f} MB")
            else:
                ws.cell(row=row_num, column=2, value=f"{total_size / (1024 * 1024 * 1024):.2f} GB")
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Export Date:")
            ws.cell(row=row_num, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            row_num += 1
            ws.cell(row=row_num, column=1, value="Source Path:")
            ws.cell(row=row_num, column=2, value=str(root))
            
            # Save workbook
            wb.save(self.output_path)
            
            self.finished.emit(self.output_path, True)
            
        except Exception as e:
            logger.error(f"Failed to export directory: {e}")
            self.finished.emit(str(e), False)


class PrintDirectoryDialog(QDialog):
    """Dialog for Print Directory options"""
    
    def __init__(self, current_path, parent=None):
        super().__init__(parent)
        self.current_path = current_path
        self.setWindowTitle("Print Directory to Excel")
        self.setModal(True)
        self.resize(550, 220)
        
        layout = QVBoxLayout(self)
        
        # Current path
        path_label = QLabel(f"<b>Directory:</b> {current_path}")
        path_label.setWordWrap(True)
        path_label.setStyleSheet("padding: 10px; background-color: #f0f0f0; border-radius: 5px;")
        layout.addWidget(path_label)
        
        layout.addSpacing(15)
        
        # Options
        options_label = QLabel("<b>Export Options:</b>")
        layout.addWidget(options_label)
        
        # Include subdirectories checkbox
        self.include_subdirs_cb = QCheckBox("Include all subdirectories and files (recursive)")
        self.include_subdirs_cb.setChecked(True)
        layout.addWidget(self.include_subdirs_cb)
        
        info_label = QLabel("📝 The directory listing will open directly in Excel as an unsaved workbook.")
        info_label.setStyleSheet("color: #666; font-size: 9pt; padding-left: 20px;")
        layout.addWidget(info_label)
        
        layout.addSpacing(10)
        
        note_label = QLabel("💡 Use 'Save As' in Excel if you want to keep the file.")
        note_label.setStyleSheet("color: #0066cc; font-size: 9pt; padding-left: 20px;")
        layout.addWidget(note_label)
        
        layout.addStretch()
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_options(self):
        """Return selected options"""
        return {
            'include_subdirs': self.include_subdirs_cb.isChecked()
        }


class FileExplorerV3(QWidget):
    """
    File Explorer with custom model showing OneDrive at top level
    Features:
    - OneDrive shortcuts at root level (like Windows Explorer)
    - System drives (C: D: etc.)
    - Lazy loading of directory contents
    - File operations (cut, copy, paste, rename, delete)
    - File preview pane with mainframe upload button
    """
    
    def __init__(self):
        super().__init__()
        self.current_file_path = None
        self.current_file_content = None
        self.current_details_folder = None  # Track current folder in details view
        self.clipboard = {"path": None, "operation": None}
        
        # Load custom quick links
        self.quick_links_file = Path.home() / ".suiteview" / "quick_links.json"
        self.custom_quick_links = self.load_quick_links()
        
        # Load hidden OneDrive paths
        self.hidden_onedrive_file = Path.home() / ".suiteview" / "hidden_onedrive.json"
        self.hidden_onedrive_paths = self.load_hidden_onedrive()
        
        # Column width settings file
        self.column_widths_file = Path.home() / ".suiteview" / "column_widths.json"
        self.column_widths = self.load_column_widths()
        
        self.init_ui()
        
    def init_ui(self):
        """Initialize the UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Create toolbar
        toolbar = self.create_toolbar()
        main_layout.addWidget(toolbar)
        
        # Create splitter for tree (left) and details (right)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Create tree panel (left side - folder navigation only)
        tree_panel = self.create_tree_panel()
        splitter.addWidget(tree_panel)
        
        # Create details panel (right side - folder contents with details)
        details_panel = self.create_details_panel()
        splitter.addWidget(details_panel)
        
        # Set initial sizes (30% tree, 70% details)
        splitter.setSizes([300, 700])
        
        main_layout.addWidget(splitter)
        
    def create_toolbar(self):
        """Create toolbar with file operations"""
        self.toolbar = QToolBar()
        self.toolbar.setMovable(False)
        
        # Bookmarks button (prominent at the start)
        self.bookmarks_action = QAction("Bookmarks", self)
        self.bookmarks_action.setToolTip("Open Bookmarks panel")
        self.bookmarks_action.triggered.connect(self.open_bookmarks_dialog)
        self.toolbar.addAction(self.bookmarks_action)
        
        self.toolbar.addSeparator()
        
        # Open in Explorer
        explorer_action = QAction("📂 Open in Explorer", self)
        explorer_action.triggered.connect(self.open_in_explorer)
        self.toolbar.addAction(explorer_action)
        
        self.toolbar.addSeparator()
        
        # Print Directory to Excel
        print_dir_action = QAction("📊 Print Directory", self)
        print_dir_action.setToolTip("Export directory structure to Excel")
        print_dir_action.triggered.connect(self.print_directory_to_excel)
        self.toolbar.addAction(print_dir_action)
        
        self.toolbar.addSeparator()
        
        # Batch Rename
        batch_rename_action = QAction("✏️📦 Batch Rename", self)
        batch_rename_action.setToolTip("Rename multiple selected files")
        batch_rename_action.triggered.connect(self.batch_rename_files)
        self.toolbar.addAction(batch_rename_action)
        
        return self.toolbar
        
    def create_tree_panel(self):
        """Create the tree view with custom model (folders only, name column only)"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header = QLabel("📁 Folders")
        header.setStyleSheet("padding: 8px; font-weight: bold; font-size: 11pt;")
        layout.addWidget(header)
        
        # Create tree view
        self.tree_view = QTreeView()
        self.tree_view.setAnimated(True)
        self.tree_view.setIndentation(20)
        self.tree_view.setHeaderHidden(True)  # Hide header for tree view
        self.tree_view.setSelectionMode(QTreeView.SelectionMode.SingleSelection)
        
        # Create custom model
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(['Name'])  # Only name column for tree
        self.model.setColumnCount(1)  # Explicitly set to 1 column only
        
        # Populate model with Quick Access and drives
        self.populate_tree_model()
        
        self.tree_view.setModel(self.model)
        
        # Ensure only the Name column is visible - hide any extra columns
        for col in range(1, 10):  # Hide columns 1-9 if they somehow exist
            self.tree_view.setColumnHidden(col, True)
        
        # Set the tree to only resize the first column
        self.tree_view.header().setStretchLastSection(False)
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        
        # Connect signals
        self.tree_view.expanded.connect(self.on_item_expanded)
        self.tree_view.clicked.connect(self.on_tree_item_clicked)
        
        # Context menu
        self.tree_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree_view.customContextMenuRequested.connect(self.show_tree_context_menu)
        
        layout.addWidget(self.tree_view)
        
        return widget
        
    def populate_tree_model(self):
        """Populate tree with OneDrive, custom links, and system drives (folders only)"""
        self.model.clear()
        self.model.setHorizontalHeaderLabels(['Name'])
        self.model.setColumnCount(1)  # Explicitly set to 1 column only
        
        # Add OneDrive folders (excluding hidden ones)
        onedrive_paths = self.get_onedrive_paths()
        for od_path in onedrive_paths:
            if od_path.exists() and str(od_path).lower() not in self.hidden_onedrive_paths:
                item = self.create_tree_folder_item(od_path, icon="⭐")
                self.model.appendRow(item)
        
        # Add custom quick links (only folders)
        for link_path in self.custom_quick_links:
            path = Path(link_path)
            if path.exists() and path.is_dir():
                item = self.create_tree_folder_item(path, icon="📌")
                # Mark as custom link
                item.setData("__CUSTOM_LINK__", Qt.ItemDataRole.UserRole + 1)
                self.model.appendRow(item)
        
        # Add separator (visual only)
        if onedrive_paths or self.custom_quick_links:
            separator = QStandardItem("─" * 30)
            separator.setEnabled(False)
            self.model.appendRow(separator)
        
        # Add system drives
        drives = self.get_system_drives()
        for drive in drives:
            item = self.create_tree_folder_item(Path(drive), icon="💾")
            self.model.appendRow(item)
        
        # Re-hide any extra columns and ensure proper column sizing
        for col in range(1, 10):
            self.tree_view.setColumnHidden(col, True)
        self.tree_view.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            
    def get_onedrive_paths(self):
        """Get all OneDrive paths (deduplicated)"""
        onedrive_paths = []
        seen_paths = set()
        
        # Check environment variables - prioritize business OneDrive
        for env_var in ['OneDriveCommercial', 'OneDrive', 'OneDriveConsumer']:
            path = os.environ.get(env_var)
            if path and os.path.exists(path):
                path_obj = Path(path).resolve()  # Resolve to absolute path
                path_str = str(path_obj).lower()  # Normalize for comparison
                if path_str not in seen_paths:
                    seen_paths.add(path_str)
                    onedrive_paths.append(path_obj)
        
        # Check common locations - only if not already found
        home = Path.home()
        possible_paths = [
            home / "OneDrive - American National Insurance Company",
            home / "OneDrive",
        ]
        
        for path in possible_paths:
            if path.exists():
                path_resolved = path.resolve()
                path_str = str(path_resolved).lower()
                if path_str not in seen_paths:
                    seen_paths.add(path_str)
                    onedrive_paths.append(path_resolved)
        
        return onedrive_paths
        
    def get_system_drives(self):
        """Get all system drives"""
        drives = []
        
        if os.name == 'nt':  # Windows
            import string
            from ctypes import windll
            
            bitmask = windll.kernel32.GetLogicalDrives()
            for letter in string.ascii_uppercase:
                if bitmask & 1:
                    drive_path = f"{letter}:\\"
                    if os.path.exists(drive_path):
                        drives.append(drive_path)
                bitmask >>= 1
        else:  # Unix-like
            drives.append("/")
        
        return drives
        
    def create_tree_folder_item(self, path, icon="📁"):
        """Create a single-column folder item for the tree view"""
        path = Path(path)
        
        # Use more descriptive icons for special folders
        folder_name = path.name.lower() if path.name else ""
        
        # Special folder icons
        if not icon or icon == "📁":  # Only override if default folder icon
            if "desktop" in folder_name:
                icon = "🖥️"
            elif "documents" in folder_name:
                icon = "📄"
            elif "downloads" in folder_name:
                icon = "⬇️"
            elif "pictures" in folder_name or "photos" in folder_name:
                icon = "🖼️"
            elif "music" in folder_name:
                icon = "🎵"
            elif "videos" in folder_name:
                icon = "🎬"
            elif "onedrive" in folder_name:
                icon = "☁️"
            elif folder_name in ["program files", "program files (x86)"]:
                icon = "⚙️"
            elif folder_name == "windows":
                icon = "🪟"
            elif folder_name == "users":
                icon = "👥"
            elif ".git" in folder_name:
                icon = "🔀"
            elif "project" in folder_name or "code" in folder_name:
                icon = "💻"
            else:
                icon = "📁"  # Default folder
        
        # Name column with icon
        name_item = QStandardItem(f"{icon} {path.name if path.name else str(path)}")
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(False)
        
        # Add placeholder child to make it expandable
        name_item.appendRow(QStandardItem("Loading..."))
        
        return name_item
    
    def create_folder_item(self, path, icon="📁"):
        """Create a row of items for a folder"""
        path = Path(path)
        
        # Use more descriptive icons for special folders
        folder_name = path.name.lower() if path.name else ""
        
        # Special folder icons
        if not icon or icon == "📁":  # Only override if default folder icon
            if "desktop" in folder_name:
                icon = "🖥️"
            elif "documents" in folder_name:
                icon = "📄"
            elif "downloads" in folder_name:
                icon = "⬇️"
            elif "pictures" in folder_name or "photos" in folder_name:
                icon = "🖼️"
            elif "music" in folder_name:
                icon = "🎵"
            elif "videos" in folder_name:
                icon = "🎬"
            elif "onedrive" in folder_name:
                icon = "☁️"
            elif folder_name in ["program files", "program files (x86)"]:
                icon = "⚙️"
            elif folder_name == "windows":
                icon = "🪟"
            elif folder_name == "users":
                icon = "👥"
            elif ".git" in folder_name:
                icon = "🔀"
            elif "project" in folder_name or "code" in folder_name:
                icon = "💻"
            else:
                icon = "📁"  # Default folder
        
        # Name column with icon
        name_item = QStandardItem(f"{icon} {path.name if path.name else str(path)}")
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(False)
        
        # Size column (empty for folders)
        size_item = QStandardItem("")
        size_item.setEditable(False)
        
        # Type column
        type_item = QStandardItem("Folder")
        type_item.setEditable(False)
        
        # Date modified - skip on network drives for speed
        try:
            # Check if network path
            if str(path).startswith('\\\\'):
                date_str = ""  # Skip date on network drives
            else:
                mtime = path.stat().st_mtime
                date_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        except:
            date_str = ""
        date_item = QStandardItem(date_str)
        date_item.setEditable(False)
        
        # Add placeholder child to make it expandable
        name_item.appendRow([
            QStandardItem("Loading..."),
            QStandardItem(""),
            QStandardItem(""),
            QStandardItem("")
        ])
        
        return [name_item, size_item, type_item, date_item]
        
    def create_file_item(self, path):
        """Create a row of items for a file"""
        path = Path(path)
        
        # Determine icon based on file type with better variety
        suffix = path.suffix.lower()
        
        # Microsoft Office files
        if suffix in ['.xlsx', '.xlsm', '.xls']:
            icon = "📊"  # Excel
        elif suffix in ['.docx', '.doc']:
            icon = "📝"  # Word
        elif suffix in ['.pptx', '.ppt']:
            icon = "📽️"  # PowerPoint
        elif suffix in ['.accdb', '.mdb']:
            icon = "🗃️"  # Access
        
        # Documents
        elif suffix == '.pdf':
            icon = "📕"  # PDF
        elif suffix in ['.txt', '.log']:
            icon = "📄"  # Text
        elif suffix in ['.md', '.markdown']:
            icon = "📋"  # Markdown
        elif suffix == '.csv':
            icon = "📑"  # CSV
        
        # Code files
        elif suffix in ['.py', '.pyw']:
            icon = "🐍"  # Python
        elif suffix in ['.js', '.jsx', '.ts', '.tsx']:
            icon = "📜"  # JavaScript/TypeScript
        elif suffix in ['.html', '.htm']:
            icon = "🌐"  # HTML
        elif suffix in ['.css', '.scss', '.sass']:
            icon = "🎨"  # CSS
        elif suffix in ['.json', '.yaml', '.yml', '.xml']:
            icon = "⚙️"  # Config
        elif suffix in ['.sql']:
            icon = "�"  # Database
        
        # Images
        elif suffix in ['.jpg', '.jpeg']:
            icon = "🖼️"  # JPEG
        elif suffix in ['.png']:
            icon = "🖼️"  # PNG
        elif suffix in ['.gif']:
            icon = "🎞️"  # GIF
        elif suffix in ['.svg']:
            icon = "🎨"  # SVG
        elif suffix in ['.bmp', '.ico']:
            icon = "🖼️"  # Bitmap
        
        # Archives
        elif suffix in ['.zip', '.rar', '.7z']:
            icon = "📦"  # Archive
        elif suffix in ['.tar', '.gz', '.bz2']:
            icon = "📦"  # Compressed
        
        # Executables
        elif suffix in ['.exe', '.msi']:
            icon = "⚙️"  # Windows executable
        elif suffix in ['.bat', '.cmd', '.ps1']:
            icon = "⚡"  # Script
        elif suffix in ['.sh']:
            icon = "🔧"  # Shell script
        
        # Media
        elif suffix in ['.mp4', '.avi', '.mkv', '.mov']:
            icon = "🎬"  # Video
        elif suffix in ['.mp3', '.wav', '.flac', '.m4a']:
            icon = "🎵"  # Audio
        
        # Other
        else:
            icon = "📃"  # Generic file
        
        # Name column
        name_item = QStandardItem(f"{icon} {path.name}")
        name_item.setData(str(path), Qt.ItemDataRole.UserRole)
        name_item.setEditable(False)
        
        # Size column
        try:
            # Skip size calculation on network drives for speed
            if str(path).startswith('\\\\'):
                size_str = ""  # Skip size on network drives
            else:
                size = path.stat().st_size
                if size < 1024:
                    size_str = f"{size} B"
                elif size < 1024 * 1024:
                    size_str = f"{size / 1024:.1f} KB"
                elif size < 1024 * 1024 * 1024:
                    size_str = f"{size / (1024 * 1024):.1f} MB"
                else:
                    size_str = f"{size / (1024 * 1024 * 1024):.2f} GB"
        except:
            size_str = ""
        size_item = QStandardItem(size_str)
        size_item.setEditable(False)
        
        # Type column
        type_item = QStandardItem(suffix.upper()[1:] if suffix else "File")
        type_item.setEditable(False)
        
        # Date modified
        try:
            # Skip date on network drives for speed
            if str(path).startswith('\\\\'):
                date_str = ""
            else:
                mtime = path.stat().st_mtime
                date_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        except:
            date_str = ""
        date_item = QStandardItem(date_str)
        date_item.setEditable(False)
        
        return [name_item, size_item, type_item, date_item]
        
    def on_item_expanded(self, index):
        """Load directory contents when expanded"""
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        # Check if this item has a placeholder child
        if item.rowCount() == 1 and item.child(0, 0).text() == "Loading...":
            # Remove placeholder
            item.removeRow(0)
            
            # Load actual contents (folders only for tree)
            path = item.data(Qt.ItemDataRole.UserRole)
            if path:
                self.load_tree_directory_contents(item, Path(path))
                
    def load_tree_directory_contents(self, parent_item, dir_path):
        """Load folder contents into the tree (folders only)"""
        try:
            dir_path = Path(dir_path)
            
            # Get only folders
            folders = [item for item in sorted(dir_path.iterdir(), key=lambda x: x.name.lower()) 
                      if item.is_dir() and not item.name.startswith('.')]
            
            for folder_path in folders:
                try:
                    item = self.create_tree_folder_item(folder_path)
                    parent_item.appendRow(item)
                except (PermissionError, OSError):
                    # Skip items we can't access
                    continue
                    
        except (PermissionError, OSError) as e:
            # Show error message
            error_item = QStandardItem(f"❌ Access denied")
            error_item.setEnabled(False)
            parent_item.appendRow(error_item)
            
    def create_details_panel(self):
        """Create the details view panel (right side) for folder contents"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        self.details_header = QLabel("� Contents")
        self.details_header.setStyleSheet("padding: 8px; font-weight: bold; font-size: 11pt;")
        layout.addWidget(self.details_header)
        
        # Create details tree view
        self.details_view = QTreeView()
        self.details_view.setAnimated(False)
        self.details_view.setRootIsDecorated(False)  # No expand arrows
        self.details_view.setIndentation(0)
        self.details_view.setHeaderHidden(False)
        self.details_view.setSelectionMode(QTreeView.SelectionMode.ExtendedSelection)  # Multi-select
        self.details_view.setSortingEnabled(True)
        
        # Create details model
        self.details_model = QStandardItemModel()
        self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified'])
        
        self.details_view.setModel(self.details_model)
        
        # Configure header
        header_view = self.details_view.header()
        header_view.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        
        # Apply saved column widths or use defaults
        default_widths = [350, 100, 120, 150]
        for col in range(4):
            width = self.column_widths.get(f'col_{col}', default_widths[col])
            self.details_view.setColumnWidth(col, width)
        
        # Connect column resize signal to save widths
        header_view.sectionResized.connect(self.on_column_resized)
        
        # Connect signals
        self.details_view.doubleClicked.connect(self.on_details_item_double_clicked)
        
        # Context menu
        self.details_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.details_view.customContextMenuRequested.connect(self.show_details_context_menu)
        
        layout.addWidget(self.details_view)
        
        return widget
    
    def on_tree_item_clicked(self, index):
        """Handle click on tree item - load folder contents in details view"""
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if path:
            self.load_folder_contents_in_details(Path(path))
    
    def load_folder_contents_in_details(self, dir_path):
        """Load folder contents into the details view - optimized for network drives"""
        try:
            dir_path = Path(dir_path)
            
            # Store current folder for breadcrumb and other operations
            self.current_details_folder = str(dir_path)
            
            # Update header
            self.details_header.setText(f"📂 {dir_path.name or str(dir_path)}")
            
            # Clear details model BUT preserve column widths
            self.details_model.clear()
            self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified'])
            
            # Restore column widths after clearing
            default_widths = [350, 100, 120, 150]
            for col in range(4):
                width = self.column_widths.get(f'col_{col}', default_widths[col])
                self.details_view.setColumnWidth(col, width)
            
            # Check if it's a network path
            is_network = str(dir_path).startswith('\\\\')
            
            # Get all items in directory - use os.scandir() which is MUCH faster
            # os.scandir() caches stat info in DirEntry objects (150x faster on network!)
            items_with_type = []
            try:
                with os.scandir(str(dir_path)) as entries:
                    for entry in entries:
                        try:
                            if entry.name.startswith('.'):
                                continue
                            # DirEntry.is_dir() uses cached stat - very fast!
                            is_directory = entry.is_dir()
                            # Convert to Path for compatibility with existing code
                            item_path = Path(entry.path)
                            items_with_type.append((item_path, is_directory))
                        except (PermissionError, OSError):
                            continue
                
                # Sort by name only (case-insensitive), folders first
                items_with_type.sort(key=lambda x: (not x[1], x[0].name.lower()))
            except Exception:
                items_with_type = []
            
            # Add items to view - use cached is_dir result
            for item_path, is_directory in items_with_type:
                try:
                    if is_directory:
                        row_items = self.create_folder_item(item_path)
                    else:
                        row_items = self.create_file_item(item_path)
                    
                    self.details_model.appendRow(row_items)
                except (PermissionError, OSError):
                    # Skip items we can't access
                    continue
                    
        except (PermissionError, OSError) as e:
            self.details_model.clear()
            self.details_model.setHorizontalHeaderLabels(['Name', 'Size', 'Type', 'Date Modified'])
            
            # Restore column widths even on error
            default_widths = [350, 100, 120, 150]
            for col in range(4):
                width = self.column_widths.get(f'col_{col}', default_widths[col])
                self.details_view.setColumnWidth(col, width)
            
            error_item = QStandardItem(f"❌ Access denied: {str(e)}")
            error_item.setEnabled(False)
            self.details_model.appendRow([error_item, QStandardItem(""), QStandardItem(""), QStandardItem("")])
    
    def on_details_item_double_clicked(self, index):
        """Handle double click in details view"""
        item = self.details_model.itemFromIndex(self.details_model.index(index.row(), 0))
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        
        path_obj = Path(path)
        
        if path_obj.is_dir():
            # Navigate into folder - update tree selection and load in details
            self.load_folder_contents_in_details(path_obj)
            # TODO: Could also expand/select this folder in the tree view
        elif path_obj.is_file():
            # Open file with default application
            try:
                if os.name == 'nt':
                    os.startfile(str(path_obj))
                elif sys.platform == 'darwin':
                    subprocess.run(['open', str(path_obj)])
                else:
                    subprocess.run(['xdg-open', str(path_obj)])
            except Exception as e:
                logger.error(f"Failed to open file: {e}")
                QMessageBox.warning(self, "Cannot Open File", f"Failed to open {path_obj.name}\n\nError: {str(e)}")
    
    def show_tree_context_menu(self, position):
        """Show context menu for tree view"""
        index = self.tree_view.indexAt(position)
        if not index.isValid():
            return
        
        item = self.model.itemFromIndex(index)
        if not item:
            return
        
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        
        # Check if this is a custom quick link
        is_custom_link = item.data(Qt.ItemDataRole.UserRole + 1) == "__CUSTOM_LINK__"
        
        menu = QMenu()
        
        # Open in Explorer
        open_explorer_action = menu.addAction("📂 Open in File Explorer")
        open_explorer_action.triggered.connect(lambda: self.open_path_in_explorer(path))
        
        menu.addSeparator()
        
        # Add to Quick Links (if not already a custom link)
        if not is_custom_link:
            add_quick_link_action = menu.addAction("📌 Add to Quick Links")
            add_quick_link_action.triggered.connect(lambda: self.add_to_quick_links(path))
        
        # Remove from Quick Links (only for custom links)
        if is_custom_link:
            remove_action = menu.addAction("📌 Remove from Quick Links")
            remove_action.triggered.connect(lambda: self.remove_quick_link_by_path(path))
        
        # Hide OneDrive folder (only for OneDrive paths, not custom links)
        if not is_custom_link and "onedrive" in str(path).lower():
            hide_action = menu.addAction("🚫 Hide This OneDrive")
            hide_action.triggered.connect(lambda: self.hide_onedrive_path(path))
        
        # Add to Bookmarks (pass None for name to auto-derive without icon)
        add_bookmark_action = menu.addAction("⭐ Add to Bookmarks")
        add_bookmark_action.triggered.connect(lambda: self.add_to_bookmarks(path, None))
        
        # Show menu
        menu.exec(self.tree_view.viewport().mapToGlobal(position))
    
    def show_details_context_menu(self, position):
        """Show context menu for details view"""
        index = self.details_view.indexAt(position)
        
        menu = QMenu()
        
        if index.isValid():
            item = self.details_model.itemFromIndex(self.details_model.index(index.row(), 0))
            if item:
                path = item.data(Qt.ItemDataRole.UserRole)
                if path:
                    path_obj = Path(path)
                    
                    # Open
                    if path_obj.is_file():
                        open_action = menu.addAction("📄 Open")
                        open_action.triggered.connect(lambda: self.open_file(path))
                    
                    menu.addSeparator()
                    
                    # Cut, Copy, Paste operations
                    cut_action = menu.addAction("✂️ Cut")
                    cut_action.triggered.connect(self.cut_file)
                    
                    copy_action = menu.addAction("📋 Copy")
                    copy_action.triggered.connect(self.copy_file)
                    
                    menu.addSeparator()
                    
                    # Rename, Delete
                    rename_action = menu.addAction("✏️ Rename")
                    rename_action.triggered.connect(self.rename_file)
                    
                    delete_action = menu.addAction("🗑️ Delete")
                    delete_action.triggered.connect(self.delete_file)
                    
                    menu.addSeparator()
                    
                    # Add to Quick Links
                    add_quick_link_action = menu.addAction("📌 Add to Quick Links")
                    add_quick_link_action.triggered.connect(lambda: self.add_to_quick_links(path))
                    
                    # Add to Bookmarks (pass None for name to auto-derive without icon)
                    add_bookmark_action = menu.addAction("⭐ Add to Bookmarks")
                    add_bookmark_action.triggered.connect(lambda: self.add_to_bookmarks(path, None))
                    
                    menu.addSeparator()
                    
                    # Upload to mainframe (files only)
                    if path_obj.is_file():
                        upload_action = menu.addAction("⬆️ Upload to Mainframe")
                        upload_action.triggered.connect(self.upload_to_mainframe)
        
        # Paste (always available if clipboard has content)
        if self.clipboard["path"]:
            if index.isValid():
                menu.addSeparator()
            paste_action = menu.addAction("📌 Paste")
            paste_action.triggered.connect(self.paste_file)
        
        # Show menu
        menu.exec(self.details_view.viewport().mapToGlobal(position))
    
    def open_file(self, path):
        """Open file with default application"""
        try:
            path_obj = Path(path)
            if os.name == 'nt':
                os.startfile(str(path_obj))
            elif sys.platform == 'darwin':
                subprocess.run(['open', str(path_obj)])
            else:
                subprocess.run(['xdg-open', str(path_obj)])
        except Exception as e:
            logger.error(f"Failed to open file: {e}")
            QMessageBox.warning(self, "Error", f"Failed to open file:\n{str(e)}")
    
    def open_path_in_explorer(self, path):
        """Open path in system file explorer"""
        try:
            path_obj = Path(path)
            if os.name == 'nt':
                subprocess.run(['explorer', '/select,', str(path_obj)])
            elif sys.platform == 'darwin':
                subprocess.run(['open', '-R', str(path_obj)])
            else:
                subprocess.run(['xdg-open', str(path_obj.parent)])
        except Exception as e:
            logger.error(f"Failed to open in explorer: {e}")
    
    def get_current_details_folder(self):
        """Get the current folder being displayed in details view"""
        # This would need to track the current folder - for now return empty
        # You could store this as self.current_details_folder when loading
        return ""
        
    def show_file_preview(self, file_path):
        """Show enhanced file preview with support for Excel, CSV, images, etc."""
        self.current_file_path = file_path
        path = Path(file_path)
        suffix = path.suffix.lower()
        
        # Clear preview
        self.preview_text.clear()
        self.upload_button.setEnabled(True)
        
        try:
            # Check file size first
            file_size = path.stat().st_size
            if file_size > 10 * 1024 * 1024:  # 10 MB
                self.preview_text.setText(
                    f"📄 {path.name}\n"
                    f"Size: {file_size / (1024 * 1024):.2f} MB\n\n"
                    f"⚠️ File too large for preview\n"
                    f"Double-click to open in default application"
                )
                return
            
            # Excel files (.xlsx, .xls, .xlsm)
            if suffix in ['.xlsx', '.xls', '.xlsm']:
                self.preview_excel_file(path)
            
            # CSV files
            elif suffix == '.csv':
                self.preview_csv_file(path)
            
            # Text files
            elif suffix in ['.txt', '.log', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', '.sql']:
                self.preview_text_file(path)
            
            # Image files
            elif suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                self.preview_image_file(path)
            
            # PDF files
            elif suffix == '.pdf':
                self.preview_text.setText(
                    f"📕 PDF Document: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in PDF viewer"
                )
            
            # Word documents
            elif suffix in ['.docx', '.doc']:
                self.preview_text.setText(
                    f"📝 Word Document: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in Microsoft Word"
                )
            
            # PowerPoint
            elif suffix in ['.pptx', '.ppt']:
                self.preview_text.setText(
                    f"📽️ PowerPoint: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in PowerPoint"
                )
            
            # Access databases
            elif suffix in ['.accdb', '.mdb']:
                self.preview_text.setText(
                    f"🗃️ Access Database: {path.name}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"📌 Double-click to open in Microsoft Access"
                )
            
            # Unknown/Binary
            else:
                self.preview_text.setText(
                    f"📃 {path.name}\n"
                    f"Type: {suffix.upper()[1:] if suffix else 'Unknown'}\n"
                    f"Size: {file_size / 1024:.1f} KB\n\n"
                    f"Preview not available for this file type\n"
                    f"Double-click to open"
                )
                
        except Exception as e:
            logger.error(f"Preview error: {e}")
            self.preview_text.setText(
                f"❌ Error previewing file:\n{str(e)}\n\n"
                f"File: {path.name}"
            )
    
    def preview_excel_file(self, path):
        """Preview Excel file showing first few rows"""
        try:
            import pandas as pd
            
            # Read first sheet using openpyxl engine
            df = pd.read_excel(path, engine='openpyxl', nrows=100)
            
            # Build preview text
            preview = f"📊 Excel File: {path.name}\n"
            preview += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
            preview += "=" * 60 + "\n\n"
            
            # Column names
            preview += "Columns:\n"
            for i, col in enumerate(df.columns, 1):
                preview += f"  {i}. {col}\n"
            
            preview += "\n" + "=" * 60 + "\n"
            preview += "First 10 Rows:\n"
            preview += "=" * 60 + "\n\n"
            
            # Show first 10 rows with formatting
            preview += df.head(10).to_string(index=True, max_colwidth=40)
            
            if len(df) > 10:
                preview += f"\n\n... and {len(df) - 10} more rows"
            
            preview += "\n\n📌 Double-click to open in Excel"
            
            self.preview_text.setText(preview)
            self.current_file_content = preview
            
        except ImportError as e:
            self.preview_text.setText(
                f"📊 Excel File: {path.name}\n\n"
                f"⚠️ Could not preview Excel file:\n"
                f"Missing optional dependency 'xlrd'. Install xlrd >= 2.0.1 for xls Excel support\n"
                f"or use openpyxl for xlsx files.\n\n"
                f"Run: pip install openpyxl\n\n"
                f"Double-click to open in Microsoft Excel"
            )
        except Exception as e:
            self.preview_text.setText(
                f"📊 Excel File: {path.name}\n\n"
                f"⚠️ Could not preview Excel file:\n{str(e)}\n\n"
                f"Double-click to open in Microsoft Excel"
            )
    
    def preview_csv_file(self, path):
        """Preview CSV file"""
        try:
            import pandas as pd
            
            # Try to read CSV
            df = pd.read_csv(path, nrows=100)
            
            preview = f"📑 CSV File: {path.name}\n"
            preview += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
            preview += "=" * 60 + "\n\n"
            
            # Column names
            preview += "Columns:\n"
            for i, col in enumerate(df.columns, 1):
                preview += f"  {i}. {col}\n"
            
            preview += "\n" + "=" * 60 + "\n"
            preview += "First 10 Rows:\n"
            preview += "=" * 60 + "\n\n"
            
            preview += df.head(10).to_string(index=False, max_colwidth=40)
            
            if len(df) > 10:
                preview += f"\n\n... and {len(df) - 10} more rows"
            
            self.preview_text.setText(preview)
            self.current_file_content = preview
            
        except Exception as e:
            # Fallback to text preview
            self.preview_text_file(path)
    
    def preview_text_file(self, path):
        """Preview text files"""
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)  # Read first 50KB
                
            preview = f"📄 Text File: {path.name}\n"
            preview += f"Size: {path.stat().st_size / 1024:.1f} KB\n"
            preview += "=" * 60 + "\n\n"
            preview += content
            
            if path.stat().st_size > 50000:
                preview += "\n\n... (file truncated for preview)"
            
            self.preview_text.setText(preview)
            self.current_file_content = content
            
        except Exception as e:
            self.preview_text.setText(
                f"Cannot read file:\n{str(e)}\n\n"
                f"File: {path.name}"
            )
    
    def preview_image_file(self, path):
        """Preview image file info"""
        try:
            from PIL import Image
            
            img = Image.open(path)
            
            preview = f"🖼️ Image: {path.name}\n"
            preview += f"Size: {path.stat().st_size / 1024:.1f} KB\n"
            preview += f"Dimensions: {img.width} × {img.height} pixels\n"
            preview += f"Format: {img.format}\n"
            preview += f"Mode: {img.mode}\n\n"
            preview += "📌 Double-click to view image"
            
            self.preview_text.setText(preview)
            
        except:
            # If PIL not available, show basic info
            self.preview_text.setText(
                f"🖼️ Image: {path.name}\n"
                f"Size: {path.stat().st_size / 1024:.1f} KB\n\n"
                f"📌 Double-click to view image"
            )
            
    def show_upload_menu(self):
        """Show mainframe upload menu"""
        if not self.current_file_path:
            return
        
        QMessageBox.information(
            self,
            "Upload to Mainframe",
            f"Upload functionality will be integrated here.\n\n"
            f"File: {Path(self.current_file_path).name}"
        )
        
    # File operations
    def get_selected_path(self):
        """Get currently selected file/folder path from details view (single selection)"""
        indexes = self.details_view.selectedIndexes()
        if not indexes:
            return None
        
        item = self.details_model.itemFromIndex(self.details_model.index(indexes[0].row(), 0))
        if item:
            return item.data(Qt.ItemDataRole.UserRole)
        return None
    
    def get_selected_paths(self):
        """Get all selected file/folder paths from details view (multi-selection)"""
        selected_rows = {}
        paths = []
        
        # Get unique rows (since selecting a row selects all columns)
        for index in self.details_view.selectedIndexes():
            row = index.row()
            key = (row, 0)
            
            if key not in selected_rows:
                selected_rows[key] = True
                item = self.details_model.itemFromIndex(self.details_model.index(row, 0))
                if item:
                    path = item.data(Qt.ItemDataRole.UserRole)
                    if path:
                        paths.append(path)
        
        return paths
        
    def cut_file(self):
        """Cut selected file/folder"""
        path = self.get_selected_path()
        if path:
            self.clipboard = {"path": path, "operation": "cut"}
            logger.info(f"Cut: {path}")
            
    def copy_file(self):
        """Copy selected file/folder"""
        path = self.get_selected_path()
        if path:
            self.clipboard = {"path": path, "operation": "copy"}
            logger.info(f"Copy: {path}")
            
    def paste_file(self):
        """Paste cut/copied file/folder"""
        if not self.clipboard["path"]:
            return
        
        dest_path = self.get_selected_path()
        if not dest_path or not os.path.isdir(dest_path):
            QMessageBox.warning(self, "Paste", "Please select a destination folder")
            return
        
        source = Path(self.clipboard["path"])
        dest = Path(dest_path) / source.name
        
        try:
            if self.clipboard["operation"] == "cut":
                shutil.move(str(source), str(dest))
            else:
                if source.is_dir():
                    shutil.copytree(str(source), str(dest))
                else:
                    shutil.copy2(str(source), str(dest))
            
            self.refresh_tree()
            QMessageBox.information(self, "Success", f"Pasted to {dest}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Paste failed: {e}")
            
    def rename_file(self):
        """Rename selected file/folder"""
        path = self.get_selected_path()
        if not path:
            return
        
        old_path = Path(path)
        new_name, ok = QInputDialog.getText(
            self, "Rename", f"Rename '{old_path.name}' to:",
            text=old_path.name
        )
        
        if ok and new_name:
            try:
                new_path = old_path.parent / new_name
                old_path.rename(new_path)
                self.refresh_tree()
                QMessageBox.information(self, "Success", f"Renamed to {new_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Rename failed: {e}")
                
    def delete_file(self):
        """Delete selected file/folder"""
        path = self.get_selected_path()
        if not path:
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete '{Path(path).name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                path_obj = Path(path)
                if path_obj.is_dir():
                    shutil.rmtree(path)
                else:
                    path_obj.unlink()
                
                self.refresh_tree()
                QMessageBox.information(self, "Success", "Deleted successfully")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Delete failed: {e}")
                
    def refresh_tree(self):
        """Refresh the tree view"""
        self.populate_tree_model()
        
    def open_in_explorer(self):
        """Open selected path in Windows Explorer"""
        path = self.get_selected_path()
        if not path:
            return
        
        try:
            path_obj = Path(path)
            if path_obj.is_file():
                path = str(path_obj.parent)
            
            if os.name == 'nt':
                subprocess.run(['explorer', path])
            else:
                subprocess.run(['xdg-open', path])
        except Exception as e:
            logger.error(f"Failed to open explorer: {e}")
            
    def show_context_menu(self, position):
        """Show context menu with quick link options"""
        menu = QMenu()
        
        # Check if item is a custom quick link and if it's a file
        indexes = self.tree_view.selectedIndexes()
        is_custom_link = False
        is_file = False
        selected_path = None
        
        if indexes:
            item = self.model.itemFromIndex(self.model.index(indexes[0].row(), 0, indexes[0].parent()))
            if item:
                if item.data(Qt.ItemDataRole.UserRole + 1) == "__CUSTOM_LINK__":
                    is_custom_link = True
                selected_path = item.data(Qt.ItemDataRole.UserRole)
                if selected_path and Path(selected_path).is_file():
                    is_file = True
        
        # Standard actions
        cut_action = menu.addAction("✂️ Cut")
        copy_action = menu.addAction("📋 Copy")
        paste_action = menu.addAction("📌 Paste")
        menu.addSeparator()
        rename_action = menu.addAction("✏️ Rename")
        delete_action = menu.addAction("🗑️ Delete")
        menu.addSeparator()
        
        # Mainframe upload (only for files)
        upload_action = None
        if is_file:
            upload_action = menu.addAction("⬆️ Upload to Mainframe")
            menu.addSeparator()
        
        # Quick link actions
        add_to_quick = menu.addAction("⭐ Add to Quick Access")
        if is_custom_link:
            remove_from_quick = menu.addAction("❌ Remove from Quick Access")
        else:
            remove_from_quick = None
        
        menu.addSeparator()
        explorer_action = menu.addAction("📂 Open in Explorer")
        
        action = menu.exec(self.tree_view.viewport().mapToGlobal(position))
        
        if action == cut_action:
            self.cut_file()
        elif action == copy_action:
            self.copy_file()
        elif action == paste_action:
            self.paste_file()
        elif action == rename_action:
            self.rename_file()
        elif action == delete_action:
            self.delete_file()
        elif action == upload_action:
            self.upload_to_mainframe()
        elif action == add_to_quick:
            self.add_to_quick_access()
        elif action == remove_from_quick:
            self.remove_from_quick_access()
        elif action == explorer_action:
            self.open_in_explorer()
    
    def upload_to_mainframe(self):
        """Upload selected file to mainframe"""
        path = self.get_selected_path()
        if not path:
            return
        
        path_obj = Path(path)
        if not path_obj.is_file():
            QMessageBox.warning(self, "Invalid Selection", "Please select a file to upload")
            return
        
        try:
            from suiteview.ui.dialogs.mainframe_upload_dialog import MainframeUploadDialog
            
            dialog = MainframeUploadDialog(str(path_obj), self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                dialog.perform_upload()
                
        except Exception as e:
            logger.error(f"Failed to show upload dialog: {e}")
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to open upload dialog:\n{str(e)}"
            )
    
    def batch_rename_files(self):
        """Batch rename multiple selected files"""
        paths = self.get_selected_paths()
        
        if not paths:
            QMessageBox.information(
                self,
                "No Selection",
                "Please select one or more files to rename.\n\n"
                "💡 Tip: Hold Ctrl and click to select multiple files"
            )
            return
        
        # Filter out folders (only rename files in batch)
        file_paths = [p for p in paths if Path(p).is_file()]
        
        if not file_paths:
            QMessageBox.warning(
                self,
                "No Files Selected",
                "Please select files (not folders) to batch rename."
            )
            return
        
        try:
            from suiteview.ui.dialogs.batch_rename_dialog import BatchRenameDialog
            
            dialog = BatchRenameDialog(file_paths, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                if dialog.perform_rename():
                    self.refresh_tree()
                    
        except Exception as e:
            logger.error(f"Failed to show batch rename dialog: {e}")
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to open batch rename dialog:\n{str(e)}"
            )
    
    def load_quick_links(self):
        """Load custom quick links from JSON file"""
        try:
            if self.quick_links_file.exists():
                with open(self.quick_links_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load quick links: {e}")
        return []
    
    def save_quick_links(self):
        """Save custom quick links to JSON file"""
        try:
            # Ensure directory exists
            self.quick_links_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.quick_links_file, 'w') as f:
                json.dump(self.custom_quick_links, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save quick links: {e}")
    
    def load_hidden_onedrive(self):
        """Load hidden OneDrive paths from JSON file"""
        try:
            if self.hidden_onedrive_file.exists():
                with open(self.hidden_onedrive_file, 'r') as f:
                    paths = json.load(f)
                    # Normalize paths to lowercase for comparison
                    return set(p.lower() for p in paths)
        except Exception as e:
            logger.error(f"Failed to load hidden OneDrive paths: {e}")
        return set()
    
    def save_hidden_onedrive(self):
        """Save hidden OneDrive paths to JSON file"""
        try:
            # Ensure directory exists
            self.hidden_onedrive_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.hidden_onedrive_file, 'w') as f:
                json.dump(list(self.hidden_onedrive_paths), f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save hidden OneDrive paths: {e}")
    
    def hide_onedrive_path(self, path):
        """Hide a OneDrive path from Quick Links"""
        path_str = str(path).lower()
        if path_str not in self.hidden_onedrive_paths:
            self.hidden_onedrive_paths.add(path_str)
            self.save_hidden_onedrive()
            # Refresh tree to remove it
            self.load_tree()
            QMessageBox.information(self, "OneDrive Hidden", f"Hidden from Quick Links:\n{Path(path).name}\n\nTo unhide, delete:\n{self.hidden_onedrive_file}")
    
    def load_column_widths(self):
        """Load column widths from JSON file"""
        try:
            if self.column_widths_file.exists():
                with open(self.column_widths_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load column widths: {e}")
        return {}
    
    def save_column_widths(self):
        """Save column widths to JSON file"""
        try:
            # Ensure directory exists
            self.column_widths_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.column_widths_file, 'w') as f:
                json.dump(self.column_widths, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save column widths: {e}")
    
    def on_column_resized(self, logical_index, old_size, new_size):
        """Handle column resize event - save new width"""
        self.column_widths[f'col_{logical_index}'] = new_size
        self.save_column_widths()
    
    def add_to_quick_access(self):
        """Add selected item to Quick Access"""
        path = self.get_selected_path()
        if not path:
            return
        
        path_str = str(Path(path).resolve())
        
        # Check if already in quick links
        if path_str in self.custom_quick_links:
            QMessageBox.information(self, "Quick Access", "This item is already in Quick Access")
            return
        
        # Add to quick links
        self.custom_quick_links.append(path_str)
        self.save_quick_links()
        self.refresh_tree()
        
        QMessageBox.information(self, "Quick Access", f"Added to Quick Access:\n{Path(path).name}")
    
    def remove_from_quick_access(self):
        """Remove selected item from Quick Access"""
        path = self.get_selected_path()
        if not path:
            return
        
        path_str = str(Path(path).resolve())
        
        if path_str in self.custom_quick_links:
            self.custom_quick_links.remove(path_str)
            self.save_quick_links()
            self.refresh_tree()
            
            QMessageBox.information(self, "Quick Access", f"Removed from Quick Access:\n{Path(path).name}")
        else:
            QMessageBox.warning(self, "Quick Access", "This item is not in Quick Access")
    
    def remove_quick_link_by_path(self, path):
        """Remove a specific path from Quick Access (used by context menu)"""
        path_str = str(Path(path).resolve())
        
        if path_str in self.custom_quick_links:
            self.custom_quick_links.remove(path_str)
            self.save_quick_links()
            self.refresh_tree()
            
            QMessageBox.information(self, "Quick Access", f"Removed from Quick Links:\n{Path(path).name}")
        else:
            QMessageBox.warning(self, "Quick Access", "This item is not in Quick Links")
    
    def add_to_quick_links(self, path):
        """Add a path to Quick Links"""
        path_str = str(Path(path).resolve())
        
        # Check if already in quick links
        if path_str in self.custom_quick_links:
            QMessageBox.information(self, "Quick Links", "This item is already in Quick Links")
            return
        
        # Add to quick links
        self.custom_quick_links.append(path_str)
        self.save_quick_links()
        self.refresh_tree()
        
        QMessageBox.information(self, "Quick Links", f"Added to Quick Links:\n{Path(path).name}")
    
    def open_bookmarks_dialog(self):
        """Open the Bookmarks panel dialog"""
        from suiteview.ui.dialogs.shortcuts_dialog import BookmarksDialog
        from PyQt6.QtCore import QPoint
        
        # Calculate position right under the Bookmarks button
        if hasattr(self, 'toolbar'):
            toolbar_geo = self.toolbar.geometry()
            # Get toolbar's global position
            toolbar_global = self.toolbar.mapToGlobal(QPoint(0, 0))
            # Position dialog at start of toolbar, below it
            button_pos = QPoint(toolbar_global.x(), toolbar_global.y() + toolbar_geo.height())
            dialog = BookmarksDialog(self, button_pos)
        else:
            dialog = BookmarksDialog(self)
            # Fallback positioning
            if self.parent():
                parent_geo = self.parent().geometry()
                dialog.move(parent_geo.left() + 10, parent_geo.top() + 50)
        
        dialog.exec()
    
    def add_to_bookmarks(self, path, name=None):
        """Add a path to Bookmarks via dialog"""
        from suiteview.ui.dialogs.shortcuts_dialog import BookmarksDialog
        
        # If no name provided, use the file/folder name
        if not name:
            name = Path(path).name
        
        # Open bookmarks dialog and add the item
        dialog = BookmarksDialog(self)
        dialog.add_bookmark_to_category(name, path)
        dialog.exec()
    
    def print_directory_to_excel(self):
        """Export current directory structure directly to Excel (no file saved)"""
        # Get current directory from selected item or use current details folder
        path = None
        
        # Try to use current details folder first
        if hasattr(self, 'current_details_folder') and self.current_details_folder:
            path = self.current_details_folder
        else:
            # Get from selection
            path = self.get_selected_path()
        
        # If no selection or selected is a file, try to find a folder
        if not path or (path and not Path(path).is_dir()):
            # Default to user's home directory or OneDrive
            onedrive_paths = self.get_onedrive_paths()
            if onedrive_paths:
                path = str(onedrive_paths[0])
            else:
                path = str(Path.home())
        
        # Show dialog
        dialog = PrintDirectoryDialog(path, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            options = dialog.get_options()
            
            # Create progress dialog
            progress = QProgressDialog("Collecting directory information...", "Cancel", 0, 100, self)
            progress.setWindowTitle("Print Directory")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            
            # Collect data in current thread (faster for small directories)
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                
                root = Path(path)
                data = []
                
                # Headers
                headers = ['Level', 'Type', 'Name', 'Full Path', 'Size', 'Modified', 'Extension']
                data.append(headers)
                
                progress.setLabelText("Scanning directory...")
                
                if options['include_subdirs']:
                    for dirpath, dirnames, filenames in os.walk(root):
                        current = Path(dirpath)
                        level = len(current.relative_to(root).parts)
                        
                        # Add directories
                        for dirname in sorted(dirnames):
                            folder_path = current / dirname
                            try:
                                stat = folder_path.stat()
                                data.append([
                                    level,
                                    'Folder',
                                    dirname,
                                    str(folder_path),
                                    '',
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                                    ''
                                ])
                            except:
                                pass
                        
                        # Add files
                        for filename in sorted(filenames):
                            file_path = current / filename
                            try:
                                stat = file_path.stat()
                                size = stat.st_size
                                size_str = f"{size:,} B" if size < 1024 else f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.1f} MB"
                                data.append([
                                    level,
                                    'File',
                                    filename,
                                    str(file_path),
                                    size_str,
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                                    file_path.suffix
                                ])
                            except:
                                pass
                        
                        progress.setLabelText(f"Scanned {len(data)} items...")
                        if progress.wasCanceled():
                            return
                else:
                    # Just immediate contents
                    for item in sorted(root.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                        try:
                            stat = item.stat()
                            if item.is_dir():
                                data.append([
                                    0, 'Folder', item.name, str(item), '',
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"), ''
                                ])
                            else:
                                size = stat.st_size
                                size_str = f"{size:,} B" if size < 1024 else f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.1f} MB"
                                data.append([
                                    0, 'File', item.name, str(item), size_str,
                                    datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"), item.suffix
                                ])
                        except:
                            pass
                
                progress.setLabelText("Opening Excel...")
                progress.setValue(50)
                
                # Open Excel using COM automation (Windows only)
                if os.name == 'nt':
                    try:
                        import win32com.client as win32
                        
                        # Create Excel instance
                        excel = win32.gencache.EnsureDispatch('Excel.Application')
                        excel.Visible = True
                        
                        # Add a new workbook
                        wb = excel.Workbooks.Add()
                        ws = wb.Worksheets(1)
                        ws.Name = "Directory Contents"
                        
                        # Write all data at once (much faster than cell-by-cell)
                        progress.setLabelText("Populating Excel...")
                        if len(data) > 0:
                            # Calculate the range needed
                            num_rows = len(data)
                            num_cols = len(data[0])
                            
                            # Define the range (e.g., "A1:G100")
                            end_col_letter = chr(64 + num_cols)  # A=65, so 64+1=A, 64+7=G, etc.
                            range_address = f"A1:{end_col_letter}{num_rows}"
                            
                            # Write entire data array at once
                            ws.Range(range_address).Value = data
                            
                            progress.setValue(80)
                        
                        # Format headers
                        header_range = ws.Range(ws.Cells(1, 1), ws.Cells(1, len(headers)))
                        header_range.Font.Bold = True
                        header_range.Font.Color = 0xFFFFFF  # White
                        header_range.Interior.Color = 0x926636  # Dark blue
                        header_range.HorizontalAlignment = -4108  # xlCenter
                        
                        progress.setValue(90)
                        
                        # Auto-fit columns
                        ws.Columns.AutoFit()
                        
                        # Freeze top row
                        ws.Range("A2").Select()
                        excel.ActiveWindow.FreezePanes = True
                        
                        # Select cell A1
                        ws.Range("A1").Select()
                        
                        progress.setValue(100)
                        progress.close()
                        
                        QMessageBox.information(
                            self,
                            "Directory Listing Opened",
                            f"Directory listing opened in Excel with {len(data)-1} items.\n\n"
                            f"The workbook is unsaved. Use 'Save As' if you want to keep it."
                        )
                        
                    except ImportError:
                        progress.close()
                        QMessageBox.critical(
                            self,
                            "Excel COM Error",
                            "Could not access Excel via COM automation.\n\n"
                            "Make sure Microsoft Excel is installed and pywin32 is available."
                        )
                    except Exception as e:
                        progress.close()
                        QMessageBox.critical(
                            self,
                            "Excel Error",
                            f"Failed to open Excel:\n\n{str(e)}"
                        )
                else:
                    progress.close()
                    QMessageBox.warning(
                        self,
                        "Not Supported",
                        "Direct Excel opening is only supported on Windows.\n\n"
                        "This feature requires Microsoft Excel and COM automation."
                    )
                    
            except Exception as e:
                progress.close()
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to collect directory information:\n\n{str(e)}"
                )
