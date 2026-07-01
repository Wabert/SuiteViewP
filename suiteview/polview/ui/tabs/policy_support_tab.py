"""
Policy Support tab - Manages a folder-based workflow for ad-hoc policy questions.

Three-column layout:
  1. Task Categories  — drag a category onto Policy Subfolders to create that task folder
  2. Policy Subfolders — mini file explorer; accepts drops from Task Categories (creates folder)
                         and from Available Tools (copies file)
  3. Available Tools   — home shows 8 pinned tool folders; double-click to navigate in;
                         drag a file onto Policy Subfolders to copy it there

Folder structure:
    <Process_Control>/Policy Support/POLICY_LIBRARY/<ProductType>/<Co>_<PolicyNo>/<TaskCategory>/
"""

import json
import os
import shutil
from datetime import date, datetime
from typing import Optional, List

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QListWidgetItem, QGroupBox, QGridLayout,
    QMessageBox, QAbstractItemView, QInputDialog, QMenu,
    QStyledItemDelegate, QSizePolicy, QLineEdit, QTableWidgetItem,
    QStackedWidget, QDialog, QTextEdit, QComboBox,
)
from PyQt6.QtCore import Qt, pyqtSignal, QPoint, QSize, QMimeData
from PyQt6.QtGui import QColor, QDrag, QPixmap, QPainter, QFont

from ..styles import (
    WHITE, GRAY_DARK, GRAY_MID, GRAY_LIGHT,
    GREEN_DARK, GREEN_PRIMARY, GREEN_SUBTLE,
    GOLD_PRIMARY, GOLD_TEXT, GOLD_DARK, GOLD_LIGHT,
    POLICY_INFO_FRAME_STYLE,
)
from ..widgets import CopyableLabel, FixedHeaderTableWidget
from .annuity_rider_tab import AnnuityRiderTab, RIDER_PLANCODE
from ....utils.excel_template import copy_as_workbook, workbook_filename
from ...services.glp_exception import (
    GlpExceptionResult,
    calculate_glp_exception,
    calculate_policy_support_forecast,
    check_forecast_availability,
    is_glp_exception_eligible,
)

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from ...models.policy_information import PolicyInformation


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _get_process_control_dir() -> str:
    username = os.environ.get("USERNAME", os.environ.get("USER", "unknown"))
    return os.path.join(
        "C:\\Users", username,
        "OneDrive - American National Insurance Company",
        "Life Product - Process_Control",
    )


def _get_onedrive_dir() -> str:
    username = os.environ.get("USERNAME", os.environ.get("USER", "unknown"))
    return os.path.join(
        "C:\\Users", username,
        "OneDrive - American National Insurance Company",
    )


def _get_abr_root_dir() -> str:
    return os.path.join(
        _get_onedrive_dir(),
        "Life Product - Accelerated_Benefits",
    )

def _get_policy_support_dir() -> str:
    return os.path.join(_get_process_control_dir(), "Policy Support")

def _get_policy_library_dir() -> str:
    return os.path.join(_get_policy_support_dir(), "POLICY_LIBRARY")

def _get_abr_dir() -> str:
    return os.path.join(
        _get_abr_root_dir(),
        "Accelerated Death Benefit (ABR11 & ABR14)",
    )


# ---------------------------------------------------------------------------
# Pinned tool folders shown on the Available Tools home screen
# Each entry: (display_label, relative_path_from_onedrive_root)
# ---------------------------------------------------------------------------

TOOL_FOLDERS = [
    ("Tools\\Illustration\\RERUN",
    r"Life Product - Process_Control\Tools\Illustration\RERUN"),
    ("Tools\\Illustration\\Product Models\\TERM",
    r"Life Product - Process_Control\Tools\Illustration\Product Models\TERM"),
    ("Task\\Policy Support\\Term Premiums Illustrations\\TERM TEMPLATE",
    r"Life Product - Process_Control\Task\Policy Support\Term Premiums Illustrations\TERM TEMPLATE"),
    ("Task\\Policy Support\\Guideline and TAMRA\\Guideline Adjust for Exception Prems",
    r"Life Product - Process_Control\Task\Policy Support\Guideline and TAMRA\Guideline Adjust for Exception Prems"),
    ("Task\\Policy Support\\Mistatement",
    r"Life Product - Process_Control\Task\Policy Support\Mistatement"),
    ("Task\\Policy Support\\Interpolated_Terminal_Reserve",
    r"Life Product - Process_Control\Task\Policy Support\Interpolated_Terminal_Reserve"),
    ("Task\\Policy Support",
    r"Life Product - Process_Control\Task\Policy Support"),
    ("Tools\\Whole Life Nonforfeiture Calc",
    r"Life Product - Process_Control\Tools\Whole Life Nonforfeiture Calc"),
    ("Cyberlife Reference Files",
    r"Life Product - Data\Cyberlife Reference Files"),
    ("Task\\Accelerated Death Benefit (ABR11 & ABR14)",
    r"Life Product - Accelerated_Benefits\Accelerated Death Benefit (ABR11 & ABR14)"),
]


# ---------------------------------------------------------------------------
# Default task categories
# ---------------------------------------------------------------------------

DEFAULT_TASK_CATEGORIES = [
    "Annual_Statement",
    "Annuity_Rider",
    "Cash_Value_Quote",
    "Decrease",
    "GLP_Exception",
    "Illustration",
    "Increase",
    "Interpolated_Terminal_Reserve",
    "Mistatement",
    "Nonforfeiture",
    "Product_Contract_and_Specs",
    "Reinstatement",
    "Time_Driven_Error",
]


# ---------------------------------------------------------------------------
# User task storage
# ---------------------------------------------------------------------------

def _get_user_tasks_path() -> str:
    app_data = os.path.join(
        os.environ.get("APPDATA", os.path.expanduser("~")), "SuiteView"
    )
    os.makedirs(app_data, exist_ok=True)
    return os.path.join(app_data, "policy_support_tasks.json")

def _load_user_tasks() -> List[str]:
    path = _get_user_tasks_path()
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return sorted(data) if isinstance(data, list) else []
    except Exception:
        return []

def _save_user_tasks(tasks: List[str]):
    path = _get_user_tasks_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(sorted(set(tasks)), f, indent=2)


def _safe_anniversary(issue_date: date, year: int) -> date:
    try:
        return issue_date.replace(year=year)
    except ValueError:
        return issue_date.replace(year=year, day=28)


GLP_HELP_TEXT = """GLP Exception quote - basic explanation

This screen estimates whether a Universal Life policy can stay in force until the target inforce date without changing its Guideline Level Premium, also called GLP.

The calculation starts with the current policy values from the valuation date. That includes the account value, premiums paid to date, accumulated withdrawals, the current GLP, the current GSP, and the current Accum GLP. Premiums paid to date includes both regular premiums and additional premiums.

If premiums were paid after the valuation date, the tool adds those premiums to create the adjusted account value and adjusted premiums paid to date. This keeps the quote from using stale values.

Next, the tool projects the policy month by month from the valuation date to the target inforce date. It applies monthly deductions, interest, loads, and any required forecast assumptions so it can estimate whether the account value survives through that period.

If the policy can stay in force with no added premium, the tool still checks the guideline limits through the target date before deciding whether an adjustment is needed. This matters when GLP is negative: crossing an anniversary can lower Accum GLP, which may still require an Accum GLP adjustment even when the required premium is $0. The monthly forecast still displays so you can see how the account value moves over time with no added premium.

If the policy needs premium to stay in force, the tool solves for the level premium needed through the target period. If the account value is negative, the first payment is the amount needed to bring the account value to $1.00, and the level premium starts after that. The Premium to get to Target Date line shows the total premium needed through the target period.

The Accum GLP on Target Date line uses the current GLP through anniversaries crossed before the target date. PremTD less AccumWD on Target Date is the adjusted premiums paid to date plus the Premium to get to Target Date, less accumulated withdrawals.

If Accum GLP on Target Date is greater than or equal to PremTD less AccumWD on Target Date, no adjustment is needed. If Accum GLP on Target Date is less than PremTD less AccumWD on Target Date and premium is needed to get to the target date, the tool shows the New GLP, Adjustment to Accum GLP, and New Accum GLP. If no premium is needed but PremTD less AccumWD is still above Accum GLP on the target date, the tool shows the force-out amount.

The goal of the quote is to answer a practical question: does this policy need an Accum GLP adjustment, and if so, how much, so the policy can remain in force up to the target date?"""


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

_CONTEXT_MENU_STYLE = f"""
    QMenu {{
        background-color: #F0F0F0;
        border: 1px solid {GRAY_DARK};
        padding: 2px;
        font-size: 11px;
    }}
    QMenu::item {{ padding: 3px 20px 3px 8px; color: #1a1a1a; }}
    QMenu::item:selected {{ background-color: {GOLD_LIGHT}; color: {GREEN_DARK}; }}
    QMenu::item:disabled {{ color: #999999; }}
    QMenu::separator {{ height: 1px; background: {GRAY_MID}; margin: 2px 4px; }}
"""

_LIST_STYLE = f"""
    QListWidget {{
        font-size: 11px;
        border: none;
        background-color: {WHITE};
        outline: none;
        padding: 0px;
        margin: 0px;
    }}
    QListWidget::item {{
        padding: 0px 4px;
        margin: 0px;
        border: none;
        min-height: 0px;
    }}
    QListWidget::item:hover   {{ background-color: {GREEN_SUBTLE}; }}
    QListWidget::item:selected {{ background-color: {GOLD_LIGHT}; color: {GREEN_DARK}; font-weight: bold; }}
"""

_ACTION_BTN_STYLE = f"""
    QPushButton {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {GOLD_TEXT}, stop:1 {GOLD_PRIMARY});
        color: {GREEN_DARK};
        border: 1px solid {GOLD_DARK};
        border-radius: 3px;
        padding: 3px 10px;
        font-size: 11px;
        font-weight: bold;
        min-height: 18px;
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {GOLD_PRIMARY}, stop:1 {GOLD_DARK});
        color: {WHITE};
    }}
    QPushButton:pressed {{ background-color: {GOLD_DARK}; }}
    QPushButton:disabled {{
        background: {GRAY_MID}; color: {GRAY_DARK}; border-color: {GRAY_MID};
    }}
"""

_FILTER_INPUT_STYLE = f"""
    QLineEdit {{
        background: {WHITE};
        color: {GRAY_DARK};
        border: 1px solid {GRAY_MID};
        border-radius: 3px;
        padding: 3px 6px;
        font-size: 11px;
        min-height: 18px;
    }}
    QLineEdit:focus {{ border-color: {GREEN_PRIMARY}; }}
"""

_MODE_BTN_ACTIVE_STYLE = f"""
    QPushButton {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {GREEN_PRIMARY}, stop:1 {GREEN_DARK});
        color: {WHITE};
        border: 2px solid {GREEN_DARK};
        border-radius: 4px;
        padding: 4px 16px;
        font-size: 12px;
        font-weight: bold;
        min-height: 22px;
    }}
"""

_MODE_BTN_INACTIVE_STYLE = f"""
    QPushButton {{
        background: {GRAY_LIGHT};
        color: {GRAY_DARK};
        border: 1px solid {GRAY_MID};
        border-radius: 4px;
        padding: 4px 16px;
        font-size: 12px;
        font-weight: bold;
        min-height: 22px;
    }}
    QPushButton:hover {{
        background: {GREEN_SUBTLE};
        color: {GREEN_DARK};
        border-color: {GREEN_PRIMARY};
    }}
    QPushButton:disabled {{
        background: {GRAY_LIGHT};
        color: {GRAY_MID};
        border-color: {GRAY_MID};
    }}
"""

_NAV_PANEL_STYLE = f"""
    QWidget#PolicySupportNavPanel {{
        background-color: {GREEN_SUBTLE};
        border: 2px solid {GREEN_PRIMARY};
        border-radius: 8px;
    }}
"""

_RESULT_VALUE_STYLE = f"""
    font-size: 11px;
    color: {GREEN_DARK};
    font-weight: normal;
    background: transparent;
    border: none;
"""

_RESULT_LABEL_STYLE = f"""
    font-size: 11px;
    color: {GRAY_DARK};
    background: transparent;
    border: none;
"""

_NAV_BTN_STYLE = f"""
    QPushButton {{
        background: {GREEN_SUBTLE};
        color: {GREEN_DARK};
        border: 1px solid {GREEN_PRIMARY};
        border-radius: 3px;
        padding: 1px 5px;
        font-size: 11px;
        font-weight: bold;
        min-height: 16px;
        max-height: 18px;
    }}
    QPushButton:hover {{ background: {GREEN_PRIMARY}; color: {WHITE}; }}
    QPushButton:disabled {{ background: {GRAY_LIGHT}; color: {GRAY_MID}; border-color: {GRAY_MID}; }}
"""

_PATH_BAR_STYLE = f"""
    font-size: 10px;
    color: {GREEN_DARK};
    background: {GREEN_SUBTLE};
    border: 1px solid {GREEN_PRIMARY};
    border-radius: 3px;
    padding: 1px 4px;
"""

_STATUS_STYLE = f"""
    font-size: 10px; color: {GRAY_DARK};
    background: transparent; border: none; padding: 2px 4px;
"""

_PATH_LABEL_STYLE = f"""
    font-size: 10px; color: {GREEN_DARK};
    background: {GREEN_SUBTLE};
    border: 1px solid {GREEN_PRIMARY};
    border-radius: 3px; padding: 3px 6px;
"""

# ── ABR (Crimson) theme overrides ──────────────────────────────────────────
# These mirror the green/gold styles above but use ABR crimson/slate colours.

_CRIMSON_DARK    = "#5C0A14"
_CRIMSON_PRIMARY = "#8B1A2A"
_CRIMSON_RICH    = "#A52535"
_CRIMSON_LIGHT   = "#C96070"
_CRIMSON_SUBTLE  = "#F9ECED"
_SLATE_PRIMARY   = "#4A6FA5"
_SLATE_TEXT      = "#B8D0F0"
_SLATE_DARK      = "#2E4F85"

_ABR_FRAME_STYLE = f"""
    QGroupBox {{
        font-size: 11px;
        font-weight: bold;
        color: {_CRIMSON_DARK};
        border: 2px solid {_CRIMSON_PRIMARY};
        border-radius: 8px;
        margin-top: 3px;
        margin-bottom: 0px;
        padding: 0px;
        background-color: {WHITE};
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 1px 10px;
        background-color: {_CRIMSON_PRIMARY};
        color: {_SLATE_TEXT};
        border-radius: 4px;
        left: 10px;
    }}
    QGroupBox QLabel {{
        font-size: 10px;
        color: {GRAY_DARK};
        border: none;
        background: transparent;
    }}
"""

_ABR_LIST_STYLE = f"""
    QListWidget {{
        font-size: 11px;
        border: none;
        background-color: {WHITE};
        outline: none;
        padding: 0px;
        margin: 0px;
    }}
    QListWidget::item {{
        padding: 0px 4px;
        margin: 0px;
        border: none;
        min-height: 0px;
    }}
    QListWidget::item:hover   {{ background-color: {_CRIMSON_SUBTLE}; }}
    QListWidget::item:selected {{ background-color: {_SLATE_PRIMARY}; color: {WHITE}; font-weight: bold; }}
"""

_ABR_NAV_BTN_STYLE = f"""
    QPushButton {{
        background: {_CRIMSON_SUBTLE};
        color: {_CRIMSON_DARK};
        border: 1px solid {_CRIMSON_PRIMARY};
        border-radius: 3px;
        padding: 1px 5px;
        font-size: 11px;
        font-weight: bold;
        min-height: 16px;
        max-height: 18px;
    }}
    QPushButton:hover {{ background: {_CRIMSON_PRIMARY}; color: {WHITE}; }}
    QPushButton:disabled {{ background: {GRAY_LIGHT}; color: {GRAY_MID}; border-color: {GRAY_MID}; }}
"""

_ABR_PATH_BAR_STYLE = f"""
    font-size: 10px;
    color: {_CRIMSON_DARK};
    background: {_CRIMSON_SUBTLE};
    border: 1px solid {_CRIMSON_PRIMARY};
    border-radius: 3px;
    padding: 1px 4px;
"""

_ABR_PATH_LABEL_STYLE = f"""
    font-size: 10px; color: {_CRIMSON_DARK};
    background: {_CRIMSON_SUBTLE};
    border: 1px solid {_CRIMSON_PRIMARY};
    border-radius: 3px; padding: 3px 6px;
"""

_ABR_ACTION_BTN_STYLE = f"""
    QPushButton {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {_SLATE_TEXT}, stop:1 {_SLATE_PRIMARY});
        color: {WHITE};
        border: 1px solid {_SLATE_DARK};
        border-radius: 3px;
        padding: 3px 10px;
        font-size: 11px;
        font-weight: bold;
        min-height: 18px;
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {_SLATE_PRIMARY}, stop:1 {_SLATE_DARK});
        color: {WHITE};
    }}
    QPushButton:pressed {{ background-color: {_SLATE_DARK}; }}
    QPushButton:disabled {{
        background: {GRAY_MID}; color: {GRAY_DARK}; border-color: {GRAY_MID};
    }}
"""

_MODE_BTN_ABR_ACTIVE_STYLE = f"""
    QPushButton {{
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 {_CRIMSON_RICH}, stop:1 {_CRIMSON_PRIMARY});
        color: {WHITE};
        border: 2px solid {_CRIMSON_DARK};
        border-radius: 4px;
        padding: 4px 16px;
        font-size: 12px;
        font-weight: bold;
        min-height: 22px;
    }}
"""

# Mime type tokens used for drag-and-drop
_MIME_CATEGORY = "application/x-suiteview-category"
_MIME_TOOL_FILE = "application/x-suiteview-toolfile"


# ---------------------------------------------------------------------------
# TightItemDelegate
# ---------------------------------------------------------------------------

class _TightItemDelegate(QStyledItemDelegate):
    ROW_H = 16
    def sizeHint(self, option, index):
        sh = super().sizeHint(option, index)
        return QSize(sh.width(), self.ROW_H)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _icon_for_ext(ext: str) -> str:
    return {
        ".py": "🐍", ".r": "📊", ".R": "📊",
        ".xlsx": "📗", ".xls": "📗", ".xlsm": "📗",
        ".sql": "🗃️", ".vbs": "⚙️", ".bat": "⚙️", ".ps1": "⚙️",
        ".txt": "📝", ".pdf": "📕", ".docx": "📘",
    }.get(ext, "📄")


# ---------------------------------------------------------------------------
# DraggableCategoryList  — Task Categories list with drag support
# ---------------------------------------------------------------------------

class _DraggableCategoryList(QListWidget):
    """QListWidget that starts a drag carrying the category name."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if not item:
            return
        mime = QMimeData()
        mime.setData(_MIME_CATEGORY, item.text().encode())
        drag = QDrag(self)
        drag.setMimeData(mime)
        # Simple drag pixmap
        pix = QPixmap(160, 18)
        pix.fill(QColor(GOLD_LIGHT))
        painter = QPainter(pix)
        painter.setFont(QFont("Segoe UI", 9))
        painter.setPen(QColor(GREEN_DARK))
        painter.drawText(4, 13, item.text())
        painter.end()
        drag.setPixmap(pix)
        drag.exec(Qt.DropAction.CopyAction)


# ---------------------------------------------------------------------------
# DraggableToolsList  — Available Tools list with drag support
# ---------------------------------------------------------------------------

class _DraggableToolsList(QListWidget):
    """QListWidget that starts a drag carrying a file path."""

    _PATH_ROLE = Qt.ItemDataRole.UserRole

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if not item:
            return
        path = item.data(self._PATH_ROLE) or ""
        if not path or not os.path.isfile(path):
            return  # only drag files, not folders
        mime = QMimeData()
        mime.setData(_MIME_TOOL_FILE, path.encode())
        drag = QDrag(self)
        drag.setMimeData(mime)
        pix = QPixmap(200, 18)
        pix.fill(QColor(GOLD_LIGHT))
        painter = QPainter(pix)
        painter.setFont(QFont("Segoe UI", 9))
        painter.setPen(QColor(GREEN_DARK))
        painter.drawText(4, 13, os.path.basename(path))
        painter.end()
        drag.setPixmap(pix)
        drag.exec(Qt.DropAction.CopyAction)


# ---------------------------------------------------------------------------
# DropTargetSubfolderList  — Policy Subfolders list that accepts drops
# ---------------------------------------------------------------------------

class _DropTargetSubfolderList(QListWidget):
    """QListWidget that accepts drops from categories and tool files."""

    category_dropped = pyqtSignal(str)   # category name
    file_dropped     = pyqtSignal(str)   # source file path

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)

    def dragEnterEvent(self, event):
        md = event.mimeData()
        if md.hasFormat(_MIME_CATEGORY) or md.hasFormat(_MIME_TOOL_FILE):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        md = event.mimeData()
        if md.hasFormat(_MIME_CATEGORY) or md.hasFormat(_MIME_TOOL_FILE):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        md = event.mimeData()
        if md.hasFormat(_MIME_CATEGORY):
            name = bytes(md.data(_MIME_CATEGORY)).decode()
            self.category_dropped.emit(name)
            event.acceptProposedAction()
        elif md.hasFormat(_MIME_TOOL_FILE):
            path = bytes(md.data(_MIME_TOOL_FILE)).decode()
            self.file_dropped.emit(path)
            event.acceptProposedAction()
        else:
            event.ignore()


# ---------------------------------------------------------------------------
# MiniExplorer  — reusable mini file-explorer panel
# ---------------------------------------------------------------------------

class MiniExplorer(QWidget):
    """Compact file explorer with Home/Up navigation and path breadcrumb.

    The list widget passed in (via list_widget_class) is used so callers
    can supply a drag-enabled subclass.
    """

    file_selected = pyqtSignal(str)

    _PATH_ROLE = Qt.ItemDataRole.UserRole

    def __init__(self, title: str = "", root_path: str = "",
                 list_widget_class=None, parent=None):
        super().__init__(parent)
        self._root_path = root_path
        self._current_path = root_path
        self._title = title
        self._at_home = True          # True when showing pinned home entries
        self._home_entries: list = [] # [(label, full_path), ...] for home view
        self._list_cls = list_widget_class or QListWidget
        self._setup_ui()

    def _setup_ui(self):
        outer = QGroupBox(self._title)
        outer.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        ol = QVBoxLayout(outer)
        ol.setContentsMargins(2, 14, 2, 2)
        ol.setSpacing(2)

        # Nav bar
        nav = QHBoxLayout()
        nav.setContentsMargins(0, 0, 0, 0)
        nav.setSpacing(2)

        self._home_btn = QPushButton("⌂")
        self._home_btn.setStyleSheet(_NAV_BTN_STYLE)
        self._home_btn.setFixedWidth(22)
        self._home_btn.setToolTip("Go home")
        self._home_btn.clicked.connect(self._go_home)

        self._up_btn = QPushButton("↑")
        self._up_btn.setStyleSheet(_NAV_BTN_STYLE)
        self._up_btn.setFixedWidth(22)
        self._up_btn.setToolTip("Go up one level")
        self._up_btn.clicked.connect(self._go_up)

        self._path_label = _DoubleClickablePathLabel("")
        self._path_label.setStyleSheet(_PATH_BAR_STYLE)
        self._path_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )

        nav.addWidget(self._home_btn)
        nav.addWidget(self._up_btn)
        nav.addWidget(self._path_label, 1)
        ol.addLayout(nav)

        # List
        self._list: QListWidget = self._list_cls()
        self._list.setStyleSheet(_LIST_STYLE)
        self._list.setItemDelegate(_TightItemDelegate(self._list))
        self._list.setUniformItemSizes(True)
        self._list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._list.itemDoubleClicked.connect(self._on_double_click)
        self._list.currentItemChanged.connect(self._on_sel_changed)
        self._list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._list.customContextMenuRequested.connect(self._on_context_menu)
        ol.addWidget(self._list, 1)

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.addWidget(outer)

    # -- Home entries (pinned folders) -------------------------------------

    def set_home_entries(self, entries: list):
        """Set the list of (label, full_path) tuples shown on the home screen."""
        self._home_entries = entries
        self._go_home()

    # -- Navigation --------------------------------------------------------

    def set_root(self, path: str):
        self._root_path = path
        self._current_path = path
        self._at_home = False
        self._refresh()

    def _go_home(self):
        self._at_home = True
        self._current_path = ""
        self._refresh()

    def _go_up(self):
        """Navigate to the parent directory. Goes home only if already at a filesystem root."""
        if self._at_home:
            return
        if not self._current_path:
            self._go_home()
            return
        parent = os.path.dirname(self._current_path)
        # If dirname returns the same path we're at a filesystem root — go home
        if parent == self._current_path:
            self._go_home()
        else:
            self._current_path = parent
            self._refresh()

    def _on_double_click(self, item: QListWidgetItem):
        """Navigate into a folder (or do nothing for files — files are drag-only)."""
        path = item.data(self._PATH_ROLE)
        if path and os.path.isdir(path):
            self._at_home = False
            self._root_path = path if not self._root_path else self._root_path
            self._current_path = path
            self._refresh()

    def _on_sel_changed(self, current: QListWidgetItem, _prev):
        if not current:
            return
        path = current.data(self._PATH_ROLE) or ""
        if os.path.isfile(path):
            self.file_selected.emit(path)

    def _on_context_menu(self, pos: QPoint):
        menu = QMenu(self)
        menu.setStyleSheet(_CONTEXT_MENU_STYLE)
        up_act = menu.addAction("↑  Go Up")
        up_act.setEnabled(not self._at_home)
        home_act = menu.addAction("⌂  Go Home")
        home_act.setEnabled(not self._at_home and bool(self._home_entries))

        item = self._list.itemAt(pos)
        open_act = None
        if item:
            path = item.data(self._PATH_ROLE) or ""
            if os.path.isdir(path):
                menu.addSeparator()
                open_act = menu.addAction("Open in Explorer")

        action = menu.exec(self._list.viewport().mapToGlobal(pos))
        if action is up_act:
            self._go_up()
        elif action is home_act:
            self._go_home()
        elif action is open_act and item:
            path = item.data(self._PATH_ROLE) or ""
            if os.path.exists(path):
                os.startfile(path)

    # -- Rendering ---------------------------------------------------------

    def _refresh(self):
        self._list.clear()
        self._up_btn.setEnabled(not self._at_home)
        self._home_btn.setEnabled(not self._at_home and bool(self._home_entries))

        if self._at_home:
            self._path_label.setText("Home")
            self._path_label.set_open_path("")  # nothing to open at home
            for label, full_path in self._home_entries:
                exists = os.path.isdir(full_path)
                icon = "📁" if exists else "⚠️"
                item = QListWidgetItem(f"{icon}  {label}")
                item.setData(self._PATH_ROLE, full_path)
                if not exists:
                    item.setForeground(QColor(GRAY_MID))
                    item.setToolTip("Folder not found")
                self._list.addItem(item)
            return

        self._update_path_label()
        if not self._current_path or not os.path.isdir(self._current_path):
            return

        try:
            entries = sorted(
                os.listdir(self._current_path),
                key=lambda e: (
                    not os.path.isdir(os.path.join(self._current_path, e)),
                    e.lower()
                )
            )
            for entry in entries:
                full = os.path.join(self._current_path, entry)
                if os.path.isdir(full):
                    item = QListWidgetItem(f"📁  {entry}")
                else:
                    ext = os.path.splitext(entry)[1].lower()
                    item = QListWidgetItem(f"{_icon_for_ext(ext)}  {entry}")
                item.setData(self._PATH_ROLE, full)
                self._list.addItem(item)
        except Exception as e:
            self._list.addItem(QListWidgetItem(f"(Error: {e})"))

    def _update_path_label(self):
        if not self._current_path:
            self._path_label.setText("")
            self._path_label.set_open_path("")
            return
        # Show last 2 path components for compactness; full path stored for open
        try:
            parts = self._current_path.replace("\\", "/").split("/")
            rel = "\\".join(parts[-2:]) if len(parts) >= 2 else self._current_path
        except Exception:
            rel = self._current_path
        self._path_label.setText(rel)
        self._path_label.set_open_path(self._current_path)

    # -- Public helpers ----------------------------------------------------

    def current_path(self) -> str:
        return self._current_path if not self._at_home else ""

    def selected_file_path(self) -> str:
        item = self._list.currentItem()
        if not item:
            return ""
        path = item.data(self._PATH_ROLE) or ""
        return path if os.path.isfile(path) else ""

    def refresh(self):
        self._refresh()


# ---------------------------------------------------------------------------
# _DoubleClickablePathLabel  — QLabel that opens a folder in Explorer on dbl-click
# ---------------------------------------------------------------------------

class _DoubleClickablePathLabel(QLabel):
    """Path-bar label that opens its associated folder in Windows Explorer on double-click.

    The displayed text may be a short/relative form; call set_open_path() to
    store the full absolute path that will actually be opened.
    """

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self._open_path: str = ""
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def set_open_path(self, path: str):
        """Set the full path that will be opened on double-click."""
        self._open_path = path
        if path:
            self.setToolTip(f"Double-click to open in Explorer:\n{path}")
        else:
            self.setToolTip("")

    def mouseDoubleClickEvent(self, event):
        path = self._open_path or self.text().strip()
        if path and os.path.isdir(path):
            os.startfile(path)
        super().mouseDoubleClickEvent(event)


class _PolicyLibraryGrid(FixedHeaderTableWidget):
    """FixedHeaderTableWidget with an additional all-column text search."""

    def __init__(self, parent=None, status_callback=None):
        super().__init__(parent=parent, filterable=True)
        self._general_filter_text = ""
        self._status_callback = status_callback

    def set_general_filter_text(self, text: str):
        self._general_filter_text = text.strip().lower()
        self._apply_filters()

    def visible_row_count(self) -> int:
        return sum(
            not self._data_table.isRowHidden(row)
            for row in range(self._data_table.rowCount())
        )

    def _apply_filters(self):
        super()._apply_filters()
        if self._general_filter_text:
            for row in range(self._data_table.rowCount()):
                if self._data_table.isRowHidden(row):
                    continue
                row_matches = False
                for col in range(self._data_table.columnCount()):
                    item = self._data_table.item(row, col)
                    if item and self._general_filter_text in item.text().lower():
                        row_matches = True
                        break
                self._data_table.setRowHidden(row, not row_matches)
        if self._status_callback:
            self._status_callback(self.visible_row_count(), self._data_table.rowCount())


# ---------------------------------------------------------------------------
# PolicyLibraryTab
# ---------------------------------------------------------------------------

class PolicyLibraryTab(QWidget):
    """Searchable index of Policy Support library folders."""

    HEADERS = ("Product Type", "Policy Number", "Category")

    def __init__(self, parent=None, root_path: str = ""):
        super().__init__(parent)
        self._root_path = root_path or _get_policy_library_dir()
        self._setup_ui()

    def _setup_ui(self):
        self.setStyleSheet(f"background-color: {WHITE};")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 4, 8, 4)
        layout.setSpacing(6)

        info_frame = QGroupBox("Policy Library")
        info_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        info_layout = QGridLayout(info_frame)
        info_layout.setContentsMargins(8, 20, 8, 6)
        info_layout.setHorizontalSpacing(8)
        info_layout.setVerticalSpacing(4)

        lbl_s = (f"font-size: 11px; font-weight: bold; color: {GREEN_DARK}; "
                 f"background: transparent; border: none;")
        folder_label = QLabel("Folder Location:")
        folder_label.setStyleSheet(lbl_s)
        info_layout.addWidget(folder_label, 0, 0)

        self._path_label = _DoubleClickablePathLabel(self._root_path)
        self._path_label.set_open_path(self._root_path)
        self._path_label.setStyleSheet(_PATH_LABEL_STYLE)
        self._path_label.setWordWrap(True)
        info_layout.addWidget(self._path_label, 0, 1)

        self._refresh_btn = QPushButton("Refresh")
        self._refresh_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._refresh_btn.clicked.connect(self.refresh)
        info_layout.addWidget(self._refresh_btn, 0, 2)

        self._status_label = QLabel("")
        self._status_label.setStyleSheet(_STATUS_STYLE)
        info_layout.addWidget(self._status_label, 1, 1, 1, 2)
        info_layout.setColumnStretch(1, 1)
        layout.addWidget(info_frame)

        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("Search product type, policy number, or category")
        self._search_input.setStyleSheet(_FILTER_INPUT_STYLE)
        self._search_input.textChanged.connect(self._apply_filters)
        layout.addWidget(self._search_input)

        self._table = _PolicyLibraryGrid(status_callback=self._update_filter_status)
        self._table.setColumnCount(len(self.HEADERS))
        self._table.setHorizontalHeaderLabels(self.HEADERS)
        self._table._data_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table._data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table._data_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table._data_table.setSortingEnabled(True)
        self._table._data_table.doubleClicked.connect(self._open_selected_folder)
        layout.addWidget(self._table, 1)

    def refresh(self):
        rows = self._scan_library()
        self._table._data_table.setSortingEnabled(False)
        self._table.setRowCount(0)
        for product_type, policy_number, category, folder_path in rows:
            row = self._table.rowCount()
            self._table._data_table.insertRow(row)
            for col, value in enumerate((product_type, policy_number, category)):
                item = QTableWidgetItem(value)
                item.setData(Qt.ItemDataRole.UserRole, folder_path)
                self._table.setItem(row, col, item, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self._table._data_table.setSortingEnabled(True)
        self._table.autoFitAllColumns()
        self._apply_filters()

    def _scan_library(self):
        rows = set()
        if not os.path.isdir(self._root_path):
            self._status_label.setText("Policy Library folder was not found")
            return []

        for product_type in sorted(os.listdir(self._root_path)):
            product_path = os.path.join(self._root_path, product_type)
            if not os.path.isdir(product_path):
                continue
            for policy_number in sorted(os.listdir(product_path)):
                policy_path = os.path.join(product_path, policy_number)
                if not os.path.isdir(policy_path):
                    continue

                found_category = False
                for current_path, dir_names, _file_names in os.walk(policy_path):
                    dir_names.sort()
                    if current_path == policy_path:
                        continue
                    category = os.path.relpath(current_path, policy_path).replace(os.sep, "\\")
                    rows.add((product_type, policy_number, category, current_path))
                    found_category = True

                if not found_category:
                    rows.add((product_type, policy_number, "", policy_path))

        self._status_label.setText(f"{len(rows)} library folders indexed")
        return sorted(rows, key=lambda row: (row[0].lower(), row[1].lower(), row[2].lower()))

    def _apply_filters(self):
        self._table.set_general_filter_text(self._search_input.text())

    def _update_filter_status(self, visible_count: int, total: int):
        if os.path.isdir(self._root_path):
            self._status_label.setText(f"Showing {visible_count} of {total} library folders")

    def _open_selected_folder(self, index):
        item = self._table.item(index.row(), 0)
        folder_path = item.data(Qt.ItemDataRole.UserRole) if item else ""
        if folder_path and os.path.isdir(folder_path):
            os.startfile(folder_path)


# ---------------------------------------------------------------------------
# PolicySupportTab
# ---------------------------------------------------------------------------

class PolicySupportTab(QWidget):
    """Policy Support tab with drag-and-drop workflow.

    - Drag a Task Category → drop on Policy Subfolders → creates that folder
    - Drag a file from Available Tools → drop on Policy Subfolders → copies the file
    - Folders from Available Tools cannot be dragged (only files)
    """

    folder_opened = pyqtSignal(str)
    policy_library_requested = pyqtSignal()
    sap_requested = pyqtSignal()
    claims_requested = pyqtSignal()
    tai_fd_requested = pyqtSignal()
    orion_pcr_requested = pyqtSignal()
    cyberlife_pdf_requested = pyqtSignal()

    # Mode constants
    MODE_POLICY_SUPPORT = "policy_support"
    MODE_ABR = "abr"
    SECTION_ANNUITY_RIDER = "annuity_rider"
    SECTION_GLP_EXCEPTION = "glp_exception"
    SECTION_FORECAST = "forecast"

    def __init__(self, parent=None):
        super().__init__(parent)
        self._policy: Optional['PolicyInformation'] = None
        self._policy_support_folder_path = ""
        self._policy_library_path = ""
        self._user_tasks: List[str] = _load_user_tasks()
        self._current_mode = self.MODE_POLICY_SUPPORT
        self._current_section = self.MODE_POLICY_SUPPORT
        self._glp_quote_cache = {}
        self._forecast_cache = {}
        self._setup_ui()

    # -- UI ----------------------------------------------------------------

    def _setup_ui(self):
        self.setStyleSheet(f"background-color: {WHITE};")
        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 4, 6, 4)
        layout.setSpacing(8)

        nav_panel = QWidget()
        nav_panel.setObjectName("PolicySupportNavPanel")
        nav_panel.setStyleSheet(_NAV_PANEL_STYLE)
        nav_panel.setFixedWidth(138)
        nav_col = QVBoxLayout(nav_panel)
        nav_col.setContentsMargins(8, 8, 8, 8)
        nav_col.setSpacing(7)

        self._btn_mode_polsup = QPushButton("Policy Support")
        self._btn_mode_polsup.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_mode_polsup.setMinimumWidth(108)
        self._btn_mode_polsup.clicked.connect(
            lambda: self._select_section(self.MODE_POLICY_SUPPORT)
        )
        nav_col.addWidget(self._btn_mode_polsup)

        self._btn_mode_abr = QPushButton("ABR")
        self._btn_mode_abr.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_mode_abr.setMinimumWidth(108)
        self._btn_mode_abr.clicked.connect(lambda: self._select_section(self.MODE_ABR))
        nav_col.addWidget(self._btn_mode_abr)

        self._btn_mode_glp_exception = QPushButton("GLP Exception")
        self._btn_mode_glp_exception.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_mode_glp_exception.setMinimumWidth(108)
        self._btn_mode_glp_exception.clicked.connect(
            lambda: self._select_section(self.SECTION_GLP_EXCEPTION)
        )
        nav_col.addWidget(self._btn_mode_glp_exception)

        self._btn_mode_forecast = QPushButton("Forecast")
        self._btn_mode_forecast.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_mode_forecast.setMinimumWidth(108)
        self._btn_mode_forecast.clicked.connect(
            lambda: self._select_section(self.SECTION_FORECAST)
        )
        nav_col.addWidget(self._btn_mode_forecast)

        self._btn_mode_annuity = QPushButton("Annuity Rider")
        self._btn_mode_annuity.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_mode_annuity.setMinimumWidth(108)
        self._btn_mode_annuity.clicked.connect(
            lambda: self._select_section(self.SECTION_ANNUITY_RIDER)
        )
        nav_col.addWidget(self._btn_mode_annuity)

        self._btn_sap = QPushButton("SAP")
        self._btn_sap.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_sap.setMinimumWidth(108)
        self._btn_sap.clicked.connect(self.sap_requested.emit)
        nav_col.addWidget(self._btn_sap)

        self._btn_claims = QPushButton("CLAIMSFILE")
        self._btn_claims.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_claims.setMinimumWidth(108)
        self._btn_claims.clicked.connect(self.claims_requested.emit)
        nav_col.addWidget(self._btn_claims)

        self._btn_tai_fd = QPushButton("TAICyberTAIFd")
        self._btn_tai_fd.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_tai_fd.setMinimumWidth(108)
        self._btn_tai_fd.clicked.connect(self.tai_fd_requested.emit)
        nav_col.addWidget(self._btn_tai_fd)

        self._btn_orion_pcr = QPushButton("orion_pcr3_r")
        self._btn_orion_pcr.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_orion_pcr.setMinimumWidth(108)
        self._btn_orion_pcr.clicked.connect(self.orion_pcr_requested.emit)
        nav_col.addWidget(self._btn_orion_pcr)

        self._btn_cyberlife_pdf = QPushButton("CYBERLIFE_PDF")
        self._btn_cyberlife_pdf.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_cyberlife_pdf.setMinimumWidth(108)
        self._btn_cyberlife_pdf.clicked.connect(self.cyberlife_pdf_requested.emit)
        nav_col.addWidget(self._btn_cyberlife_pdf)
        nav_col.addStretch(1)

        layout.addWidget(nav_panel)

        self._content_stack = QStackedWidget()
        self._content_stack.setStyleSheet(f"background-color: {WHITE};")

        self._workspace_page = QWidget()
        workspace_layout = QVBoxLayout(self._workspace_page)
        workspace_layout.setContentsMargins(0, 0, 0, 0)
        workspace_layout.setSpacing(4)

        # ── Top info bar ──────────────────────────────────────────────────
        self._info_frame = QGroupBox("Policy Support")
        self._info_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        ig = QGridLayout(self._info_frame)
        ig.setContentsMargins(8, 18, 8, 6)
        ig.setHorizontalSpacing(6)
        ig.setVerticalSpacing(4)

        lbl_s = (f"font-size: 11px; font-weight: bold; color: {GREEN_DARK}; "
                 f"background: transparent; border: none;")
        val_s = (f"font-size: 11px; color: {GRAY_DARK}; background: transparent; "
                 f"border: none;")

        ig.addWidget(self._mk_lbl("Policy Folder:", lbl_s), 0, 0)
        self._policy_folder_label = QLabel("No policy loaded")
        self._policy_folder_label.setStyleSheet(val_s)
        ig.addWidget(self._policy_folder_label, 0, 1)

        # OneDrive / SharePoint warning — shown when Process_Control folder is missing
        self._onedrive_warning_label = QLabel(
            "⚠️  SharePoint library not linked to OneDrive.\n"
            "To use Policy Support, sync the 'Life Product - Process_Control'\n"
            "library to your OneDrive."
        )
        self._onedrive_warning_label.setStyleSheet(
            "font-size: 10px; font-weight: bold; color: #7B3F00; "
            "background: #FFF3CD; border: 1px solid #FFC107; "
            "border-radius: 4px; padding: 4px 8px;"
        )
        self._onedrive_warning_label.setWordWrap(True)
        self._onedrive_warning_label.setVisible(False)
        ig.addWidget(self._onedrive_warning_label, 0, 2, 2, 1)  # span rows 0–1

        self._policy_library_btn = QPushButton("Policy Library")
        self._policy_library_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._policy_library_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._policy_library_btn.clicked.connect(self.policy_library_requested.emit)
        ig.addWidget(self._policy_library_btn, 0, 3)

        ig.addWidget(self._mk_lbl("Library Path:", lbl_s), 1, 0)
        self._library_path_label = _DoubleClickablePathLabel("")
        self._library_path_label.setStyleSheet(_PATH_LABEL_STYLE)
        self._library_path_label.setWordWrap(True)
        self._library_path_label.setToolTip("Double-click to open in Explorer")
        ig.addWidget(self._library_path_label, 1, 1)

        ig.addWidget(self._mk_lbl("Status:", lbl_s), 2, 0)
        self._status_label = QLabel("")
        self._status_label.setStyleSheet(_STATUS_STYLE)
        ig.addWidget(self._status_label, 2, 1)

        self._create_folder_btn = QPushButton("📁  Create Policy Folder")
        self._create_folder_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._create_folder_btn.clicked.connect(self._on_create_policy_folder)
        self._create_folder_btn.setVisible(False)
        ig.addWidget(self._create_folder_btn, 2, 3)

        ig.setColumnStretch(1, 1)
        workspace_layout.addWidget(self._info_frame)

        # Check OneDrive linkage immediately so the warning shows even before a policy loads
        self._check_onedrive()

        # ── Hint bar ──────────────────────────────────────────────────────
        hint = QLabel(
            "💡  Drag a category → Policy Subfolders to create a task folder  |  "
            "Drag a file from Available Tools → Policy Subfolders to copy it"
        )
        hint.setStyleSheet(
            f"font-size: 10px; color: {GRAY_DARK}; background: transparent; "
            f"border: none; padding: 0px 2px;"
        )
        workspace_layout.addWidget(hint)

        # ── Three-column middle ───────────────────────────────────────────
        cols = QWidget()
        cl = QHBoxLayout(cols)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(4)

        # Column 1 — Task Categories (draggable)
        self._cat_frame = QGroupBox("Task Categories")
        self._cat_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        cat_fl = QVBoxLayout(self._cat_frame)
        cat_fl.setContentsMargins(2, 14, 2, 2)
        cat_fl.setSpacing(0)

        self._category_list = _DraggableCategoryList()
        self._category_list.setStyleSheet(_LIST_STYLE)
        self._category_list.setItemDelegate(_TightItemDelegate(self._category_list))
        self._category_list.setUniformItemSizes(True)
        self._category_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._category_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._category_list.customContextMenuRequested.connect(self._on_category_context_menu)
        cat_fl.addWidget(self._category_list)

        cl.addWidget(self._cat_frame, 3)

        # Column 2 — Policy Subfolders (drop target + mini explorer)
        sub_col = QVBoxLayout()
        sub_col.setSpacing(0)

        # Drop-target list at the top (shows current subfolder contents)
        self._subfolder_explorer = _SubfolderExplorer(title="Policy Subfolders")
        self._subfolder_explorer.category_dropped.connect(self._on_category_dropped)
        self._subfolder_explorer.file_dropped.connect(self._on_file_dropped)
        sub_col.addWidget(self._subfolder_explorer, 1)
        cl.addLayout(sub_col, 5)

        # Column 3 — Available Tools (draggable mini explorer)
        self._tools_explorer = MiniExplorer(
            title="Available Tools",
            list_widget_class=_DraggableToolsList,
        )
        # Build home entries from the current user's OneDrive root.
        onedrive_root = _get_onedrive_dir()
        home_entries = [
            (label, os.path.join(onedrive_root, rel_path))
            for label, rel_path in TOOL_FOLDERS
        ]
        self._tools_explorer.set_home_entries(home_entries)
        cl.addWidget(self._tools_explorer, 3)

        workspace_layout.addWidget(cols, 1)

        self._annuity_rider_tab = AnnuityRiderTab()
        self._glp_exception_page = self._build_glp_exception_page()
        self._forecast_page = self._build_forecast_page()

        self._content_stack.addWidget(self._workspace_page)
        self._content_stack.addWidget(self._annuity_rider_tab)
        self._content_stack.addWidget(self._glp_exception_page)
        self._content_stack.addWidget(self._forecast_page)
        layout.addWidget(self._content_stack, 1)

        self._refresh_section_buttons()

    @staticmethod
    def _mk_lbl(text: str, style: str) -> QLabel:
        l = QLabel(text)
        l.setStyleSheet(style)
        return l

    def _build_glp_exception_page(self) -> QWidget:
        page = QWidget()
        page.setStyleSheet(f"background-color: {WHITE};")
        page.setAutoFillBackground(True)
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        status_frame = QGroupBox("GLP Exception")
        status_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        sg = QGridLayout(status_frame)
        sg.setContentsMargins(8, 18, 8, 6)
        sg.setHorizontalSpacing(6)
        sg.setVerticalSpacing(4)

        lbl_s = (f"font-size: 11px; font-weight: bold; color: {GREEN_DARK}; "
                 f"background: transparent; border: none;")
        sg.addWidget(self._mk_lbl("Forecast Status:", lbl_s), 0, 0)
        self._glp_forecast_status_label = QLabel("Load an eligible policy to check forecast data")
        self._glp_forecast_status_label.setStyleSheet(_STATUS_STYLE)
        self._glp_forecast_status_label.setWordWrap(True)
        sg.addWidget(self._glp_forecast_status_label, 0, 1, 1, 3)

        sg.addWidget(self._mk_lbl("Target Inforce Date:", lbl_s), 1, 0)
        self._glp_target_date = QLineEdit()
        self._glp_target_date.setPlaceholderText("MM/DD/YYYY")
        self._glp_target_date.setStyleSheet(_FILTER_INPUT_STYLE)
        sg.addWidget(self._glp_target_date, 1, 1)

        self._glp_calculate_btn = QPushButton("Calculate")
        self._glp_calculate_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._glp_calculate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._glp_calculate_btn.clicked.connect(self._on_calculate_glp_exception)
        sg.addWidget(self._glp_calculate_btn, 1, 2)

        button_bar = QHBoxLayout()
        button_bar.setContentsMargins(0, 0, 0, 0)
        button_bar.setSpacing(6)
        self._glp_export_btn = QPushButton("Export")
        self._glp_export_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._glp_export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._glp_export_btn.setEnabled(False)
        self._glp_export_btn.clicked.connect(self._export_glp_quote)
        self._glp_print_details_btn = QPushButton("Print Details to Folder")
        self._glp_print_details_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._glp_print_details_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._glp_print_details_btn.setEnabled(False)
        self._glp_print_details_btn.clicked.connect(self._print_glp_quote_to_folder)
        self._glp_helper_btn = QPushButton("Helper")
        self._glp_helper_btn.setStyleSheet(_MODE_BTN_INACTIVE_STYLE)
        self._glp_helper_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._glp_helper_btn.clicked.connect(self._show_glp_helper)
        button_bar.addWidget(self._glp_export_btn)
        button_bar.addWidget(self._glp_print_details_btn)
        button_bar.addWidget(self._glp_helper_btn)
        button_bar.addStretch(1)
        sg.addLayout(button_bar, 1, 3)

        self._glp_note_label = QLabel(
            "The New Accum GLP will cover the policy up to but not including the target date. "
            "The New Accum GLP will need to be recalculated if a premium is paid or policy anniversary is crossed since this quote was calculated."
        )
        self._glp_note_label.setWordWrap(True)
        self._glp_note_label.setStyleSheet(
            f"font-size: 10px; color: {GRAY_DARK}; background: transparent; "
            f"border: none; padding: 2px 4px;"
        )
        sg.addWidget(self._glp_note_label, 2, 0, 1, 4)
        sg.setColumnStretch(3, 1)
        layout.addWidget(status_frame)

        results_frame = QGroupBox("Calculation")
        results_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        results_frame.setMinimumWidth(430)
        rg = QGridLayout(results_frame)
        rg.setContentsMargins(10, 18, 10, 8)
        rg.setHorizontalSpacing(18)
        rg.setVerticalSpacing(4)
        self._glp_result_labels = {}
        self._glp_result_name_labels = {}
        rows = [
            ("Current Valuation Date", "current_valuation_date", "date"),
            ("Account Value", "account_value", "money"),
            ("Premiums Paid To Date", "premiums_paid_to_date", "money"),
            ("Premiums since Val Date", "premiums_since_valuation_date", "money"),
            ("Adjusted Account Value", "adjusted_account_value", "money"),
            ("Adjusted Premiums Paid To Date", "adjusted_premiums_paid_to_date", "money"),
            ("Accum Withdrawals", "accumulated_withdrawals", "money"),
            ("GLP", "glp", "money"),
            ("GSP", "gsp", "money"),
            ("Accum GLP (Current)", "accumulated_glp", "money"),
            ("", "glp_timing_separator", "thin_separator"),
            ("Total Required Premium to stay inforce to Target Date (before load)", "total_required_premium_before_load", "money"),
            ("Premium to get to Target Date", "total_required_premium_after_load", "money"),
            ("Accum GLP on Target Date", "accumulated_glp_prior_to_target", "money"),
            ("PremTD less AccumWD on Target Date", "premium_td_on_target_date", "money"),
            ("Adjustment to Accum GLP pre calc", "adjustment_to_accum_glp_pre_calc", "money"),
            ("", "glp_decision_separator", "separator"),
            ("New GLP", "new_glp", "money"),
            ("Adjustment to Accum GLP", "adjustment_to_accum_glp", "money"),
            ("New Accum GLP", "new_accum_glp", "money"),
            ("NO ADJUSTMENT NEEDED", "glp_adjustment_message", "message"),
            ("FORCE-OUT REQUIRED", "force_out_required", "message"),
            ("Force-out Amount", "force_out_amount", "money"),
        ]
        self._glp_result_rows = rows
        for row, (label, key, kind) in enumerate(rows):
            name = QLabel(label) if kind in {"separator", "thin_separator"} else CopyableLabel(label)
            if kind in {"separator", "thin_separator"}:
                name.setFixedHeight(10 if kind == "separator" else 6)
                name.setStyleSheet(
                    f"background: transparent; border: none; "
                    f"border-top: {'2px dashed' if kind == 'separator' else '1px solid'} {GREEN_PRIMARY}; "
                    f"margin-top: {'6px' if kind == 'separator' else '2px'};"
                )
                value = QLabel("", results_frame)
                rg.addWidget(name, row, 0, 1, 2)
                self._glp_result_name_labels[key] = name
                self._glp_result_labels[key] = (value, kind)
                continue
            if kind == "message":
                name.setStyleSheet(_RESULT_LABEL_STYLE + f"font-weight: bold; color: {GREEN_DARK};")
            elif key in {"new_glp", "new_accum_glp", "adjustment_to_accum_glp"}:
                name.setStyleSheet(_RESULT_LABEL_STYLE + f"font-weight: bold; color: {GREEN_DARK};")
            else:
                name.setStyleSheet(_RESULT_LABEL_STYLE)
            value = CopyableLabel("-")
            value.setStyleSheet(_RESULT_VALUE_STYLE)
            value.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            value.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            rg.addWidget(name, row, 0)
            rg.addWidget(value, row, 1)
            self._glp_result_name_labels[key] = name
            self._glp_result_labels[key] = (value, kind)
        rg.setColumnStretch(1, 1)

        self._glp_forecast_frame = QGroupBox("Monthly Forecast")
        self._glp_forecast_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        self._glp_forecast_frame.setMinimumHeight(180)
        forecast_layout = QVBoxLayout(self._glp_forecast_frame)
        forecast_layout.setContentsMargins(6, 18, 6, 6)
        forecast_layout.setSpacing(0)
        self._glp_forecast_table = FixedHeaderTableWidget()
        self._glp_forecast_table.setAutoFillBackground(True)
        self._glp_forecast_table._data_table.viewport().setAutoFillBackground(True)
        self._glp_forecast_table.setColumnCount(7)
        self._glp_forecast_table.setHorizontalHeaderLabels([
            "Date", "Year", "Month", "Interest Credited", "Premium", "Monthly Deduction", "Account Value"
        ])
        self._glp_forecast_table._data_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._glp_forecast_table._data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._glp_forecast_table._data_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        forecast_layout.addWidget(self._glp_forecast_table, 1)
        self._glp_forecast_frame.setVisible(False)

        body = QWidget()
        body.setAutoFillBackground(True)
        body.setStyleSheet(f"background-color: {WHITE};")
        body_layout = QHBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(6)
        body_layout.addWidget(results_frame, 3)
        body_layout.addWidget(self._glp_forecast_frame, 4)
        layout.addWidget(body, 1)
        return page

    def _build_forecast_page(self) -> QWidget:
        page = QWidget()
        page.setStyleSheet(f"background-color: {WHITE};")
        page.setAutoFillBackground(True)
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        status_frame = QGroupBox("Forecast")
        status_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        sg = QGridLayout(status_frame)
        sg.setContentsMargins(8, 18, 8, 6)
        sg.setHorizontalSpacing(6)
        sg.setVerticalSpacing(4)

        lbl_s = (f"font-size: 11px; font-weight: bold; color: {GREEN_DARK}; "
                 f"background: transparent; border: none;")
        sg.addWidget(self._mk_lbl("Forecast Status:", lbl_s), 0, 0)
        self._forecast_status_label = QLabel("Load an eligible policy to check forecast data")
        self._forecast_status_label.setStyleSheet(_STATUS_STYLE)
        self._forecast_status_label.setWordWrap(True)
        sg.addWidget(self._forecast_status_label, 0, 1, 1, 5)

        sg.addWidget(self._mk_lbl("Target Date:", lbl_s), 1, 0)
        self._forecast_target_date = QLineEdit()
        self._forecast_target_date.setPlaceholderText("MM/DD/YYYY")
        self._forecast_target_date.setStyleSheet(_FILTER_INPUT_STYLE)
        sg.addWidget(self._forecast_target_date, 1, 1)

        sg.addWidget(self._mk_lbl("Premium Amount:", lbl_s), 1, 2)
        self._forecast_premium_amount = QLineEdit()
        self._forecast_premium_amount.setPlaceholderText("0.00")
        self._forecast_premium_amount.setStyleSheet(_FILTER_INPUT_STYLE)
        sg.addWidget(self._forecast_premium_amount, 1, 3)

        sg.addWidget(self._mk_lbl("Mode:", lbl_s), 1, 4)
        self._forecast_premium_mode = QComboBox()
        self._forecast_premium_mode.addItems(["Monthly", "Quarterly", "Semi-Annual", "Annual"])
        self._forecast_premium_mode.setStyleSheet(_FILTER_INPUT_STYLE)
        sg.addWidget(self._forecast_premium_mode, 1, 5)

        self._forecast_calculate_btn = QPushButton("Calculate")
        self._forecast_calculate_btn.setStyleSheet(_ACTION_BTN_STYLE)
        self._forecast_calculate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._forecast_calculate_btn.clicked.connect(self._on_calculate_forecast)
        sg.addWidget(self._forecast_calculate_btn, 1, 6)
        sg.setColumnStretch(6, 1)
        layout.addWidget(status_frame)

        self._forecast_table = FixedHeaderTableWidget()
        self._forecast_table.setAutoFillBackground(True)
        self._forecast_table._data_table.viewport().setAutoFillBackground(True)
        self._forecast_table.setColumnCount(13)
        self._forecast_table.setHorizontalHeaderLabels([
            "Date", "Year", "Month", "Interest", "GLP", "Accum GLP", "PremTD",
            "AccumWD", "ForceOut", "Premium", "AV bf MD", "Monthly Deduction", "Account Value"
        ])
        self._forecast_table._data_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._forecast_table._data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._forecast_table._data_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        layout.addWidget(self._forecast_table, 1)
        return page

    # -- OneDrive check ----------------------------------------------------

    def _check_onedrive(self) -> bool:
        """Check if the active SharePoint library is synced to OneDrive.

        Shows a warning label inside this tab if not found.
        Returns True if the folder exists, False otherwise.
        Note: this only affects the Policy Support tab — the rest of PolView
        works fine without the OneDrive link.
        """
        if self._current_mode == self.MODE_ABR:
            required_dir = _get_abr_root_dir()
            library_name = "Life Product - Accelerated_Benefits"
        else:
            required_dir = _get_process_control_dir()
            library_name = "Life Product - Process_Control"

        exists = os.path.isdir(required_dir)
        self._onedrive_warning_label.setText(
            "⚠️  SharePoint library not linked to OneDrive.\n"
            f"To use Policy Support, sync the '{library_name}'\n"
            "library to your OneDrive."
        )
        self._onedrive_warning_label.setVisible(not exists)
        # Hide library path row when there's nothing useful to show
        self._library_path_label.setVisible(exists)
        return exists

    # -- Data loading -------------------------------------------------------

    def load_data_from_policy(self, policy: 'PolicyInformation'):
        self._policy = policy
        has_annuity_rider = self._has_annuity_rider(policy)
        self._btn_mode_annuity.setEnabled(has_annuity_rider)
        glp_eligible = is_glp_exception_eligible(policy)
        self._btn_mode_glp_exception.setEnabled(glp_eligible)
        self._btn_mode_forecast.setEnabled(glp_eligible)
        if self._current_section == self.SECTION_GLP_EXCEPTION:
            self._refresh_glp_exception_status()
        elif self._current_section == self.SECTION_FORECAST:
            self._refresh_forecast_status()
        elif hasattr(self, "_glp_forecast_status_label"):
            self._clear_glp_exception_results()
            self._glp_forecast_status_label.setText(
                "Select GLP Exception to check forecast data"
                if glp_eligible else
                "GLP Exception is available only for UL policies using Guideline Premium."
            )
            self._glp_calculate_btn.setEnabled(False)
            self._clear_forecast_results()
            self._forecast_status_label.setText(
                "Select Forecast to check forecast data"
                if glp_eligible else
                "Forecast is available only for UL policies using Guideline Premium."
            )
            self._forecast_calculate_btn.setEnabled(False)

        if has_annuity_rider:
            self._annuity_rider_tab.load_data_from_policy(policy)
        elif self._current_section == self.SECTION_ANNUITY_RIDER:
            self._select_section(self.MODE_POLICY_SUPPORT, reload_policy=False)

        if not policy or not policy.exists:
            self._policy_folder_label.setText("No policy loaded")
            self._library_path_label.setText("")
            self._status_label.setText("")
            self._create_folder_btn.setVisible(False)
            self._category_list.clear()
            self._subfolder_explorer.set_root("")
            self._btn_mode_abr.setEnabled(False)
            self._btn_mode_glp_exception.setEnabled(False)
            self._btn_mode_forecast.setEnabled(False)
            self._refresh_section_buttons()
            return

        self._btn_mode_abr.setEnabled(True)

        # Bail out early if the SharePoint library isn't synced to OneDrive.
        # The rest of PolView works fine without it — only this tab needs it.
        if not self._check_onedrive():
            self._policy_folder_label.setText("")
            self._library_path_label.setText("")
            self._status_label.setText("")
            self._create_folder_btn.setVisible(False)
            self._subfolder_explorer.set_root("")
            self._refresh_section_buttons()
            return

        # Populate categories (disabled in ABR mode)
        if self._current_mode == self.MODE_ABR:
            self._category_list.clear()
            self._category_list.setEnabled(False)
        else:
            self._category_list.setEnabled(True)
            self._populate_category_list()

        product_type = policy.product_type
        company_code = policy.company_code
        policy_number = policy.policy_number

        if self._current_mode == self.MODE_ABR:
            # ABR mode: point to the ABR Policies folder
            abr_dir = _get_abr_dir()
            self._policy_library_path = os.path.join(abr_dir, "Policies")
            self._policy_support_folder_path = os.path.join(
                self._policy_library_path, policy_number
            )
        else:
            # Policy Support mode: original behavior
            ps_dir = _get_policy_support_dir()
            self._policy_library_path = os.path.join(ps_dir, "POLICY_LIBRARY", product_type)
            self._policy_support_folder_path = os.path.join(
                self._policy_library_path, f"{company_code}_{policy_number}"
            )

        self._library_path_label.setText(self._policy_library_path)
        folder_exists = os.path.isdir(self._policy_support_folder_path)

        if folder_exists:
            folder_name = policy_number if self._current_mode == self.MODE_ABR else f"{company_code}_{policy_number}"
            self._policy_folder_label.setText(folder_name)
            self._policy_folder_label.setStyleSheet(
                f"font-size: 11px; color: {GREEN_DARK}; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._create_folder_btn.setVisible(False)
            self._status_label.setText("✓ Policy folder exists")
            self._status_label.setStyleSheet(
                f"font-size: 10px; color: {GREEN_DARK}; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._subfolder_explorer.set_root(self._policy_support_folder_path)
        else:
            self._policy_folder_label.setText("No Folder Found")
            self._policy_folder_label.setStyleSheet(
                f"font-size: 11px; color: #C00000; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._create_folder_btn.setVisible(True)
            self._status_label.setText(
                "Policy folder does not exist — drag a category here or click Create"
            )
            self._status_label.setStyleSheet(
                f"font-size: 10px; color: {GOLD_DARK}; "
                f"background: transparent; border: none;"
            )
            self._subfolder_explorer.set_root("")

        self._refresh_section_buttons()

    # -- Mode switching -----------------------------------------------------

    def _set_mode(self, mode: str, *, reload_policy: bool = True):
        """Switch between Policy Support and ABR modes."""
        if mode == self._current_mode:
            if reload_policy and self._policy:
                self.load_data_from_policy(self._policy)
            return
        self._current_mode = mode

        if mode == self.MODE_POLICY_SUPPORT:
            self._info_frame.setTitle("Policy Support")
            self._policy_library_btn.setVisible(True)
            self._apply_green_theme()
        else:
            self._info_frame.setTitle("ABR Policy Support")
            self._policy_library_btn.setVisible(False)
            self._apply_abr_theme()

        if reload_policy and self._policy:
            self.load_data_from_policy(self._policy)

    def _select_section(self, section: str, *, reload_policy: bool = True):
        if section == self.SECTION_ANNUITY_RIDER and not self._btn_mode_annuity.isEnabled():
            return
        if section == self.MODE_ABR and not self._btn_mode_abr.isEnabled():
            return
        if section == self.SECTION_GLP_EXCEPTION and not self._btn_mode_glp_exception.isEnabled():
            return
        if section == self.SECTION_FORECAST and not self._btn_mode_forecast.isEnabled():
            return

        self._current_section = section
        if section == self.SECTION_ANNUITY_RIDER:
            self._content_stack.setCurrentWidget(self._annuity_rider_tab)
            if reload_policy and self._policy and self._has_annuity_rider(self._policy):
                self._annuity_rider_tab.load_data_from_policy(self._policy)
        elif section == self.SECTION_GLP_EXCEPTION:
            self._content_stack.setCurrentWidget(self._glp_exception_page)
            self._refresh_glp_exception_status()
        elif section == self.SECTION_FORECAST:
            self._content_stack.setCurrentWidget(self._forecast_page)
            self._refresh_forecast_status()
        else:
            self._content_stack.setCurrentWidget(self._workspace_page)
            self._set_mode(section, reload_policy=reload_policy)

        self._refresh_section_buttons()

    def show_annuity_rider(self):
        self._select_section(self.SECTION_ANNUITY_RIDER)

    def _refresh_section_buttons(self):
        self._btn_mode_polsup.setStyleSheet(
            _MODE_BTN_ACTIVE_STYLE
            if self._current_section == self.MODE_POLICY_SUPPORT
            else _MODE_BTN_INACTIVE_STYLE
        )
        self._btn_mode_abr.setStyleSheet(
            _MODE_BTN_ABR_ACTIVE_STYLE
            if self._current_section == self.MODE_ABR
            else _MODE_BTN_INACTIVE_STYLE
        )
        self._btn_mode_glp_exception.setStyleSheet(
            _MODE_BTN_ACTIVE_STYLE
            if self._current_section == self.SECTION_GLP_EXCEPTION
            else _MODE_BTN_INACTIVE_STYLE
        )
        self._btn_mode_forecast.setStyleSheet(
            _MODE_BTN_ACTIVE_STYLE
            if self._current_section == self.SECTION_FORECAST
            else _MODE_BTN_INACTIVE_STYLE
        )
        self._btn_mode_annuity.setStyleSheet(
            _MODE_BTN_ACTIVE_STYLE
            if self._current_section == self.SECTION_ANNUITY_RIDER
            else _MODE_BTN_INACTIVE_STYLE
        )

    def _refresh_glp_exception_status(self):
        if not hasattr(self, "_glp_forecast_status_label"):
            return
        self._clear_glp_exception_results()
        if not self._policy or not self._policy.exists:
            self._set_glp_status("Load an eligible policy to check forecast data")
            self._glp_calculate_btn.setEnabled(False)
            return
        if not is_glp_exception_eligible(self._policy):
            self._set_glp_status(
                "GLP Exception is available only for UL policies using Guideline Premium."
            )
            self._glp_calculate_btn.setEnabled(False)
            return
        availability = check_forecast_availability(self._policy)
        self._set_glp_status(availability.message, is_error=not availability.available)
        self._glp_calculate_btn.setEnabled(availability.available)
        cached_quote = self._glp_quote_cache.get(self._glp_policy_cache_key(self._policy))
        if cached_quote:
            self._glp_target_date.setText(cached_quote["target_text"])
            self._set_glp_status(cached_quote["status_text"])
            self._display_glp_exception_result(cached_quote["result"])
        else:
            self._glp_target_date.setText(self._default_glp_target_date_text(self._policy))

    def _refresh_forecast_status(self):
        if not hasattr(self, "_forecast_status_label"):
            return
        self._clear_forecast_results()
        if not self._policy or not self._policy.exists:
            self._set_forecast_status("Load an eligible policy to check forecast data")
            self._forecast_calculate_btn.setEnabled(False)
            return
        if not is_glp_exception_eligible(self._policy):
            self._set_forecast_status(
                "Forecast is available only for UL policies using Guideline Premium."
            )
            self._forecast_calculate_btn.setEnabled(False)
            return
        availability = check_forecast_availability(self._policy)
        self._set_forecast_status(availability.message, is_error=not availability.available)
        self._forecast_calculate_btn.setEnabled(availability.available)
        cached_forecast = self._forecast_cache.get(self._glp_policy_cache_key(self._policy))
        if cached_forecast:
            self._forecast_target_date.setText(cached_forecast["target_text"])
            self._forecast_premium_amount.setText(cached_forecast["premium_text"])
            self._forecast_premium_mode.setCurrentText(cached_forecast["premium_mode"])
            self._set_forecast_status(cached_forecast["status_text"])
            self._display_forecast_rows(cached_forecast["result"])
        else:
            self._forecast_target_date.setText(self._default_glp_target_date_text(self._policy))
            self._forecast_premium_amount.setText("0.00")
            self._forecast_premium_mode.setCurrentText("Monthly")

    def _on_calculate_glp_exception(self):
        if not self._policy:
            return
        target_text = self._glp_target_date.text().strip()
        if not target_text:
            self._set_glp_status("Enter a Target Inforce Date before calculating", is_error=True)
            self._clear_glp_exception_results()
            return
        try:
            target_date = datetime.strptime(target_text, "%m/%d/%Y").date()
        except ValueError:
            self._set_glp_status("Target Inforce Date must be entered as MM/DD/YYYY", is_error=True)
            self._clear_glp_exception_results()
            return
        try:
            result = calculate_glp_exception(self._policy, target_date)
        except Exception as exc:
            self._set_glp_status(str(exc), is_error=True)
            self._clear_glp_exception_results()
            return
        status_text = "Data for forecasting is available"
        self._set_glp_status(status_text)
        self._display_glp_exception_result(result)
        self._glp_quote_cache[self._glp_policy_cache_key(self._policy)] = {
            "target_text": target_text,
            "status_text": status_text,
            "result": result,
        }

    def _on_calculate_forecast(self):
        if not self._policy:
            return
        target_text = self._forecast_target_date.text().strip()
        if not target_text:
            self._set_forecast_status("Enter a Target Date before calculating", is_error=True)
            self._clear_forecast_results()
            return
        try:
            target_date = datetime.strptime(target_text, "%m/%d/%Y").date()
        except ValueError:
            self._set_forecast_status("Target Date must be entered as MM/DD/YYYY", is_error=True)
            self._clear_forecast_results()
            return

        premium_text = self._forecast_premium_amount.text().strip().replace(",", "").replace("$", "")
        try:
            premium_amount = float(premium_text or 0.0)
        except ValueError:
            self._set_forecast_status("Premium Amount must be numeric", is_error=True)
            self._clear_forecast_results()
            return

        premium_mode = self._forecast_premium_mode.currentText()
        try:
            result = calculate_policy_support_forecast(
                self._policy,
                target_date,
                premium_amount,
                premium_mode,
            )
        except Exception as exc:
            self._set_forecast_status(str(exc), is_error=True)
            self._clear_forecast_results()
            return

        status_text = "Data for forecasting is available"
        self._set_forecast_status(status_text)
        self._display_forecast_rows(result)
        self._forecast_cache[self._glp_policy_cache_key(self._policy)] = {
            "target_text": target_text,
            "premium_text": f"{premium_amount:.2f}",
            "premium_mode": premium_mode,
            "status_text": status_text,
            "result": result,
        }

    def _set_glp_status(self, text: str, *, is_error: bool = False):
        if is_error:
            self._glp_forecast_status_label.setStyleSheet(
                "font-size: 10px; color: #C00000; font-weight: bold; "
                "background: transparent; border: none; padding: 2px 4px;"
            )
        else:
            self._glp_forecast_status_label.setStyleSheet(_STATUS_STYLE)
        self._glp_forecast_status_label.setText(text)

    def _set_forecast_status(self, text: str, *, is_error: bool = False):
        if is_error:
            self._forecast_status_label.setStyleSheet(
                "font-size: 10px; color: #C00000; font-weight: bold; "
                "background: transparent; border: none; padding: 2px 4px;"
            )
        else:
            self._forecast_status_label.setStyleSheet(_STATUS_STYLE)
        self._forecast_status_label.setText(text)

    @staticmethod
    def _default_glp_target_date_text(policy: Optional['PolicyInformation']) -> str:
        allowed_dates = PolicySupportTab._glp_allowed_target_dates(policy)
        return allowed_dates[0].strftime("%m/%d/%Y") if allowed_dates else ""

    @staticmethod
    def _glp_allowed_target_dates(policy: Optional['PolicyInformation']) -> List[date]:
        if not policy:
            return []
        issue_date = getattr(policy, "issue_date", None)
        today = date.today()
        if not issue_date:
            return []
        anniversary = _safe_anniversary(issue_date, today.year)
        if anniversary <= today:
            anniversary = _safe_anniversary(issue_date, today.year + 1)
        return [
            anniversary,
            _safe_anniversary(issue_date, anniversary.year + 1),
        ]

    @staticmethod
    def _glp_policy_cache_key(policy: Optional['PolicyInformation']) -> tuple[str, str, str, str]:
        if not policy:
            return "", "", "", ""
        return (
            str(getattr(policy, "region", "") or "").strip().upper(),
            str(getattr(policy, "system_code", "") or "").strip().upper(),
            str(getattr(policy, "company_code", "") or "").strip().upper(),
            str(getattr(policy, "policy_number", "") or "").strip().upper(),
        )

    def _clear_glp_exception_results(self):
        if not hasattr(self, "_glp_result_labels"):
            return
        for key, (value_label, _kind) in self._glp_result_labels.items():
            value_label.setText("-")
            value_label.setVisible(True)
            self._glp_result_name_labels[key].setVisible(True)
        self._glp_forecast_table.setRowCount(0)
        self._glp_forecast_frame.setVisible(False)
        self._glp_export_btn.setEnabled(False)
        self._glp_print_details_btn.setEnabled(False)

    def _clear_forecast_results(self):
        if hasattr(self, "_forecast_table"):
            self._forecast_table.setRowCount(0)

    def _display_glp_exception_result(self, result: GlpExceptionResult):
        for key, (value_label, kind) in self._glp_result_labels.items():
            name_label = self._glp_result_name_labels[key]
            visible = self._glp_result_row_visible(result, key)
            name_label.setVisible(visible)
            value_label.setVisible(visible)
            if kind in {"separator", "thin_separator"}:
                continue
            if kind == "message":
                value_label.setText("")
                continue
            value = getattr(result, key)
            if value is None:
                value_label.setText("-")
            elif kind == "money":
                value_label.setText(f"${float(value):,.2f}")
            elif kind == "percent":
                value_label.setText(f"{float(value) * 100:.2f}%")
            elif kind == "date":
                value_label.setText(value.strftime("%m/%d/%Y") if value else "-")
            else:
                value_label.setText(f"{value:,}")
        self._display_glp_forecast_rows(result)
        self._glp_export_btn.setEnabled(True)
        self._glp_print_details_btn.setEnabled(True)

    def _glp_result_row_visible(self, result: GlpExceptionResult, key: str) -> bool:
        no_adjustment_needed = result.accumulated_glp_prior_to_target >= result.premium_td_on_target_date - 0.005
        if key == "glp_adjustment_message":
            return no_adjustment_needed
        if key in {"force_out_required", "force_out_amount"}:
            return result.force_out_required
        if key in {"new_glp", "adjustment_to_accum_glp", "new_accum_glp"}:
            if no_adjustment_needed or result.force_out_required:
                return False
            return key != "new_glp" or result.new_glp is not None
        return True

    def _has_glp_quote_to_export(self) -> bool:
        return any(value_label.text() != "-" for value_label, _kind in self._glp_result_labels.values())

    def _build_glp_quote_workbook(self):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GLP Quote"

        title_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_font = Font(bold=True, color="006100")
        bold_font = Font(bold=True)

        policy_text = ""
        if self._policy:
            policy_text = " - ".join(
                part for part in (
                    str(getattr(self._policy, "region", "") or "").strip(),
                    str(getattr(self._policy, "company_code", "") or "").strip(),
                    str(getattr(self._policy, "policy_number", "") or "").strip(),
                ) if part
            )

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        title_cell = ws.cell(row=1, column=1, value=f"GLP Exception Quote{f' - {policy_text}' if policy_text else ''}")
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=3, column=1, value="Target Inforce Date").font = bold_font
        ws.cell(row=3, column=2, value=self._glp_target_date.text().strip())
        ws.cell(row=4, column=1, value="Forecast Status").font = bold_font
        ws.cell(row=4, column=2, value=self._glp_forecast_status_label.text())

        row_num = 6
        ws.cell(row=row_num, column=1, value="Calculation").fill = header_fill
        ws.cell(row=row_num, column=1).font = header_font
        ws.cell(row=row_num, column=2).fill = header_fill
        row_num += 1
        for label, key, _kind in self._glp_result_rows:
            if not self._glp_result_name_labels[key].isVisible():
                continue
            value_label, _ = self._glp_result_labels[key]
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value_label.text())
            if key == "new_accum_glp":
                ws.cell(row=row_num, column=1).font = bold_font
                ws.cell(row=row_num, column=2).font = bold_font
            row_num += 1

        row_num += 2
        ws.cell(row=row_num, column=1, value="Monthly Forecast").fill = header_fill
        ws.cell(row=row_num, column=1).font = header_font
        row_num += 1
        for col_index in range(self._glp_forecast_table.columnCount()):
            header_item = self._glp_forecast_table._data_table.horizontalHeaderItem(col_index)
            cell = ws.cell(row=row_num, column=col_index + 1, value=header_item.text() if header_item else "")
            cell.fill = header_fill
            cell.font = header_font
        row_num += 1
        for table_row in range(self._glp_forecast_table.rowCount()):
            for col_index in range(self._glp_forecast_table.columnCount()):
                item = self._glp_forecast_table.item(table_row, col_index)
                ws.cell(row=row_num, column=col_index + 1, value=item.text() if item else "")
            row_num += 1

        for col_index in range(1, ws.max_column + 1):
            max_length = max(
                len(str(ws.cell(row=row_index, column=col_index).value or ""))
                for row_index in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 55)

        return wb

    def _export_glp_quote(self):
        if not any(value_label.text() != "-" for value_label, _kind in self._glp_result_labels.values()):
            QMessageBox.information(self, "Export GLP Quote", "Calculate a GLP quote before exporting.")
            return
        try:
            import tempfile
            wb = self._build_glp_quote_workbook()
            temp_file = tempfile.NamedTemporaryFile(mode="w", suffix=".xlsx", prefix="SuiteView_GLP_Quote_", delete=False)
            temp_path = temp_file.name
            temp_file.close()
            wb.save(temp_path)
            os.startfile(temp_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", f"Failed to export GLP quote:\n{exc}")

    def _print_glp_quote_to_folder(self):
        if not self._has_glp_quote_to_export():
            QMessageBox.information(self, "Print Details to Folder", "Calculate a GLP quote before printing details.")
            return

        folder_path = self._glp_exception_folder_path()
        if not folder_path or not os.path.isdir(folder_path) or not os.access(folder_path, os.W_OK):
            self._show_glp_folder_inaccessible(folder_path)
            return

        file_path = os.path.join(folder_path, self._glp_exception_file_name())
        confirm = QMessageBox.question(
            self,
            "Print Details to Folder",
            f"The GLP detail file will be saved to:\n\n{file_path}\n\nContinue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        try:
            wb = self._build_glp_quote_workbook()
            wb.save(file_path)
        except Exception as exc:
            self._show_glp_folder_inaccessible(folder_path, str(exc))
            return

        QMessageBox.information(
            self,
            "Print Details to Folder",
            f"GLP detail file saved to:\n\n{file_path}",
        )

    def _glp_exception_folder_path(self) -> str:
        if not self._policy:
            return ""
        product_type = str(getattr(self._policy, "product_type", "") or "").strip()
        company_code = str(getattr(self._policy, "company_code", "") or "").strip()
        policy_number = str(getattr(self._policy, "policy_number", "") or "").strip()
        if not product_type or not company_code or not policy_number:
            return ""
        return os.path.join(
            _get_policy_library_dir(),
            product_type,
            f"{company_code}_{policy_number}",
            "GLP_Exception",
        )

    def _glp_exception_file_name(self) -> str:
        company_code = str(getattr(self._policy, "company_code", "") or "").strip()
        policy_number = str(getattr(self._policy, "policy_number", "") or "").strip()
        timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
        return f"GLP_Exception_{company_code}_{policy_number}_{timestamp}.xlsx"

    def _show_glp_folder_inaccessible(self, folder_path: str, detail: Optional[str] = None):
        message = (
            "The GLP Exception folder is not accessible:\n\n"
            f"{folder_path or '(policy folder could not be determined)'}\n\n"
            "You can click the Export button to get the file."
        )
        if detail:
            message += f"\n\nDetails: {detail}"
        QMessageBox.warning(self, "Print Details to Folder", message)

    def _show_glp_helper(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("GLP Exception Helper")
        dialog.resize(620, 520)
        layout = QVBoxLayout(dialog)
        text = QTextEdit()
        text.setReadOnly(True)
        text.setPlainText(GLP_HELP_TEXT)
        text.setStyleSheet(
            f"font-size: 11px; color: {GRAY_DARK}; background: {WHITE}; "
            f"border: 1px solid {GRAY_MID};"
        )
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(_ACTION_BTN_STYLE)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(text, 1)
        layout.addWidget(close_btn, 0, Qt.AlignmentFlag.AlignRight)
        dialog.exec()

    def _display_glp_forecast_rows(self, result: GlpExceptionResult):
        rows = result.forecast_rows
        self._glp_forecast_frame.setVisible(bool(rows))
        self._glp_forecast_table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            values = [
                row.forecast_date.strftime("%m/%d/%Y") if row.forecast_date else "-",
                f"{row.policy_year:,}",
                f"{row.policy_month:,}",
                f"${row.interest_credited:,.2f}",
                f"${row.premium:,.2f}",
                f"${row.monthly_deduction:,.2f}",
                f"${row.account_value:,.2f}",
            ]
            for col_index, text in enumerate(values):
                alignment = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                if col_index == 0:
                    alignment = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                self._glp_forecast_table.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem(text),
                    alignment,
                )
        self._glp_forecast_table.autoFitAllColumns()

    def _display_forecast_rows(self, result):
        rows = result.rows
        self._forecast_table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            values = [
                row.forecast_date.strftime("%m/%d/%Y") if row.forecast_date else "-",
                f"{row.policy_year:,}",
                f"{row.policy_month:,}",
                f"${row.interest_credited:,.2f}",
                f"${row.glp:,.2f}",
                f"${row.accumulated_glp:,.2f}",
                f"${row.premiums_paid_to_date:,.2f}",
                f"${row.accumulated_withdrawals:,.2f}",
                f"${row.force_out:,.2f}",
                f"${row.premium:,.2f}",
                f"${row.account_value_before_monthly_deduction:,.2f}",
                f"${row.monthly_deduction:,.2f}",
                f"${row.account_value:,.2f}",
            ]
            for col_index, text in enumerate(values):
                alignment = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                if col_index == 0:
                    alignment = Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                self._forecast_table.setItem(
                    row_index,
                    col_index,
                    QTableWidgetItem(text),
                    alignment,
                )
        self._forecast_table.autoFitAllColumns()

    @staticmethod
    def _has_annuity_rider(policy: Optional['PolicyInformation']) -> bool:
        if not policy or not policy.exists:
            return False
        try:
            return any(
                str(getattr(coverage, "plancode", "")).strip().upper() == RIDER_PLANCODE
                for coverage in policy.get_coverages()
            )
        except Exception:
            return False

    def _apply_green_theme(self):
        """Restore the default green/gold Policy Support theme to all 4 group panels."""
        # 1. Info frame
        self._info_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        self._library_path_label.setStyleSheet(_PATH_LABEL_STYLE)
        self._create_folder_btn.setStyleSheet(_ACTION_BTN_STYLE)

        # 2. Task Categories
        self._cat_frame.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        self._category_list.setStyleSheet(_LIST_STYLE)

        # 3. Policy Subfolders
        self._apply_explorer_theme(
            self._subfolder_explorer,
            POLICY_INFO_FRAME_STYLE, _LIST_STYLE, _NAV_BTN_STYLE, _PATH_BAR_STYLE,
        )

        # 4. Available Tools
        self._apply_explorer_theme(
            self._tools_explorer,
            POLICY_INFO_FRAME_STYLE, _LIST_STYLE, _NAV_BTN_STYLE, _PATH_BAR_STYLE,
        )

    def _apply_abr_theme(self):
        """Apply the crimson/slate ABR theme to all 4 group panels."""
        # 1. Info frame
        self._info_frame.setStyleSheet(_ABR_FRAME_STYLE)
        self._library_path_label.setStyleSheet(_ABR_PATH_LABEL_STYLE)
        self._create_folder_btn.setStyleSheet(_ABR_ACTION_BTN_STYLE)

        # 2. Task Categories
        self._cat_frame.setStyleSheet(_ABR_FRAME_STYLE)
        self._category_list.setStyleSheet(_ABR_LIST_STYLE)

        # 3. Policy Subfolders
        self._apply_explorer_theme(
            self._subfolder_explorer,
            _ABR_FRAME_STYLE, _ABR_LIST_STYLE, _ABR_NAV_BTN_STYLE, _ABR_PATH_BAR_STYLE,
        )

        # 4. Available Tools
        self._apply_explorer_theme(
            self._tools_explorer,
            _ABR_FRAME_STYLE, _ABR_LIST_STYLE, _ABR_NAV_BTN_STYLE, _ABR_PATH_BAR_STYLE,
        )

    @staticmethod
    def _apply_explorer_theme(explorer, frame_style, list_style, nav_style, path_style):
        """Apply a theme to a MiniExplorer or _SubfolderExplorer widget."""
        # Group box
        gb = explorer.findChild(QGroupBox)
        if gb:
            gb.setStyleSheet(frame_style)
        # Nav buttons
        if hasattr(explorer, '_home_btn'):
            explorer._home_btn.setStyleSheet(nav_style)
        if hasattr(explorer, '_up_btn'):
            explorer._up_btn.setStyleSheet(nav_style)
        # Path label
        if hasattr(explorer, '_path_label'):
            explorer._path_label.setStyleSheet(path_style)
        # List widget
        if hasattr(explorer, '_list'):
            explorer._list.setStyleSheet(list_style)

    # -- Category helpers --------------------------------------------------

    def _all_categories(self) -> List[str]:
        return sorted(set(DEFAULT_TASK_CATEGORIES) | set(self._user_tasks))

    def _populate_category_list(self):
        self._category_list.clear()
        for cat in self._all_categories():
            item = QListWidgetItem(cat)
            if cat in self._user_tasks:
                item.setForeground(QColor(GOLD_DARK))
            self._category_list.addItem(item)

    # -- Category context menu ---------------------------------------------

    def _on_category_context_menu(self, pos: QPoint):
        menu = QMenu(self)
        menu.setStyleSheet(_CONTEXT_MENU_STYLE)
        add_action = menu.addAction("Add New Task Category...")

        item = self._category_list.itemAt(pos)
        sel_name = item.text() if item else ""
        is_user = sel_name in self._user_tasks

        rename_action = delete_action = None
        if sel_name:
            menu.addSeparator()
            rename_action = menu.addAction(f"Rename '{sel_name}'...")
            delete_action = menu.addAction(f"Delete '{sel_name}'")
            if not is_user:
                rename_action.setEnabled(False)
                delete_action.setEnabled(False)
                rename_action.setToolTip("Cannot modify default categories")
                delete_action.setToolTip("Cannot delete default categories")

        action = menu.exec(self._category_list.viewport().mapToGlobal(pos))
        if action is add_action:
            self._add_task_category()
        elif action is rename_action and sel_name:
            self._rename_task_category(sel_name)
        elif action is delete_action and sel_name:
            self._delete_task_category(sel_name)

    def _add_task_category(self):
        name, ok = QInputDialog.getText(
            self, "Add Task Category",
            "Enter new task category name\n(use underscores for spaces):",
        )
        if not ok or not name.strip():
            return
        name = name.strip().replace(" ", "_")
        if name in self._all_categories():
            QMessageBox.information(self, "Duplicate", f"'{name}' already exists.")
            return
        self._user_tasks.append(name)
        _save_user_tasks(self._user_tasks)
        self._populate_category_list()
        self._status_label.setText(f"✓ Added task category: {name}")

    def _rename_task_category(self, old_name: str):
        new_name, ok = QInputDialog.getText(
            self, "Rename Task Category", f"Rename '{old_name}' to:", text=old_name
        )
        if not ok or not new_name.strip() or new_name.strip() == old_name:
            return
        new_name = new_name.strip().replace(" ", "_")
        if new_name in self._all_categories():
            QMessageBox.information(self, "Duplicate", f"'{new_name}' already exists.")
            return
        self._user_tasks = [new_name if t == old_name else t for t in self._user_tasks]
        _save_user_tasks(self._user_tasks)
        self._populate_category_list()
        self._status_label.setText(f"✓ Renamed '{old_name}' → '{new_name}'")

    def _delete_task_category(self, name: str):
        reply = QMessageBox.question(
            self, "Delete Task Category",
            f"Delete '{name}' from the task list?\n\n(This does not delete any folders on disk.)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self._user_tasks = [t for t in self._user_tasks if t != name]
        _save_user_tasks(self._user_tasks)
        self._populate_category_list()
        self._status_label.setText(f"✓ Removed task category: {name}")

    # -- Drop handlers -----------------------------------------------------

    def _on_category_dropped(self, category_name: str):
        """A category was dragged onto the Policy Subfolders panel."""
        dest_dir = self._subfolder_explorer.current_path()
        if not dest_dir:
            dest_dir = self._policy_support_folder_path

        if not dest_dir:
            QMessageBox.information(self, "No Policy Folder",
                                    "Load a policy first.")
            return

        task_folder = os.path.join(dest_dir, category_name)
        if os.path.isdir(task_folder):
            self._status_label.setText(f"ℹ️  '{category_name}' folder already exists")
            return

        try:
            # Auto-create policy folder if needed
            if not os.path.isdir(self._policy_support_folder_path):
                os.makedirs(self._policy_support_folder_path, exist_ok=True)
                self._subfolder_explorer.set_root(self._policy_support_folder_path)
                self._create_folder_btn.setVisible(False)
                if self._policy:
                    cc, pn = self._policy.company_code, self._policy.policy_number
                    self._policy_folder_label.setText(f"{cc}_{pn}")
                    self._policy_folder_label.setStyleSheet(
                        f"font-size: 11px; color: {GREEN_DARK}; font-weight: bold; "
                        f"background: transparent; border: none;"
                    )

            os.makedirs(task_folder, exist_ok=True)
            self._status_label.setText(f"✓ Created task folder: {category_name}")
            self._status_label.setStyleSheet(
                f"font-size: 10px; color: {GREEN_DARK}; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._subfolder_explorer.refresh()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to create folder:\n{e}")

    def _on_file_dropped(self, source_path: str):
        """A tool file was dragged onto the Policy Subfolders panel."""
        if not source_path or not os.path.isfile(source_path):
            return

        dest_dir = self._subfolder_explorer.current_path()
        if not dest_dir or not os.path.isdir(dest_dir):
            QMessageBox.information(
                self, "No Destination",
                "Please navigate into a subfolder in Policy Subfolders first."
            )
            return

        filename = workbook_filename(os.path.basename(source_path))
        if self._policy:
            dest_filename = f"{self._policy.policy_number} - {filename}"
        else:
            dest_filename = filename

        dest_path = os.path.join(dest_dir, dest_filename)
        if os.path.exists(dest_path):
            QMessageBox.information(
                self, "File Exists",
                f"'{dest_filename}' already exists here.\nOperation cancelled."
            )
            return

        try:
            copy_as_workbook(source_path, dest_path)
            self._status_label.setText(f"✓ Copied: {dest_filename}")
            self._status_label.setStyleSheet(
                f"font-size: 10px; color: {GREEN_DARK}; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._subfolder_explorer.refresh()
        except Exception as e:
            QMessageBox.warning(self, "Copy Error", f"Failed to copy file:\n{e}")

    # -- Policy folder actions ---------------------------------------------

    def _on_create_policy_folder(self):
        if not self._policy_support_folder_path:
            return
        try:
            os.makedirs(self._policy_support_folder_path, exist_ok=True)
            self._status_label.setText("✓ Policy folder created")
            self._status_label.setStyleSheet(
                f"font-size: 10px; color: {GREEN_DARK}; font-weight: bold; "
                f"background: transparent; border: none;"
            )
            self._create_folder_btn.setVisible(False)
            if self._policy:
                cc, pn = self._policy.company_code, self._policy.policy_number
                self._policy_folder_label.setText(f"{cc}_{pn}")
                self._policy_folder_label.setStyleSheet(
                    f"font-size: 11px; color: {GREEN_DARK}; font-weight: bold; "
                    f"background: transparent; border: none;"
                )
            self._subfolder_explorer.set_root(self._policy_support_folder_path)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to create policy folder:\n{e}")



# ---------------------------------------------------------------------------
# _SubfolderExplorer  — Policy Subfolders panel (drop target + mini explorer)
# ---------------------------------------------------------------------------

class _SubfolderExplorer(QWidget):
    """Mini explorer for the policy folder that also accepts drag-and-drop."""

    category_dropped = pyqtSignal(str)
    file_dropped     = pyqtSignal(str)

    _PATH_ROLE = Qt.ItemDataRole.UserRole

    def __init__(self, title: str = "", parent=None):
        super().__init__(parent)
        self._root_path = ""
        self._current_path = ""
        self._title = title
        self._setup_ui()

    def _setup_ui(self):
        outer = QGroupBox(self._title)
        outer.setStyleSheet(POLICY_INFO_FRAME_STYLE)
        ol = QVBoxLayout(outer)
        ol.setContentsMargins(2, 14, 2, 2)
        ol.setSpacing(2)

        # Nav bar
        nav = QHBoxLayout()
        nav.setContentsMargins(0, 0, 0, 0)
        nav.setSpacing(2)

        self._home_btn = QPushButton("⌂")
        self._home_btn.setStyleSheet(_NAV_BTN_STYLE)
        self._home_btn.setFixedWidth(22)
        self._home_btn.setToolTip("Go to policy root folder")
        self._home_btn.clicked.connect(self._go_home)

        self._up_btn = QPushButton("↑")
        self._up_btn.setStyleSheet(_NAV_BTN_STYLE)
        self._up_btn.setFixedWidth(22)
        self._up_btn.setToolTip("Go up one level")
        self._up_btn.clicked.connect(self._go_up)

        self._path_label = _DoubleClickablePathLabel("")
        self._path_label.setStyleSheet(_PATH_BAR_STYLE)
        self._path_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )

        nav.addWidget(self._home_btn)
        nav.addWidget(self._up_btn)
        nav.addWidget(self._path_label, 1)
        ol.addLayout(nav)

        self._list = _DropTargetSubfolderList()
        self._list.setStyleSheet(_LIST_STYLE)
        self._list.setItemDelegate(_TightItemDelegate(self._list))
        self._list.setUniformItemSizes(True)
        self._list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._list.itemDoubleClicked.connect(self._on_double_click)
        self._list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._list.customContextMenuRequested.connect(self._on_context_menu)
        # Wire drop signals
        self._list.category_dropped.connect(self.category_dropped)
        self._list.file_dropped.connect(self.file_dropped)
        ol.addWidget(self._list, 1)

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.addWidget(outer)

    def set_root(self, path: str):
        self._root_path = path
        self._current_path = path
        self._refresh()

    def _go_home(self):
        if self._root_path:
            self._current_path = self._root_path
            self._refresh()

    def _go_up(self):
        if not self._root_path or not self._current_path:
            return
        parent = os.path.dirname(self._current_path)
        root_norm = os.path.normpath(self._root_path)
        if os.path.normpath(parent).startswith(root_norm):
            self._current_path = parent
            self._refresh()

    def _on_double_click(self, item: QListWidgetItem):
        path = item.data(self._PATH_ROLE) or ""
        if os.path.isdir(path):
            self._current_path = path
            self._refresh()
        elif os.path.isfile(path):
            os.startfile(path)

    def _on_context_menu(self, pos: QPoint):
        menu = QMenu(self)
        menu.setStyleSheet(_CONTEXT_MENU_STYLE)

        can_up = (bool(self._root_path) and bool(self._current_path) and
                  os.path.normpath(self._current_path) != os.path.normpath(self._root_path))
        up_act = menu.addAction("↑  Go Up")
        up_act.setEnabled(can_up)
        home_act = menu.addAction("⌂  Go to Policy Root")
        home_act.setEnabled(can_up)

        item = self._list.itemAt(pos)
        open_act = rename_act = delete_act = None
        entry_name = ""
        if item:
            path = item.data(self._PATH_ROLE) or ""
            entry_name = os.path.basename(path)
            if entry_name:
                menu.addSeparator()
                open_act = menu.addAction(f"Open '{entry_name}'")
                rename_act = menu.addAction(f"Rename '{entry_name}'...")
                delete_act = menu.addAction(f"Delete '{entry_name}'")

        action = menu.exec(self._list.viewport().mapToGlobal(pos))
        if action is up_act:
            self._go_up()
        elif action is home_act:
            self._go_home()
        elif action is open_act and entry_name:
            path = item.data(self._PATH_ROLE) or ""
            if os.path.exists(path):
                os.startfile(path)
        elif action is rename_act and entry_name:
            self._rename_entry(item)
        elif action is delete_act and entry_name:
            self._delete_entry(item)

    def _rename_entry(self, item: QListWidgetItem):
        path = item.data(self._PATH_ROLE) or ""
        old_name = os.path.basename(path)
        new_name, ok = QInputDialog.getText(
            self, "Rename", f"Rename '{old_name}' to:", text=old_name
        )
        if not ok or not new_name.strip() or new_name.strip() == old_name:
            return
        new_path = os.path.join(os.path.dirname(path), new_name.strip())
        if os.path.exists(new_path):
            QMessageBox.information(self, "Exists", f"'{new_name}' already exists.")
            return
        try:
            os.rename(path, new_path)
            self._refresh()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to rename:\n{e}")

    def _delete_entry(self, item: QListWidgetItem):
        path = item.data(self._PATH_ROLE) or ""
        name = os.path.basename(path)
        is_dir = os.path.isdir(path)
        kind = "folder" if is_dir else "file"
        reply = QMessageBox.question(
            self, f"Delete {kind.title()}",
            f"Permanently delete {kind} '{name}'?\n\nThis cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            if is_dir:
                shutil.rmtree(path)
            else:
                os.remove(path)
            self._refresh()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to delete:\n{e}")

    def _refresh(self):
        self._list.clear()
        can_up = (bool(self._root_path) and bool(self._current_path) and
                  os.path.normpath(self._current_path) != os.path.normpath(self._root_path or ""))
        self._up_btn.setEnabled(can_up)
        self._home_btn.setEnabled(can_up)

        if not self._current_path:
            self._path_label.setText("(no policy folder)")
            self._path_label.set_open_path("")
            return

        # Show path relative to root; store full path for double-click open
        try:
            rel = os.path.relpath(self._current_path, os.path.dirname(self._root_path))
        except ValueError:
            rel = self._current_path
        self._path_label.setText(rel)
        self._path_label.set_open_path(self._current_path)

        if not os.path.isdir(self._current_path):
            return

        try:
            entries = sorted(
                os.listdir(self._current_path),
                key=lambda e: (
                    not os.path.isdir(os.path.join(self._current_path, e)),
                    e.lower()
                )
            )
            for entry in entries:
                full = os.path.join(self._current_path, entry)
                if os.path.isdir(full):
                    item = QListWidgetItem(f"📁  {entry}")
                else:
                    ext = os.path.splitext(entry)[1].lower()
                    item = QListWidgetItem(f"{_icon_for_ext(ext)}  {entry}")
                item.setData(self._PATH_ROLE, full)
                self._list.addItem(item)
        except Exception as e:
            self._list.addItem(QListWidgetItem(f"(Error: {e})"))

    def current_path(self) -> str:
        return self._current_path

    def refresh(self):
        self._refresh()
