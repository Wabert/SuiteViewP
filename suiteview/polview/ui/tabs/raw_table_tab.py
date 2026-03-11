"""
Raw Table tab – transposed/normal data view with Excel export.
"""

from datetime import datetime

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidgetItem, QMessageBox,
)

from suiteview.core.db2_connection import DB2Connection
from ..styles import BLUE_LIGHT, BLUE_DARK, BLUE_PRIMARY, GOLD_LIGHT, GOLD_PRIMARY
from ..widgets import TableDataWidget


class RawTableTab(QWidget):
    """Tab for viewing raw table data with transpose and export functionality."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_transposed = True  # Default to transposed view (fields as rows)
        self._current_cols = []
        self._current_rows = []
        self._current_table_name = ""
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # Header row with label and buttons
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(4, 4, 4, 0)

        self.table_label = QLabel("Select a table from the left panel")
        self.table_label.setStyleSheet("font-weight: bold;")
        header_layout.addWidget(self.table_label)

        header_layout.addStretch()

        # Transpose button
        self.transpose_btn = QPushButton("⇄ Transpose")
        self.transpose_btn.setToolTip("Toggle between transposed and normal view")
        self.transpose_btn.setFixedSize(90, 24)
        self.transpose_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {BLUE_LIGHT};
                color: {BLUE_DARK};
                border: 1px solid {BLUE_PRIMARY};
                border-radius: 3px;
                font-size: 11px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {GOLD_LIGHT};
                border-color: {GOLD_PRIMARY};
            }}
            QPushButton:pressed {{
                background-color: {GOLD_PRIMARY};
            }}
        """)
        self.transpose_btn.clicked.connect(self._toggle_transpose)
        header_layout.addWidget(self.transpose_btn)

        # Export button (green with spreadsheet icon)
        self.export_btn = QPushButton("📊")
        self.export_btn.setToolTip("Export to Excel")
        self.export_btn.setFixedSize(28, 24)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: 1px solid #1e7e34;
                border-radius: 3px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #34ce57;
                border-color: #28a745;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)
        self.export_btn.clicked.connect(self._export_to_excel)
        header_layout.addWidget(self.export_btn)

        layout.addLayout(header_layout)

        self.data_table = TableDataWidget()
        layout.addWidget(self.data_table)

    # ── clear / reset ────────────────────────────────────────────────────

    def clear(self):
        """Reset the tab to its initial empty state."""
        self._current_cols = []
        self._current_rows = []
        self._current_table_name = ""
        self.table_label.setText("Select a table from the left panel")
        self.data_table.clear()
        self.data_table.setRowCount(0)
        self.data_table.setColumnCount(0)

    # ── view toggling ────────────────────────────────────────────────────

    def _toggle_transpose(self):
        """Toggle between transposed and normal view."""
        self._is_transposed = not self._is_transposed
        if self._current_cols and self._current_rows:
            self._display_data()

    def _display_data(self):
        """Display data in current mode (transposed or normal)."""
        if not self._current_cols or not self._current_rows:
            return

        if self._is_transposed:
            # Fields as rows, records as columns (current behavior)
            self.data_table.load_data(self._current_cols, self._current_rows)
        else:
            # Normal view: fields as columns, records as rows
            self.data_table.clear()
            self.data_table.setRowCount(len(self._current_rows))
            self.data_table.setColumnCount(len(self._current_cols))
            self.data_table.setHorizontalHeaderLabels(self._current_cols)
            self.data_table.horizontalHeader().setVisible(True)
            self.data_table.verticalHeader().setVisible(False)

            for row_idx, row in enumerate(self._current_rows):
                for col_idx, value in enumerate(row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.data_table.setItem(row_idx, col_idx, item)

            # Resize columns to content
            self.data_table.resizeColumnsToContents()

    # ── Excel export ─────────────────────────────────────────────────────

    def _export_to_excel(self):
        """Export current table data to a new Excel file."""
        if not self._current_cols or not self._current_rows:
            QMessageBox.information(self, "Export", "No data to export")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self, "Export",
                "openpyxl is required for Excel export.\nInstall with: pip install openpyxl",
            )
            return

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._current_table_name[:31] if self._current_table_name else "Data"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write data based on current view mode
        if self._is_transposed:
            ws.cell(row=1, column=1, value="Field").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).border = thin_border

            for col_idx in range(len(self._current_rows)):
                cell = ws.cell(row=1, column=col_idx + 2, value=f"Row {col_idx + 1}")
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, field_name in enumerate(self._current_cols):
                ws.cell(row=row_idx + 2, column=1, value=field_name).border = thin_border
                for col_idx, row_data in enumerate(self._current_rows):
                    value = row_data[row_idx] if row_idx < len(row_data) else ""
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 2, value=value)
                    cell.border = thin_border
        else:
            for col_idx, col_name in enumerate(self._current_cols):
                cell = ws.cell(row=1, column=col_idx + 1, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, row_data in enumerate(self._current_rows):
                for col_idx, value in enumerate(row_data):
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)
                    cell.border = thin_border

        # Auto-fit column widths
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to temp file and open
        import tempfile
        import os
        import subprocess

        temp_dir = tempfile.gettempdir()
        filename = f"{self._current_table_name or 'Export'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)

        wb.save(filepath)

        # Open in Excel
        try:
            os.startfile(filepath)
        except Exception:
            subprocess.Popen(["start", filepath], shell=True)

        self.status_message = f"Exported to {filepath}"

    # ── data loading ─────────────────────────────────────────────────────

    def load_table(self, db: DB2Connection, table_name: str, where_clause: str,
                   policy_id: str = None, company_code: str = None):
        """Load a specific table.

        Args:
            db: Database connection
            table_name: Name of the table to load
            where_clause: Standard WHERE clause (with CK_SYS_CD)
            policy_id: Policy ID for FH_ tables (no CK_SYS_CD)
            company_code: Company code for FH_ tables
        """
        self.table_label.setText(f"Table: {table_name}")
        self._current_table_name = table_name

        try:
            # FH_ tables don't have CK_SYS_CD column
            if table_name.startswith("FH_") and policy_id and company_code:
                fh_where_clause = f"TCH_POL_ID = '{policy_id}' AND CK_CMP_CD = '{company_code}'"
                sql = f"SELECT * FROM DB2TAB.{table_name} WHERE {fh_where_clause}"
            else:
                sql = f"SELECT * FROM DB2TAB.{table_name} WHERE {where_clause}"

            cols, rows = db.execute_query_with_headers(sql)

            if rows:
                self._current_cols = cols
                self._current_rows = rows
                self._display_data()
            else:
                self._current_cols = []
                self._current_rows = []
                self.data_table.clear()
                self.data_table.setRowCount(1)
                self.data_table.setColumnCount(1)
                self.data_table.setItem(0, 0, QTableWidgetItem("No data found"))

        except Exception as e:
            self._current_cols = []
            self._current_rows = []
            self.data_table.clear()
            self.data_table.setRowCount(1)
            self.data_table.setColumnCount(1)
            self.data_table.setItem(0, 0, QTableWidgetItem(f"Error: {e}"))
