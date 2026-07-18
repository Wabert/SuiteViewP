"""Check (and optionally fix) the mandatory comparison toggles in a RERUN workbook.

All RERUN comparison testing must run with:
  - sINPUT_TEFRA_Force        = TRUE   (guideline premium cap forced on)
  - sINPUT_Exact_Days_Boolean = FALSE  (no exact-days interest)

For each requirement this script reports, for one workbook:
  - the LIVE defined-name cell value (cached, openpyxl), and
  - the toggle's row across ALL Saved Cases columns (case number, CaseID,
    value, and whether it complies).

With {"fix": true} it rewrites every non-compliant Saved Cases cell (and live
cell) to the required value via Excel COM and SAVES the workbook — only use on
the "local" workbook, never on the pristine production copy.

Usage (single JSON arg):
    venv\\Scripts\\python.exe tools/check_comparison_inputs.py '<json>'

    {"workbook": "docs/Illustration_UL/RERUN (v20.0) local.xlsm",
     "fix": false}
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

# name -> required value. Single source of truth for the comparison policy
# (rerun_com.assert_comparison_inputs imports this).
REQUIRED = {
    "sINPUT_TEFRA_Force": True,
    "sINPUT_Exact_Days_Boolean": False,
}


def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _complies(value, required: bool) -> bool:
    return _as_bool(value) is required


def scan(workbook: Path) -> dict:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Saved Cases"]

    # Locate each toggle's Saved Cases row + the CaseID row.
    name_rows = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is not None:
            name_rows[str(a).strip()] = r
    case_id_row = name_rows.get("sINPUT_CaseID")

    out = {"requirements": {}}
    all_ok = True
    for name, required in REQUIRED.items():
        req = {"required": required, "live": {"value": None, "ok": False,
                                              "error": None}}
        # Live defined-name cell (cached value from last save).
        try:
            dn = wb.defined_names[name]
            (sheet, coord), = dn.destinations
            req["live"]["ref"] = f"{sheet}!{coord}"
            req["live"]["value"] = wb[sheet][coord.replace("$", "")].value
            req["live"]["ok"] = _complies(req["live"]["value"], required)
        except Exception as exc:  # noqa: BLE001
            req["live"]["error"] = str(exc)

        row = name_rows.get(name)
        req["saved_cases_row"] = row
        cases = []
        if row is None:
            req["error"] = f"{name} row not found in Saved Cases"
        else:
            for c in range(3, ws.max_column + 1):
                header = ws.cell(row=1, column=c).value
                if header is None:
                    continue
                val = ws.cell(row=row, column=c).value
                cases.append({
                    "case": header,
                    "col": openpyxl.utils.get_column_letter(c),
                    "case_id": (ws.cell(row=case_id_row, column=c).value
                                if case_id_row else None),
                    "value": val,
                    "ok": _complies(val, required),
                })
        req["cases"] = cases
        req["all_cases_ok"] = bool(cases) and all(c["ok"] for c in cases)
        req["bad_cases"] = [c["case"] for c in cases if not c["ok"]]
        all_ok = all_ok and req["all_cases_ok"] and req["live"]["ok"]
        out["requirements"][name] = req
    out["compliant"] = all_ok
    wb.close()
    return out


def fix(workbook: Path, scan_result: dict) -> dict:
    """COM-write the required value into every non-compliant cell + live cell."""
    from rerun_com import _open_excel, XL_CALC_MANUAL
    from openpyxl.utils import column_index_from_string

    xl = _open_excel()
    written = []
    try:
        wb = xl.Workbooks.Open(str(workbook), UpdateLinks=0, ReadOnly=False)
        if wb.ReadOnly:
            # Excel silently opens read-only when another instance holds the
            # file (e.g. the user has it open) — Save() would no-op.
            wb.Close(SaveChanges=False)
            raise RuntimeError(
                f"{workbook.name} opened READ-ONLY (already open elsewhere?) — "
                "close it in Excel and rerun the fix.")
        xl.Calculation = XL_CALC_MANUAL
        ws = wb.Worksheets("Saved Cases")
        for name, req in scan_result["requirements"].items():
            required = req["required"]
            row = req["saved_cases_row"]
            for c in req["cases"]:
                if c["ok"] or row is None:
                    continue
                col = column_index_from_string(c["col"])
                before = ws.Cells(row, col).Value
                ws.Cells(row, col).Value = required
                written.append({"name": name, "case": c["case"], "col": c["col"],
                                "before": before, "after": required})
            if not req["live"]["ok"]:
                rng = wb.Names(name).RefersToRange
                written.append({"name": name, "live_cell": req["live"].get("ref"),
                                "before": rng.Value, "after": required})
                rng.Value = required
        wb.Save()
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
    return {"written": written}


def main():
    cmd = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    workbook = Path(cmd.get("workbook")
                    or "docs/Illustration_UL/RERUN (v20.0) local.xlsm").resolve()
    if not workbook.exists():
        print(json.dumps({"ok": False, "error": f"workbook not found: {workbook}"}))
        sys.exit(1)

    out = {"ok": True, "workbook": str(workbook)}
    out["before"] = scan(workbook)

    if cmd.get("fix"):
        if out["before"]["compliant"]:
            out["fix"] = "nothing to fix"
        else:
            out["fix"] = fix(workbook, out["before"])
            out["after"] = scan(workbook)

    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
