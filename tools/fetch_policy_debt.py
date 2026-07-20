r"""Fetch current total policy debt (loan balance) for a batch of policies.

Reads a GLP batch workbook (Policy col B, Company col A, Valuation Date col U),
resolves each policy's TCH_POL_ID from ``LH_BAS_POL``, then aggregates loan
balances from ``LH_CSH_VAL_LOAN`` (traditional) and ``LH_FND_VAL_LOAN``
(advanced) exactly the way ``PolicyInformation.get_loans`` /
``total_loan_balance`` do:

  * principal = LN_PRI_AMT, skipping rows with principal <= 0
  * trad accrued = POL_LN_ITS_AMT only when LN_ITS_AMT_TYP_CD == '2'
  * adv  accrued = POL_LN_ITS_AMT
  * total debt   = sum(principal + accrued)

All lookups are batched with parameterized IN-lists (no one-at-a-time calls).

Usage:
    venv\Scripts\python.exe tools/fetch_policy_debt.py "<workbook>" [--sheet Batch]
        [--region CKPR] [--limit N] [--out <path>]

Emits a JSON summary to stdout; writes the full record array to --out (or the
default path next to the workbook) as {"generated", "region", "source", "count",
"records":[{company, policy, valuation_date, total_policy_debt, found}]}.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, date
from decimal import Decimal

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

# Make the suiteview package importable when run from repo root.
sys.path.insert(0, ".")
from suiteview.core.db2_connection import DB2Connection  # noqa: E402


CHUNK = 400


def _norm_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value)


def read_workbook(path: str, sheet: str | None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2):
        company = row[0].value if len(row) > 0 else None
        policy = row[1].value if len(row) > 1 else None
        valuation = row[20].value if len(row) > 20 else None
        if policy is None:
            continue
        policy_str = str(policy).strip()
        if not policy_str:
            continue
        records.append({
            "company": str(company).strip() if company is not None else None,
            "policy": policy_str,
            "valuation_date": _norm_date(valuation),
        })
    return records


def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _sql_literal(value: str) -> str:
    """Single-quote a string value for inline SQL (this DB2 ODBC driver rejects
    parameter binding, so all callers inline literals — see policy_data.py)."""
    return "'" + str(value).replace("'", "''") + "'"


def resolve_pol_ids(db: DB2Connection, by_company: dict, system_code: str = "I") -> dict:
    """Return {(company, policy): tch_pol_id} for policies found in LH_BAS_POL."""
    resolved = {}
    for company, policies in by_company.items():
        for chunk in _chunks(policies, CHUNK):
            in_list = ",".join(_sql_literal(p) for p in chunk)
            sql = (
                "SELECT CK_CMP_CD, CK_POLICY_NBR, TCH_POL_ID "
                "FROM DB2TAB.LH_BAS_POL "
                f"WHERE CK_SYS_CD = {_sql_literal(system_code)} "
                f"AND CK_CMP_CD = {_sql_literal(company)} "
                f"AND CK_POLICY_NBR IN ({in_list})"
            )
            _, rows = db.execute_query_with_headers(sql)
            for cmp_cd, pol_nbr, tch in rows:
                resolved[(str(cmp_cd).strip(), str(pol_nbr).strip())] = str(tch)
    return resolved


def aggregate_loans(db: DB2Connection, table: str, tch_ids: list, trad: bool) -> dict:
    """Return {tch_pol_id: Decimal debt} summed from a loan table."""
    debt = {}
    for chunk in _chunks(tch_ids, CHUNK):
        in_list = ",".join(_sql_literal(t) for t in chunk)
        sql = (
            "SELECT TCH_POL_ID, LN_PRI_AMT, POL_LN_ITS_AMT, LN_ITS_AMT_TYP_CD "
            f"FROM DB2TAB.{table} "
            f"WHERE TCH_POL_ID IN ({in_list})"
        )
        _, rows = db.execute_query_with_headers(sql)
        for tch, ln_pri, pol_its, its_typ in rows:
            principal = Decimal(str(ln_pri or 0))
            if principal <= 0:
                continue
            if trad:
                accrued = Decimal(str(pol_its or 0)) if str(its_typ) == "2" else Decimal("0")
            else:
                accrued = Decimal(str(pol_its or 0))
            debt[str(tch)] = debt.get(str(tch), Decimal("0")) + principal + accrued
    return debt


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--region", default="CKPR")
    ap.add_argument("--limit", type=int, default=None,
                    help="only process the first N rows (probe mode)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    records = read_workbook(args.workbook, args.sheet)
    if args.limit:
        records = records[:args.limit]

    # Group unique policies by company.
    by_company: dict = {}
    for r in records:
        by_company.setdefault(r["company"], set()).add(r["policy"])
    by_company = {c: sorted(p) for c, p in by_company.items()}

    db = DB2Connection(region=args.region)
    resolved = resolve_pol_ids(db, by_company)

    tch_ids = list(resolved.values())
    trad_debt = aggregate_loans(db, "LH_CSH_VAL_LOAN", tch_ids, trad=True)
    adv_debt = aggregate_loans(db, "LH_FND_VAL_LOAN", tch_ids, trad=False)

    debt_by_tch = {}
    for tch in tch_ids:
        total = trad_debt.get(tch, Decimal("0")) + adv_debt.get(tch, Decimal("0"))
        debt_by_tch[tch] = total

    out_records = []
    found_n = 0
    for r in records:
        key = (r["company"], r["policy"])
        tch = resolved.get(key)
        if tch is None:
            out_records.append({**r, "total_policy_debt": None, "found": False})
            continue
        found_n += 1
        total = debt_by_tch.get(tch, Decimal("0"))
        out_records.append({
            **r,
            "total_policy_debt": f"{total:.2f}",
            "found": True,
        })

    payload = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "region": args.region,
        "source": args.workbook,
        "count": len(out_records),
        "found": found_n,
        "not_found": len(out_records) - found_n,
        "records": out_records,
    }

    out_path = args.out
    if out_path:
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, ensure_ascii=False)

    summary = {k: payload[k] for k in ("generated", "region", "count", "found", "not_found")}
    summary["out"] = out_path
    summary["sample"] = out_records[:5]
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
