r"""Append a header label to the first empty column of row 1 (idempotent).

If the label is already present anywhere in row 1, nothing changes. Otherwise it
is written in the column just past the last non-empty header. Single-purpose
helper for extending a batch sheet's columns. Usage:

    venv\Scripts\python.exe tools/add_xlsx_header.py "path\to\book.xlsx" "Def Life Ins" [Sheet]
"""
from __future__ import annotations

import json
import sys

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def main() -> None:
    if len(sys.argv) < 3:
        print(json.dumps({"error": "usage: add_xlsx_header.py <workbook> <label> [sheet]"}))
        sys.exit(1)
    path, label = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(path)
    ws = wb[sys.argv[3]] if len(sys.argv) > 3 else wb[wb.sheetnames[0]]

    existing = {str(c.value).strip().lower(): c.column_letter
                for c in ws[1] if c.value is not None}
    if label.strip().lower() in existing:
        print(json.dumps({"workbook": path, "sheet": ws.title, "label": label,
                          "added": False, "column": existing[label.strip().lower()]}))
        return

    last_col = max((c.column for c in ws[1] if c.value is not None), default=0)
    target = get_column_letter(last_col + 1)
    ws[f"{target}1"] = label
    wb.save(path)
    print(json.dumps({"workbook": path, "sheet": ws.title, "label": label,
                      "added": True, "column": target}))


if __name__ == "__main__":
    main()
