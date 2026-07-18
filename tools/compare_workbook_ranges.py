"""Diff cached cell values of one or more ranges between two workbooks.

Usage (single JSON arg):
    venv\\Scripts\\python.exe tools/compare_workbook_ranges.py '<json>'

    {"a": "<workbook A>", "b": "<workbook B>",
     "ranges": [{"sheet": "Rates_Control", "ref": "NP12:NY1147"}, ...],
     "tol": 1e-9, "max_diffs": 20}

Numeric cells compare within tol; everything else compares as-is (None == ""
is treated as equal). Prints a JSON summary with per-range diff counts and the
first few differing cells.
"""
from __future__ import annotations

import json
import sys

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter


def _norm(v):
    if v is None or v == "":
        return None
    return v


def _equal(a, b, tol):
    a, b = _norm(a), _norm(b)
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= tol
    return a == b


def main():
    cmd = json.loads(sys.argv[1])
    tol = float(cmd.get("tol", 1e-9))
    max_diffs = int(cmd.get("max_diffs", 20))

    wa = openpyxl.load_workbook(cmd["a"], read_only=True, data_only=True)
    wb = openpyxl.load_workbook(cmd["b"], read_only=True, data_only=True)

    out = {"a": cmd["a"], "b": cmd["b"], "ranges": []}
    try:
        for spec in cmd["ranges"]:
            sa = wa[spec["sheet"]]
            sb = wb[spec["sheet"]]
            c1, r1, c2, r2 = range_boundaries(spec["ref"])
            cells = 0
            diffs = []
            rows_a = sa.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2)
            rows_b = sb.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2)
            for row_a, row_b in zip(rows_a, rows_b):
                for ca, cb in zip(row_a, row_b):
                    cells += 1
                    if not _equal(ca.value, cb.value, tol):
                        if len(diffs) < max_diffs:
                            diffs.append({
                                "cell": f"{get_column_letter(ca.column)}{ca.row}",
                                "a": ca.value, "b": cb.value,
                            })
                        else:
                            diffs.append("...")
                            break
                else:
                    continue
                break
            n_diffs = sum(1 for d in diffs if d != "...")
            out["ranges"].append({
                **spec, "cells": cells,
                "diff_count": n_diffs if "..." not in diffs else f">{n_diffs}",
                "diffs": diffs,
            })
    finally:
        wa.close()
        wb.close()
    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
