"""Drive the RERUN workbook via Excel COM: load a Saved Case, recalc, dump CalcEngine.

RERUN is a single-case engine: the ``Saved Cases`` sheet stores one case per
column (col A = the named input cell, e.g. ``sINPUT_Issue_Age`` / repeated
``vINPUT_Specified_Amount`` for vectors).  Loading a case = writing those
name->value pairs back onto their defined-name ranges, then recalculating.

This tool ALWAYS operates on a temp copy so the 38 MB source-of-truth workbook is
never mutated.

Usage (single JSON arg):
    venv\\Scripts\\python.exe tools/rerun_com.py '<json>'

Modes:
  probe  - open temp copy, (optionally) recalc, read cells. De-risks COM.
    {"mode":"probe","workbook":"<path>","recalc":true,
     "reads":[{"sheet":"CalcEngine","cell":"AL6"},{"sheet":"CalcEngine","cell":"G6"}]}

  run    - load a Saved Case, recalc, dump CalcEngine rows to CSV.
    {"mode":"run","workbook":"<path>","case":1,"out_csv":"<path>",
     "max_month":120,"cols":["B","C","D","G","AL"]}
      case: 1-based case number (Saved Cases row-1 header) OR a CaseID string.
      cols: CalcEngine column letters to dump (default: a small key set).
            CalcEngine data rows start at row 6 = month 1 (row r = month r-5).
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


SAVED_CASES = "Saved Cases"
CALC_FIRST_ROW = 6  # CalcEngine row 6 == projection month 1
DEFAULT_COLS = ["B", "C", "D", "G", "AL", "AM"]

# msoAutomationSecurityForceDisable — open with macros disabled (no prompts).
MSO_AUTOMATION_SECURITY_FORCE_DISABLE = 3
XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105


# ── Saved Cases reading (openpyxl, no COM) ─────────────────────────────────

def _resolve_case_column(ws, case) -> int:
    """Return the 1-based column index for the requested case.

    ``case`` is a 1-based case number (matched against the row-1 header) or a
    CaseID string (matched against the sINPUT_CaseID row in col A).
    """
    # Build a name->row index from col A for CaseID matching.
    name_rows = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a:
            name_rows[str(a).strip()] = r

    # Numeric case -> match row-1 header.
    try:
        case_num = int(case)
    except (TypeError, ValueError):
        case_num = None

    for c in range(3, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if case_num is not None and header is not None:
            try:
                if int(header) == case_num:
                    return c
            except (TypeError, ValueError):
                pass

    # String CaseID -> match the sINPUT_CaseID row.
    cid_row = name_rows.get("sINPUT_CaseID")
    if cid_row is not None:
        for c in range(3, ws.max_column + 1):
            if str(ws.cell(row=cid_row, column=c).value).strip() == str(case).strip():
                return c

    raise ValueError(f"Could not resolve Saved Cases column for case={case!r}")


def read_case_inputs(workbook: str, case) -> list[tuple[str, list]]:
    """Return ordered [(defined_name, [values...]), ...] for a Saved Case.

    Consecutive identical names (e.g. vINPUT_Specified_Amount) are grouped into a
    single vector entry, preserving order.
    """
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[SAVED_CASES]
    col = _resolve_case_column(ws, case)

    pairs: list[tuple[str, list]] = []
    last_name = None
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        value = ws.cell(row=r, column=col).value
        if name == last_name:
            pairs[-1][1].append(value)
        else:
            pairs.append((name, [value]))
            last_name = name
    wb.close()
    return pairs, col


# ── COM lifecycle ──────────────────────────────────────────────────────────

def _open_excel():
    """Spawn a brand-new, ISOLATED Excel instance (never the user's session).

    Uses CoCreateInstance with CLSCTX_LOCAL_SERVER so we get a dedicated process
    and late-bound dispatch — Quit() then only closes our own instance.
    """
    import win32com.client

    # DispatchEx forces a NEW out-of-process instance (CLSCTX_LOCAL_SERVER),
    # so we never attach to the user's running Excel.  Falls back to late-bound
    # dynamic dispatch when no makepy cache exists (avoids gen_py issues).
    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.ScreenUpdating = False
    try:
        xl.EnableEvents = False
    except Exception:
        pass
    try:
        xl.AutomationSecurity = MSO_AUTOMATION_SECURITY_FORCE_DISABLE
    except Exception:
        pass
    return xl


def _write_named_range(wb, name: str, values: list) -> None:
    """Write value(s) onto a defined-name range in ONE COM call (block write)."""
    rng = wb.Names(name).RefersToRange
    nrows = int(rng.Rows.Count)
    ncols = int(rng.Columns.Count)
    if nrows * ncols == 1:
        rng.Value = _coerce(values[0])
        return
    data = []
    for r in range(nrows):
        row_vals = []
        for c in range(ncols):
            idx = r * ncols + c
            row_vals.append(_coerce(values[idx]) if idx < len(values) else None)
        data.append(tuple(row_vals))
    rng.Value = tuple(data)


def _temp_copy(workbook: str) -> Path:
    src = Path(workbook)
    tmp = Path(tempfile.gettempdir()) / f"rerun_com_{src.stem}.xlsm"
    shutil.copyfile(src, tmp)
    return tmp


def _enable_iteration(xl) -> None:
    # UL account-value formulas have AV<->interest circular references; RERUN
    # relies on iterative calculation.  Enable it defensively (Excel ignores it
    # when there are no circular refs).
    try:
        xl.Iteration = True
        xl.MaxIterations = 1000
        xl.MaxChange = 1e-9
    except Exception:
        pass


def _coerce(v):
    """openpyxl value -> Excel-writable value ('' becomes a blank cell)."""
    if v is None or v == "":
        return None
    return v


# ── Modes ──────────────────────────────────────────────────────────────────

def mode_probe(cmd):
    tmp = _temp_copy(cmd["workbook"])
    xl = _open_excel()
    out = {"temp": str(tmp), "reads": []}
    try:
        _enable_iteration(xl)
        wb = xl.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False)
        xl.Calculation = XL_CALC_MANUAL
        if cmd.get("recalc", True):
            xl.CalculateFull()
        for spec in cmd.get("reads", []):
            ws = wb.Worksheets(spec["sheet"])
            val = ws.Range(spec["cell"]).Value
            out["reads"].append({**spec, "value": val})
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
        try:
            tmp.unlink()
        except OSError:
            pass
    print(json.dumps(out, indent=2, default=str))


def mode_run(cmd):
    workbook = cmd["workbook"]
    pairs, src_col = read_case_inputs(workbook, cmd["case"])
    cols = cmd.get("cols", DEFAULT_COLS)
    max_month = int(cmd.get("max_month", 120))
    out_csv = cmd.get("out_csv")

    tmp = _temp_copy(workbook)
    xl = _open_excel()
    report = {
        "case": cmd["case"], "saved_cases_col": get_column_letter(src_col),
        "inputs_written": 0, "input_failures": [], "cols": cols,
        "max_month": max_month, "out_csv": out_csv,
    }
    try:
        _enable_iteration(xl)
        wb = xl.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False)
        # Manual calc BEFORE writing inputs — otherwise each cell write triggers a
        # full recalc of the 729-col workbook (a recalc storm that hangs Excel).
        xl.Calculation = XL_CALC_MANUAL

        # ── Load the case: write each name's value(s) onto its range ──
        for name, values in pairs:
            try:
                _write_named_range(wb, name, values)
                report["inputs_written"] += 1
            except Exception as exc:
                report["input_failures"].append({"name": name, "error": str(exc)})

        # ── Scenario overrides (e.g. construct a face increase/decrease or DBO
        # change on top of a loaded case). Each: {"target": "INPUT!J14:J126" or a
        # defined name, "value": x}. A scalar fills the whole range. ──
        for ov in cmd.get("overrides", []):
            target = ov["target"]
            try:
                if "!" in target:
                    sheet, addr = target.split("!", 1)
                    rng = wb.Worksheets(sheet).Range(addr)
                else:
                    rng = wb.Names(target).RefersToRange
                rng.Value = ov["value"]
                report["overrides_applied"] = report.get("overrides_applied", 0) + 1
            except Exception as exc:
                report.setdefault("override_failures", []).append(
                    {"target": target, "error": str(exc)})

        xl.CalculateFull()

        # ── Dump CalcEngine output rows ──
        ce = wb.Worksheets("CalcEngine")
        col_idx = [column_index_from_string(c) for c in cols]
        c_lo, c_hi = min(col_idx), max(col_idx)
        r_lo, r_hi = CALC_FIRST_ROW, CALC_FIRST_ROW + max_month - 1
        block = ce.Range(ce.Cells(r_lo, c_lo), ce.Cells(r_hi, c_hi)).Value
        # block is a tuple of row-tuples; index by (col - c_lo).
        rows = []
        for r_off, row in enumerate(block):
            rec = {"month": r_off + 1}
            for c, ci in zip(cols, col_idx):
                rec[c] = row[ci - c_lo]
            rows.append(rec)
        report["rows_dumped"] = len(rows)

        if out_csv:
            import csv
            with open(out_csv, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(["month"] + cols)
                for rec in rows:
                    w.writerow([rec["month"]] + [rec[c] for c in cols])
        else:
            report["rows"] = rows[:24]

        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
        try:
            tmp.unlink()
        except OSError:
            pass
    print(json.dumps(report, indent=2, default=str))


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "missing JSON arg"}))
        sys.exit(1)
    cmd = json.loads(sys.argv[1])
    mode = cmd.get("mode", "probe")
    {"probe": mode_probe, "run": mode_run}[mode](cmd)


if __name__ == "__main__":
    main()
