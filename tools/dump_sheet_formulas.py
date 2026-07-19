"""Print cell formulas (not values) from a workbook sheet via openpyxl.

Usage:
    venv\\Scripts\\python.exe tools/dump_sheet_formulas.py '<json>'
    {"workbook": "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm",
     "sheet": "Debug File", "cells": ["D13","F13","K13"]}
    or {"sheet": "Debug File", "row": 13, "cols": ["A","AU"]}  # col range
"""
from __future__ import annotations

import json
import sys

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def main():
    cmd = json.loads(sys.argv[1])
    wb = openpyxl.load_workbook(cmd["workbook"], read_only=True, data_only=False)
    ws = wb[cmd["sheet"]]
    out = {}
    cells = cmd.get("cells") or []
    if cmd.get("row") and cmd.get("cols"):
        lo, hi = (column_index_from_string(c) for c in cmd["cols"])
        cells += [f"{get_column_letter(c)}{cmd['row']}" for c in range(lo, hi + 1)]
    for addr in cells:
        v = ws[addr].value
        # openpyxl wraps CSE/dynamic-array formulas in ArrayFormula objects.
        out[addr] = getattr(v, "text", v)
    wb.close()
    print(json.dumps(out, indent=2, default=str))


if __name__ == "__main__":
    main()
