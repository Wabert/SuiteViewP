r"""Batch "Billable to MD" hand-off dates over a policy list (GLP Limit Calc v2).

Reads a policy list from a workbook (Company in col A, Policy in col B, header
in row 1) and, for each policy, runs it on the "Billable to MD" premium type —
pay the policy's current billable premium (on its billing mode) from the
current policy year to maturity, letting the engine hand off to the Monthly
Deduction premium the first month the billable premium can no longer keep the
policy in force, then to GP exception premiums once the guideline room runs
out. Two columns are written, matched to the header row by label:

  * "Billable to MD - MD Date"        — the date Monthly Deduction premiums
                                         begin, or "Maturity" if they never do.
  * "Billable to MD - Exception Date" — the date GP exception premiums begin,
                                         or "Maturity" if they never do.

Missing output columns are appended to the header row. Bypassed policies
(active shadow account / MD diff / missing rates / CVAT) leave both dates blank;
the workbook's existing Run Status column explains why.

The per-policy forecast logic lives in
``suiteview/illustration/core/batch_runner.run_billable_to_md_policy`` (shared
with the GLP forecast set); this tool owns only the workbook I/O around it.

Live data: this reads policy data through PolicyInformation / DB2 like the rest
of the app. It does NOT enable local SQLite fixtures.

Usage (PowerShell):
    venv\Scripts\python.exe tools/run_billable_to_md_batch.py ^
        "docs\Illustration_UL\GLP Limit Calc v2.xlsx" --limit 10

Flags:
    --region CKPR   DB2 region (default: CKPR)
    --sheet  NAME   sheet name (default: first sheet)
    --limit  N      process at most N policy rows from the top
    --lumpsum-to-next  solve a Lumpsum to Next Premium bridge and hold the MD
                       hand-off off until the next billable premium
    --skip-loans    bypass any policy carrying a loan balance
    --dry-run       compute and print, but do not save the workbook

Crash safety: every completed row is appended to a per-row JSONL sidecar next
to the workbook (flushed + fsync'd), so a crash or Ctrl-C never loses prior
work. If the final workbook save fails (e.g. the file is open in Excel), rebuild
it from the sidecar with ``--replay-sidecar PATH``.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: "
          "venv\\Scripts\\python.exe -m pip install openpyxl")
    sys.exit(1)

from suiteview.illustration.core.batch_runner import (  # noqa: E402
    BILLABLE_TO_MD_COLUMNS, run_billable_to_md_policy,
)

DATE_FMT = "m/d/yyyy"

# Header label (lower-cased, tolerant of the "exceptoin" typo) -> field key.
_HEADER_ALIASES = {
    "billable to md - md date": "b2md_md_date",
    "billable to md - exception date": "b2md_exc_date",
    "billable to md - exceptoin date": "b2md_exc_date",
}


def _resolve_columns(ws) -> dict:
    """Map field keys to existing column letters from the header row."""
    col: dict = {}
    for cell in ws[1]:
        key = _HEADER_ALIASES.get(str(cell.value or "").strip().lower())
        if key and key not in col:
            col[key] = cell.column_letter
    return col


def _ensure_columns(ws, col: dict) -> None:
    """Append any Billable-to-MD output column missing from the header row."""
    next_col = ws.max_column + 1
    for key, label in BILLABLE_TO_MD_COLUMNS:
        if key in col:
            continue
        cell = ws.cell(row=1, column=next_col)
        cell.value = label
        col[key] = cell.column_letter
        next_col += 1


def _write(ws, row: int, letter: str, value) -> None:
    cell = ws[f"{letter}{row}"]
    cell.value = value
    if isinstance(value, date):
        cell.number_format = DATE_FMT


def _json_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _default_sidecar_path(wb_path: Path, sheet: str, limit) -> Path:
    """Sidecar JSONL path next to the workbook, stamped with sheet/limit/time."""
    parts = [sheet]
    if limit:
        parts.append(f"limit{limit}")
    parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    suffix = ".".join(re.sub(r"[^A-Za-z0-9_.-]+", "_", str(p)) for p in parts)
    return wb_path.with_name(f"{wb_path.stem}.b2md.{suffix}.results.jsonl")


def _write_sidecar(handle, record: dict) -> None:
    """Append one crash-safe JSON line (flushed + fsync'd so a crash keeps it)."""
    if handle is None:
        return
    handle.write(json.dumps(record, default=str) + "\n")
    handle.flush()
    os.fsync(handle.fileno())


def _row_record(row: int, policy: str, company, result) -> dict:
    """The per-row result captured to the sidecar."""
    return {
        "row": row,
        "company": _json_value(company),
        "policy": policy,
        "status": result.status,
        "error": result.error,
        "b2md_md_date": _json_value(result.values.get("b2md_md_date")),
        "b2md_exc_date": _json_value(result.values.get("b2md_exc_date")),
        "lumpsum": result.values.get("lumpsum"),
    }


def _replay_value(value):
    """Turn a sidecar JSON date string back into a date; keep text labels."""
    if value in (None, ""):
        return None
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value        # a text label ("Maturity", ...)
    return value


def _replay_sidecar(ws, col: dict, sidecar_path: Path) -> dict:
    """Write a prior sidecar JSONL back into the workbook's output columns."""
    counts: dict = {}
    rows = []
    with sidecar_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            row = int(record["row"])
            _write(ws, row, col["b2md_md_date"],
                   _replay_value(record.get("b2md_md_date")))
            _write(ws, row, col["b2md_exc_date"],
                   _replay_value(record.get("b2md_exc_date")))
            status = str(record.get("status") or "").strip() or "(blank)"
            counts[status] = counts.get(status, 0) + 1
            rows.append(row)
    return {
        "sidecar": str(sidecar_path),
        "processed": len(rows),
        "start_row": min(rows) if rows else None,
        "end_row": max(rows) if rows else None,
        "status_counts": counts,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook")
    parser.add_argument("--region", default="CKPR")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--start-row", type=int, default=None,
        help="first worksheet row to process (default: 2, the first data row)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--lumpsum-to-next", action="store_true",
        help="solve a Lumpsum to Next Premium bridge and hold the MD hand-off "
             "off until the next billable premium")
    parser.add_argument(
        "--skip-loans", action="store_true",
        help="bypass any policy carrying a loan balance")
    parser.add_argument(
        "--sidecar", default=None,
        help="write per-row JSONL results here (default: next to the workbook)")
    parser.add_argument(
        "--replay-sidecar", default=None,
        help="write a prior sidecar JSONL back into the workbook and save")
    args = parser.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"Workbook not found: {wb_path}")
        return 1

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    col = _resolve_columns(ws)
    _ensure_columns(ws, col)

    # ── Replay mode: rebuild the workbook from a prior sidecar and stop ──
    if args.replay_sidecar:
        replay = _replay_sidecar(ws, col, Path(args.replay_sidecar))
        if not args.dry_run:
            try:
                wb.save(wb_path)
            except PermissionError:
                print(f"Could not save {wb_path} — is it open in Excel?")
                return 1
        print(json.dumps({"workbook": str(wb_path), "replayed": True,
                          "saved": not args.dry_run, **replay}, indent=2))
        return 0

    # ── Open the crash-safe sidecar (per-row JSONL, flushed each row) ──
    sidecar_handle = None
    sidecar_path = None
    if not args.dry_run:
        sidecar_path = (Path(args.sidecar) if args.sidecar
                        else _default_sidecar_path(wb_path, ws.title, args.limit))
        sidecar_path.parent.mkdir(parents=True, exist_ok=True)
        sidecar_handle = sidecar_path.open("w", encoding="utf-8")
        print(f"Writing per-row sidecar JSONL: {sidecar_path}")

    processed = 0
    try:
        processed = _run_rows(ws, col, args, sidecar_handle)
    finally:
        if sidecar_handle is not None:
            sidecar_handle.close()

    if args.dry_run:
        print(f"\nDry run — {processed} policies processed, workbook not saved.")
        return 0

    try:
        wb.save(wb_path)
    except PermissionError:
        print(f"\nCould not save {wb_path} — is it open in Excel? "
              f"Close it and replay the sidecar with:\n"
              f"  venv\\Scripts\\python.exe tools/run_billable_to_md_batch.py "
              f"\"{wb_path}\" --replay-sidecar \"{sidecar_path}\"")
        return 1
    print(f"\nSaved {processed} policies to {wb_path}")
    if sidecar_path is not None:
        print(f"Sidecar retained: {sidecar_path}")
    return 0


def _run_rows(ws, col: dict, args, sidecar_handle) -> int:
    """Process each policy row, writing cells and a crash-safe sidecar line."""
    processed = 0
    start_row = args.start_row if args.start_row is not None else 2
    for row in range(start_row, ws.max_row + 1):
        company = ws[f"A{row}"].value
        policy = ws[f"B{row}"].value
        if not policy:
            continue
        if args.limit is not None and processed >= args.limit:
            break
        processed += 1

        company_code = str(company).strip() if company not in (None, "") else None
        policy_number = str(policy).strip()

        result = run_billable_to_md_policy(
            policy_number, company=company_code, region=args.region,
            lumpsum_to_next=args.lumpsum_to_next, skip_loans=args.skip_loans)

        md_date = result.values.get("b2md_md_date")
        exc_date = result.values.get("b2md_exc_date")
        lump = result.values.get("lumpsum")
        lump_txt = f"  Lump={lump:,.2f}" if lump else ""
        print(f"row {row:>4}  {policy_number:<12}  {result.status:<28}  "
              f"MD={md_date!s:<12}  Exception={exc_date!s}{lump_txt}"
              + (f"  [{result.error}]" if result.error else ""))

        if not args.dry_run:
            _write(ws, row, col["b2md_md_date"], md_date)
            _write(ws, row, col["b2md_exc_date"], exc_date)
            # Persist this row before moving on — a crash/Ctrl-C keeps every
            # completed row and the workbook can be rebuilt via --replay-sidecar.
            _write_sidecar(
                sidecar_handle,
                _row_record(row, policy_number, company_code, result))

    return processed


if __name__ == "__main__":
    raise SystemExit(main())
