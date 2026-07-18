r"""Create a minimal policy-list workbook (Company col A, Policy col B).

Single-purpose helper for building test inputs for the batch forecast tools,
which resolve/append their own output headers. Takes one JSON argument:

    venv\Scripts\python.exe tools/make_policy_list_workbook.py ^
        "{\"out\": \"path\\to\\book.xlsx\", \"policies\": [[\"01\", \"UL054426\"]]}"

Keys: out (path), policies (list of [company, policy]), sheet (optional name).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: make_policy_list_workbook.py '<json>'"}))
        sys.exit(1)
    cmd = json.loads(sys.argv[1])
    out = Path(cmd["out"])
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cmd.get("sheet", "Sheet1")
    ws["A1"] = "Company"
    ws["B1"] = "Policy"
    for i, (company, policy) in enumerate(cmd["policies"], start=2):
        ws[f"A{i}"] = company
        ws[f"B{i}"] = policy
    wb.save(out)
    print(json.dumps({"out": str(out), "sheet": ws.title,
                      "rows": len(cmd["policies"])}))


if __name__ == "__main__":
    main()
