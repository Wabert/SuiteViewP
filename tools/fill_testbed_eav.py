"""Run as-is (current billable premium) forecasts for workbook testbed policies.

Writes column E (Termination Date) and column F (EAV at termination). The
termination date is the first projected lapsed state, using the same lapse flag
as the illustration UI; for SV-lapse plans this is driven by surrender value
after surrender charges and policy debt. If the policy never lapses, writes the
maturity date and EAV at maturity. If present, writes the "Plancodes for Riders
and Benefits" column with comma-delimited active rider plancodes and
supplemental benefit codes, and writes "MD Diff" as CyberLife monthly deduction
minus calculated monthly deduction on the valuation date.

Usage:
    venv\\Scripts\\python.exe tools/fill_testbed_eav.py --sheet TestBed2 --limit 10 --exact-days false
    venv\\Scripts\\python.exe tools/fill_testbed_eav.py --sheet FullTest2 --limit 50 --output "docs/Illustration_UL/Test Matrix prior 2000 - filled.xlsx"
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import openpyxl
from PyQt6.QtWidgets import QApplication

from suiteview.core.policy_service import get_policy_info, clear_cache
from suiteview.illustration.core.calc_engine import IllustrationEngine
from suiteview.illustration.core.illustration_policy_service import (
    active_rider_benefit_codes,
    build_illustration_data,
)
from suiteview.illustration.core.scenario_builder import build_illustration_scenario
from suiteview.illustration.ui.inputs_tab import IllustrationInputsTab

XLSX = ROOT / "docs" / "Illustration_UL" / "Test Matrix prior 2000.xlsx"
SHEET = "TestBed"
REGION = "CKPR"
TERMINATION_DATE_HEADER = "Termination Date"
EAV_AT_TERMINATION_HEADER = "EAV at termination"
RIDERS_BENEFITS_HEADER = "Plancodes for Riders and Benefits"
MD_DIFF_HEADER = "MD Diff"


def _excel_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, time())
    return value


def save_or_update_open_workbook(wb, path: Path, sheet_name: str, updates: list[tuple[int, int, Any, str | None]]) -> str:
    try:
        wb.save(path)
        return "openpyxl"
    except PermissionError:
        import win32com.client  # type: ignore[import-not-found]

        try:
            excel_wb = win32com.client.GetObject(str(path))
            excel_ws = excel_wb.Worksheets(sheet_name)
            for row, col, value, number_format in updates:
                cell = excel_ws.Cells(row, col)
                cell.Value = _excel_scalar(value)
                if number_format is not None:
                    cell.NumberFormat = number_format
            excel_wb.Save()
            return "excel-com-getobject"
        except Exception:
            pass

        excel = win32com.client.GetActiveObject("Excel.Application")
        workbooks = excel.Workbooks
        target = str(path).lower()
        for index in range(1, workbooks.Count + 1):
            excel_wb = workbooks.Item(index)
            if str(excel_wb.FullName).lower() != target:
                continue
            excel_ws = excel_wb.Worksheets(sheet_name)
            for row, col, value, number_format in updates:
                cell = excel_ws.Cells(row, col)
                cell.Value = _excel_scalar(value)
                if number_format is not None:
                    cell.NumberFormat = number_format
            excel_wb.Save()
            return "excel-com"
        raise


def find_header_column(ws, header: str) -> int | None:
    for cell in ws[1]:
        if str(cell.value or "").strip() == header:
            return cell.column
    return None


def forecast_termination(policy_number: str, *, exact_days_interest: bool):
    """Return (date, eav/status, esv, rider_benefit_codes, md_diff, outcome)."""
    clear_cache()
    pi = get_policy_info(policy_number, REGION)
    if pi is None or not getattr(pi, "exists", False):
        return None, "NOT FOUND", None, "", None, "not_found"

    rider_benefit_codes = active_rider_benefit_codes(pi)

    policy_data = build_illustration_data(policy_number, region=REGION, company_code=pi.company_code)

    tab = IllustrationInputsTab()
    tab.load_data_from_policy(pi)
    tab.exact_days_check.setChecked(exact_days_interest)

    scenario = build_illustration_scenario(
        policy_data,
        inforce_overrides=tab.export_inforce_overrides(),
        future_inputs=tab.export_input_set(),
    )
    months = tab._months_to_maturity(scenario.projectable_policy)

    engine = IllustrationEngine()
    options = tab.export_options()
    results = engine.project(
        scenario.projectable_policy,
        months=months,
        future_inputs=scenario.future_inputs,
        options=options,
        stop_on_lapse=False,
    )
    inforce = results[0]
    md_diff = (
        float(inforce.system_monthly_deduction or 0.0)
        - float(inforce.md_check_calculated_deduction or 0.0)
    )

    for st in results[1:]:
        if st.lapsed:
            return st.date, round(st.av_end_of_month, 2), round(st.ending_sv, 2), rider_benefit_codes, md_diff, "termination"
    maturity = results[-1]
    return maturity.date, round(maturity.av_end_of_month, 2), round(maturity.ending_sv, 2), rider_benefit_codes, md_diff, "maturity"


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill illustration testbed EAV results.")
    parser.add_argument("--sheet", default=SHEET, help="Workbook sheet to fill")
    parser.add_argument("--limit", type=int, default=None, help="Maximum policies to process")
    parser.add_argument("--start-policy", default=None, help="Policy number to start processing from")
    parser.add_argument("--output", default=None, help="Optional output workbook path; defaults to updating the source workbook")
    parser.add_argument(
        "--exact-days",
        choices=("true", "false"),
        default="false",
        help="Use exact-days interest instead of monthly compounding",
    )
    args = parser.parse_args()

    _app = QApplication.instance() or QApplication([])

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[args.sheet]
    rider_benefit_col = find_header_column(ws, RIDERS_BENEFITS_HEADER)
    md_diff_col = find_header_column(ws, MD_DIFF_HEADER)
    exact_days_interest = args.exact_days == "true"
    processed = 0
    found_start_policy = args.start_policy is None
    updates: list[tuple[int, int, Any, str | None]] = []

    ws.cell(row=1, column=5).value = TERMINATION_DATE_HEADER
    ws.cell(row=1, column=6).value = EAV_AT_TERMINATION_HEADER
    updates.extend([
        (1, 5, TERMINATION_DATE_HEADER, None),
        (1, 6, EAV_AT_TERMINATION_HEADER, None),
    ])

    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if raw is None or not str(raw).strip():
            continue
        policy = str(raw).strip()
        if not found_start_policy:
            if policy.upper() != args.start_policy.upper():
                continue
            found_start_policy = True
        if args.limit is not None and processed >= args.limit:
            break
        try:
            d, val, esv, rider_benefit_codes, md_diff, outcome = forecast_termination(
                policy,
                exact_days_interest=exact_days_interest,
            )
        except Exception as exc:  # noqa: BLE001 - record per-policy failure
            ws.cell(row=row, column=5).value = "ERROR"
            ws.cell(row=row, column=6).value = str(exc)[:120]
            updates.extend([
                (row, 5, "ERROR", None),
                (row, 6, str(exc)[:120], None),
            ])
            if rider_benefit_col is not None:
                ws.cell(row=row, column=rider_benefit_col).value = None
                updates.append((row, rider_benefit_col, None, None))
            if md_diff_col is not None:
                ws.cell(row=row, column=md_diff_col).value = None
                updates.append((row, md_diff_col, None, None))
            print(f"{policy}: ERROR {exc}")
            processed += 1
            continue

        if rider_benefit_col is not None:
            ws.cell(row=row, column=rider_benefit_col).value = rider_benefit_codes or None
            updates.append((row, rider_benefit_col, rider_benefit_codes or None, None))
        if md_diff_col is not None:
            ws.cell(row=row, column=md_diff_col).value = round(md_diff, 2) if md_diff is not None else None
            if md_diff is not None:
                ws.cell(row=row, column=md_diff_col).number_format = "#,##0.00"
            updates.append((row, md_diff_col, round(md_diff, 2) if md_diff is not None else None, "#,##0.00" if md_diff is not None else None))

        if outcome == "not_found":
            ws.cell(row=row, column=5).value = val
            ws.cell(row=row, column=6).value = None
            updates.extend([
                (row, 5, val, None),
                (row, 6, None, None),
            ])
            print(f"{policy}: {val}  Riders/Benefits={rider_benefit_codes or '(blank)'}")
        else:
            ws.cell(row=row, column=5).value = d
            ws.cell(row=row, column=5).number_format = "mm/dd/yyyy"
            ws.cell(row=row, column=6).value = val
            ws.cell(row=row, column=6).number_format = "#,##0.00"
            updates.extend([
                (row, 5, d, "mm/dd/yyyy"),
                (row, 6, val, "#,##0.00"),
            ])
            esv_label = f"  ESV={esv:,.2f}" if esv is not None else ""
            md_label = f"  MD Diff={md_diff:,.2f}" if md_diff is not None else ""
            print(f"{policy}: Termination Date={d:%m/%d/%Y}  EAV at termination={val:,.2f}{esv_label}{md_label}  Riders/Benefits={rider_benefit_codes or '(blank)'}")
        processed += 1

    if not found_start_policy:
        raise SystemExit(f"Start policy not found on {args.sheet}: {args.start_policy}")

    output_path = Path(args.output) if args.output else XLSX
    if not output_path.is_absolute():
        output_path = ROOT / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    save_method = save_or_update_open_workbook(wb, output_path, args.sheet, updates)
    print(f"\nSaved {processed} policy row(s) on {args.sheet} via {save_method} -> {output_path}")


if __name__ == "__main__":
    main()
