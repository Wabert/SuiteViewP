r"""Append a 'Total Policy Debt' column to the GLP batch workbook.

Reads the debt values produced by ``fetch_policy_debt.py`` (JSON) and writes
them into a new column immediately after the last populated header column of the
``Batch`` sheet, matching rows by (Company col A, Policy col B). Debt is written
as a numeric value (2dp); rows whose policy wasn't found get left blank.

Usage:
    venv\Scripts\python.exe tools/append_debt_column.py <workbook> <debt_json> [--sheet Batch]
"""
from __future__ import annotations

import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


HEADER = "Total Policy Debt"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("debt_json")
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    with open(args.debt_json, "r", encoding="utf-8") as fh:
        payload = json.load(fh)

    debt_by_key = {}
    for rec in payload["records"]:
        key = (str(rec.get("company")).strip() if rec.get("company") is not None else None,
               str(rec.get("policy")).strip())
        debt_by_key[key] = rec.get("total_policy_debt")

    wb = openpyxl.load_workbook(args.workbook)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    # Find the last populated header column in row 1.
    last_col = 0
    for cell in ws[1]:
        if cell.value is not None:
            last_col = cell.column
    target_col = last_col + 1

    ws.cell(row=1, column=target_col, value=HEADER)

    written = 0
    blank = 0
    for r in range(2, ws.max_row + 1):
        policy = ws.cell(row=r, column=2).value  # B
        if policy is None or not str(policy).strip():
            continue
        company = ws.cell(row=r, column=1).value  # A
        key = (str(company).strip() if company is not None else None,
               str(policy).strip())
        debt = debt_by_key.get(key)
        if debt is None:
            blank += 1
            continue
        ws.cell(row=r, column=target_col, value=round(float(debt), 2))
        written += 1

    wb.save(args.workbook)

    print(json.dumps({
        "workbook": args.workbook,
        "sheet": ws.title,
        "header": HEADER,
        "column_index": target_col,
        "column_letter": openpyxl.utils.get_column_letter(target_col),
        "rows_written": written,
        "rows_blank": blank,
    }, indent=2))


if __name__ == "__main__":
    main()
