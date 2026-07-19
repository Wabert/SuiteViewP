r"""Batch "Min Level to Exception" solve over a list of policies in a workbook.

Reads a policy list from a workbook (Company in col A, Policy in col B, header
in row 1), runs the Min Level to Exception solve for each (the lowest modal
level premium that keeps the GPT policy in force to maturity, with GLP
exception premiums allowed), and writes the solved premium plus a snapshot of
the policy back into the same sheet — one row per policy.

The per-policy solve logic lives in
``suiteview/illustration/core/batch_runner.py`` (shared with the Illustration
app's Batch tab); this tool owns only the workbook I/O around it.

Columns are matched to the workbook by HEADER LABEL at runtime (see HEADERS), so
the tool follows whatever order the headers are in and tolerates reordering. A
header that isn't present is simply skipped (and reported). Output fields:
    Plancode, Form, Active Riders and Benefits, Run Status, Absolute Max Prem,
    Absolute Max AV, Min Level Prem, Total Prem Paid, Exception Duration,
    Current Duration (policy year), MD (CyberLife), MD Diff (CyberLife - calc),
    Loan Amount, Exception Date, Face Amount, Issue date, Maturity Date,
    Maturity Duration (years), Maturity Age, Issue Age, Attained Age, GSP, GLP,
    AccumLP, PremTD, AccumWD, Valuation Date, Suspense Code, Def Life Ins.
(Company / Policy are read from the first two columns as input.)

Absolute Max Prem: a forecast funded to the guideline maximum every month with
GLP exceptions OFF (premium 999,999,999, which the cap clips). "Maturity" if the
policy still reaches its maturity age, otherwise the policy year it lapses. This
runs for every loaded policy, independent of Run Status. Absolute Max AV is the
account value at maturity in that scenario (blank when it lapses before maturity).

Total Prem Paid: lifetime premiums to maturity in the Min Level to Exception solve
scenario — the cumulative applied premium plus GP exception premiums plus loan
repayments. Blank when the solve doesn't run (bypassed / error).

Run Status decides whether the Min Level to Exception solve runs, and its outcome:
    "bypass (<codes>)" - the solve was skipped; the codes name the reason(s):
        A            = active shadow account (benefit type "A")
        MD           = MD diff != 0 (our engine doesn't match CyberLife)
        rates missing= an active rider/benefit charge has no loaded rate
        CVAT         = CVAT policy (solve is GPT-only)
        no solution  = no level premium keeps it in force to maturity
      Multiple reasons combine, e.g. "bypass (A, MD)". Solve columns left blank.
      A policy loan is NOT a bypass: it is solved with Apply Premium to Loan First
      (the level premium repays the loan before funding the account value).
    "Error" - the policy could not be loaded or rates/MD could not be evaluated.
    Otherwise the solve ran and the status is:
      "Matured w/o Except Prem" - the level premium endows on its own.
      "Except Prem Required"    - it rides the GLP exception period to maturity.

The solve uses each policy's own billing mode, with Lumpsum to Next Prem always
layered on top of the solved premium. Every column except the solve set (Min
Level Prem / Total Prem Paid / Exception Duration / Exception Date) is the
policy's current inforce snapshot and is written regardless of Run Status.

Live data: this reads policy data through PolicyInformation / DB2 like the rest
of the app. It does NOT enable local SQLite fixtures — set SUITEVIEW_LOCAL_DATA=1
in the environment yourself only for deliberate offline testing of the bundled
fixture policies.

Usage (friendly — quoted path + flags, works in PowerShell):
    venv\Scripts\python.exe tools/run_min_level_to_exception_batch.py ^
        "docs\Illustration_UL\Exception prem testing.xlsx"

    # test run — first policy only, do not save:
    venv\Scripts\python.exe tools/run_min_level_to_exception_batch.py ^
        "docs\Illustration_UL\Exception prem testing.xlsx" --limit 1 --dry-run

Flags:
    --region CKPR     DB2 region (default: CKPR)
    --sheet  Sheet1   sheet name (default: first sheet)
    --limit  N        process at most N policy rows from the top
    --start-policy P  start at the first row whose policy in col B matches P
    --start-row N     start at 1-based sheet row N
    --rows   2,3,4    explicit 1-based sheet rows to run (overrides --limit)
    --dry-run         compute and print, but do not save the workbook
    --summarize-only  read matching rows and print saved Run Status counts
    --sidecar PATH    write per-row JSONL results here (default: next to workbook)
    --replay-sidecar PATH
                      write a prior sidecar JSONL back into the workbook

Also accepts a single JSON blob as the only argument (keys: workbook, sheet,
region, rows, start_policy, start_row, limit, dry_run, summarize_only, sidecar,
replay_sidecar) for programmatic callers.
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: "
                      "venv\\Scripts\\python.exe -m pip install openpyxl"}))
    sys.exit(1)

from suiteview.illustration.core.batch_runner import (  # noqa: E402
    FMT_DATE, FMT_INT, FMT_MONEY, MINLEVEL_COLUMNS, MINLEVEL_FORMATS,
    MINLEVEL_SOLVE_KEYS, STATUS_ERROR, STATUS_EXCEPT, STATUS_MATURED,
)

# Field key -> exact header text in row 1 (shared with the Batch tab via
# batch_runner.MINLEVEL_COLUMNS). Columns are resolved by label at runtime (see
# _resolve_columns) so the tool follows the user's column order. Missing output
# columns are appended. A=Company / B=Policy are read directly, not here.
HEADERS = dict(MINLEVEL_COLUMNS)

# Resolved at runtime: field key -> column letter (only headers actually found).
COL: dict = {}

MONEY_FMT = "#,##0.00"
DATE_FMT = "m/d/yyyy"
INT_FMT = "0"

MONEY_KEYS = {k for k, f in MINLEVEL_FORMATS.items() if f == FMT_MONEY}
DATE_KEYS = {k for k, f in MINLEVEL_FORMATS.items() if f == FMT_DATE}
INT_KEYS = {k for k, f in MINLEVEL_FORMATS.items() if f == FMT_INT}


def _resolve_columns(ws) -> list:
    """Map field keys to column letters from the header row. Returns the list of
    expected headers that were NOT found (so the caller can surface them)."""
    COL.clear()
    label_to_key = {label.strip().lower(): key for key, label in HEADERS.items()}
    for cell in ws[1]:
        key = label_to_key.get(str(cell.value or "").strip().lower())
        if key:
            COL[key] = cell.column_letter
    return [HEADERS[k] for k in HEADERS if k not in COL]


def _put(ws, row: int, key: str, value, fmt: Optional[str] = None) -> None:
    letter = COL.get(key)
    if letter is None:          # header not present in this workbook — skip
        return
    cell = ws[f"{letter}{row}"]
    cell.value = value
    if fmt and value is not None:
        cell.number_format = fmt


def _cell_format(key: str) -> Optional[str]:
    fmt = MINLEVEL_FORMATS.get(key)
    if fmt == FMT_MONEY:
        return MONEY_FMT
    if fmt == FMT_DATE:
        return DATE_FMT
    if fmt == FMT_INT:
        return INT_FMT
    return None


def _write_result(ws, row: int, result) -> None:
    """Write one PolicyResult's values into the row (clearing solve cells first
    so bypassed/error rows never keep stale solve values)."""
    for key in MINLEVEL_SOLVE_KEYS:
        _put(ws, row, key, None)
    for key, value in result.values.items():
        if key.startswith("_"):
            continue
        _put(ws, row, key, value, _cell_format(key))


def _ensure_output_columns(ws) -> None:
    """Append every known output header missing from row 1."""
    _resolve_columns(ws)
    next_col = ws.max_column + 1
    for key, label in HEADERS.items():
        if key in COL:
            continue
        cell = ws.cell(row=1, column=next_col)
        cell.value = label
        COL[key] = cell.column_letter
        next_col += 1


def _json_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _row_record(ws, row: int, *, workbook: str, region: str, extras: Optional[dict] = None) -> dict:
    record = {
        "workbook": workbook,
        "sheet": ws.title,
        "row": row,
        "region": region,
        "company": _json_value(ws[f"A{row}"].value),
        "policy": _json_value(ws[f"B{row}"].value),
        "columns": {},
    }
    for key, label in HEADERS.items():
        letter = COL.get(key)
        record["columns"][label] = _json_value(ws[f"{letter}{row}"].value) if letter else None
    if extras:
        record.update(extras)
    return record


def _default_sidecar_path(workbook_path: str, sheet: str, start_policy: str, start_row: Optional[int], limit) -> Path:
    workbook = Path(workbook_path)
    suffix_parts = [sheet]
    if start_policy:
        suffix_parts.append(start_policy)
    elif start_row:
        suffix_parts.append(f"row{start_row}")
    if limit:
        suffix_parts.append(f"limit{limit}")
    suffix_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    suffix = ".".join(re.sub(r"[^A-Za-z0-9_.-]+", "_", str(part)) for part in suffix_parts)
    return workbook.with_name(f"{workbook.stem}.{suffix}.results.jsonl")


def _write_sidecar(handle, record: dict) -> None:
    if handle is None:
        return
    handle.write(json.dumps(record, default=str) + "\n")
    handle.flush()
    os.fsync(handle.fileno())


def _replay_value(key: str, value):
    if value in (None, ""):
        return None
    if key in DATE_KEYS and isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value
    return value


def _replay_format(key: str) -> Optional[str]:
    if key in MONEY_KEYS:
        return MONEY_FMT
    if key in DATE_KEYS:
        return DATE_FMT
    if key in INT_KEYS:
        return INT_FMT
    return None


def _replay_sidecar(ws, sidecar_path: Path) -> dict:
    _ensure_output_columns(ws)
    label_to_key = {label: key for key, label in HEADERS.items()}
    counts = {}
    rows = []
    with sidecar_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            row = int(record["row"])
            columns = record.get("columns") or {}
            for label, value in columns.items():
                key = label_to_key.get(label)
                if key is None:
                    continue
                _put(ws, row, key, _replay_value(key, value), _replay_format(key))
            status = str(columns.get(HEADERS["run_status"]) or "").strip() or "(blank)"
            counts[status] = counts.get(status, 0) + 1
            rows.append(row)
    return {
        "sidecar": str(sidecar_path),
        "processed": len(rows),
        "start_row": min(rows) if rows else None,
        "end_row": max(rows) if rows else None,
        "status_counts": counts,
    }


def _parse_args(argv: list) -> dict:
    """Two call styles: a single JSON blob, or a workbook path + --flags.

    PowerShell mangles embedded JSON quotes, so the friendly style is:
        run_..._batch.py "path\\to\\workbook.xlsx" [--region CKPR] [--sheet S]
                         [--limit N] [--rows 2,3,4] [--dry-run]
    """
    if not argv:
        return {}
    if argv[0].lstrip().startswith("{"):
        return json.loads(argv[0])

    cmd = {"workbook": argv[0]}
    i = 1
    while i < len(argv):
        flag = argv[i]
        if flag == "--dry-run":
            cmd["dry_run"] = True
            i += 1
            continue
        if flag == "--summarize-only":
            cmd["summarize_only"] = True
            i += 1
            continue
        value = argv[i + 1] if i + 1 < len(argv) else None
        if flag == "--region":
            cmd["region"] = value
        elif flag == "--sheet":
            cmd["sheet"] = value
        elif flag == "--limit":
            cmd["limit"] = int(value)
        elif flag == "--start-policy":
            cmd["start_policy"] = value
        elif flag == "--start-row":
            cmd["start_row"] = int(value)
        elif flag == "--rows":
            cmd["rows"] = [int(r) for r in str(value).replace(",", " ").split()]
        elif flag == "--sidecar":
            cmd["sidecar"] = value
        elif flag == "--replay-sidecar":
            cmd["replay_sidecar"] = value
        else:
            print(json.dumps({"error": f"unknown flag: {flag}"}))
            sys.exit(1)
        i += 2
    return cmd


def main() -> None:
    cmd = _parse_args(sys.argv[1:])
    workbook_path = cmd.get("workbook")
    if not workbook_path:
        print(json.dumps({"error": "missing required arg: workbook path"}))
        sys.exit(1)

    region = cmd.get("region", "CKPR")
    explicit_rows = cmd.get("rows")
    start_policy = str(cmd.get("start_policy") or "").strip().upper()
    start_row = cmd.get("start_row")
    limit = cmd.get("limit")
    dry_run = bool(cmd.get("dry_run", False))
    summarize_only = bool(cmd.get("summarize_only", False))
    sidecar_arg = cmd.get("sidecar")
    replay_sidecar = cmd.get("replay_sidecar")

    from suiteview.illustration.core.batch_runner import run_min_level_policy
    from suiteview.illustration.core.calc_engine import IllustrationEngine

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[cmd["sheet"]] if cmd.get("sheet") else wb[wb.sheetnames[0]]

    if replay_sidecar:
        replay_result = _replay_sidecar(ws, Path(replay_sidecar))
        saved = False
        if not dry_run:
            try:
                wb.save(workbook_path)
                saved = True
            except PermissionError:
                print(json.dumps({
                    "error": "Could not save — the workbook is open in Excel. "
                             "Close it and re-run.",
                    "workbook": workbook_path,
                    **replay_result,
                }, indent=2))
                sys.exit(1)
        print(json.dumps({
            "workbook": workbook_path,
            "sheet": ws.title,
            "saved": saved,
            "dry_run": dry_run,
            **replay_result,
        }, indent=2))
        return

    if summarize_only:
        _resolve_columns(ws)
    else:
        _ensure_output_columns(ws)

    # Collect candidate data rows (policy number present in col B, below header).
    data_rows = [
        r for r in range(2, ws.max_row + 1)
        if str(ws[f"B{r}"].value or "").strip()
    ]
    if explicit_rows:
        wanted = {int(r) for r in explicit_rows}
        data_rows = [r for r in data_rows if r in wanted]
    else:
        if start_policy:
            matches = [
                r for r in data_rows
                if str(ws[f"B{r}"].value or "").strip().upper() == start_policy
            ]
            if not matches:
                print(json.dumps({
                    "error": f"start policy not found: {start_policy}",
                    "workbook": workbook_path,
                    "sheet": ws.title,
                }, indent=2))
                sys.exit(1)
            start_row = matches[0]
        if start_row:
            data_rows = [r for r in data_rows if r >= int(start_row)]
        if limit:
            data_rows = data_rows[: int(limit)]

    if summarize_only:
        status_col = COL.get("run_status")
        if status_col is None:
            print(json.dumps({"error": "Run Status header not found"}, indent=2))
            sys.exit(1)
        counts = {}
        for row in data_rows:
            status = str(ws[f"{status_col}{row}"].value or "").strip()
            status = status or "(blank)"
            counts[status] = counts.get(status, 0) + 1
        print(json.dumps({
            "workbook": workbook_path,
            "sheet": ws.title,
            "start_row": data_rows[0] if data_rows else None,
            "end_row": data_rows[-1] if data_rows else None,
            "processed": len(data_rows),
            "status_counts": counts,
        }, indent=2))
        return

    sidecar_path = None
    sidecar_handle = None
    if not dry_run or sidecar_arg:
        sidecar_path = Path(sidecar_arg) if sidecar_arg else _default_sidecar_path(
            workbook_path, ws.title, start_policy, start_row, limit)
        sidecar_path.parent.mkdir(parents=True, exist_ok=True)
        sidecar_handle = sidecar_path.open("w", encoding="utf-8")
        print(f"Writing per-row sidecar JSONL: {sidecar_path}")

    def persist_row(row: int, extras: Optional[dict] = None) -> None:
        _write_sidecar(
            sidecar_handle,
            _row_record(ws, row, workbook=workbook_path, region=region, extras=extras),
        )

    engine = IllustrationEngine()
    results = []
    try:
        for row in data_rows:
            company = str(ws[f"A{row}"].value or "").strip() or None
            policy_number = str(ws[f"B{row}"].value or "").strip()

            result = run_min_level_policy(
                policy_number, company=company, region=region, engine=engine)
            _write_result(ws, row, result)

            values = result.values
            abs_max = values.get("abs_max")
            record = {"row": row, "policy": policy_number,
                      "status": result.status,
                      "abs_max": abs_max,
                      "abs_max_av": values.get("abs_max_av")}
            if result.status in (STATUS_MATURED, STATUS_EXCEPT):
                record.update({
                    "plancode": values.get("plancode"),
                    "form": values.get("form"),
                    "riders": values.get("riders"),
                    "md_diff": values.get("md_diff"),
                    "mode": values.get("_mode"),
                    "min_level_prem": values.get("min_prem"),
                    "total_premium_paid": values.get("total_prem_paid"),
                    "lumpsum_to_next": values.get("lumpsum"),
                    "enters_exception": values.get("_enters_exception"),
                    "exception_date": str(values.get("exc_date")) if values.get("exc_date") else None,
                    "exception_duration": values.get("exc_dur"),
                    "iterations": values.get("_iterations"),
                })
                results.append(record)
                persist_row(row, {
                    "mode": values.get("_mode"),
                    "lumpsum_to_next": values.get("lumpsum"),
                    "enters_exception": values.get("_enters_exception"),
                    "iterations": values.get("_iterations"),
                })
                lumpsum_amount = values.get("lumpsum")
                lump_txt = (f"  lumpsum {lumpsum_amount:>10,.2f}"
                            if lumpsum_amount else "")
                exc_txt = (f"exc {values.get('exc_date')} (yr {values.get('exc_dur')})"
                           if values.get("_enters_exception") else "endows")
                print(f"row {row:>4}  {policy_number:<12}  {values.get('_mode')} "
                      f"min level {values.get('min_prem'):>12,.2f}{lump_txt}  {exc_txt}  "
                      f"total paid {values.get('total_prem_paid'):>12,.2f}  "
                      f"[abs-max: {abs_max}]")
            elif result.status == STATUS_ERROR:
                record.update({"error": result.error,
                               "riders": values.get("riders")})
                results.append(record)
                persist_row(row, {"error": result.error})
                print(f"row {row:>4}  {policy_number:<12}  ERROR: {result.error}")
            else:  # bypass (<codes>)
                record.update({"reasons": result.error,
                               "riders": values.get("riders"),
                               "md_diff": values.get("md_diff")})
                results.append(record)
                persist_row(row, {"reasons": result.error})
                print(f"row {row:>4}  {policy_number:<12}  {result.status}: "
                      f"{result.error}  [abs-max: {abs_max}]")
    finally:
        if sidecar_handle is not None:
            sidecar_handle.close()

    saved = False
    if not dry_run:
        try:
            wb.save(workbook_path)
            saved = True
        except PermissionError:
            print(json.dumps({
                "error": "Could not save — the workbook is open in Excel. "
                         "Close it and re-run.",
                "workbook": workbook_path,
                "processed": len(results),
                "sidecar": str(sidecar_path) if sidecar_path else None,
            }, indent=2))
            sys.exit(1)

    counts = {}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1

    print(json.dumps({
        "workbook": workbook_path,
        "sheet": ws.title,
        "processed": len(results),
        "status_counts": counts,
        "saved": saved,
        "dry_run": dry_run,
        "sidecar": str(sidecar_path) if sidecar_path else None,
        "results": results,
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
