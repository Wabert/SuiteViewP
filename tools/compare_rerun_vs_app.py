"""Orchestrator: run RERUN Saved Cases + the SuiteView engine, compare in one workbook.

For each specified Saved Case this tool:
  1. Loads the case into a temp copy of RERUN via Excel COM, recalcs, and reads
     the **Debug File** sheet (RERUN's curated ~45-field monthly projection).
  2. Runs the SuiteView illustration engine on the SAME case (local SQLite,
     SUITEVIEW_LOCAL_DATA=1), reusing tools/run_engine_case.py.
  3. Aligns the two by month (RERUN ``vID`` == engine ``duration``) and writes a
     comparison workbook: one sheet per case with RERUN | App | Δ triplets per
     Debug File field (out-of-tolerance Δ highlighted), plus a Summary sheet.

RERUN is fully self-contained (Saved Cases + Rates_Control), so this runs on any
machine with Excel — no product/rate DB needed.  Mapping of Debug File columns to
engine fields lives in tools/rerun_debug_map.py (refine there, not here).

Usage (optional single JSON arg; all keys optional):
    venv\\Scripts\\python.exe tools/compare_rerun_vs_app.py
    venv\\Scripts\\python.exe tools/compare_rerun_vs_app.py '{"cases":[1,2,3,4],"months":750}'

    {"workbook": "docs/Illustration_UL/RERUN (v20.0).xlsm",
     "cases":    [1,2,3,4],          # case numbers or CaseID strings (default: 4 baselines)
     "months":   750,                # months to project/compare (engine caps at maturity)
     "company":  "01", "region": "CKPR",
     "out_dir":  "<Documents>/SuiteView_DevTest",
     "open":     true}               # open the saved .xlsx in Excel when done
"""
from __future__ import annotations

import datetime
import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TOOLS = Path(__file__).resolve().parent
for _p in (str(ROOT), str(TOOLS)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

# Must be set before importing engine modules (run_engine_case sets it too).
os.environ.setdefault("SUITEVIEW_LOCAL_DATA", "1")

from rerun_debug_map import (  # noqa: E402
    DEBUG_COLUMNS, KIND_TOL, LABEL_ROW, LAST_COL,
)

DEFAULT_WORKBOOK = ROOT / "docs" / "Illustration_UL" / "RERUN (v20.0).xlsm"
DEFAULT_CASES = [1, 2, 3, 4]
DEFAULT_MONTHS = 750
DEFAULT_OUT_DIR = Path.home() / "Documents" / "SuiteView_DevTest"

# Columns shown as the single identity block (not triplets).
IDENTITY_COLS = {"A", "B", "C"}
# Blank spacer columns in the sheet.
SKIP_COLS = {c["col"] for c in DEBUG_COLUMNS if not c["label"]}


# ── small coercion helpers ──────────────────────────────────────────────────

def _as_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t")


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _engine_value(row: dict, engine):
    """Resolve a Debug-map engine spec (name | [names to sum] | None) on a row."""
    if engine is None:
        return None
    if isinstance(engine, (list, tuple)):
        total = 0.0
        for f in engine:
            n = _num(row.get(f))
            if n is None:
                return None
            total += n
        return total
    return _num(row.get(engine))


def _excel_date(v):
    """Coerce a Saved-Cases date cell (Excel serial float or datetime) to date."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    n = _num(v)
    if n is None:
        return None
    return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)).date()


def _anniv(issue: datetime.date, year: int) -> datetime.date:
    """Anniversary date starting policy year `year` (1-based)."""
    y = issue.year + year - 1
    try:
        return issue.replace(year=y)
    except ValueError:  # Feb 29 issue
        return issue.replace(year=y, day=28)


def _app_cmd_from_case(pairs, months, company, region):
    """Translate a Saved Case's inputs into a run_engine_case command.

    Baselines: policy + billable premium + GPT/TAMRA/exact-days toggles.  Premium
    overrides REPLACE the billed premium from year 1 on (mirrors RERUN's vector).
    Scenario cases: transitions in the face/DBO vectors become dated
    PolicyChangeEvents at the anniversary starting that policy year; non-zero
    withdrawal-vector rows become dated withdrawals at the same anniversary
    (RERUN takes withdrawals at BOY).
    """
    d = {}
    for name, vals in pairs:
        d.setdefault(name, vals)

    def first(name):
        v = d.get(name)
        return v[0] if v else None

    policy = str(first("sCyberlifePolicyNumber") or "").strip()
    prem = _num(first("sINPUT_BillablePrem")) or 0.0
    mode = str(first("sINPUT_BillingMode") or "M").strip() or "M"
    cmd = {
        "policy": policy, "company": company, "region": region, "months": months,
        "tamra": _as_bool(first("sINPUT_TAMRA_Force")),
        "tefra": _as_bool(first("sINPUT_TEFRA_Force")),
        "exact_days": _as_bool(first("sINPUT_Exact_Days_Boolean")),
        "premiums": [{"year": 1, "amount": prem, "mode": mode}],
    }

    issue = _excel_date(first("sINPUT_Issue_Date"))
    if issue is not None:
        changes, withdrawals = [], []
        sa = d.get("vINPUT_Specified_Amount") or []
        for i in range(1, len(sa)):
            prev, cur = _num(sa[i - 1]), _num(sa[i])
            if prev is not None and cur is not None and abs(cur - prev) > 1e-6:
                changes.append({"kind": "face_amount", "value": cur,
                                "date": _anniv(issue, i + 1).isoformat()})
        dbo = d.get("vINPUT_DBO") or []
        for i in range(1, len(dbo)):
            prev = str(dbo[i - 1] or "").strip().upper()
            cur = str(dbo[i] or "").strip().upper()
            if prev and cur and cur != prev:
                changes.append({"kind": "db_option", "value": cur,
                                "date": _anniv(issue, i + 1).isoformat()})
        for i, v in enumerate(d.get("vINPUT_Withdrawal") or []):
            amt = _num(v)
            if amt:
                withdrawals.append({"amount": amt,
                                    "date": _anniv(issue, i + 1).isoformat()})
        if changes:
            cmd["changes"] = changes
        if withdrawals:
            cmd["withdrawals"] = withdrawals

    meta = {
        "case_id": str(first("sINPUT_CaseID") or "").strip(),
        "description": str(first("sINPUT_CaseDescription") or "").strip(),
        "policy": policy,
    }
    return cmd, meta


# ── RERUN side (Excel COM) ──────────────────────────────────────────────────

def _read_debug_file(wb, months):
    """Read the Debug File block; return (labels, {vID: {col_letter: value}})."""
    ws = wb.Worksheets("Debug File")
    r_lo, r_hi = LABEL_ROW, LABEL_ROW + months  # label row + `months` data rows
    block = ws.Range(ws.Cells(r_lo, 1), ws.Cells(r_hi, LAST_COL)).Value
    labels = block[0]
    rows = {}
    for raw in block[1:]:
        vid = raw[0] if raw else None
        if vid is None:
            break
        vid = int(round(float(vid)))
        rows[vid] = {c["col"]: (raw[i] if i < len(raw) else None)
                     for i, c in enumerate(DEBUG_COLUMNS)}
    return labels, rows


def run_rerun_side(workbook, cases, months):
    """Open Excel once; for each case load+recalc+read Debug File."""
    from rerun_com import (
        _enable_iteration, _open_excel, _temp_copy, _write_named_range,
        assert_comparison_inputs, read_case_inputs, XL_CALC_MANUAL,
    )

    xl = _open_excel()
    tmp = _temp_copy(str(workbook))
    results = {}
    try:
        _enable_iteration(xl)
        # One workbook OPEN per case: the AV chain is iterative, so an Excel
        # error (#N/A etc.) from one case latches in the circular cells and
        # contaminates every case loaded after it in the same session.
        for case in cases:
            wb = xl.Workbooks.Open(str(tmp), UpdateLinks=0, ReadOnly=False)
            xl.Calculation = XL_CALC_MANUAL  # avoid a recalc storm on each cell write
            pairs, src_col = read_case_inputs(workbook, case)
            written, failed = 0, []
            for name, vals in pairs:
                try:
                    _write_named_range(wb, name, vals)
                    written += 1
                except Exception as exc:  # noqa: BLE001
                    failed.append({"name": name, "error": str(exc)})
            # Comparison policy: every case must run TEFRA-forced and with
            # exact-days OFF; fail loudly before the (expensive) recalc if a
            # stray case says otherwise.
            assert_comparison_inputs(wb, case)
            xl.CalculateFull()
            labels, rrows = _read_debug_file(wb, months)
            # RERUN injects the current shadow account value here at valuation; feed
            # it to the app so both sides seed the shadow from the same value.
            try:
                shadow_seed = wb.Names("sInput_CurrentShadowAV").RefersToRange.Value
            except Exception:  # noqa: BLE001
                shadow_seed = None
            results[case] = {
                "pairs": pairs, "labels": labels, "rows": rrows,
                "shadow_seed": _num(shadow_seed),
                "inputs_written": written, "input_failures": failed,
            }
            wb.Close(SaveChanges=False)
    finally:
        xl.Quit()
        try:
            tmp.unlink()
        except OSError:
            pass
    return results


# ── App side (engine) ───────────────────────────────────────────────────────

def run_app_side(app_cmds):
    from run_engine_case import run_engine_case

    out = {}
    for case, cmd in app_cmds.items():
        res = run_engine_case(cmd)
        by_vid = {}
        snapshot_vid = None
        for row in res["rows"]:
            dur = row.get("duration")
            if dur is None:
                continue
            vid = int(round(float(dur)))
            by_vid[vid] = row
            if row.get("row") == 0:  # row 0 = inforce snapshot (placeholder end-of-month fields)
                snapshot_vid = vid
        out[case] = {"summary": res["summary"], "by_vid": by_vid,
                     "snapshot_vid": snapshot_vid,
                     "has_shadow": bool(res["summary"].get("has_shadow_account"))}
    return out


# ── Workbook output (openpyxl) ──────────────────────────────────────────────

def _label_mismatches(labels):
    out = []
    for i, c in enumerate(DEBUG_COLUMNS):
        if not c["label"]:
            continue
        got = labels[i] if i < len(labels) else None
        if str(got or "").strip() != c["label"]:
            out.append({"col": c["col"], "expected": c["label"], "got": got})
    return out


def _build_workbook(cases, rerun, app, months, case_meta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    VAL_FMT, RATE_FMT, INT_FMT = "#,##0.00", "0.0000000", "0"
    fmt_for = {"val": VAL_FMT, "rate": RATE_FMT, "int": INT_FMT, "key": INT_FMT}

    blue = PatternFill("solid", fgColor="1E5BA8")
    sub_rerun = PatternFill("solid", fgColor="D9E1F2")
    sub_app = PatternFill("solid", fgColor="E2EFDA")
    sub_delta = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F4CCCC")
    green = PatternFill("solid", fgColor="D9EAD3")
    grey = PatternFill("solid", fgColor="EFEFEF")
    snap_fill = PatternFill("solid", fgColor="DDEBF7")  # inforce snapshot row: shown, not scored
    noshadow_fill = PatternFill("solid", fgColor="FCE4D6")  # CCV col on a no-shadow policy: shown, not scored
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Fields that get RERUN|App|Δ triplets (everything meaningful except identity).
    tri_fields = [c for c in DEBUG_COLUMNS
                  if c["col"] not in IDENTITY_COLS and c["col"] not in SKIP_COLS]

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_records = []  # (case, field, max_abs, vid_at_max, tol, status)
    case_status = []

    for case in cases:
        meta = case_meta[case]
        rrows = rerun[case]["rows"]
        avid = app[case]["by_vid"]
        title = f"Case{case}_{meta['policy']}"[:31]
        ws = wb.create_sheet(title=title)

        # Banner
        desc = f" — {meta['description']}" if meta["description"] else ""
        shadow_note = ("   |   CCV cols NOT scored (no shadow account)"
                       if not app[case].get("has_shadow", False) else "")
        ws.cell(row=1, column=1,
                value=(f"Case {case}: {meta['policy']}{desc}   |   align: RERUN vID = "
                       f"engine duration   |   Δ = App − RERUN   |   red = |Δ| > tol   "
                       f"|   grey = RERUN-only (no app field)   "
                       f"|   blue row = inforce snapshot (shown, not scored)"
                       f"{shadow_note}")).font = bold

        # Header rows 2 (group) + 3 (sub).  Identity block first.
        ws.cell(row=2, column=1, value="vID").fill = blue
        ws.cell(row=2, column=1).font = white_bold
        for off, lbl in enumerate(("vID", "Year", "Month")):
            cc = ws.cell(row=3, column=1 + off, value=lbl)
            cc.font = white_bold
            cc.fill = blue
            cc.alignment = center
        col = 4
        field_cols = {}  # df col letter -> (rerun_col, app_col, delta_col)
        for c in tri_fields:
            ref_only = c["engine"] is None
            grp = ws.cell(row=2, column=col,
                          value=c["label"] + (" (ref)" if ref_only else ""))
            grp.font = white_bold
            grp.fill = blue
            grp.alignment = center
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
            for j, (txt, fill) in enumerate((("RERUN", sub_rerun), ("App", sub_app), ("Δ", sub_delta))):
                sc = ws.cell(row=3, column=col + j, value=txt)
                sc.font = bold
                sc.fill = fill
                sc.alignment = center
            field_cols[c["col"]] = (col, col + 1, col + 2)
            col += 3
        last_col = col - 1

        # Data rows, one per RERUN vID (the reference set), aligned to the app.
        per_field_max = {c["col"]: [0.0, None] for c in tri_fields}  # max_abs, vid
        snap_vid = app[case].get("snapshot_vid")
        has_shadow = app[case].get("has_shadow", False)  # benefit type "A"
        out_row = 4
        for vid in sorted(rrows):
            rrow = rrows[vid]
            arow = avid.get(vid)
            is_snap = (vid == snap_vid)
            ws.cell(row=out_row, column=1, value=vid).number_format = INT_FMT
            ws.cell(row=out_row, column=2,
                    value=_num(rrow.get("B"))).number_format = INT_FMT
            ws.cell(row=out_row, column=3,
                    value=_num(rrow.get("C"))).number_format = INT_FMT
            for c in tri_fields:
                rc, ac, dc = field_cols[c["col"]]
                fmt = fmt_for.get(c["kind"], VAL_FMT)
                rv = _num(rrow.get(c["col"]))
                ws.cell(row=out_row, column=rc, value=rv).number_format = fmt
                if c["engine"] is None:
                    ws.cell(row=out_row, column=ac).fill = grey
                    ws.cell(row=out_row, column=dc).fill = grey
                    continue
                # CCV columns on a policy with no shadow account: RERUN computes a
                # spurious load on nothing; show both sides but don't score.
                skip_shadow = (not has_shadow) and c["label"].startswith("CCV")
                av = _engine_value(arow, c["engine"]) if arow is not None else None
                ws.cell(row=out_row, column=ac, value=av).number_format = fmt
                if rv is not None and av is not None:
                    delta = av - rv
                    dcell = ws.cell(row=out_row, column=dc, value=delta)
                    dcell.number_format = fmt
                    tol = KIND_TOL.get(c["kind"], 0.01)
                    if skip_shadow:
                        dcell.fill = noshadow_fill
                    elif is_snap:
                        dcell.fill = snap_fill  # snapshot row: shown but not scored
                    else:
                        dcell.fill = red if abs(delta) > tol else green
                        if abs(delta) > per_field_max[c["col"]][0]:
                            per_field_max[c["col"]] = [abs(delta), vid]
                elif arow is None:
                    ws.cell(row=out_row, column=dc, value="no app row").fill = grey
            out_row += 1

        ws.freeze_panes = "D4"
        ws.column_dimensions["A"].width = 6
        for ci in range(2, last_col + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 12

        # Roll up per-case status.
        fails = []
        for c in tri_fields:
            if c["engine"] is None:
                continue
            if (not has_shadow) and c["label"].startswith("CCV"):
                continue  # no shadow account → CCV differences ignored
            mx, vid_at = per_field_max[c["col"]]
            tol = KIND_TOL.get(c["kind"], 0.01)
            status = "FAIL" if mx > tol else "ok"
            summary_records.append((case, meta["policy"], c["label"], mx, vid_at, tol, status))
            if status == "FAIL":
                fails.append((c["label"], mx, vid_at))
        worst = max(fails, key=lambda x: x[1]) if fails else None
        case_status.append({
            "case": case, "policy": meta["policy"], "desc": meta["description"],
            "vids": len(rrows), "fail_fields": len(fails),
            "worst": worst, "ok": not fails,
        })

    # ── Summary sheet ──
    sw = summary_ws
    sw.cell(row=1, column=1, value="RERUN vs SuiteView App — comparison summary").font = Font(bold=True, size=13)
    sw.cell(row=2, column=1, value=f"Generated {datetime.datetime.now():%Y-%m-%d %H:%M:%S}  ·  months={months}")
    hdr = ["Case", "Policy", "Description", "vIDs", "Fail fields", "Worst field", "Worst |Δ|", "@vID", "Status"]
    for j, h in enumerate(hdr, 1):
        cc = sw.cell(row=4, column=j, value=h)
        cc.font = white_bold
        cc.fill = blue
    r = 5
    for cs in case_status:
        w = cs["worst"]
        sw.cell(row=r, column=1, value=cs["case"])
        sw.cell(row=r, column=2, value=cs["policy"])
        sw.cell(row=r, column=3, value=cs["desc"])
        sw.cell(row=r, column=4, value=cs["vids"])
        sw.cell(row=r, column=5, value=cs["fail_fields"])
        sw.cell(row=r, column=6, value=(w[0] if w else "—"))
        sw.cell(row=r, column=7, value=(w[1] if w else 0)).number_format = "#,##0.0000"
        sw.cell(row=r, column=8, value=(w[2] if w else "—"))
        st = sw.cell(row=r, column=9, value=("PASS" if cs["ok"] else "FAIL"))
        st.fill = green if cs["ok"] else red
        st.font = bold
        r += 1

    # Per-field detail below.
    r += 1
    sw.cell(row=r, column=1, value="Per-field detail (max |Δ| over all months)").font = bold
    r += 1
    for j, h in enumerate(("Case", "Policy", "Field", "Max |Δ|", "@vID", "Tol", "Status"), 1):
        cc = sw.cell(row=r, column=j, value=h)
        cc.font = white_bold
        cc.fill = blue
    r += 1
    for rec in summary_records:
        case, policy, field, mx, vid_at, tol, status = rec
        sw.cell(row=r, column=1, value=case)
        sw.cell(row=r, column=2, value=policy)
        sw.cell(row=r, column=3, value=field)
        sw.cell(row=r, column=4, value=mx).number_format = "#,##0.0000000"
        sw.cell(row=r, column=5, value=(vid_at if vid_at is not None else "—"))
        sw.cell(row=r, column=6, value=tol).number_format = "0.0000000"
        sc = sw.cell(row=r, column=7, value=status)
        sc.fill = red if status == "FAIL" else green
        r += 1
    for ci, wdt in enumerate((6, 11, 26, 14, 8, 12, 10), 1):
        sw.column_dimensions[get_column_letter(ci)].width = wdt
    sw.freeze_panes = "A5"
    return wb, case_status, summary_records


# ── main ────────────────────────────────────────────────────────────────────

def main():
    cmd = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    workbook = Path(cmd.get("workbook") or DEFAULT_WORKBOOK)
    cases = cmd.get("cases") or DEFAULT_CASES
    months = int(cmd.get("months", DEFAULT_MONTHS))
    company = cmd.get("company", "01")
    region = cmd.get("region", "CKPR")
    out_dir = Path(cmd.get("out_dir") or DEFAULT_OUT_DIR)
    do_open = cmd.get("open", True)

    if not workbook.exists():
        print(json.dumps({"ok": False, "error": f"workbook not found: {workbook}"}))
        return 1

    # 1) RERUN side (Excel COM) — also yields each case's inputs for the app side.
    rerun = run_rerun_side(workbook, cases, months)

    # 2) Build app commands from the Saved Case inputs, then run the engine.
    app_cmds, case_meta = {}, {}
    for case in cases:
        app_cmds[case], case_meta[case] = _app_cmd_from_case(
            rerun[case]["pairs"], months, company, region)
        seed = rerun[case].get("shadow_seed")
        if seed:
            app_cmds[case]["shadow_av"] = seed
    app = run_app_side(app_cmds)

    # 3) Compare + build workbook.
    wb, case_status, summary_records = _build_workbook(
        cases, rerun, app, months, case_meta)

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"rerun_vs_app_{ts}.xlsx"
    wb.save(str(out_path))

    if do_open:
        try:
            os.startfile(str(out_path))  # noqa: SLF001 (Windows)
        except Exception:  # noqa: BLE001
            pass

    label_warnings = {case: _label_mismatches(rerun[case]["labels"]) for case in cases}
    fails_by_case = {}
    for (case, _policy, field, mx, vid_at, tol, status) in summary_records:
        if status == "FAIL":
            fails_by_case.setdefault(case, []).append(
                {"field": field, "abs_delta": round(mx, 6), "vid": vid_at, "tol": tol})
    print(json.dumps({
        "ok": True,
        "out_path": str(out_path),
        "cases": cases,
        "months": months,
        "case_status": [
            {"case": c["case"], "policy": c["policy"], "desc": c["desc"],
             "vids": c["vids"], "fail_fields": c["fail_fields"],
             "fails": fails_by_case.get(c["case"], []),
             "ok": c["ok"]}
            for c in case_status
        ],
        "label_warnings": {str(k): v for k, v in label_warnings.items() if v},
        "app_inputs": {str(k): app_cmds[k] for k in cases},
    }, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
