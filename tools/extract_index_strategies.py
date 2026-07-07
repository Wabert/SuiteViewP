"""Port the RERUN IUL index-strategy tables to the illustration app's JSON bundle.

Reads the "Illustration Values" sheet of the RERUN workbook:
  - mIndex_Illustrated_Rates   (BS78:CF94) — per-plancode, per-strategy AG49
    maximum illustrated rates (the INPUT "Available Rate" lookup)
  - mIndex_Strategy_Parameters (BS52:CB68) — current caps / participation
    per plancode per strategy (informational)
  - Multiplier constants       (BY70:BZ72) — IP/IR asset fees, multipliers,
    maximum asset-charge fees
and writes suiteview/illustration/plancodes/index_strategies.json.

The MB column is folded into M1: MB ("low volatility") was never implemented —
M1 (MARC5) is the low-volatility strategy in use, and RERUN's IUL21/23/GIUL2x
rows carry their only nonzero index rate in the MB column.

Usage:
    venv\\Scripts\\python.exe tools/extract_index_strategies.py
    venv\\Scripts\\python.exe tools/extract_index_strategies.py '{"workbook": "path/to/RERUN.xlsm"}'
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT / "docs" / "Illustration_UL" / "RERUN (v20.0).xlsm"
OUT_PATH = ROOT / "suiteview" / "illustration" / "plancodes" / "index_strategies.json"

SHEET = "Illustration Values"

# Ordered as on the RERUN INPUT allocations block (U1 first, M1 last).
# column = mIndex_Illustrated_Rates column letter on "Illustration Values".
STRATEGIES = [
    {"fund_id": "U1", "label": "Fixed Strategy", "column": "CD"},
    {"fund_id": "IS", "label": "1 Yr PtP w/ Specified Rate", "column": "BW"},
    {"fund_id": "IX", "label": "1 Yr PtP w/ Cap", "column": "BU"},
    {"fund_id": "IC", "label": "1 Yr PtP 1.5% Floor w/ Cap", "column": "BX"},
    {"fund_id": "IF", "label": "1 Yr PtP Uncapped", "column": "BV"},
    {"fund_id": "IP", "label": "Index w/ Low Multiplier", "column": "BY"},
    {"fund_id": "IR", "label": "Index w/ High Multiplier", "column": "BZ"},
    {"fund_id": "NX", "label": "NASDAQ-100 1 Yr PtP w/ Cap", "column": "CA"},
    {"fund_id": "M1", "label": "MARC5 1 Yr PtP w/ Participation", "column": "CB"},
]
MB_COLUMN = "CC"  # folded into M1

RATES_ROWS = range(78, 95)      # mIndex_Illustrated_Rates data rows
PARAMS_ROWS = range(52, 69)     # mIndex_Strategy_Parameters data rows
PLANCODE_COL = "BS"
PRODUCT_COL = "BR"


def _num(cell) -> float | None:
    value = cell.value
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _text(cell) -> str:
    return str(cell.value or "").strip()


def _read_block(ws, rows) -> dict:
    """{plancode: {fund_id: rate}} for one table block, MB folded into M1."""
    out: dict[str, dict] = {}
    for row in rows:
        plancode = _text(ws[f"{PLANCODE_COL}{row}"])
        if not plancode:
            continue
        rates = {}
        for strat in STRATEGIES:
            value = _num(ws[f"{strat['column']}{row}"])
            if value is not None:
                rates[strat["fund_id"]] = value
        mb = _num(ws[f"{MB_COLUMN}{row}"])
        if mb:  # nonzero MB supersedes an absent/zero M1
            if not rates.get("M1"):
                rates["M1"] = mb
        if rates:
            out[plancode] = {"product": _text(ws[f"{PRODUCT_COL}{row}"]), "rates": rates}
    return out


def main():
    workbook = DEFAULT_WORKBOOK
    if len(sys.argv) > 1:
        cmd = json.loads(sys.argv[1])
        workbook = Path(cmd.get("workbook") or DEFAULT_WORKBOOK)

    wb = openpyxl.load_workbook(workbook, read_only=False, data_only=True)
    ws = wb[SHEET]

    illustrated = _read_block(ws, RATES_ROWS)
    parameters = _read_block(ws, PARAMS_ROWS)

    multiplier = {
        "IP": {
            "asset_charge": _num(ws["BY70"]),
            "multiplier": _num(ws["BY71"]),
            "max_asset_charge": _num(ws["BY72"]),
        },
        "IR": {
            "asset_charge": _num(ws["BZ70"]),
            "multiplier": _num(ws["BZ71"]),
            "max_asset_charge": _num(ws["BZ72"]),
        },
    }
    wb.close()

    plancodes = {}
    for plancode, entry in illustrated.items():
        plancodes[plancode] = {
            "product": entry["product"],
            "illustrated_rates": entry["rates"],
            "strategy_parameters": parameters.get(plancode, {}).get("rates", {}),
        }

    payload = {
        "_source": f"{workbook.name} — '{SHEET}' mIndex_Illustrated_Rates / "
                   "mIndex_Strategy_Parameters (MB column folded into M1)",
        "strategies": STRATEGIES,
        "multiplier_strategies": multiplier,
        # RERUN Rates_Control CP79/CP80: index = MAX(2, date-based tier);
        # variable-loan credit spread by AG49 index (1-based CHOOSE list).
        "ag49": {
            "default_index": 2,
            "loan_credit_spread_by_index": [0.0, 0.01, 0.005, 0.005],
        },
        "plancodes": plancodes,
    }

    OUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "out_file": str(OUT_PATH),
        "plancode_count": len(plancodes),
        "plancodes": sorted(plancodes),
    }, indent=2))


if __name__ == "__main__":
    main()
