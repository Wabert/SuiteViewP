"""Introspect the RERUN CalcEngine workbook: sheets, column map, formulas.

Usage:
    venv\\Scripts\\python.exe tools/extract_calcengine.py '<json>'

Commands (single JSON arg):
  {"mode":"sheets","workbook":"<path>"}
      -> list every sheet with max_row / max_col

  {"mode":"map","workbook":"<path>","sheet":"CalcEngine",
   "header_rows":[3,4,5],"formula_row":12,"value_rows":[12],
   "max_col":820,"out_file":"<path.tsv>","keywords":["glp","force","loan"]}
      -> write full per-column map (idx, letter, header, formula, values)
         to out_file as TSV; print only columns whose header/formula match
         any keyword (so the token cost stays small).

  {"mode":"cols","workbook":"<path>","sheet":"CalcEngine",
   "cols":["A","F","UL"],"header_rows":[5],"min_row":6,"max_row":20}
      -> for the given columns, dump header + per-row formula and value
         across the row range (trace a calc down the months).
"""
import sys
import json

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def _clean(v):
    if v is None:
        return ""
    return str(v).replace("\r", " ").replace("\n", " ").strip()


def _load_rows(path, sheet, needed_rows, max_col, data_only):
    """Return {row_index: [values...]} for needed rows, one streamed pass."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    ws = wb[sheet]
    ncols = max_col or ws.max_column
    want = set(needed_rows)
    top = max(needed_rows)
    cache = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=top, max_col=ncols), start=1):
        if r in want:
            cache[r] = [c.value for c in row]
    wb.close()
    return cache, ncols


def mode_sheets(cmd):
    wb = openpyxl.load_workbook(cmd["workbook"], read_only=True, data_only=False)
    out = [{"name": ws.title, "max_row": ws.max_row, "max_col": ws.max_column}
           for ws in wb.worksheets]
    wb.close()
    print(json.dumps({"sheets": out}, indent=2))


def mode_map(cmd):
    path = cmd["workbook"]
    sheet = cmd["sheet"]
    header_rows = cmd.get("header_rows", [5])
    formula_row = cmd.get("formula_row", 12)
    value_rows = cmd.get("value_rows", [])
    max_col = cmd.get("max_col")
    out_file = cmd.get("out_file")
    keywords = [k.lower() for k in cmd.get("keywords", [])]

    f_needed = sorted(set(header_rows + [formula_row]))
    fcache, ncols = _load_rows(path, sheet, f_needed, max_col, data_only=False)

    vcache = {}
    if value_rows:
        vcache, _ = _load_rows(path, sheet, value_rows, ncols, data_only=True)

    def cell(cache, r, ci):
        v = cache.get(r, [])
        return v[ci - 1] if ci - 1 < len(v) else None

    lines = []
    matches = []
    for ci in range(1, ncols + 1):
        letter = get_column_letter(ci)
        headers = [_clean(cell(fcache, hr, ci)) for hr in header_rows]
        header_join = " / ".join(h for h in headers if h)
        formula = cell(fcache, formula_row, ci)
        formula_s = _clean(formula)
        vals = [_clean(cell(vcache, vr, ci)) for vr in value_rows]
        lines.append("\t".join([str(ci), letter, header_join, formula_s] + vals))
        blob = (header_join + " " + formula_s).lower()
        if keywords and any(k in blob for k in keywords):
            matches.append({
                "col": ci, "letter": letter, "header": header_join,
                "formula": formula_s, "values": vals,
            })

    if out_file:
        with open(out_file, "w", encoding="utf-8") as fh:
            fh.write("idx\tletter\theader\tformula\t"
                     + "\t".join(f"val_r{r}" for r in value_rows) + "\n")
            fh.write("\n".join(lines) + "\n")

    print(json.dumps({
        "sheet": sheet, "ncols": ncols, "out_file": out_file,
        "header_rows": header_rows, "formula_row": formula_row,
        "match_count": len(matches), "matches": matches,
    }, indent=2, default=str))


def mode_cols(cmd):
    path = cmd["workbook"]
    sheet = cmd["sheet"]
    cols = cmd["cols"]
    header_rows = cmd.get("header_rows", [5])
    min_row = cmd.get("min_row", 6)
    max_row = cmd.get("max_row", 20)
    col_idx = [column_index_from_string(c) for c in cols]

    needed = sorted(set(header_rows + list(range(min_row, max_row + 1))))
    fcache, _ = _load_rows(path, sheet, needed, max(col_idx), data_only=False)
    vcache, _ = _load_rows(path, sheet, needed, max(col_idx), data_only=True)

    def cell(cache, r, ci):
        v = cache.get(r, [])
        return v[ci - 1] if ci - 1 < len(v) else None

    columns = []
    for c, ci in zip(cols, col_idx):
        headers = [_clean(cell(fcache, hr, ci)) for hr in header_rows]
        rows = []
        for r in range(min_row, max_row + 1):
            rows.append({
                "row": r,
                "formula": _clean(cell(fcache, r, ci)),
                "value": _clean(cell(vcache, r, ci)),
            })
        columns.append({"letter": c, "header": " / ".join(h for h in headers if h), "rows": rows})

    print(json.dumps({"sheet": sheet, "columns": columns}, indent=2, default=str))


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "missing JSON arg"}))
        sys.exit(1)
    cmd = json.loads(sys.argv[1])
    mode = cmd.get("mode", "sheets")
    {"sheets": mode_sheets, "map": mode_map, "cols": mode_cols}[mode](cmd)


if __name__ == "__main__":
    main()
