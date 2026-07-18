"""Build a RERUN Saved Case column for a policy from the local offline DBs.

Offline replacement for RERUN's ``mdl_GetCyberlifePolicy.PopulateInputSheet``
(which pulls a policy from DB2 and fills the INPUT sheet): the same field
mapping is computed in Python from ``build_illustration_data()`` (which reads
bundled_data/dev/policy_records.sqlite when SUITEVIEW_LOCAL_DATA=1) and written
as a NEW column on the ``Saved Cases`` sheet, so the existing
rerun_com/compare_rerun_vs_app pipeline can load it like any hand-saved case.

Fields RERUN sources from its own UI/options (TAMRA/TEFRA toggles, exact-days,
loan behaviour switches, SigTerm durations, ...) are inherited from a template
case column and listed in the report under "inherited" — nothing is filled
silently. Fields this builder cannot source offline are also reported.

Usage (single JSON arg):
    venv\\Scripts\\python.exe tools/rerun_build_case_inputs.py '<json>'

    {"policy": "UE013383",
     "region": "CKPR", "company": null,
     "workbook": "docs/Illustration_UL/RERUN (v20.0) local.xlsm",
     "template_case": 1,          # case whose values seed unmapped inputs
     "case_id": null,             # default: policy number
     "description": null,
     "dry_run": false}            # true = print mapping report, no COM write

Note: this writes the *inputs*. Rates for the policy's plancodes must be
loaded separately (tools/rerun_load_local_rates.py) before running the case.
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_WORKBOOK = ROOT / "docs" / "Illustration_UL" / "RERUN (v20.0) local.xlsm"

# RERUN benefit-name translation (cls_PolicyInformation.TranslateBenefitTypeToText).
_BENEFIT_NAMES = {
    "1": "ADB", "2": "ADnD", "3": "PWoC", "4": "PWoT", "7": "GIO",
    "9": "PPB", "A": "CCV", "U": "COLA", "B": "LTC",
}
_ABR_SUBTYPES = {"1": "ABRTM", "2": "ABRCT", "3": "ABRCH",
                 "4": "ABRTM", "5": "ABRCT", "6": "ABRCH", "L": "ABRLN"}
_GCO_SUBTYPES = {"1": "GCO15", "2": "GCO20", "3": "GCO25"}

_BILLING_LETTER = {1: "M", 3: "Q", 6: "S", 12: "A"}

_FUND_IDS = ("U1", "IS", "IX", "IC", "IF", "IP", "IR", "NX", "M1")


def _serial(d) -> object:
    """date -> Excel serial number ('' when None)."""
    if d is None:
        return ""
    return (d - _dt.date(1899, 12, 30)).days


def _benefit_name(ben) -> str:
    btype = str(ben.benefit_type or "").strip()
    bsub = str(ben.benefit_subtype or "").strip()
    if btype == "#":
        return _ABR_SUBTYPES.get(bsub, "ABR" if not bsub else btype + bsub)
    if btype == "V":
        return _GCO_SUBTYPES.get(bsub, btype + bsub)
    return _BENEFIT_NAMES.get(btype, btype)


def _years_between(d0, d1) -> int:
    if d0 is None or d1 is None:
        return 0
    years = d1.year - d0.year
    if (d1.month, d1.day) < (d0.month, d0.day):
        years -= 1
    return max(years, 0)


def build_mapping(data, warnings: list[str], skip_benefits: bool = False) -> dict:
    """RERUN input name -> value, mirroring PopulateInputSheet.

    skip_benefits=True leaves every benefit boolean False (useful because
    RERUN's tBenefitDefinitionFile only carries 1A/1G/1S plancodes — an active
    waiver/GIR/ADB on a plancode absent from that file #N/As the whole run).
    """
    m: dict[str, object] = {}

    m["TimeStamp"] = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    m["sUID"] = "Offline case builder"
    m["sINPUT_CaseDescription"] = ""
    m["sINPUT_CaseID"] = data.policy_number
    m["sCyberlifePolicyNumber"] = data.policy_number

    # ── Plan / demographics ──────────────────────────────────────────
    m["sPlancode"] = data.plancode
    m["sFormNumber"] = data.form_number
    dob = data.insured_birth_date
    if dob is None:
        # Synthesize from issue date - issue age: RERUN derives the issue age of
        # face-INCREASE segments from sINPUT_DOB (CalcEngine FF YEARFRAC), so a
        # blank DOB turns any illustrated increase into a ~168-age lookup #N/A.
        # Exact for ALB products; ANB can be off a year near the half-birthday.
        try:
            dob = data.issue_date.replace(year=data.issue_date.year - int(data.issue_age))
        except ValueError:  # Feb 29
            dob = data.issue_date.replace(
                year=data.issue_date.year - int(data.issue_age), day=28)
        warnings.append(
            f"sINPUT_DOB synthesized as issue_date - issue_age = {dob} "
            "(insured_birth_date not in local export)")
    m["sINPUT_DOB"] = _serial(dob)
    m["sINPUT_Issue_Date"] = _serial(data.issue_date)
    m["sINPUT_Issue_Age"] = data.issue_age
    # 1U135900 rates were keyed with sex "X" though CyberLife says "U" (VBA quirk).
    m["sINPUT_Rate_Sex"] = "X" if data.plancode == "1U135900" else data.rate_sex
    m["sINPUT_TrueSex"] = data.rate_sex  # TODO: verify true sex source for unisex-rated policies
    m["sINPUT_State"] = data.issue_state
    m["sINPUT_Rateclass"] = data.rate_class

    seg1 = data.segments[0] if data.segments else None
    m["sINPUT_Table_Rating"] = seg1.table_rating if seg1 else 0
    m["sINPUT_Flat_1_Amount"] = seg1.flat_extra if seg1 else 0
    m["sINPUT_Flat_1_CeaseAge"] = (
        seg1.issue_age + _years_between(seg1.issue_date, seg1.flat_cease_date)
        if seg1 and seg1.flat_extra and seg1.flat_cease_date else 0)
    m["sINPUT_Flat_2_Amount"] = 0
    m["sINPUT_Flat_2_CeaseAge"] = 0

    m["sINPUT_1035_Exchange_Amount"] = 0
    m["sINPUT_1035_Exchange_Loan"] = 0
    m["sINPUT_1035_Exchange_CostBasis"] = 0
    m["sINPUT_DialToPremium"] = ""
    m["sINPUT_DialToAge"] = 0
    m["sINPUT_PayDuration"] = ""

    # ── Billing / premium ────────────────────────────────────────────
    mode = _BILLING_LETTER.get(int(data.billing_frequency or 1), "M")
    m["sINPUT_BillablePrem"] = data.modal_premium
    m["sINPUT_BillingMode"] = mode
    m["sINPUT_PremiumYTD"] = data.premiums_ytd
    m["sINPUT_PremiumTD"] = data.premiums_paid_to_date
    m["sINPUT_WithdrawalTD"] = data.withdrawals_to_date
    m["sINPUT_CostBasis"] = data.cost_basis
    m["sINPUT_Inforce_Lumpsum"] = 0

    # ── Allocations / IUL rates ──────────────────────────────────────
    # IUL test: the plancode has index strategies (product_type strings vary).
    from suiteview.illustration.models.index_strategies import load_index_strategies

    plan_strat = load_index_strategies(data.plancode)
    is_iul = plan_strat is not None
    allocs = {f: 0.0 for f in _FUND_IDS}
    if is_iul:
        for fund, pct in (data.premium_allocations or {}).items():
            if fund in allocs:
                allocs[fund] = float(pct) / 100.0
    else:
        allocs["U1"] = 1
    for fund in _FUND_IDS:
        m[f"sINPUT_PremAllocation{fund}"] = allocs[fund]

    # Illustrated rates: the app engine's per-plancode defaults, so both sides
    # of a comparison run the same crediting assumption.
    if plan_strat is not None:
        rates = plan_strat.default_rates()
        for fund in _FUND_IDS[1:]:  # U1 is the fixed account
            m[f"sINPUT_{fund}_IllustratedRate"] = rates.get(fund, 0)
        for s in plan_strat.strategies:
            if s.fund_id in ("IP", "IR"):
                m[f"s{s.fund_id}_Multiplier"] = s.multiplier
                m[f"s{s.fund_id}_AssetFee"] = s.asset_charge
    # RERUN defaults the fixed rate to GINT so nothing illustrates too high; for
    # comparison parity we use the current declared rate the app engine credits.
    m["sINPUT_Fixed_Int_Rate"] = data.current_interest_rate

    # ── Loans ────────────────────────────────────────────────────────
    var_ln = data.variable_loan_principal + data.variable_loan_accrued
    fix_ln = (data.regular_loan_principal + data.regular_loan_accrued
              + data.preferred_loan_principal + data.preferred_loan_accrued)
    if var_ln > 0:
        m["sINPUT_Loan_Type"] = "Variable"
    elif fix_ln > 0:
        m["sINPUT_Loan_Type"] = "Fixed"
    m["sINPUT_FixedLn_Principle"] = data.regular_loan_principal
    m["sINPUT_FixedLn_Accrued"] = data.regular_loan_accrued
    m["sINPUT_PreflLn_Principle"] = data.preferred_loan_principal
    m["sINPUT_PreflLn_Accrued"] = data.preferred_loan_accrued
    m["sINPUT_VblLn_Principle"] = data.variable_loan_principal
    m["sINPUT_VblLn_Accrued"] = data.variable_loan_accrued
    if data.variable_loan_charge_rate:
        m["sINPUT_Variable_Loan_Rate"] = data.variable_loan_charge_rate

    # ── Base coverages ───────────────────────────────────────────────
    active_segs = [s for s in data.segments if s.status != "T"]
    total_sa = sum(s.face_amount for s in active_segs)
    for i in (1, 2, 3):
        seg = data.segments[i - 1] if len(data.segments) >= i else None
        live = seg is not None and seg.status != "T"
        m[f"sINPUT_OriginalSA{i}"] = seg.original_face_amount if live else 0
        m[f"sINPUT_CurrentSA{i}"] = seg.face_amount if live else 0
        m[f"sINPUT_Original_Band_Cov_{i}"] = seg.original_band if live else 0
        m[f"sINPUT_MonthsTerminatedCov{i}"] = seg.months_since_terminated if seg else 0
        if i >= 2:
            m[f"sINPUT_Rateclass_Cov_{i}"] = seg.rate_class if live else ""
            m[f"sINPUT_TableRatingSA{i}"] = seg.table_rating if live else 0
            m[f"sINPUT_IssueDateSA{i}"] = _serial(seg.issue_date) if live else 0
            m[f"sINPUT_IssueAgeSA{i}"] = seg.issue_age if live else 0

    # ── Benefits ─────────────────────────────────────────────────────
    for name in ("PW", "PWST", "CCV", "GIR", "ADB"):
        m[f"sINPUT_{name}_Boolean"] = False
        m[f"sINPUT_{name}_BenefitCode"] = ""
    for name in ("ShadowBenefit_Active", "ABRTM_Active", "ABRCT_Active",
                 "ABRCH_Active", "GCO15_Active", "GCO20_Active", "GCO25_Active",
                 "COLA_Active"):
        m[f"sINPUT_{name}"] = False
    m["sINPUT_ADB_Units"] = 0
    m["sInput_CurrentShadowAV"] = 0
    m["sINPUT_CCV_Units"] = ""

    val_date = data.valuation_date
    adb_units = 0.0
    for ben in data.benefits:
        if not ben.is_active:
            continue
        if skip_benefits:
            continue
        if ben.cease_date is not None and val_date is not None and ben.cease_date <= val_date:
            continue
        name = _benefit_name(ben)
        # RERUN benefit codes are TYPE+SUBTYPE (keys into tBenefitDefinitionFile,
        # e.g. plancode & "39" for a PWoC subtype 9).
        ben_code = f"{str(ben.benefit_type).strip()}{str(ben.benefit_subtype).strip()}"
        if name == "GIO":
            m["sINPUT_GIR_Boolean"] = True
            m["sINPUT_GIR_Units"] = ben.benefit_amount / 1000
            m["sINPUT_GIR_BenefitCode"] = ben_code
        elif name == "PWoT":
            m["sINPUT_PWST_Boolean"] = True
            m["sINPUT_PWST_Premium"] = ben.benefit_amount
            m["sINPUT_PWST_BenefitCode"] = ben_code
        elif name == "PWoC":
            m["sINPUT_PW_Boolean"] = True
            m["sINPUT_PW_BenefitCode"] = ben_code
        elif name == "ADB":
            m["sINPUT_ADB_Boolean"] = True
            adb_units += ben.benefit_amount / 1000
            m["sINPUT_ADB_Units"] = adb_units
            m["sINPUT_ADB_BenefitCode"] = ben_code
        elif name == "CCV":
            m["sINPUT_ShadowBenefit_Active"] = True
            m["sINPUT_CCV_BenefitCode"] = ben_code
            m["sInput_CurrentShadowAV"] = data.shadow_account_value
            if data.ccv_active and ben.benefit_amount / 1000 != total_sa / 1000:
                m["sINPUT_CCV_Units"] = ben.benefit_amount / 1000
            from suiteview.illustration.models.plancode_config import load_plancode
            try:
                cfg = load_plancode(data.plancode)
                if str(getattr(cfg, "shadow_availability", "")).strip() == "Rider":
                    m["sINPUT_CCV_Boolean"] = True
            except Exception as exc:
                warnings.append(f"CCV: could not check ShadowAvailability ({exc})")
        elif name in ("ABRTM", "ABRLN"):
            m["sINPUT_ABRTM_Active"] = True
        elif name == "ABRCT":
            m["sINPUT_ABRCT_Active"] = True
        elif name == "ABRCH":
            m["sINPUT_ABRCH_Active"] = True
        elif name == "COLA":
            m["sINPUT_COLA_Active"] = True
        elif name in ("GCO15", "GCO20", "GCO25"):
            m[f"sINPUT_{name}_Active"] = True

    for flag, code_name in (("sINPUT_PW_Boolean", "sINPUT_PW_BenefitCode"),
                            ("sINPUT_PWST_Boolean", "sINPUT_PWST_BenefitCode"),
                            ("sINPUT_GIR_Boolean", "sINPUT_GIR_BenefitCode"),
                            ("sINPUT_ADB_Boolean", "sINPUT_ADB_BenefitCode")):
        if m.get(flag):
            warnings.append(
                f"{flag} set with code {m.get(code_name)!r}: RERUN looks up "
                f"tBenefitDefinitionFile key {data.plancode}{m.get(code_name)} — "
                "a missing key #N/As the run (skip_benefits compares without benefits)")

    # ── Riders ───────────────────────────────────────────────────────
    for i in (1, 2, 3):
        m[f"sINPUT_R{i}_Boolean"] = False
        m[f"sINPUT_R{i}_SigTerm_Boolean"] = False
    m["sINPUT_CTR_Boolean"] = False
    m["sINPUT_APB_Boolean"] = False
    m["sINPUT_APB_Face"] = 0
    m["sINPUT_Original_APB_SA"] = 0

    slot = 0
    ctr_units = 0.0
    apb_face = 0.0
    for rdr in data.riders:
        if not rdr.is_active:
            continue
        cov = (rdr.cov_type or "").strip().upper()
        is_anico_ctr = cov == "CTR" and data.company_code == "01"
        if cov in ("", "0", "LTR", "STR", "SIGTERM") or (cov == "CTR" and not is_anico_ctr):
            slot += 1
            if slot > 3:
                warnings.append(f"rider {rdr.plancode}: more than 3 term riders; slot dropped")
                continue
            m[f"sINPUT_R{slot}_Boolean"] = True
            m[f"sINPUT_R{slot}_Plancode"] = rdr.plancode
            m[f"sINPUT_R{slot}_Issue_Age"] = rdr.issue_age
            m[f"sINPUT_R{slot}_Gender"] = rdr.rate_sex
            m[f"sINPUT_R{slot}_Rateclass"] = rdr.rate_class
            m[f"sINPUT_R{slot}_Table_Rating"] = rdr.table_rating
            m[f"sINPUT_R{slot}_Face"] = rdr.face_amount
            m[f"sINPUT_R{slot}_Primary_Boolean"] = rdr.on_primary_insured
            m[f"sINPUT_R{slot}_QAB_Boolean"] = True  # TODO: verify QAB source offline
            m[f"sINPUT_R{slot}_Flat_Extra"] = rdr.flat_extra
            m[f"sINPUT_R{slot}_Flat_CeaseAge"] = 0
            m[f"sINPUT_R{slot}_IssueDate"] = _serial(rdr.issue_date)
            m[f"sINPUT_R{slot}_SigTerm_Boolean"] = cov == "SIGTERM"
        elif is_anico_ctr:
            m["sINPUT_CTR_Boolean"] = True
            ctr_units += rdr.face_amount / 1000
            m["sINPUT_CTR_Units"] = ctr_units
        elif cov == "APB":
            m["sINPUT_APB_Boolean"] = True
            apb_face = rdr.face_amount
            m["sINPUT_APB_Face"] = rdr.face_amount
            m["sINPUT_Original_APB_SA"] = rdr.face_amount  # TODO: original APB SA source

    # ── Inforce valuation block ──────────────────────────────────────
    m["sINPUT_InforceIndicator"] = "Y"
    m["sINPUT_ValuationDate"] = _serial(val_date)
    m["sINPUT_DBOption"] = data.db_option
    m["sInput_CurrentAV"] = data.account_value
    m["sInput_DeemedCashValue"] = data.deemed_cash_value
    m["sINPUT_SWAM"] = data.swam if is_iul else 0
    m["sINPUT_Accum_Min"] = data.accumulated_mtp
    m["sINPUT_MonthlyMTP"] = data.mtp
    m["sINPUT_CTP"] = data.ctp
    m["sINPUT_MAP_CeaseDate"] = _serial(data.map_cease_date)
    m["sINPUT_Guideline"] = data.def_of_life_ins
    m["sINPUT_GLP"] = data.glp
    m["sINPUT_GSP"] = data.gsp
    m["sINPUT_AccumGLP"] = data.accumulated_glp

    # ── TAMRA / MEC ──────────────────────────────────────────────────
    m["sINPUT_IsMEC"] = data.is_mec
    m["sINPUT_7PayStartDate"] = _serial(data.tamra_7pay_start_date)
    m["sINPUT_7PayPremium"] = data.tamra_7pay_level
    m["sINPUT_7PayCashValue"] = data.tamra_7pay_start_av
    m["sINPUT_7YrLowestDB"] = data.tamra_7year_lowest_db
    for i in range(1, 8):
        m[f"sINPUT_TAMRA_Contribution_Yr_{i}"] = data.tamra_7year_contributions[i - 1]

    # ── Inforce change slots (none for a fresh pull) ─────────────────
    for name in ("sCTR_Change_Units", "sCTR_Change_Date", "sPW_Change_Active",
                 "sPW_Change_Date", "sPWST_Change_Amount", "sPWST_Change_Date",
                 "sADB_Change_Units", "sADB_Change_Date", "sGIR_Change_Units",
                 "sGIR_Change_Date", "sINPUT_R1_Change_Amount", "sINPUT_R1_Change_Date",
                 "sINPUT_R2_Change_Amount", "sINPUT_R2_Change_Date",
                 "sINPUT_R3_Change_Amount", "sINPUT_R3_Change_Date",
                 "sINPUT_CCV_Change_Date", "sINPUT_APB_Change_Amount",
                 "sINPUT_APB_Change_Date", "sINPUT_Rateclass_Change_NewRateclassCode",
                 "sINPUT_Rateclass_Change_Date"):
        m[name] = ""

    # ── Report data (fund AV / allocation display) ───────────────────
    fund_values = data.fund_values or {}
    m["sReportData_SW_AV"] = fund_values.get("SW", 0)
    for fund in _FUND_IDS:
        av = fund_values.get(fund, 0)
        if fund == "U1" and not is_iul:
            av = data.account_value
        m[f"sReportData_{fund}_AV"] = av
        m[f"sReportData_{fund}_pct"] = allocs[fund]

    # ── Input vectors (121 policy years) ─────────────────────────────
    vec_sa = total_sa + (apb_face if m["sINPUT_APB_Boolean"] else 0)
    m["vINPUT_Specified_Amount"] = [vec_sa] * 121
    m["vINPUT_DBO"] = [data.db_option] * 121
    m["vINPUT_Premium_Amount"] = [data.modal_premium] * 121
    m["vINPUT_Premium_Mode"] = [mode[0]] * 121
    m["vINPUT_Loans"] = [0] * 121
    m["vINPUT_Loan_Mode"] = ["A"] * 121
    m["vINPUT_Loan_Repayment"] = [0] * 121
    m["vINPUT_Withdrawal"] = [0] * 121

    return m


# ── Scenario overrides (policy-change testing) ──────────────────────────────

def _safe_anniv(issue, year: int):
    try:
        return issue.replace(year=year)
    except ValueError:  # Feb 29 issue
        return issue.replace(year=year, day=28)


def current_policy_year(data) -> int:
    """1-based policy year containing the valuation date."""
    yrs = data.valuation_date.year - data.issue_date.year
    if _safe_anniv(data.issue_date, data.issue_date.year + yrs) > data.valuation_date:
        yrs -= 1
    return yrs + 1


def apply_scenario(m: dict, data, scenario: dict, warnings: list) -> str:
    """Mutate the input vectors for a policy-change scenario; return a summary.

    scenario = {"label": "FaceInc",
                "face":       {"year_offset": 2, "delta": 25000 | "pct": 0.5 | "set": X},
                "dbo":        {"year_offset": 2, "value": "B" | "flip"},
                "withdrawal": {"year_offset": 2, "amount": 2500, "years": 3}}

    year_offset is relative to the CURRENT policy year (projection start);
    the override applies from that policy year's row onward (levels) or for
    `years` rows (withdrawals).  Vectors are 121 rows, one per policy year.
    """
    py = current_policy_year(data)
    parts = []

    def abs_year(spec, default_offset=2) -> int:
        yr = py + int(spec.get("year_offset", default_offset))
        if not 1 <= yr <= 121:
            raise ValueError(f"scenario year {yr} outside 1..121 (policy year {py})")
        return yr

    face = scenario.get("face")
    if face:
        yr = abs_year(face)
        vec = list(m["vINPUT_Specified_Amount"])
        base = float(vec[yr - 2])
        if "set" in face:
            new = float(face["set"])
        elif "pct" in face:
            new = round(base * (1.0 + float(face["pct"])) / 1000.0) * 1000.0
        else:
            new = base + float(face["delta"])
        for i in range(yr - 1, 121):
            vec[i] = new
        m["vINPUT_Specified_Amount"] = vec
        parts.append(f"Face {base:,.0f}->{new:,.0f} @yr{yr}")

    dbo = scenario.get("dbo")
    if dbo:
        yr = abs_year(dbo)
        vec = list(m["vINPUT_DBO"])
        base = str(vec[yr - 2]).strip().upper()
        new = dbo["value"].strip().upper()
        if new == "FLIP":
            new = "B" if base == "A" else "A"
        if new == base:
            warnings.append(f"dbo scenario: already {base}; no-op")
        for i in range(yr - 1, 121):
            vec[i] = new
        m["vINPUT_DBO"] = vec
        parts.append(f"DBO {base}->{new} @yr{yr}")

    wd = scenario.get("withdrawal")
    if wd:
        yr = abs_year(wd)
        amount = float(wd["amount"])
        years = int(wd.get("years", 1))
        vec = list(m["vINPUT_Withdrawal"])
        for i in range(yr - 1, min(yr - 1 + years, 121)):
            vec[i] = amount
        m["vINPUT_Withdrawal"] = vec
        parts.append(f"WD {amount:,.0f}/yr x{years} @yr{yr}")

    label = scenario.get("label") or "scenario"
    m["sINPUT_CaseID"] = f"{data.policy_number}-{label}"
    return "; ".join(parts)


# ── Saved Cases column assembly ─────────────────────────────────────────────

def read_saved_case_rows(workbook: Path, template_case) -> tuple[list, int, int]:
    """Return ([(row, name, template_value)], next_case_number, next_free_col)."""
    import openpyxl
    from rerun_com import _resolve_case_column

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Saved Cases"]
    tcol = _resolve_case_column(ws, template_case)

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            continue
        rows.append((r, str(name).strip(), ws.cell(row=r, column=tcol).value))

    max_case, max_col = 0, 2
    for c in range(3, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if header is None:
            continue
        max_col = max(max_col, c)
        try:
            max_case = max(max_case, int(header))
        except (TypeError, ValueError):
            pass
    wb.close()
    return rows, max_case + 1, max_col + 1


def assemble_column(rows, mapping: dict, case_desc: str,
                    inherited: list, missing: list) -> list:
    """Values for the new column, one per Saved Cases row, in row order."""
    consumed: dict[str, int] = {}
    out = []
    for _r, name, template_val in rows:
        if name == "sINPUT_CaseDescription" and case_desc:
            out.append(case_desc)
            continue
        if name in mapping:
            val = mapping[name]
            if isinstance(val, list):
                idx = consumed.get(name, 0)
                consumed[name] = idx + 1
                out.append(val[idx] if idx < len(val) else val[-1])
            else:
                out.append(val)
        else:
            out.append(template_val)
            if name not in inherited:
                inherited.append(name)
    for name in mapping:
        if name not in {n for _r, n, _v in rows}:
            missing.append(name)
    return out


def write_column(workbook: Path, rows, values: list, case_number: int, col: int) -> None:
    from rerun_com import _open_excel, XL_CALC_MANUAL

    xl = _open_excel()
    try:
        wb = xl.Workbooks.Open(str(workbook), UpdateLinks=0, ReadOnly=False)
        xl.Calculation = XL_CALC_MANUAL
        ws = wb.Worksheets("Saved Cases")
        ws.Cells(1, col).Value = case_number
        for (r, _name, _tv), val in zip(rows, values):
            ws.Cells(r, col).Value = "" if val is None else val
        wb.Save()
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()


def main():
    cmd = json.loads(sys.argv[1])
    os.environ["SUITEVIEW_LOCAL_DATA"] = "1"

    from suiteview.illustration.core.illustration_policy_service import build_illustration_data

    workbook = Path(cmd.get("workbook") or DEFAULT_WORKBOOK).resolve()
    if not workbook.exists():
        print(json.dumps({"ok": False, "error": f"workbook not found: {workbook}"}))
        sys.exit(1)

    policy = cmd["policy"]
    data = build_illustration_data(
        policy, region=cmd.get("region", "CKPR"), company_code=cmd.get("company"))

    warnings: list[str] = []
    mapping = build_mapping(data, warnings, skip_benefits=bool(cmd.get("skip_benefits")))
    if cmd.get("case_id"):
        mapping["sINPUT_CaseID"] = cmd["case_id"]

    scenario_desc = ""
    if cmd.get("scenario"):
        scenario_desc = apply_scenario(mapping, data, cmd["scenario"], warnings)

    rows, case_number, col = read_saved_case_rows(workbook, cmd.get("template_case", 1))

    # Comparison policy: ALL RERUN testing runs with sINPUT_TEFRA_Force TRUE
    # and sINPUT_Exact_Days_Boolean FALSE. Never inherit these toggles from the
    # template — emit the required values unconditionally.
    from check_comparison_inputs import REQUIRED, _as_bool
    for name, required in REQUIRED.items():
        template_val = next((tv for _r, n, tv in rows if n == name), None)
        mapping[name] = required
        if _as_bool(template_val) is not required:
            warnings.append(
                f"{name} forced to {required} (template case had "
                f"{template_val!r}) — mandatory for all RERUN comparison cases")

    inherited: list[str] = []
    missing: list[str] = []
    desc = cmd.get("description") or scenario_desc or (
        f"Offline build from local DB ({_dt.date.today().isoformat()})")
    values = assemble_column(rows, mapping, desc, inherited, missing)

    report = {
        "ok": True,
        "policy": policy,
        "plancode": data.plancode,
        "product_type": data.product_type,
        "case_number": case_number,
        "scenario": scenario_desc or None,
        "current_policy_year": current_policy_year(data),
        "comparison_toggles": dict(REQUIRED),
        "saved_cases_rows": len(rows),
        "mapped": len(mapping) - len(missing),
        "inherited_from_template": sorted(inherited),
        "mapped_but_not_in_saved_cases": sorted(missing),
        "warnings": warnings,
    }

    if cmd.get("dry_run"):
        report["dry_run"] = True
        preview = {}
        for (_r, name, _tv), val in zip(rows, values):
            if name not in preview:
                preview[name] = val
        report["preview"] = preview
        print(json.dumps(report, indent=2, default=str))
        return

    write_column(workbook, rows, values, case_number, col)
    report["workbook"] = str(workbook)
    print(json.dumps(report, indent=2, default=str))


if __name__ == "__main__":
    main()
