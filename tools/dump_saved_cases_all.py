"""Dump ALL Saved Cases from a RERUN workbook in one workbook load.

Like dump_saved_case_summary.py but loads the (38 MB) workbook once and
emits every case column: scalars verbatim, vectors run-length summarized.
Names whose values are blank across the board can be skipped to keep the
output focused on what a case actually specifies.

Usage:
    venv\\Scripts\\python.exe tools/dump_saved_cases_all.py '<json>'
    {"workbook": "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm",
     "skip_blank": true, "cases": [8, 9, 12]}   # cases optional -> all
"""
from __future__ import annotations

import json
import sys

import openpyxl
from openpyxl.utils import get_column_letter


def _rle(values: list) -> str:
    runs = []
    for v in values:
        s = "" if v is None else str(v)
        if runs and runs[-1][0] == s:
            runs[-1][1] += 1
        else:
            runs.append([s, 1])
    return ", ".join(f"{s or '(blank)'} x{n}" for s, n in runs)


def main() -> None:
    cmd = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    workbook = cmd.get("workbook") or "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm"
    want_cases = set(cmd.get("cases") or [])
    skip_blank = bool(cmd.get("skip_blank", True))

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Saved Cases"]

    # One streaming pass over the sheet.
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    case_cols = []  # (0-based col index, case number)
    for i, h in enumerate(header[2:], start=2):
        if h is None or str(h).strip() == "":
            continue
        try:
            num = int(h)
        except (TypeError, ValueError):
            continue
        if not want_cases or num in want_cases:
            case_cols.append((i, num))

    for col, num in case_cols:
        print(f"===== Case {num} (Saved Cases col {get_column_letter(col + 1)}) =====")
        pairs: list[tuple[str, list]] = []
        last_name = None
        for row in rows[1:]:
            name = row[0] if len(row) > 0 else None
            if name is None or str(name).strip() == "":
                continue
            name = str(name).strip()
            value = row[col] if col < len(row) else None
            if name == last_name:
                pairs[-1][1].append(value)
            else:
                pairs.append((name, [value]))
                last_name = name
        for name, values in pairs:
            if skip_blank and all(v is None or str(v).strip() == "" for v in values):
                continue
            if len(values) == 1:
                print(f"  {name} = {values[0]}")
            else:
                print(f"  {name} [{len(values)}] = {_rle(values)}")
        print()


if __name__ == "__main__":
    main()
