"""List every case column on a RERUN workbook's Saved Cases sheet.

For each populated case column prints: column letter, row-1 case number,
CaseID, policy number, and company — enough to enumerate the case set
without dumping full inputs (use dump_saved_case_summary.py for that).

Usage:
    venv\\Scripts\\python.exe tools/list_saved_cases.py '<json>'
    {"workbook": "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm"}
"""
from __future__ import annotations

import json
import sys

import openpyxl
from openpyxl.utils import get_column_letter

INTEREST_ROWS = ("sINPUT_CaseID", "sINPUT_Policy_Number", "sINPUT_Company",
                 "sINPUT_Company_Code", "sINPUT_Plan_Code")


def main() -> None:
    cmd = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    workbook = cmd.get("workbook") or "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm"

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Saved Cases"]

    name_rows: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a and str(a).strip() not in name_rows:
            name_rows[str(a).strip()] = r

    cases = []
    for c in range(3, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if header is None or str(header).strip() == "":
            continue
        entry = {"col": get_column_letter(c), "case": header}
        for name in INTEREST_ROWS:
            row = name_rows.get(name)
            if row is not None:
                entry[name.replace("sINPUT_", "").lower()] = ws.cell(row=row, column=c).value
        cases.append(entry)
    wb.close()

    print(json.dumps({"workbook": workbook, "count": len(cases), "cases": cases}, indent=1, default=str))


if __name__ == "__main__":
    main()
