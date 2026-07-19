"""Merge RERUN's issue-date-dependent band boundary rule into plancode_table.json.

RERUN Rates_Control column CZ ("Use Band Table 2 by Issue Date"):
    CZ9        = the cutoff issue date (2018-10-01)
    CZ12:CZ32  = the plancodes the rule applies to
    CZ6        = AND(MATCH(sPlancode, CZ12:CZ32), sINPUT_Issue_Date >= CZ9)
    sBandTableCode (CG30) = IF(CZ6, 2, <plancode table col U "Band Table">)

Listed plancodes default to mBandTable1 (band 3 starts at 250,001); issues
on/after the cutoff are overridden to mBandTable2 (band 3 starts at 250,000 —
the thresholds stored in UL_Rates BANDSPECS). This tool stamps the cutoff as
"BandTable2IssueDate" on the matching rows of plancode_table.json; all other
plancodes get no key (banding independent of issue date).

Usage:
    venv\\Scripts\\python.exe tools/merge_band_table2_date.py '<json>'

JSON arg:
    {"workbook": "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm",
     "table": "suiteview/illustration/plancodes/plancode_table.json",
     "write": true}

With "write": false (default) it only reports what would change.
"""
import json
import sys
from datetime import date, datetime

import openpyxl


def main():
    cmd = json.loads(sys.argv[1]) if len(sys.argv) > 1 else {}
    workbook = cmd.get("workbook", "docs/Illustration_UL/RERUN (v20.0) local IUL.xlsm")
    table_path = cmd.get(
        "table", "suiteview/illustration/plancodes/plancode_table.json"
    )
    write = bool(cmd.get("write", False))

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Rates_Control"]
    cutoff = ws["CZ9"].value
    if isinstance(cutoff, datetime):
        cutoff = cutoff.date()
    if not isinstance(cutoff, date):
        raise SystemExit(f"Rates_Control!CZ9 is not a date: {cutoff!r}")
    plancodes = []
    for row in ws.iter_rows(min_row=12, max_row=32, min_col=104, max_col=104):
        v = str(row[0].value or "").strip()
        if v:
            plancodes.append(v)
    wb.close()

    with open(table_path, "r") as f:
        data = json.load(f)

    cutoff_str = cutoff.isoformat()
    updated, missing_in_json = [], []
    json_rows = {
        str(e.get("Plancode", "")).strip(): e for e in data.get("Plancodes", [])
    }
    for plancode in plancodes:
        entry = json_rows.get(plancode)
        if entry is None:
            missing_in_json.append(plancode)
            continue
        if entry.get("BandTable2IssueDate") != cutoff_str:
            entry["BandTable2IssueDate"] = cutoff_str
            updated.append(plancode)
    stale = [
        p for p, e in json_rows.items()
        if e.get("BandTable2IssueDate") and p not in plancodes
    ]
    for p in stale:
        del json_rows[p]["BandTable2IssueDate"]

    if write:
        with open(table_path, "w") as f:
            json.dump(data, f, indent=2)
            f.write("\n")

    print(json.dumps({
        "cutoff_date": cutoff_str,
        "cz_plancodes": plancodes,
        "updated": updated,
        "stale_removed": stale,
        "missing_in_json": missing_in_json,
        "written": write,
    }, indent=2))


if __name__ == "__main__":
    main()
