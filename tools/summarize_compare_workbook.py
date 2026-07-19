"""Per-field delta onset summary for a rerun_vs_app comparison workbook.

For each RERUN|App|Δ triplet on a case sheet, report: first vID where |Δ| >
tol, the RERUN/App values there, the max |Δ| and its vID, and a short list of
sample rows around the onset — the month-by-month evidence needed to classify
a discrepancy without opening Excel.

Usage:
    venv\\Scripts\\python.exe tools/summarize_compare_workbook.py '<json>'
    {"workbook": "docs/Illustration_UL/rerun_vs_app_iul_blend_x.xlsx",
     "sheet": "Case9_UE013383",          # default: every Case* sheet
     "fields": ["MTP", "EPU Rate"],      # default: all failing fields
     "context": 3}                        # rows printed at onset/max
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from rerun_debug_map import KIND_TOL, DEBUG_COLUMNS  # noqa: E402

KIND_BY_LABEL = {c["label"]: c["kind"] for c in DEBUG_COLUMNS if c["label"]}


def summarize_sheet(ws, fields_filter, context):
    # Row 2: merged group labels at each triplet start; row 3: RERUN/App/Δ.
    # Identity cols 1..3 (vID/Year/Month), triplets from col 4.
    out = []
    col = 4
    while col <= ws.max_column:
        label = ws.cell(row=2, column=col).value
        if label is None:
            col += 1
            continue
        label = str(label).replace(" (ref)", "").strip()
        ref_only = str(ws.cell(row=2, column=col).value or "").endswith("(ref)")
        rc, ac, dc = col, col + 1, col + 2
        col += 3
        if ref_only:
            continue
        if fields_filter and label not in fields_filter:
            continue

        tol = KIND_TOL.get(KIND_BY_LABEL.get(label, "val"), 0.01)
        first = None
        mx = (0.0, None)
        rows = []
        for r in range(4, ws.max_row + 1):
            vid = ws.cell(row=r, column=1).value
            if vid is None:
                break
            d = ws.cell(row=r, column=dc).value
            if not isinstance(d, (int, float)):
                continue
            rows.append((int(vid), r))
            if abs(d) > tol:
                if first is None:
                    first = (int(vid), r)
                if abs(d) > mx[0]:
                    mx = (abs(d), (int(vid), r))
        if first is None:
            continue

        def sample(anchor_row):
            recs = []
            for r in range(max(4, anchor_row - context), anchor_row + context + 1):
                vid = ws.cell(row=r, column=1).value
                if vid is None:
                    continue
                recs.append({
                    "vid": int(vid),
                    "rerun": ws.cell(row=r, column=rc).value,
                    "app": ws.cell(row=r, column=ac).value,
                    "delta": ws.cell(row=r, column=dc).value,
                })
            return recs

        entry = {
            "field": label, "tol": tol,
            "first_fail_vid": first[0], "max_abs": round(mx[0], 6),
            "max_vid": mx[1][0] if mx[1] else None,
            "onset": sample(first[1]),
        }
        if mx[1] and mx[1][0] != first[0]:
            entry["at_max"] = sample(mx[1][1])
        out.append(entry)
    return out


def main():
    cmd = json.loads(sys.argv[1])
    wb = openpyxl.load_workbook(cmd["workbook"], read_only=False, data_only=True)
    sheets = [cmd["sheet"]] if cmd.get("sheet") else [
        s for s in wb.sheetnames if s.startswith("Case")]
    fields = cmd.get("fields") or None
    context = int(cmd.get("context", 3))
    result = {}
    for name in sheets:
        result[name] = summarize_sheet(wb[name], fields, context)
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
