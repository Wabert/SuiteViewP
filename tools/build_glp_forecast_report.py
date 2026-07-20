r"""Build a management-facing HTML exhibit from the GLP forecast batch workbook.

Reads the workbook produced by tools/run_glp_forecast_batch.py (headers matched
by label via batch_runner.GLP_COLUMNS), classifies every policy into one of four
MUTUALLY EXCLUSIVE categories, and writes a single self-contained HTML page:

    Sustained           current premium carries the policy to maturity
    Increase fixes it   lapses at current premium, but the solved level premium
                        endows with NO exception premiums (exc date = none)
    Exceptions required the minimum level premium crosses the guideline limit
                        (exc date is a real date). Sub-metric: the "front-
                        loading only" subset — policies whose absolute-max
                        (guideline-capped) funding run still reaches maturity,
                        i.e. exceptions could be avoided by funding hard now.
    Not Classified      bypassed by the engine (MD mismatch, missing rates,
                        shadow account, CVAT) — outcome unknown, not analyzed —
                        OR no level premium keeps the policy in force to maturity

Label vocabulary accepted in the mixed date/text forecast columns (older runs
used different labels): "Maturity"; "(none)" / "not needed"; "no solution".
Rows not yet run (blank Run Status) are reported in the footer, never silently
dropped.

The exhibit: KPI row, severity funnel, an INTERACTIVE by-form breakdown map
(click a category chip — or a funnel segment — to see that category's
concentration by policy form), the lapse-timeline stacked columns (toggle
policies/face) with a cumulative face-at-risk line, and the exception-premium
onset timeline. All inline SVG/HTML — no external assets — emails and prints
as-is. Open with #cat=D (etc.) in the URL to preselect a form-map view.

Usage:
    venv\Scripts\python.exe tools/build_glp_forecast_report.py ^
        "docs\Illustration_UL\GLP Limit Calc v2.xlsx"

Flags:
    --sheet NAME   sheet to read (default: first sheet)
    --out PATH     output HTML path (default: "<workbook dir>\GLP Funding Outlook.html")

Prints a JSON summary (row coverage, category counts, distinct labels seen) so
the classification can be audited against the workbook.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: "
                      "venv\\Scripts\\python.exe -m pip install openpyxl"}))
    sys.exit(1)

from suiteview.illustration.core.batch_runner import GLP_COLUMNS  # noqa: E402

# ── Label vocabulary (current + legacy) ─────────────────────────────────────
MATURITY_LABELS = {"maturity"}
NO_EXCEPTION_LABELS = {"(none)", "not needed"}
NO_SOLUTION_LABELS = {"no solution"}

# The four mutually exclusive categories. "NC" (Not Classified) is gray and
# deliberately last; the gray IS the message (no data), not a series color.
# No-solution policies (no level premium keeps them in force) are folded into
# "Not Classified" — see classify().
CAT_META = [
    ("A", "Sustained", "Current premium carries the policy to maturity", "#0ca30c"),
    ("B", "Increase fixes it", "Reaches maturity with a premium increase — no exception premiums", "#fab219"),
    ("D", "Exceptions required", "Minimum level premium crosses the guideline limit (GLP exceptions needed)", "#d03b3b"),
    ("NC", "Not Classified", "Bypassed by the engine (MD mismatch, missing rates, shadow account, CVAT), or no level premium keeps the policy in force to maturity", "#898781"),
]
CAT_COLOR = {k: c for k, _, _, c in CAT_META}
CAT_SHORT = {k: s for k, s, _, _ in CAT_META}
CAT_LONG = {k: lg for k, _, lg, _ in CAT_META}
PROBLEM_CATS = ("B", "D")               # lapse at current premium

INK = "#0b0b0b"
INK2 = "#52514e"
MUTED = "#898781"
GRID = "#e1e0d9"
BASELINE = "#c3c2b7"
SURFACE = "#fcfcfb"
PAGE = "#f9f9f7"
FADE = "#e7e6e1"                        # de-emphasised segment in category views


# ── Workbook reading ────────────────────────────────────────────────────────

def _as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v.strip())
        except ValueError:
            return None
    return None


def _label(v) -> str:
    return str(v).strip().lower() if isinstance(v, str) else ""


def _num(v) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def read_rows(workbook_path: str, sheet: Optional[str]) -> Tuple[List[dict], str]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    label_to_key = {lbl.strip().lower(): key for key, lbl in GLP_COLUMNS}
    # Loan/debt is not part of GLP_COLUMNS but the workbook carries it; accept
    # the known header labels for it under the "loan" key.
    for alias in ("total policy debt", "loan amount", "loan balance"):
        label_to_key.setdefault(alias, "loan")
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or ()
    col_of = {}
    for idx, cell in enumerate(header):
        key = label_to_key.get(str(cell or "").strip().lower())
        if key:
            col_of[key] = idx

    def get(tup, key):
        idx = col_of.get(key)
        return tup[idx] if idx is not None and idx < len(tup) else None

    out = []
    for sheet_row, tup in enumerate(rows_iter, start=2):
        policy = str(tup[1] if len(tup) > 1 else "" or "").strip()
        if not policy or policy.lower() == "none":
            continue
        out.append({
            "row": sheet_row,
            "company": str(tup[0] or "").strip(),
            "policy": policy,
            **{key: get(tup, key) for key in col_of},
        })
    title = ws.title
    wb.close()
    return out, title


# ── Classification ──────────────────────────────────────────────────────────

def classify(rows: List[dict]) -> dict:
    """Assign every row to exactly one category; derive per-policy metrics."""
    cats: Dict[str, List[dict]] = {k: [] for k, *_ in CAT_META}
    bypass_reasons: Dict[str, int] = {}
    not_run = 0
    unclassified: List[dict] = []
    labels_seen: Dict[str, set] = {"lapse_cur_prem": set(), "exc_date": set(),
                                   "lapse_abs_max": set()}
    missing_abs_max = 0

    for r in rows:
        status = str(r.get("run_status") or "").strip()
        form = str(r.get("form") or "").strip() or "(unknown form)"
        face = _num(r.get("face")) or 0.0
        has_loan = (_num(r.get("loan")) or 0.0) > 0.0

        if not status:
            not_run += 1
            continue
        if status != "Complete":
            m = re.search(r"\((.*)\)", status)
            for reason in (m.group(1).split(",") if m else [status]):
                reason = reason.strip()
                bypass_reasons[reason] = bypass_reasons.get(reason, 0) + 1
            cats["NC"].append({"row": r["row"], "policy": r["policy"],
                               "form": form, "face": face, "cat": "NC",
                               "has_loan": has_loan})
            continue

        lapse_cur = r.get("lapse_cur_prem")
        exc = r.get("exc_date")
        abs_max = r.get("lapse_abs_max")
        for key, v in (("lapse_cur_prem", lapse_cur), ("exc_date", exc),
                       ("lapse_abs_max", abs_max)):
            if isinstance(v, str):
                labels_seen[key].add(v.strip())

        lapse_cur_d = _as_date(lapse_cur)
        exc_d = _as_date(exc)
        front_load_ok = False

        if _label(lapse_cur) in MATURITY_LABELS:
            cat = "A"
        elif exc_d is not None:
            cat = "D"
            front_load_ok = _label(abs_max) in MATURITY_LABELS
            if abs_max in (None, ""):
                missing_abs_max += 1
        elif _label(exc) in NO_SOLUTION_LABELS:
            cat = "NC"          # folded into Not Classified — no viable solution
        elif _label(exc) in NO_EXCEPTION_LABELS:
            cat = "B"
        else:
            unclassified.append({"row": r["row"], "policy": r["policy"],
                                 "lapse_cur_prem": str(lapse_cur),
                                 "exc_date": str(exc)})
            continue

        billing = _num(r.get("billing_prem"))
        level = _num(r.get("level_prem"))
        cats[cat].append({
            "row": r["row"], "company": r["company"], "policy": r["policy"],
            "form": form, "cat": cat, "face": face,
            "has_loan": has_loan,
            "lapse_date": lapse_cur_d,
            "lapse_year": lapse_cur_d.year if lapse_cur_d else None,
            "exc_date": exc_d,
            "exc_year": exc_d.year if exc_d else None,
            "front_load_ok": front_load_ok,
            "multiple": (round(level / billing, 2)
                         if level is not None and billing and billing > 0 else None),
        })

    return {
        "cats": cats, "bypass_reasons": bypass_reasons, "not_run": not_run,
        "unclassified": unclassified,
        "labels_seen": {k: sorted(v) for k, v in labels_seen.items()},
        "missing_abs_max": missing_abs_max,
    }


# ── Formatting helpers ──────────────────────────────────────────────────────

def money_compact(v: Optional[float]) -> str:
    if v is None:
        return "–"
    a = abs(v)
    if a >= 1e9:
        return f"${v / 1e9:,.2f}B"
    if a >= 1e6:
        return f"${v / 1e6:,.1f}M"
    if a >= 1e3:
        return f"${v / 1e3:,.0f}K"
    return f"${v:,.0f}"


def n_fmt(v: int) -> str:
    return f"{v:,}"


def pct(part: float, whole: float) -> str:
    return f"{100.0 * part / whole:.0f}%" if whole else "–"


def pct1(part: float, whole: float) -> str:
    return f"{100.0 * part / whole:.1f}%" if whole else "–"


def _tt(lines: List[str]) -> str:
    return escape("\n".join(lines), quote=True)


def median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    s = sorted(vals)
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2


# ── SVG chart builders ──────────────────────────────────────────────────────

def _col_path(x: float, y: float, w: float, h: float, r: float) -> str:
    """Column with rounded top corners, square baseline."""
    r = min(r, w / 2, h)
    return (f"M{x:.1f},{y + h:.1f} L{x:.1f},{y + r:.1f} "
            f"Q{x:.1f},{y:.1f} {x + r:.1f},{y:.1f} "
            f"L{x + w - r:.1f},{y:.1f} Q{x + w:.1f},{y:.1f} {x + w:.1f},{y + r:.1f} "
            f"L{x + w:.1f},{y + h:.1f} Z")


def _nice_step(max_v: float, target_ticks: int = 4) -> float:
    if max_v <= 0:
        return 1.0
    raw = max_v / target_ticks
    mag = 10 ** int(len(str(int(raw))) - 1) if raw >= 1 else 1
    for mult in (1, 2, 2.5, 5, 10):
        if mag * mult >= raw:
            return mag * mult
    return mag * 10


def _y_axis(max_v: float, plot_h: float, top: float, left: float, width: float,
            money: bool) -> Tuple[str, float]:
    step = _nice_step(max_v)
    ticks = []
    v = 0.0
    while v <= max_v + step * 0.999:
        ticks.append(v)
        v += step
    adj_max = ticks[-1]
    parts = []
    for tv in ticks:
        y = top + plot_h - (tv / adj_max) * plot_h
        if tv > 0:
            parts.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left + width}" '
                         f'y2="{y:.1f}" stroke="{GRID}" stroke-width="1"/>')
        lab = money_compact(tv) if money else n_fmt(int(tv))
        parts.append(f'<text x="{left - 8}" y="{y + 4:.1f}" text-anchor="end" '
                     f'class="tick">{lab}</text>')
    parts.append(f'<line x1="{left}" y1="{top + plot_h}" x2="{left + width}" '
                 f'y2="{top + plot_h}" stroke="{BASELINE}" stroke-width="1"/>')
    return "".join(parts), adj_max


def stacked_columns_svg(cats: List[str], series: Dict[str, List[float]],
                        *, money: bool, seg_label: Optional[str] = None,
                        width: int = 1010, plot_h: int = 230) -> str:
    """Stacked columns; series keyed by category (drawn bottom-up B,D,E)."""
    if not cats:
        return '<div class="note">No policies in this view.</div>'
    left, right, top, bottom = 62, 14, 12, 34
    plot_w = width - left - right
    n = len(cats)
    totals = [sum(series[t][i] for t in series) for i in range(n)]
    max_v = max(totals) if totals else 1.0
    grid, adj_max = _y_axis(max_v, plot_h, top, left, plot_w, money)
    slot = plot_w / n
    bar_w = min(16.0, max(slot - 3.0, 2.0))
    gap = 2.0
    marks = []
    order = [t for t in ("B", "D", "E") if t in series]
    for i, cat in enumerate(cats):
        x = left + i * slot + (slot - bar_w) / 2
        y_cursor = top + plot_h
        segs = [(t, series[t][i]) for t in order if series[t][i] > 0]
        for j, (t, v) in enumerate(segs):
            h = (v / adj_max) * plot_h
            is_top = j == len(segs) - 1
            h_draw = max(h - (0 if is_top else gap), 0.8)
            y = y_cursor - h
            val = money_compact(v) if money else n_fmt(int(v))
            if seg_label:
                lines = [f"{cat} · {seg_label}", f"{CAT_SHORT[t]}: {val}"]
            else:
                lines = [f"{cat} · {CAT_SHORT[t]}", f"{val}"]
            if is_top:
                d = _col_path(x, y, bar_w, h_draw, 3)
                marks.append(f'<path d="{d}" fill="{CAT_COLOR[t]}" '
                             f'class="mk" data-tt="{_tt(lines)}"/>')
            else:
                marks.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" '
                             f'height="{h_draw:.1f}" fill="{CAT_COLOR[t]}" '
                             f'class="mk" data-tt="{_tt(lines)}"/>')
            y_cursor = y
    lab_every = max(1, (n + 15) // 16)
    labels = []
    for i, cat in enumerate(cats):
        digits = re.sub(r"\D", "", cat)
        year = int(digits) if digits else None
        # Label round half-decades (…2030, 2035…) plus the first and last bar;
        # fall back to even spacing when the labels aren't years.
        if year is not None:
            show = i == 0 or i == n - 1 or year % 5 == 0
        else:
            show = i % lab_every == 0 or i == n - 1
        if not show:
            continue
        cx = left + i * slot + slot / 2
        labels.append(f'<text x="{cx:.1f}" y="{top + plot_h + 18}" '
                      f'text-anchor="middle" class="tick">{escape(cat)}</text>')
    h_total = top + plot_h + bottom
    return (f'<svg viewBox="0 0 {width} {h_total}" role="img" '
            f'style="width:100%;height:auto">{grid}{"".join(marks)}{"".join(labels)}</svg>')


def funnel_svg(counts: Dict[str, int], faces: Dict[str, float],
               loans: Optional[Dict[str, int]] = None,
               width: int = 1010) -> str:
    """One horizontal stacked bar over all classified policies. Segments are
    clickable — they select the matching form-map view."""
    loans = loans or {}
    total = sum(counts.values())
    if total == 0:
        return ""
    h, bar_h, top = 44, 26, 8
    x = 0.0
    parts = []
    segs = [(k, counts[k]) for k, *_ in CAT_META if counts[k] > 0]
    for i, (k, c) in enumerate(segs):
        w = (c / total) * width
        w_draw = max(w - (2 if i < len(segs) - 1 else 0), 1.0)
        lines = [CAT_SHORT[k], f"{n_fmt(c)} policies · {pct(c, total)}",
                 f"{money_compact(faces.get(k, 0.0))} face amount",
                 f"{pct(loans.get(k, 0), c)} have a loan",
                 "(click to see by form)"]
        parts.append(f'<rect x="{x:.1f}" y="{top}" width="{w_draw:.1f}" '
                     f'height="{bar_h}" rx="3" fill="{CAT_COLOR[k]}" '
                     f'class="mk fseg" data-cat="{k}" data-tt="{_tt(lines)}"/>')
        if w > 52:
            tx = x + w / 2
            fill = "#ffffff" if k in ("D", "E", "NC") else INK
            parts.append(f'<text x="{tx:.1f}" y="{top + bar_h / 2 + 4}" '
                         f'text-anchor="middle" class="seg" fill="{fill}" '
                         f'pointer-events="none">{pct(c, total)}</text>')
        x += w
    return (f'<svg viewBox="0 0 {width} {h}" role="img" '
            f'style="width:100%;height:auto">{"".join(parts)}</svg>')


# ── By-form breakdown map ───────────────────────────────────────────────────

MAX_FORM_ROWS = 24


def build_form_map(cats: Dict[str, List[dict]]) -> dict:
    """forms[form][cat] = {n, face}; folds the long tail into 'Other forms'."""
    forms: Dict[str, Dict[str, dict]] = {}
    fl: Dict[str, int] = {}
    for k in cats:
        for p in cats[k]:
            f = forms.setdefault(p["form"], {c: {"n": 0, "face": 0.0}
                                             for c, *_ in CAT_META})
            f[k]["n"] += 1
            f[k]["face"] += p["face"]
            if k == "D" and p.get("front_load_ok"):
                fl[p["form"]] = fl.get(p["form"], 0) + 1
    order = sorted(forms, key=lambda f: -sum(forms[f][c]["n"] for c, *_ in CAT_META))
    if len(order) > MAX_FORM_ROWS:
        other = {c: {"n": 0, "face": 0.0} for c, *_ in CAT_META}
        fl_other = 0
        for f in order[MAX_FORM_ROWS:]:
            for c, *_ in CAT_META:
                other[c]["n"] += forms[f][c]["n"]
                other[c]["face"] += forms[f][c]["face"]
            fl_other += fl.get(f, 0)
        order = order[:MAX_FORM_ROWS] + ["Other forms"]
        forms["Other forms"] = other
        fl["Other forms"] = fl_other
    return {"forms": forms, "order": order, "front_load": fl}


def _form_total(fdata: Dict[str, dict]) -> int:
    return sum(fdata[c]["n"] for c, *_ in CAT_META)


def form_map_all_view(fm: dict) -> str:
    """One 100%-stacked bar per form, all four categories."""
    rows = []
    for form in fm["order"]:
        fdata = fm["forms"][form]
        total = _form_total(fdata)
        if total == 0:
            continue
        segs = []
        for k, *_ in CAT_META:
            n = fdata[k]["n"]
            if n == 0:
                continue
            share = 100.0 * n / total
            lines = [f"{form} · {CAT_SHORT[k]}",
                     f"{n_fmt(n)} policies · {pct1(n, total)} of form",
                     f"{money_compact(fdata[k]['face'])} face"]
            if k == "D" and fm["front_load"].get(form):
                lines.append(f"front-load-capable: {n_fmt(fm['front_load'][form])}")
            inner = (f'<span class="fm-pct">{share:.0f}%</span>'
                     if share >= 9 else "")
            segs.append(f'<div class="fm-seg" data-tt="{_tt(lines)}" '
                        f'style="width:{share:.2f}%;background:{CAT_COLOR[k]};'
                        f'color:{"#fff" if k in ("D", "E", "NC") else INK}">'
                        f'{inner}</div>')
        rows.append(f'<div class="fm-row"><div class="fm-name">{escape(form)}</div>'
                    f'<div class="fm-bar">{"".join(segs)}</div>'
                    f'<div class="fm-n">{n_fmt(total)}</div></div>')
    return "".join(rows)


def form_map_cat_view(fm: dict, cat: str) -> str:
    """One bar per form: that category's share of the form's policies."""
    cat_total = sum(fm["forms"][f][cat]["n"] for f in fm["order"])
    ranked = sorted((f for f in fm["order"] if fm["forms"][f][cat]["n"] > 0),
                    key=lambda f: -fm["forms"][f][cat]["n"])
    if not ranked:
        return ('<div class="note">No policies in this category.</div>')
    rows = []
    for form in ranked:
        fdata = fm["forms"][form]
        n = fdata[cat]["n"]
        total = _form_total(fdata)
        share = 100.0 * n / total
        lines = [f"{form} · {CAT_SHORT[cat]}",
                 f"{n_fmt(n)} of {n_fmt(total)} policies · {pct1(n, total)} of form",
                 f"{pct1(n, cat_total)} of all {CAT_SHORT[cat]} policies",
                 f"{money_compact(fdata[cat]['face'])} face"]
        if cat == "D" and fm["front_load"].get(form):
            lines.append(f"front-load-capable: {n_fmt(fm['front_load'][form])}")
        rows.append(
            f'<div class="fm-row"><div class="fm-name">{escape(form)}</div>'
            f'<div class="fm-bar fm-track" data-tt="{_tt(lines)}">'
            f'<div class="fm-fill" style="width:{share:.2f}%;'
            f'background:{CAT_COLOR[cat]}"></div></div>'
            f'<div class="fm-n wide">{n_fmt(n)} · {pct(n, total)}</div></div>')
    note = ('<div class="note" style="margin-top:8px">Bar length = share of the '
            'form’s policies in this category. Forms with none are omitted.</div>')
    return "".join(rows) + note


# ── Aggregation for the timelines ───────────────────────────────────────────

# Fixed lapse/exception timeline window: one bar per year from 2026 through
# 2075. Years before the start fold into the first bar; years after the end
# fold into the last bar (kept small — the block matures out well before then).
TIMELINE_START = 2026
TIMELINE_END = 2100


def _timeline_years() -> List[int]:
    return list(range(TIMELINE_START, TIMELINE_END + 1))


def _timeline_labels(years: List[int]) -> List[str]:
    labels = [str(y) for y in years]
    labels[0] = f"\u2264{years[0]}"      # ≤2026 — prior years folded in
    labels[-1] = f"{years[-1]}+"         # 2075+ — later years folded in
    return labels


def _timeline_slot(year: int) -> int:
    return min(max(year, TIMELINE_START), TIMELINE_END) - TIMELINE_START


def build_wave(problem: List[dict]) -> dict:
    years = _timeline_years()
    cats = _timeline_labels(years)
    n = len(years)
    count_s = {t: [0.0] * n for t in PROBLEM_CATS}
    face_s = {t: [0.0] * n for t in PROBLEM_CATS}
    for p in problem:
        if not p["lapse_year"]:
            continue
        i = _timeline_slot(p["lapse_year"])
        count_s[p["cat"]][i] += 1
        face_s[p["cat"]][i] += p["face"]
    cum = []
    running = 0.0
    for i in range(n):
        running += sum(face_s[t][i] for t in PROBLEM_CATS)
        cum.append(running)
    return {"cats": cats, "count_series": count_s, "face_series": face_s,
            "cum_face": cum}


def build_exception_wave(d_rows: List[dict]) -> dict:
    pol = [p for p in d_rows if p["exc_year"]]
    years = _timeline_years()
    cats = _timeline_labels(years)
    n = len(years)
    count_s = {"D": [0.0] * n}
    for p in pol:
        i = _timeline_slot(p["exc_year"])
        count_s["D"][i] += 1
    return {"cats": cats, "count_series": count_s, "n": len(pol)}


# ── HTML assembly ───────────────────────────────────────────────────────────

def build_html(cls: dict, *, workbook_name: str, region_note: str) -> str:
    cats = cls["cats"]
    counts = {k: len(v) for k, v in cats.items()}
    faces = {k: sum(p["face"] for p in v) for k, v in cats.items()}
    loans = {k: sum(1 for p in v if p.get("has_loan")) for k, v in cats.items()}
    classified_total = sum(counts.values())          # includes Not Classified
    analyzed = classified_total - counts["NC"]
    loan_analyzed = sum(v for k, v in loans.items() if k != "NC")
    problem = [p for t in PROBLEM_CATS for p in cats[t]]
    total_at_risk = sum(p["face"] for p in problem)
    front_load_n = sum(1 for p in cats["D"] if p["front_load_ok"])
    front_load_face = sum(p["face"] for p in cats["D"] if p["front_load_ok"])
    hard = counts["D"]
    hard_face = faces["D"]
    mults = [p["multiple"] for p in cats["B"] if p["multiple"] is not None]
    med_mult = median(mults)

    wave = build_wave(problem)
    exc_wave = build_exception_wave(cats["D"])
    fm = build_form_map(cats)

    onset = {}
    for frac in (0.25, 0.5):
        for i, v in enumerate(wave["cum_face"]):
            if total_at_risk and v >= frac * total_at_risk:
                onset[frac] = wave["cats"][i]
                break
    onset_line = ""
    if onset.get(0.25) and onset.get(0.5):
        onset_line = (f"A quarter of the at-risk face amount lapses by "
                      f"<strong>{onset[0.25]}</strong>; half by "
                      f"<strong>{onset[0.5]}</strong>.")

    cat_legend = ""
    for k, *_ in CAT_META:
        fn = ' <sup class="fn-mark">†</sup>' if k == "D" and counts["D"] else ''
        cat_legend += (
            f'<div class="leg-row"><span class="sw" style="background:{CAT_COLOR[k]}"></span>'
            f'<span class="leg-name">{escape(CAT_SHORT[k])}{fn}</span>'
            f'<span class="leg-detail">{escape(CAT_LONG[k])}</span>'
            f'<span class="leg-num">{n_fmt(counts[k])} policies · '
            f'{money_compact(faces[k])} face · {pct(counts[k], classified_total)} · '
            f'{pct(loans[k], counts[k])} have a loan</span></div>')

    cat_footnote = ""
    if counts["D"]:
        cat_footnote = (
            f'<div class="leg-fn"><span class="fn-mark">†</span> '
            f'<strong>Front-loading only:</strong> {n_fmt(front_load_n)} of the '
            f'{n_fmt(counts["D"])} Exceptions-required policies '
            f'({money_compact(front_load_face)} face) could still reach maturity — '
            f'without exceptions — by funding to the guideline maximum starting now.</div>')

    kpis = f"""
    <div class="kpi-row">
      <div class="kpi"><div class="kpi-label">Total ULs inforce issued &le; 1/1/2000 (all companies)</div>
        <div class="kpi-value">{n_fmt(20122)}</div>
        <div class="kpi-sub">of these, {n_fmt(classified_total + cls['not_run'])} of which are company 01 (ANICO)</div></div>
      <div class="kpi"><div class="kpi-label">Policies analyzed</div>
        <div class="kpi-value">{n_fmt(analyzed)}</div>
        <div class="kpi-sub">{n_fmt(classified_total + cls['not_run'])} in scope but {n_fmt(counts['NC'])} could not be analyzed · {n_fmt(loan_analyzed)} ({pct(loan_analyzed, analyzed)}) have a loan</div></div>
      <div class="kpi"><div class="kpi-label">Will lapse at current premium</div>
        <div class="kpi-value">{n_fmt(len(problem))}</div>
        <div class="kpi-sub">{pct(len(problem), analyzed)} of analyzed · {money_compact(total_at_risk)} face</div></div>
      <div class="kpi accent-bad"><div class="kpi-label">Exceptions required</div>
        <div class="kpi-value">{n_fmt(hard)}</div>
        <div class="kpi-sub">{pct(hard, analyzed)} of analyzed · {money_compact(hard_face)} face · {n_fmt(front_load_n)} could front-load instead</div></div>
      <div class="kpi"><div class="kpi-label">Fixable with premium increase</div>
        <div class="kpi-value">{n_fmt(counts['B'])}</div>
        <div class="kpi-sub">median increase {f"{med_mult:.1f}× current premium" if med_mult else "n/a"}</div></div>
    </div>"""

    chips = '<button class="chip on" data-cat="ALL">All categories</button>'
    for k, *_ in CAT_META:
        chips += (f'<button class="chip" data-cat="{k}">'
                  f'<span class="sw" style="background:{CAT_COLOR[k]}"></span>'
                  f'{escape(CAT_SHORT[k])} <span class="chip-n">{n_fmt(counts[k])}</span></button>')

    fm_views = f'<div class="fm-view" data-cat="ALL">{form_map_all_view(fm)}</div>'
    for k, *_ in CAT_META:
        fm_views += (f'<div class="fm-view" data-cat="{k}" hidden>'
                     f'{form_map_cat_view(fm, k)}</div>')

    wave_svg_count = stacked_columns_svg(wave["cats"], wave["count_series"],
                                         money=False, seg_label="Lapses projected")
    wave_svg_face = stacked_columns_svg(wave["cats"], wave["face_series"],
                                        money=True, seg_label="Lapses projected")
    exc_svg = (stacked_columns_svg(exc_wave["cats"], exc_wave["count_series"],
                                   money=False) if exc_wave["n"] else "")
    chart_legend_problem = "".join(
        f'<span class="cl"><span class="sw" style="background:{CAT_COLOR[t]}"></span>{escape(CAT_SHORT[t])}</span>'
        for t in PROBLEM_CATS if counts[t])

    bypass_bits = " · ".join(f"{escape(k)} {n_fmt(v)}"
                             for k, v in sorted(cls["bypass_reasons"].items(),
                                                key=lambda kv: -kv[1]))
    gen_date = date.today().strftime("%B %d, %Y")

    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>UL Block — Funding-to-Maturity Outlook</title>
<style>
  :root {{ color-scheme: light; }}
  * {{ box-sizing: border-box; margin: 0; }}
  body {{ background:{PAGE}; color:{INK}; font-family: system-ui, -apple-system,
         "Segoe UI", sans-serif; font-size:14px; line-height:1.45; }}
  .wrap {{ max-width:1080px; margin:0 auto; padding:28px 24px 40px; }}
  header h1 {{ font-size:23px; font-weight:650; letter-spacing:-0.01em; }}
  header .sub {{ color:{INK2}; margin-top:4px; }}
  .card {{ background:{SURFACE}; border:1px solid rgba(11,11,11,0.10);
           border-radius:10px; padding:18px 20px 16px; margin-top:18px; }}
  .card h2 {{ font-size:15px; font-weight:650; }}
  .card .note {{ color:{INK2}; font-size:13px; margin:2px 0 10px; }}
  .kpi-row {{ display:grid; grid-template-columns:repeat(5,1fr); gap:12px;
              margin-top:18px; }}
  .kpi {{ background:{SURFACE}; border:1px solid rgba(11,11,11,0.10);
          border-radius:10px; padding:14px 16px; }}
  .kpi-label {{ font-size:12.5px; color:{INK2}; }}
  .kpi-value {{ font-size:34px; font-weight:600; margin-top:2px; }}
  .kpi-sub {{ font-size:12px; color:{MUTED}; margin-top:2px; }}
  .kpi.accent-bad {{ border-color:#d03b3b55; }}
  .kpi.accent-bad .kpi-value {{ color:#b22a2a; }}
  .leg-row {{ display:grid; grid-template-columns:14px 170px 1fr auto; gap:10px;
              align-items:baseline; padding:5px 0; border-top:1px solid {GRID}; }}
  .leg-row:first-child {{ border-top:none; }}
  .fn-mark {{ color:#d03b3b; font-weight:700; }}
  .leg-fn {{ margin-top:8px; padding-top:8px; border-top:1px solid {GRID};
             color:{INK2}; font-size:12.5px; }}
  .sw {{ width:12px; height:12px; border-radius:3px; display:inline-block;
         position:relative; top:1px; flex:none; }}
  .leg-name {{ font-weight:600; font-size:13px; }}
  .leg-detail {{ color:{INK2}; font-size:12.5px; }}
  .leg-num {{ font-size:12.5px; color:{INK2}; white-space:nowrap;
              font-variant-numeric: tabular-nums; }}
  .tick {{ font-size:11px; fill:{MUTED}; font-family:inherit;
           font-variant-numeric: tabular-nums; }}
  .seg {{ font-size:11.5px; font-weight:600; font-family:inherit; }}
  .dlab {{ font-size:11.5px; font-weight:600; fill:{INK2}; font-family:inherit; }}
  .mk {{ cursor:default; }}
  .mk:hover {{ opacity:0.82; }}
  .fseg {{ cursor:pointer; }}
  .chart-head {{ display:flex; align-items:center; gap:14px; margin-bottom:6px;
                 flex-wrap:wrap; }}
  .chart-legend {{ display:flex; gap:12px; margin-left:auto; }}
  .cl {{ display:inline-flex; align-items:center; gap:5px; font-size:12px;
         color:{INK2}; white-space:nowrap; }}
  .toggle {{ display:inline-flex; border:1px solid rgba(11,11,11,0.14);
             border-radius:7px; overflow:hidden; }}
  .toggle button {{ border:none; background:{SURFACE}; padding:4px 12px;
                    font:inherit; font-size:12.5px; color:{INK2}; cursor:pointer; }}
  .toggle button[aria-pressed="true"] {{ background:#e8eef8; color:{INK};
                                         font-weight:600; }}
  .chips {{ display:flex; gap:8px; flex-wrap:wrap; margin:10px 0 14px; }}
  .chip {{ display:inline-flex; align-items:center; gap:6px; border:1px solid
           rgba(11,11,11,0.14); border-radius:16px; background:{SURFACE};
           padding:4px 12px; font:inherit; font-size:12.5px; color:{INK2};
           cursor:pointer; }}
  .chip:hover {{ background:{PAGE}; }}
  .chip.on {{ background:#e8eef8; border-color:#2a78d6; color:{INK};
              font-weight:600; }}
  .chip-n {{ color:{MUTED}; font-variant-numeric: tabular-nums; }}
  .fm-row {{ display:flex; align-items:center; gap:10px; padding:2px 0; }}
  .fm-name {{ flex:0 0 150px; font-size:12.5px; text-align:right; color:{INK2};
              white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  .fm-bar {{ flex:1; display:flex; height:18px; border-radius:4px;
             overflow:hidden; gap:2px; background:transparent; }}
  .fm-seg {{ height:100%; border-radius:2px; display:flex; align-items:center;
             justify-content:center; min-width:2px; }}
  .fm-seg:hover, .fm-track:hover {{ opacity:0.85; }}
  .fm-pct {{ font-size:10.5px; font-weight:600; }}
  .fm-track {{ background:{FADE}; border-radius:4px; overflow:hidden; }}
  .fm-fill {{ height:100%; border-radius:4px 2px 2px 4px; }}
  .fm-n {{ flex:0 0 56px; font-size:12px; color:{INK2}; text-align:right;
           font-variant-numeric: tabular-nums; white-space:nowrap; }}
  .fm-n.wide {{ flex-basis:92px; }}
  .tt {{ position:fixed; pointer-events:none; background:#1c1c1b; color:#fff;
         padding:7px 10px; border-radius:7px; font-size:12px; line-height:1.4;
         white-space:pre-line; z-index:10; display:none; max-width:290px;
         box-shadow:0 4px 14px rgba(0,0,0,0.25); }}
  .foot {{ color:{MUTED}; font-size:12px; margin-top:20px; line-height:1.6; }}
  @media print {{ .tt {{ display:none !important; }} body {{ background:#fff; }}
                  .card {{ break-inside:avoid; }} }}
</style></head><body><div class="wrap">

<header>
  <h1>UL Block — Funding-to-Maturity Outlook</h1>
  <div class="sub">Older universal life policies: ability to stay in force to maturity
  under Guideline Premium Test limits · {escape(region_note)} · prepared {gen_date}</div>
</header>

{kpis}

<div class="card">
  <h2>How big is the problem?</h2>
  <div class="note">Every policy, classified into four mutually exclusive categories.
  Click a segment to see its breakdown by policy form.</div>
  {funnel_svg(counts, faces, loans)}
  <div style="margin-top:6px">{cat_legend}</div>
  {cat_footnote}
</div>

<div class="card">
  <div class="chart-head">
    <h2>When do policies lapse (at current premiums)</h2>
    <div class="chart-legend">{chart_legend_problem}</div>
    <div class="toggle" role="group" aria-label="Measure">
      <button aria-pressed="true" data-view="count">Policies</button>
      <button aria-pressed="false" data-view="face">Face amount</button>
    </div>
  </div>
  <div class="note">Year each policy is projected to lapse if premiums stay exactly
  as billed today. Colors show what it would take to save the policy. {onset_line}</div>
  <div class="view-count">{wave_svg_count}</div>
  <div class="view-face" hidden>{wave_svg_face}</div>
</div>

{f'''<div class="card">
  <h2>When do exception premiums begin?</h2>
  <div class="note">For the {n_fmt(exc_wave['n'])} Exceptions-required policies: the
  year GLP exception premiums would need to start if each policy is funded at its
  minimum level premium from now on.</div>
  {exc_svg}
</div>''' if exc_wave['n'] else ''}

<div class="card" id="formmap">
  <h2>Where is it concentrated? — by policy form</h2>
  <div class="note">All categories: each form&rsquo;s policy mix. Pick a category to
  rank forms by how many of their policies fall in it.</div>
  <div class="chips" id="fm-chips">{chips}</div>
  {fm_views}
</div>

<div class="foot">
  {n_fmt(analyzed)} analyzed of {n_fmt(classified_total + cls['not_run'])} in scope ·
  {n_fmt(counts['NC'])} not classified{f" ({bypass_bits})" if bypass_bits else ""} ·
  {n_fmt(cls['not_run'])} not yet run{f" · {n_fmt(len(cls['unclassified']))} unclassifiable" if cls['unclassified'] else ""}.
  Forecasts run on a guideline-only basis (Conform to TEFRA on, Conform to TAMRA off)
  from each policy&rsquo;s inforce values as of its valuation date; a bridging lumpsum
  to the next premium due date is applied where needed.
  Source: {escape(workbook_name)} · SuiteView illustration engine GLP forecast batch.
</div>
</div>

<div class="tt" id="tt"></div>
<script>
(function () {{
  var tt = document.getElementById('tt');
  document.querySelectorAll('[data-tt]').forEach(function (el) {{
    el.addEventListener('pointermove', function (e) {{
      tt.textContent = el.getAttribute('data-tt');
      tt.style.display = 'block';
      var x = e.clientX + 14, y = e.clientY + 12;
      var r = tt.getBoundingClientRect();
      if (x + r.width > window.innerWidth - 8) x = e.clientX - r.width - 10;
      if (y + r.height > window.innerHeight - 8) y = e.clientY - r.height - 10;
      tt.style.left = x + 'px'; tt.style.top = y + 'px';
    }});
    el.addEventListener('pointerleave', function () {{ tt.style.display = 'none'; }});
  }});
  document.querySelectorAll('.toggle').forEach(function (tg) {{
    var card = tg.closest('.card');
    tg.querySelectorAll('button').forEach(function (btn) {{
      btn.addEventListener('click', function () {{
        tg.querySelectorAll('button').forEach(function (b) {{
          b.setAttribute('aria-pressed', b === btn ? 'true' : 'false');
        }});
        card.querySelector('.view-count').hidden = btn.dataset.view !== 'count';
        card.querySelector('.view-face').hidden = btn.dataset.view !== 'face';
      }});
    }});
  }});
  function selectCat(cat) {{
    document.querySelectorAll('#fm-chips .chip').forEach(function (c) {{
      c.classList.toggle('on', c.dataset.cat === cat);
    }});
    document.querySelectorAll('.fm-view').forEach(function (v) {{
      v.hidden = v.dataset.cat !== cat;
    }});
  }}
  document.querySelectorAll('#fm-chips .chip').forEach(function (c) {{
    c.addEventListener('click', function () {{ selectCat(c.dataset.cat); }});
  }});
  document.querySelectorAll('.fseg').forEach(function (s) {{
    s.addEventListener('click', function () {{
      selectCat(s.dataset.cat);
      document.getElementById('formmap').scrollIntoView({{behavior: 'smooth'}});
    }});
  }});
  var m = (location.hash || '').match(/cat=(\\w+)/);
  if (m) selectCat(m[1]);
}})();
</script>
</body></html>"""


# ── Entry point ─────────────────────────────────────────────────────────────

def main() -> None:
    argv = sys.argv[1:]
    if not argv:
        print(json.dumps({"error": "usage: build_glp_forecast_report.py "
                          "<workbook.xlsx> [--sheet NAME] [--out PATH]"}))
        sys.exit(1)
    workbook_path = argv[0]
    sheet = None
    out_path = None
    i = 1
    while i < len(argv):
        flag, val = argv[i], argv[i + 1] if i + 1 < len(argv) else None
        if flag == "--sheet":
            sheet = val
        elif flag == "--out":
            out_path = val
        else:
            print(json.dumps({"error": f"unknown flag: {flag}"}))
            sys.exit(1)
        i += 2

    rows, sheet_name = read_rows(workbook_path, sheet)
    cls = classify(rows)
    html = build_html(cls, workbook_name=Path(workbook_path).name,
                      region_note="Inforce block, CyberLife UL")
    out = Path(out_path) if out_path else (
        Path(workbook_path).parent / "GLP Funding Outlook.html")
    out.write_text(html, encoding="utf-8")

    cats = cls["cats"]
    front_load_n = sum(1 for p in cats["D"] if p["front_load_ok"])
    print(json.dumps({
        "workbook": workbook_path,
        "sheet": sheet_name,
        "rows_with_policy": len(rows),
        "not_run": cls["not_run"],
        "bypass_reasons": cls["bypass_reasons"],
        "unclassified": cls["unclassified"][:10],
        "unclassified_count": len(cls["unclassified"]),
        "missing_abs_max": cls["missing_abs_max"],
        "labels_seen": cls["labels_seen"],
        "category_counts": {CAT_SHORT[k]: len(v) for k, v in cats.items()},
        "front_loading_only_subset_of_exceptions": front_load_n,
        "category_face": {CAT_SHORT[k]: round(sum(p["face"] for p in v), 2)
                          for k, v in cats.items()},
        "forms": len({p["form"] for k in cats for p in cats[k]}),
        "output": str(out),
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
