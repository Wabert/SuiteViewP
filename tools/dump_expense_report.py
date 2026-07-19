"""Dump the RERUN "Expense Report" sheet's J:Y values table (headers, formulas, values).

Purpose: establish the J..Y column mapping (label -> formula -> sample values)
so the illustration report's supplemental Expense Report page can reproduce it
from engine outputs. Read-only openpyxl; never opens Excel COM.

Usage:
    venv\\Scripts\\python.exe tools/dump_expense_report.py [<workbook>] [<sheet>]

Defaults: "docs/Illustration_UL/RERUN (v20.0) local.xlsm", sheet "Expense Report".

Output (text):
  1. Header rows (first ~12 rows of J:Y, labels/values).
  2. Formulas for the first 3 data rows of each column J..Y.
  3. Cached values for a handful of spread-out data rows.
"""
import sys

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

DEFAULT_WB = r"docs/Illustration_UL/RERUN (v20.0) local.xlsm"
DEFAULT_SHEET = "Expense Report"
COL_FIRST, COL_LAST = 10, 25  # J..Y


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    sheet = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET

    wbf = openpyxl.load_workbook(path, read_only=True, data_only=False)
    wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wbf.sheetnames:
        print(f"Sheet {sheet!r} not found. Sheets: {wbf.sheetnames}")
        sys.exit(1)
    wsf, wsv = wbf[sheet], wbv[sheet]
    print(f"sheet={sheet} dims={wsf.calculate_dimension()} "
          f"max_row={wsf.max_row} max_col={wsf.max_column}")

    # 1. Header block: rows 1..14, all columns A..Y (context included).
    print("\n== HEADER BLOCK (rows 1-14, cols A..Y) ==")
    for r in range(1, 15):
        for c in range(1, COL_LAST + 1):
            cf = wsf.cell(row=r, column=c)
            cv = wsv.cell(row=r, column=c)
            formula = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else ""
            val = cv.value
            if formula == "" and (val is None or val == ""):
                continue
            print(f"{cf.coordinate} | {formula[:200]} | {str(val)[:60]}")

    # Find the first data row: first row >= 5 where col J holds a number/formula.
    first_data = None
    for r in range(5, min(wsf.max_row, 80) + 1):
        v = wsv.cell(row=r, column=COL_FIRST).value
        if isinstance(v, (int, float)):
            first_data = r
            break
    if first_data is None:
        print("\nNo numeric data row found in column J (rows 5-80).")
        wbf.close(); wbv.close()
        return
    print(f"\nfirst_data_row={first_data}")

    # 2. Formulas for the first 3 data rows of each column J..Y.
    print("\n== FORMULAS (first 3 data rows) ==")
    for c in range(COL_FIRST, COL_LAST + 1):
        letter = get_column_letter(c)
        for r in range(first_data, first_data + 3):
            cf = wsf.cell(row=r, column=c)
            f = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else repr(cf.value)
            print(f"{letter}{r} | {str(f)[:240]}")

    # 3. Cached values for spread-out rows.
    print("\n== SAMPLE VALUES ==")
    sample_rows = [first_data, first_data + 1, first_data + 4, first_data + 9,
                   first_data + 19, first_data + 39]
    header = "row  " + "".join(f"{get_column_letter(c):>16}" for c in range(COL_FIRST, COL_LAST + 1))
    print(header)
    for r in sample_rows:
        if r > wsf.max_row:
            continue
        cells = []
        for c in range(COL_FIRST, COL_LAST + 1):
            v = wsv.cell(row=r, column=c).value
            if isinstance(v, float):
                cells.append(f"{v:>16,.2f}")
            else:
                cells.append(f"{str(v):>16}")
        print(f"{r:<5}" + "".join(cells))

    wbf.close()
    wbv.close()


if __name__ == "__main__":
    main()
