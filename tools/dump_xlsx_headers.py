r"""Inspect a workbook's columns (and optionally a data row) as JSON.

Single-purpose helper for inspecting a sheet before/after wiring a batch tool to
it. With ``--row N`` it instead prints that row's values keyed by the row-1
header label, so you can confirm a tool wrote the right columns. Usage:

    venv\Scripts\python.exe tools/dump_xlsx_headers.py "path\to\book.xlsx" [Sheet]
    venv\Scripts\python.exe tools/dump_xlsx_headers.py "path\to\book.xlsx" --row 2
"""
from __future__ import annotations

import json
import sys

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: dump_xlsx_headers.py <workbook> [sheet|--row N]"}))
        sys.exit(1)
    path = sys.argv[1]
    rest = sys.argv[2:]
    row = None
    sheet = None
    if "--row" in rest:
        idx = rest.index("--row")
        row = int(rest[idx + 1])
        rest = rest[:idx] + rest[idx + 2:]
    if rest:
        sheet = rest[0]

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    header_by_letter = {
        cell.column_letter: str(cell.value)
        for cell in ws[1]
        if cell.value is not None
    }
    if row is None:
        print(json.dumps({"sheet": ws.title, "sheets": wb.sheetnames,
                          "headers": header_by_letter}, indent=2, ensure_ascii=False))
        return

    values = {}
    for cell in ws[row]:
        label = header_by_letter.get(cell.column_letter)
        if label is not None:
            values[label] = cell.value
    print(json.dumps({"sheet": ws.title, "row": row, "values": values},
                     indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
