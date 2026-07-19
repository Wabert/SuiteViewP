"""Generate docs/Illustration_UL/Testing/TEST_MATRIX.xlsx from test_matrix.json.

The JSON file (docs/Illustration_UL/Testing/test_matrix.json) is the source of
truth for the illustration-engine baseline test matrix; the workbook is a
regenerable, auditor-facing rendering of it. Never edit the xlsx by hand.

openpyxl (batch/headless) is correct here per the Excel-export convention:
this is file generation for the repo, not an interactive "Dump to Excel".

Columns (fixed):
    Company | Policy | Form | DB Option | Description | Forecast Inputs |
    PASS/FAIL | Comments | GLP/GSP | Contributions | Distributions |
    Policy Debt | EAV | ESV

Per-quantity cells: "EXACT" (green), a number -> rendered "D <max abs delta>"
(amber), "FAIL" (red), "-" (grey, N/A or not yet compared).
PASS/FAIL cell: PASS (green) / FAIL (red) / PARTIAL (amber) / PENDING (grey).

Usage:
    venv\\Scripts\\python.exe tools/build_test_matrix.py
Writes the workbook, re-reads it, and prints a JSON self-check summary.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
TESTING_DIR = ROOT / "docs" / "Illustration_UL" / "Testing"
JSON_PATH = TESTING_DIR / "test_matrix.json"
XLSX_PATH = TESTING_DIR / "TEST_MATRIX.xlsx"

HEADERS = [
    "Company", "Policy", "Form", "DB Option", "Description", "Forecast Inputs",
    "PASS/FAIL", "Comments", "GLP/GSP", "Contributions", "Distributions",
    "Policy Debt", "EAV", "ESV",
]
ROW_KEYS = [
    "company", "policy", "form", "db_option", "description", "forecast_inputs",
    "status", "comments", "glp_gsp", "contributions", "distributions",
    "policy_debt", "eav", "esv",
]
QUANTITY_COLS = set(range(9, 15))  # 1-based columns I..N
STATUS_COL = 7

# House header styling (SuiteView blue/gold).
HEADER_FILL = PatternFill("solid", fgColor="0D3A7A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GOLD_SIDE = Side(style="thin", color="D4A017")

# Status color scheme.
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FONT_GREEN = Font(color="006100", size=9)
FILL_AMBER = PatternFill("solid", fgColor="FFEB9C")
FONT_AMBER = Font(color="9C6500", size=9)
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FONT_RED = Font(color="9C0006", size=9)
FILL_GREY = PatternFill("solid", fgColor="D9D9D9")
FONT_GREY = Font(color="595959", size=9)
FONT_BODY = Font(size=9)

# Text columns get fixed generous widths + wrapping; the rest autofit.
WRAP_WIDTHS = {5: 52, 6: 46, 8: 70}
NA_STRINGS = {"-", "—", ""}


def quantity_display(value):
    """Return (text, fill, font) for a per-quantity cell value."""
    if isinstance(value, (int, float)):
        text = f"Δ {value:,.2f}"
        return text, FILL_AMBER, FONT_AMBER
    s = str(value).strip()
    if s == "EXACT":
        return s, FILL_GREEN, FONT_GREEN
    if s == "FAIL":
        return s, FILL_RED, FONT_RED
    if s in NA_STRINGS:
        return "—", FILL_GREY, FONT_GREY
    raise ValueError(f"Unknown quantity value: {value!r}")


def status_display(value):
    s = str(value).strip().upper()
    styles = {
        "PASS": (FILL_GREEN, FONT_GREEN),
        "FAIL": (FILL_RED, FONT_RED),
        "PARTIAL": (FILL_AMBER, FONT_AMBER),
        "PENDING": (FILL_GREY, FONT_GREY),
    }
    if s not in styles:
        raise ValueError(f"Unknown status: {value!r}")
    fill, font = styles[s]
    return s, fill, Font(bold=True, color=font.color.rgb, size=9)


def build() -> dict:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    rows = data["rows"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Matrix"

    for c, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=GOLD_SIDE)
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(ROW_KEYS, start=1):
            value = row.get(key, "")
            if c in QUANTITY_COLS:
                text, fill, font = quantity_display(value)
            elif c == STATUS_COL:
                text, fill, font = status_display(value)
            else:
                text, fill, font = str(value), None, FONT_BODY
            cell = ws.cell(row=r, column=c, value=text)
            cell.font = font
            if fill is not None:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(
                    vertical="top", wrap_text=(c in WRAP_WIDTHS))

    # Column widths: wrap columns fixed, others autofit to content (capped).
    for c in range(1, len(HEADERS) + 1):
        letter = get_column_letter(c)
        if c in WRAP_WIDTHS:
            ws.column_dimensions[letter].width = WRAP_WIDTHS[c]
            continue
        longest = max(
            [len(str(ws.cell(row=r, column=c).value or ""))
             for r in range(1, len(rows) + 2)])
        ws.column_dimensions[letter].width = min(max(longest + 2, 9), 24)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)

    # Self-check: re-read and verify shape + status tallies.
    check = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    cws = check["Test Matrix"]
    grid = list(cws.iter_rows(values_only=True))
    check.close()
    assert len(grid) == len(rows) + 1, (
        f"row count {len(grid)} != {len(rows) + 1}")
    assert len(grid[0]) == len(HEADERS), (
        f"column count {len(grid[0])} != {len(HEADERS)}")
    assert list(grid[0]) == HEADERS, "header row mismatch"

    tally: dict[str, int] = {}
    for g in grid[1:]:
        tally[str(g[STATUS_COL - 1])] = tally.get(str(g[STATUS_COL - 1]), 0) + 1
    return {
        "workbook": str(XLSX_PATH),
        "rows": len(rows),
        "columns": len(HEADERS),
        "status_tally": tally,
        "self_check": "ok",
    }


if __name__ == "__main__":
    try:
        print(json.dumps(build(), indent=1))
    except Exception as exc:  # noqa: BLE001 - report and fail loudly
        print(json.dumps({"error": str(exc)}))
        sys.exit(1)
