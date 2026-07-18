"""Bridge between RERUN's VBA and the local offline SQLite exports.

RERUN's ``mdl_LocalData`` VBA module (installed by
tools/rerun_install_local_vba.py) shells this script when INPUT!sDataSource
starts with "Local".  It reads an args JSON file (path in argv[1]), writes a
result .xlsx the VBA pastes from, and a status file the VBA checks:

  policy mode  — args {"mode":"policy","policy":"UE209026","region":"CKPR",
                       "skip_benefits":false,"out":"<x.xlsx>","status":"<s.txt>"}
      out.xlsx sheet INPUTS: name | value | kind ("scalar" writes Range(name),
      "vector1" writes Range(name).Rows(1)); sheet META: A1 warnings,
      A2.. plancodes involved.

  rates mode   — args {"mode":"rates","plancodes":["1U147500",...],"state":"AA",
                       "out":"<x.xlsx>","status":"<s.txt>"}
      out.xlsx: one sheet per Span_* block (raw rows, no header) ready to paste
      onto the same-named range; sheet META: A1 warnings, A2.. target-family
      plancodes (for vPlancodesPresent).

The status file gets {"ok": true/false, "error": ...} — written LAST so the
VBA can poll for completion.  All heavy lifting reuses the verified logic in
tools/rerun_build_case_inputs.py and tools/rerun_load_local_rates.py.
"""
from __future__ import annotations

import json
import os
import sys
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ["SUITEVIEW_LOCAL_DATA"] = "1"

# The 8 vINPUT vectors: VBA writes row 1 only; PopulateInputFormulas propagates.
_VECTOR_NAMES = {
    "vINPUT_Specified_Amount", "vINPUT_DBO", "vINPUT_Premium_Amount",
    "vINPUT_Premium_Mode", "vINPUT_Loans", "vINPUT_Loan_Mode",
    "vINPUT_Loan_Repayment", "vINPUT_Withdrawal",
}
# Saved-Cases-only row labels that are not workbook defined names.
_NOT_RANGE_NAMES = {"TimeStamp", "sUID"}


def _policy_mode(cmd: dict) -> tuple[Path, list[str]]:
    from rerun_build_case_inputs import build_mapping
    from suiteview.illustration.core.illustration_policy_service import build_illustration_data
    from suiteview.core.policy_service import get_policy_info

    policy = cmd["policy"]
    region = cmd.get("region") or "CKPR"
    warnings: list[str] = []

    data = build_illustration_data(policy, region=region, company_code=cmd.get("company"))
    mapping = build_mapping(data, warnings, skip_benefits=bool(cmd.get("skip_benefits")))

    # INPUT-only fields PopulateInputSheet sets that aren't Saved Cases rows.
    pi = get_policy_info(policy, region=region)
    mapping["sCompany"] = data.company_code
    mapping["sStatus"] = (pi.status_code if pi is not None else "") or ""
    mapping["sMD_From_Cyberlife"] = (
        data.system_coi_charge + data.system_expense_charge + data.system_other_charge)
    mapping["sblnPrintMode"] = False
    mapping["sINPUT_VaryingAssumeRate"] = "FALSE"
    mapping["sReportFileName"] = ""
    mapping["sINPUT_LoanDuration"] = ""
    mapping["sBaseCovCount"] = len(data.segments)
    mapping["sTermRiderCount"] = sum(
        1 for i in (1, 2, 3) if mapping.get(f"sINPUT_R{i}_Boolean"))

    rows = []
    for name, val in mapping.items():
        if name in _NOT_RANGE_NAMES:
            continue
        if isinstance(val, list):
            if name in _VECTOR_NAMES:
                rows.append((name, val[0], "vector1"))
            else:
                warnings.append(f"{name}: unexpected vector; skipped")
        else:
            rows.append((name, val, "scalar"))

    plancodes = [data.plancode] + [
        str(mapping.get(f"sINPUT_R{i}_Plancode") or "").strip()
        for i in (1, 2, 3) if mapping.get(f"sINPUT_R{i}_Boolean")
    ]

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INPUTS"
    ws.append(("name", "value", "kind"))
    for r in rows:
        ws.append(r)
    meta = wb.create_sheet("META")
    meta["A1"] = "; ".join(warnings)
    for i, pc in enumerate(p for p in plancodes if p):
        meta.cell(row=i + 2, column=1).value = pc
    out = Path(cmd["out"])
    wb.save(out)
    return out, warnings


def _rates_mode(cmd: dict) -> tuple[Path, list[str]]:
    from rerun_load_local_rates import expand_plancodes, build_blocks

    plancodes = [str(p).strip() for p in cmd["plancodes"] if str(p).strip() not in ("", "0")]
    state = str(cmd.get("state") or "AA").strip() or "AA"

    fams, warnings = expand_plancodes(plancodes)
    blocks, w2 = build_blocks(fams, state)
    warnings.extend(w2)

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for span, rows in blocks.items():
        ws = wb.create_sheet(span)
        for row in rows:
            ws.append(tuple(row))
    meta = wb.create_sheet("META")
    meta["A1"] = "; ".join(warnings)
    for i, pc in enumerate(sorted(fams["targets"])):
        meta.cell(row=i + 2, column=1).value = pc
    out = Path(cmd["out"])
    wb.save(out)
    return out, warnings


def main():
    args_path = Path(sys.argv[1])
    cmd = json.loads(args_path.read_text(encoding="utf-8-sig"))
    status_path = Path(cmd["status"])
    try:
        if cmd["mode"] == "policy":
            out, warnings = _policy_mode(cmd)
        elif cmd["mode"] == "rates":
            out, warnings = _rates_mode(cmd)
        else:
            raise ValueError(f"unknown mode {cmd['mode']!r}")
        status = {"ok": True, "out": str(out), "warnings": warnings}
    except Exception as exc:
        status = {"ok": False, "error": f"{exc}", "trace": traceback.format_exc()[-1500:]}
    status_path.write_text(json.dumps(status, indent=2, default=str), encoding="utf-8")
    print(json.dumps(status, indent=2, default=str))


if __name__ == "__main__":
    main()
