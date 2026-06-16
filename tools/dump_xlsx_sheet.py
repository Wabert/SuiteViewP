"""Dump a sheet from an .xlsx/.xlsm file as JSON, including formulas AND cached values.

Usage:
    dump_xlsx_sheet.py <file>                       -> list sheet names
    dump_xlsx_sheet.py <file> <sheet>               -> dump non-empty cells (formula + value)
    dump_xlsx_sheet.py <file> <sheet> <max_rows>    -> limit rows scanned (default 400)

For each non-empty cell emits: coord, row, col, formula (if the cell holds one),
and value (the cached computed value). openpyxl must be loaded twice because a
single load can give either formulas (data_only=False) or cached values
(data_only=True), never both.
"""
import sys
import json

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: venv\\Scripts\\python.exe -m pip install openpyxl"}))
    sys.exit(1)


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: dump_xlsx_sheet.py <file> [sheet] [max_rows]"}))
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    max_rows = int(sys.argv[3]) if len(sys.argv) > 3 else 400

    wb_f = openpyxl.load_workbook(file_path, data_only=False, read_only=True)

    if sheet_name is None:
        print(json.dumps({"sheets": wb_f.sheetnames}, indent=2))
        return

    wb_v = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    cells = []
    for row_f, row_v in zip(ws_f.iter_rows(max_row=max_rows), ws_v.iter_rows(max_row=max_rows)):
        for cf, cv in zip(row_f, row_v):
            formula = cf.value
            value = cv.value
            if formula is None and value is None:
                continue
            entry = {"coord": cf.coordinate, "row": cf.row, "col": cf.column}
            if isinstance(formula, str) and formula.startswith("="):
                entry["formula"] = formula
                entry["value"] = value
            else:
                entry["value"] = formula  # plain literal / label
            cells.append(entry)

    print(json.dumps({
        "sheet": sheet_name,
        "max_dim": ws_f.calculate_dimension(),
        "cell_count": len(cells),
        "cells": cells,
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
