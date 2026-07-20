"""Extract Policy number + Valuation Date for every row of a GLP batch workbook.

Reads the ``Batch`` sheet (Policy in col B, Valuation Date in col U) and emits a
JSON list of ``{"company", "policy", "valuation_date"}`` records to stdout. This
is the input list for a follow-up batch policy-debt lookup — the workbook itself
carries no debt column.

Usage:
    venv\\Scripts\\python.exe tools/extract_glp_policy_valuation.py "path\\to\\book.xlsx" [Sheet]
"""
from __future__ import annotations

import json
import sys

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def _norm_date(value) -> str | None:
    if value is None:
        return None
    try:
        return value.date().isoformat()
    except AttributeError:
        return str(value)


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: extract_glp_policy_valuation.py <workbook> [sheet]"}))
        sys.exit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    records = []
    for row in ws.iter_rows(min_row=2):
        company = row[0].value if len(row) > 0 else None   # A
        policy = row[1].value if len(row) > 1 else None     # B
        valuation = row[20].value if len(row) > 20 else None  # U
        if policy is None:
            continue
        policy_str = str(policy).strip()
        if not policy_str:
            continue
        records.append({
            "company": str(company).strip() if company is not None else None,
            "policy": policy_str,
            "valuation_date": _norm_date(valuation),
        })

    print(json.dumps({"sheet": ws.title, "count": len(records),
                      "records": records}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
