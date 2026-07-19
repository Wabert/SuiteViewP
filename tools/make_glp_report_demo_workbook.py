r"""Write a synthetic GLP-forecast workbook for exercising the funding report.

Generates a workbook with the same headers as the real GLP batch (via
batch_runner.GLP_COLUMNS) and a plausible mix of outcomes, so
tools/build_glp_forecast_report.py can be developed and visually verified on
the minipc before the real batch runs on the work laptop. Seeded — output is
deterministic. TEST DATA ONLY; policy numbers are fake.

Usage:
    venv\Scripts\python.exe tools/make_glp_report_demo_workbook.py <out.xlsx> [n_rows]
"""
from __future__ import annotations

import json
import random
import sys
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

from suiteview.illustration.core.batch_runner import GLP_COLUMNS  # noqa: E402

VAL_DATE = date(2026, 6, 1)


def rand_date(rng, y0: int, y1: int) -> date:
    d0, d1 = date(y0, 1, 1), date(y1, 12, 31)
    return d0 + timedelta(days=rng.randint(0, (d1 - d0).days))


def hump_year(rng, lo: int, peak: int, hi: int) -> int:
    return int(round(rng.triangular(lo, hi, peak)))


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: make_glp_report_demo_workbook.py "
                          "<out.xlsx> [n_rows]"}))
        sys.exit(1)
    out = Path(sys.argv[1])
    n = int(sys.argv[2]) if len(sys.argv) > 2 else 2400
    rng = random.Random(20260718)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet3"
    ws.append(["Company", "Policy"] + [label for _, label in GLP_COLUMNS])
    key_col = {key: i + 3 for i, (key, _) in enumerate(GLP_COLUMNS)}

    plancodes = ["1U130M29", "1U133300", "1S133B2X", "1U130N2X", "1U1F4M00"]
    # (form, pick-weight, cumulative category weights: bypass/A/B/C/D/E) —
    # categories correlate with form so the by-form map has structure. "C" here
    # means exc date + abs-max "Maturity" (the front-loading-only subset of D).
    form_profiles = [
        ("FPL83",   22, (0.03, 0.30, 0.28, 0.04, 0.31, 0.04)),
        ("FUL86",   18, (0.02, 0.40, 0.30, 0.03, 0.24, 0.01)),
        ("FPL85",   14, (0.02, 0.52, 0.28, 0.02, 0.155, 0.005)),
        ("UL-100",  13, (0.02, 0.62, 0.24, 0.02, 0.10, 0.00)),
        ("UL-90",   10, (0.01, 0.72, 0.20, 0.01, 0.06, 0.00)),
        ("FUL92",    9, (0.02, 0.78, 0.16, 0.01, 0.03, 0.00)),
        ("ISL-88",   8, (0.01, 0.86, 0.11, 0.00, 0.02, 0.00)),
        ("UL-2000",  6, (0.01, 0.93, 0.055, 0.00, 0.005, 0.00)),
    ]
    form_names = [f for f, _, _ in form_profiles]
    form_weights = [w for _, w, _ in form_profiles]
    form_mix = {f: mix for f, _, mix in form_profiles}
    counts = {}
    for i in range(n):
        r = ws.max_row + 1
        policy = f"UL{i:06d}"
        face = round(rng.choice([25, 50, 50, 75, 100, 100, 150, 250, 500]) * 1000
                     * rng.uniform(0.9, 1.1), 0)
        billing = round(face * rng.uniform(0.006, 0.016) / 12, 2)
        issue = rand_date(rng, 1983, 1996)
        issue_age = rng.randint(25, 60)
        mat_age = rng.choice([95, 95, 96, 100])
        maturity = date(issue.year + (mat_age - issue_age), issue.month, 1)

        form = rng.choices(form_names, weights=form_weights, k=1)[0]
        w_byp, w_a, w_b, w_c, w_d, w_e = form_mix[form]
        u = rng.random()
        if u < w_byp:
            tier = "bypass"
        elif u < w_byp + w_a:
            tier = "A"
        elif u < w_byp + w_a + w_b:
            tier = "B"
        elif u < w_byp + w_a + w_b + w_c:
            tier = "C"
        elif u < w_byp + w_a + w_b + w_c + w_d:
            tier = "D"
        else:
            tier = "E"
        counts[tier] = counts.get(tier, 0) + 1

        ws.cell(row=r, column=1, value="01")
        ws.cell(row=r, column=2, value=policy)

        def put(key, value):
            ws.cell(row=r, column=key_col[key], value=value)

        put("plancode", rng.choice(plancodes))
        put("form", form)
        put("issue_date", issue)
        put("issue_age", issue_age)
        put("maturity_date", maturity)
        put("maturity_age", mat_age)
        put("face", face)
        put("db_option", rng.choice(["A - Level", "B - Increasing"]))
        put("def_life_ins", "GPT")
        put("billing_prem", billing)
        put("billing_mode", rng.choice(["Monthly", "Quarterly", "Annual"]))
        put("gsp", round(face * 0.09, 2))
        put("glp", round(face * 0.0085, 2))
        put("accum_lp", round(face * 0.0085 * rng.uniform(10, 40), 2))
        put("prem_td", round(billing * 12 * rng.uniform(8, 35), 2))
        put("accum_wd", 0)
        put("valuation_date", VAL_DATE)
        put("suspense_code", "0 - Active")

        if tier == "bypass":
            put("run_status", rng.choice(["bypass (MD)", "bypass (rates missing)",
                                          "bypass (A)", "bypass (CVAT)",
                                          "bypass (MD, rates missing)"]))
            continue
        put("run_status", "Complete")

        if tier == "A":
            put("lapse_no_prem", rand_date(rng, 2027, 2040) if rng.random() < 0.7
                else "Maturity")
            put("lapse_cur_prem", "Maturity")
            put("exc_date", "not needed" if rng.random() < 0.5 else "(none)")
            put("level_prem", round(billing * rng.uniform(0.5, 0.98), 2))
            put("lapse_abs_max", "Maturity")
            continue

        lapse_year = hump_year(rng, 2026, 2036, 2058)
        lapse = rand_date(rng, lapse_year, lapse_year)
        put("lapse_no_prem", rand_date(rng, 2026, max(2027, lapse_year - 2)))
        put("lapse_cur_prem", lapse)
        if rng.random() < 0.25:
            put("lumpsum", round(rng.uniform(200, 6000), 2))

        if tier == "B":
            mult = rng.lognormvariate(0.9, 0.55)
            put("level_prem", round(billing * max(mult, 1.05), 2))
            put("exc_date", "(none)")
            put("lapse_abs_max", "Maturity")
        elif tier == "C":
            put("level_prem", round(billing * rng.uniform(2.5, 8), 2))
            put("exc_date", rand_date(rng, max(2026, lapse_year - 4), lapse_year))
            put("lapse_abs_max", "Maturity")
        elif tier == "D":
            put("level_prem", round(billing * rng.uniform(2, 12), 2))
            exc_year = hump_year(rng, 2026, 2033, min(lapse_year, 2050))
            put("exc_date", rand_date(rng, exc_year, exc_year))
            put("lapse_abs_max", rand_date(rng, lapse_year,
                                           min(lapse_year + 12, 2070)))
        else:  # E
            put("exc_date", "no solution")
            put("lapse_abs_max", rand_date(rng, lapse_year,
                                           min(lapse_year + 6, 2070)))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(json.dumps({"output": str(out), "rows": n, "mix": counts}))


if __name__ == "__main__":
    main()
