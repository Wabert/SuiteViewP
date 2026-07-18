"""Dump the Saved Cases input-name list (col A) and one case's values from RERUN.

Usage:
    dump_saved_case_names.py <workbook> [<case_number>]

Prints one line per Saved Cases row: row | name | value-for-case (if requested).
Vector names repeat (e.g. vINPUT_Premium_Amount) — repeats are shown as-is.
"""
import sys
import json

import openpyxl

sys.path.insert(0, __file__.rsplit("\\", 1)[0].rsplit("/", 1)[0])
from rerun_com import _resolve_case_column  # noqa: E402


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: dump_saved_case_names.py <workbook> [<case>]"}))
        sys.exit(1)
    path = sys.argv[1]
    case = sys.argv[2] if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Saved Cases"]
    col = _resolve_case_column(ws, case) if case else None

    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            continue
        val = ws.cell(row=r, column=col).value if col else ""
        print(f"{r} | {str(name).strip()} | {val}")
    wb.close()


if __name__ == "__main__":
    main()
