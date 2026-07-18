"""Dump one row (or a row range) of a sheet from an .xlsx/.xlsm: formulas + cached values.

Usage:
    dump_xlsx_row.py <file> <sheet> <row> [<end_row>] [<max_col_letter>]

Emits one line per non-empty cell: coord | formula | value  (formula truncated to 160 chars).
Complements tools/dump_xlsx_sheet.py (which scans from the top and is noisy for deep rows).
"""
import sys
import json

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def main():
    if len(sys.argv) < 4:
        print(json.dumps({"error": "Usage: dump_xlsx_row.py <file> <sheet> <row> [<end_row>] [<max_col_letter>]"}))
        sys.exit(1)
    path, sheet = sys.argv[1], sys.argv[2]
    r0 = int(sys.argv[3])
    r1 = int(sys.argv[4]) if len(sys.argv) > 4 else r0
    max_col = column_index_from_string(sys.argv[5]) if len(sys.argv) > 5 else None

    wbf = openpyxl.load_workbook(path, read_only=True, data_only=False)
    wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
    wsf, wsv = wbf[sheet], wbv[sheet]
    ncols = max_col or wsf.max_column

    for r in range(r0, r1 + 1):
        for c in range(1, ncols + 1):
            cf = wsf.cell(row=r, column=c)
            cv = wsv.cell(row=r, column=c)
            formula = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else ""
            val = cv.value
            if formula == "" and (val is None or val == ""):
                continue
            print(f"{cf.coordinate} | {formula[:160]} | {str(val)[:50]}")
    wbf.close()
    wbv.close()


if __name__ == "__main__":
    main()
