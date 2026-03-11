"""Check Row 31 specifically."""
from openpyxl import load_workbook

wb = load_workbook("docs/DB2_Field_Mapping_Review.xlsx")
ws = wb.active

# Row 31 data is in Excel row 33 (header + 1-indexed)
for r in range(2, ws.max_row + 1):
    num = ws.cell(r, 1).value
    if num == 31:
        print(f"Row {num}:")
        for c in range(1, 9):
            h = ws.cell(1, c).value
            v = ws.cell(r, c).value
            print(f"  Col {c} ({h}): {v!r}")
        break
