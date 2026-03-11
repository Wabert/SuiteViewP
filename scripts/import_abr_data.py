"""
One-time script to extract all rate data from the ABR Quote Excel workbook
and populate:
  1. suiteview/abrquote/models/vbt_2008.py  (embedded VBT mortality data)
  2. ~/.suiteview/abr_quote.db              (term rates, interest rates, per diem)

Usage:
    python scripts/import_abr_data.py

Requires the workbook at:
    ABRQuote/ABR Quote System Signature Term(v5.6) - WIP.xlsm
"""

import os
import sys
import textwrap

# Add project root to path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

import openpyxl


WB_PATH = os.path.join(project_root, "ABRQuote",
                       "ABR Quote System Signature Term(v5.6) - WIP.xlsm")

VBT_OUTPUT = os.path.join(project_root, "suiteview", "abrquote", "models", "vbt_2008.py")


def extract_vbt_2008(wb):
    """Extract 2008 VBT mortality table and generate vbt_2008.py source file.

    Sheet: '2008 VBT'
    Layout: 4 blocks stacked vertically, each 121 data rows × 100 age cols.
    Block positions (verified by probing):
      MN: label row 1,   header row 2,   data rows 3-123,   cols B(2)–CW(101) = ages 0-99
      FN: label row 125,  header row 126, data rows 127-247, cols B(2)–CW(101)
      MS: label row 249,  header row 250, data rows 251-371, cols B(2)–CW(101)
      FS: label row 373,  header row 374, data rows 375-495, cols B(2)–CW(101)
    Col A = duration number (1-121), Cols B-CW = issue ages 0-99
    """
    ws = wb["2008 VBT"]
    print("Extracting 2008 VBT mortality table...")

    blocks = {
        "MN": {"start_row": 3,   "start_col": 2},     # col B = 2, ages 0-99 in cols 2-101
        "FN": {"start_row": 127, "start_col": 2},
        "MS": {"start_row": 251, "start_col": 2},
        "FS": {"start_row": 375, "start_col": 2},
    }

    # First, let's detect the actual layout by scanning for data
    # Check header row for issue ages (should be 0-99)
    data = {}

    for block_name, pos in blocks.items():
        sr = pos["start_row"]
        sc = pos["start_col"]
        block_data = []

        for dur in range(121):  # durations 1-121 (rows)
            row_idx = sr + dur
            row_data = []
            for age in range(100):  # issue ages 0-99 (cols)
                col_idx = sc + age
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    val = 0.0
                row_data.append(float(val))
            block_data.append(row_data)

        data[block_name] = block_data
        # Verify: check a few values aren't all zero
        nonzero = sum(1 for row in block_data for v in row if v > 0)
        print(f"  {block_name}: 121×100, {nonzero} non-zero values")

    # Generate Python source file
    print(f"Writing {VBT_OUTPUT}...")

    with open(VBT_OUTPUT, "w", encoding="utf-8") as f:
        f.write('"""\n')
        f.write("2008 Valuation Basic Table (VBT) — Select mortality rates.\n\n")
        f.write("Auto-generated from ABR Quote System Signature Term (v5.6) workbook.\n")
        f.write("DO NOT EDIT — regenerate with: python scripts/import_abr_data.py\n\n")
        f.write("Blocks:\n")
        f.write("    MN = Male Non-smoker    FN = Female Non-smoker\n")
        f.write("    MS = Male Smoker        FS = Female Smoker\n\n")
        f.write("Indexing: VBT_DATA[block][duration_year - 1][issue_age]\n")
        f.write("    duration_year: 1-121\n")
        f.write("    issue_age: 0-99\n")
        f.write("    rates: per 1,000 lives\n")
        f.write('"""\n\n')
        f.write("from typing import Optional\n\n\n")

        # Write the data as a compact dict
        f.write("# 2008 VBT Select mortality rates: 4 blocks × 121 durations × 100 issue ages\n")
        f.write("# Rates are per 1,000 lives\n")
        f.write("VBT_DATA: dict[str, list[list[float]]] = {\n")

        for block_name in ["MN", "FN", "MS", "FS"]:
            block = data[block_name]
            f.write(f'    "{block_name}": [\n')
            for dur_idx, row in enumerate(block):
                # Compact format: one row per duration
                vals = ", ".join(f"{v:.6f}" if v > 0 else "0" for v in row)
                f.write(f"        [{vals}],\n")
            f.write("    ],\n")

        f.write("}\n\n\n")

        # Write lookup function
        f.write(textwrap.dedent('''\
            def get_qx(block: str, issue_age: int, duration_year: int) -> float:
                """Look up annual mortality rate (qx) from 2008 VBT.

                Args:
                    block: "MN", "FN", "MS", or "FS"
                    issue_age: 0-99
                    duration_year: 1-121

                Returns:
                    Mortality rate per 1,000 lives.
                    Returns 1000.0 (certainty of death) if duration >= 121 or age >= 100.
                """
                if duration_year >= 121 or issue_age >= 100:
                    return 1000.0
                if duration_year < 1 or issue_age < 0:
                    return 0.0
                if block not in VBT_DATA:
                    raise ValueError(f"Unknown VBT block: {block!r}. Use MN, FN, MS, or FS.")
                return VBT_DATA[block][duration_year - 1][issue_age]


            def get_qx_decimal(block: str, issue_age: int, duration_year: int) -> float:
                """Like get_qx() but returns rate as a decimal (0-1) instead of per-1000."""
                return get_qx(block, issue_age, duration_year) / 1000.0
        '''))

    file_size = os.path.getsize(VBT_OUTPUT)
    print(f"  Generated {VBT_OUTPUT}")
    print(f"  File size: {file_size:,} bytes")
    return data


def extract_term_rates(wb, db):
    """Extract BaseTermPremiumRates into abr_quote.db."""
    ws = wb["BaseTermPremiumRates"]
    print("\nExtracting BaseTermPremiumRates...")

    # Rows 1-2 = headers (duplicate header rows)
    # Data starts at row 3
    # Layout: col1=Key, col2=sort-key, col3=Product, col4=LevelPeriod,
    #         col5=Sex, col6=RateClass, col7=Band, col8=Age,
    #         cols 9-85 = rate years 1-77
    rows = []
    row_idx = 3
    while True:
        key = ws.cell(row=row_idx, column=1).value
        if key is None:
            break

        plancode = ws.cell(row=row_idx, column=3).value or ""
        sex = ws.cell(row=row_idx, column=5).value or ""
        rate_class = ws.cell(row=row_idx, column=6).value or ""
        band = ws.cell(row=row_idx, column=7).value or 0
        issue_age = ws.cell(row=row_idx, column=8).value or 0

        # Rates in columns 9 through 85 (years 1-77), pad to 82 with zeros
        rates = []
        for col in range(9, 86):  # 77 rate columns (years 1-77)
            val = ws.cell(row=row_idx, column=col).value
            if val is None or (isinstance(val, str) and val.startswith('#')):
                rates.append(0.0)
            else:
                try:
                    rates.append(float(val))
                except (ValueError, TypeError):
                    rates.append(0.0)
        # Pad to 82 columns to match schema
        while len(rates) < 82:
            rates.append(0.0)

        row_tuple = (str(key), str(plancode), str(sex), str(rate_class),
                     int(band), int(issue_age), *rates)
        rows.append(row_tuple)
        row_idx += 1

        if row_idx % 5000 == 0:
            print(f"  Read {row_idx - 2} rows...")

    print(f"  Total rows: {len(rows)}")

    if rows:
        db.bulk_insert_term_rates(rows)
        db.update_import_metadata("term_rates", len(rows), WB_PATH)
        print(f"  Inserted {len(rows)} rows into term_rates")

    return len(rows)


def extract_interest_rates(wb, db):
    """Extract ABR Interest Rate tab into abr_quote.db."""
    ws = wb["ABR Interest Rate"]
    print("\nExtracting ABR Interest Rates...")

    # Row 5 = headers ("Date", "Rate", "IUL Var Loan Rate")
    # Data starts at row 6 (rows 1-4 are text/notes)
    rows = []
    row_idx = 6
    while True:
        date_val = ws.cell(row=row_idx, column=1).value

        if date_val is None:
            break

        rate_val = ws.cell(row=row_idx, column=2).value

        # Convert date to string format
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime("%Y-%m")
        elif date_val is not None:
            date_str = str(date_val)
        else:
            row_idx += 1
            continue

        rate = float(rate_val) if rate_val is not None else 0.0

        # IUL var loan rate in column 3
        iul = ws.cell(row=row_idx, column=3).value
        iul_val = float(iul) if iul is not None else None

        rows.append((date_str, rate, iul_val))
        row_idx += 1

    print(f"  Total rows: {len(rows)}")

    if rows:
        db.bulk_insert_interest_rates(rows)
        db.update_import_metadata("interest_rates", len(rows), WB_PATH)
        print(f"  Inserted {len(rows)} rows into interest_rates")

    return len(rows)


def extract_per_diem(wb, db):
    """Extract Per Diem data from the ABR Interest Rate sheet."""
    ws = wb["ABR Interest Rate"]
    print("\nExtracting Per Diem limits...")

    # Per Diem data is in cols H-J (8-10) of the ABR Interest Rate sheet
    # col 8 = year, col 9 = daily limit, col 10 = annual limit
    rows = []
    for r in range(6, 20):
        yr = ws.cell(row=r, column=8).value
        if isinstance(yr, (int, float)) and 2020 <= yr <= 2030:
            daily = ws.cell(row=r, column=9).value
            annual = ws.cell(row=r, column=10).value
            if daily is not None and annual is not None:
                rows.append((int(yr), float(daily), float(annual)))

    if not rows:
        # Hardcode known values from workbook extraction
        print("  Using hardcoded per diem values")
        rows = [
            (2022, 390, 142_350),
            (2024, 410, 150_060),
            (2025, 420, 153_300),
            (2026, 430, 156_950),
        ]

    print(f"  Total rows: {len(rows)}")

    if rows:
        db.bulk_insert_per_diem(rows)
        db.update_import_metadata("per_diem", len(rows), WB_PATH)
        print(f"  Inserted {len(rows)} rows into per_diem")

    return len(rows)


def main():
    print("=" * 70)
    print("ABR Quote Data Import")
    print(f"Source: {WB_PATH}")
    print("=" * 70)

    if not os.path.exists(WB_PATH):
        print(f"\nERROR: Workbook not found at {WB_PATH}")
        sys.exit(1)

    print("\nLoading workbook (data_only=True for values)...")
    wb = openpyxl.load_workbook(WB_PATH, data_only=True, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # 1. Extract VBT (generates Python source file)
    # Need the non-read-only workbook for VBT since values might be there
    wb_full = openpyxl.load_workbook(WB_PATH, data_only=True)
    vbt_data = extract_vbt_2008(wb_full)
    wb_full.close()

    # 2. Extract term rates + interest rates into SQLite
    from suiteview.abrquote.models.abr_database import ABRDatabase

    db = ABRDatabase()
    db.initialize_schema()
    print(f"\nDatabase: {db.db_path}")

    # Use read-only workbook for the large tables
    # But we need regular mode for cell access
    wb.close()
    wb2 = openpyxl.load_workbook(WB_PATH, data_only=True)

    extract_term_rates(wb2, db)
    extract_interest_rates(wb2, db)
    extract_per_diem(wb2, db)

    # Summary
    print("\n" + "=" * 70)
    print("IMPORT COMPLETE")
    print(f"  VBT 2008:       4 blocks × 121 × 100 → {VBT_OUTPUT}")
    print(f"  Term rates:     {db.term_rate_count():,} rows → {db.db_path}")
    print(f"  Interest rates: {db.interest_rate_count():,} rows → {db.db_path}")
    meta = db.get_import_metadata()
    for tbl, info in meta.items():
        print(f"    {tbl}: {info['row_count']} rows, imported {info['imported_at']}")
    print("=" * 70)

    wb2.close()
    db.close()


if __name__ == "__main__":
    main()
