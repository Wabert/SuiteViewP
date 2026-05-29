"""Create an Excel workbook for forward annuity rider value verification."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_PATH = Path("docs/polview/Annuity_Rider_Forward_Verification.xlsx")
MAX_EVENTS = 900
MAX_START_BUCKETS = 30
MAX_TRANSACTIONS = 250


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(bold=True, color="1F4E78")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    inputs = wb.active
    inputs.title = "Inputs"
    buckets = wb.create_sheet("Starting Buckets")
    transactions = wb.create_sheet("Transactions")
    output = wb.create_sheet("Projection")
    calc = wb.create_sheet("Calc")

    build_inputs(inputs)
    build_starting_buckets(buckets)
    build_transactions(transactions)
    build_projection(output)
    build_calc(calc)

    calc.sheet_state = "hidden"
    output_path = save_workbook(wb, OUTPUT_PATH)

    # Verify the generated workbook can be opened by openpyxl.
    load_workbook(output_path, data_only=False).close()
    print(output_path)


def save_workbook(wb: Workbook, output_path: Path) -> Path:
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_updated{output_path.suffix}")
        wb.save(fallback)
        return fallback


def title(ws, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, text)
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center")


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = SUBHEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def build_inputs(ws) -> None:
    title(ws, "Annuity Rider Forward Verification Inputs", 4)
    rows = [
        ("Starting Date", date(2008, 6, 16), "Enter the date you want to roll forward from."),
        ("Starting Cash Value", 0, "Cash value as of the starting date."),
        ("New Money Annual Rate", 0.06, "New premiums earn this rate until their roll date."),
        ("Portfolio Annual Rate", 0.04, "Rolled money and portfolio value earn this rate."),
        ("Projection End Date", "=TODAY()", "Projection runs through this date."),
        ("Withdrawal Bucket Rule", "Portfolio only", "Withdrawals reduce portfolio value only; active new-money buckets are not reduced."),
    ]
    for idx, (label, value, note) in enumerate(rows, start=3):
        ws.cell(idx, 1, label).font = HEADER_FONT
        ws.cell(idx, 2, value)
        ws.cell(idx, 2).fill = INPUT_FILL if idx not in (8,) else CALC_FILL
        ws.cell(idx, 3, note)
        ws.cell(idx, 1).border = BORDER
        ws.cell(idx, 2).border = BORDER
        ws.cell(idx, 3).border = BORDER

    ws["B3"].number_format = "m/d/yyyy"
    ws["B4"].number_format = "$#,##0.00"
    ws["B5"].number_format = "0.00%"
    ws["B6"].number_format = "0.00%"
    ws["B7"].number_format = "m/d/yyyy"

    ws["A11"] = "How to use"
    ws["A11"].font = HEADER_FONT
    notes = [
        "1. Enter the starting date and cash value above.",
        "2. Enter active new-money buckets on the Starting Buckets sheet. Each bucket needs either a bucket effective date or a roll-date override.",
        "3. Enter premiums and withdrawals on the Transactions sheet as positive amounts with Type = Premium or Withdrawal.",
        "4. Projection shows every month-end and transaction date from the starting date through the projection end date.",
        "5. The hidden Calc sheet keeps bucket-level state. Unhide it if you want to audit every bucket lane.",
    ]
    for row_offset, note in enumerate(notes, start=12):
        ws.cell(row_offset, 1, note)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 92
    ws.freeze_panes = "A3"


def build_starting_buckets(ws) -> None:
    title(ws, "Starting New Money Buckets", 6)
    headers = [
        "Bucket Effective Date", "Roll Date Override", "Starting Bucket Value",
        "Calculated Roll Date", "Bucket Label", "Notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col, header)
    style_header_row(ws, 3, 1, len(headers))

    for row in range(4, 4 + MAX_START_BUCKETS):
        ws.cell(row, 4, f'=IF(B{row}<>"",B{row},IF(A{row}<>"",EOMONTH(A{row},12),""))')
        ws.cell(row, 5, f'=IF(C{row}<>"","Start Bucket "&ROW()-3,"")')
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).border = BORDER
        ws.cell(row, 1).fill = INPUT_FILL
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 4).fill = CALC_FILL
        ws.cell(row, 5).fill = CALC_FILL
        ws.cell(row, 1).number_format = "m/d/yyyy"
        ws.cell(row, 2).number_format = "m/d/yyyy"
        ws.cell(row, 3).number_format = "$#,##0.00"
        ws.cell(row, 4).number_format = "m/d/yyyy"

    ws["A4"] = date(2008, 5, 20)
    ws["C4"] = 0
    ws["A4"].comment = Comment(
        "Example only. Replace with the original bucket date. The starting value is the bucket value as of Inputs!Starting Date.",
        "Copilot",
    )
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40
    ws.freeze_panes = "A4"


def build_transactions(ws) -> None:
    title(ws, "Premiums and Withdrawals", 6)
    headers = ["Effective Date", "Type", "Amount", "Code/Label", "Notes", "Calculated Roll Date"]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col, header)
    style_header_row(ws, 3, 1, len(headers))

    dv = DataValidation(type="list", formula1='"Premium,Withdrawal"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B4:B{3 + MAX_TRANSACTIONS}")

    for row in range(4, 4 + MAX_TRANSACTIONS):
        ws.cell(row, 6, f'=IF(AND(A{row}<>"",B{row}="Premium"),EOMONTH(A{row},12),"")')
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).border = BORDER
        for col in (1, 2, 3, 4, 5):
            ws.cell(row, col).fill = INPUT_FILL
        ws.cell(row, 6).fill = CALC_FILL
        ws.cell(row, 1).number_format = "m/d/yyyy"
        ws.cell(row, 3).number_format = "$#,##0.00"
        ws.cell(row, 6).number_format = "m/d/yyyy"

    ws["A4"] = date(2008, 7, 18)
    ws["B4"] = "Premium"
    ws["C4"] = 15
    ws["D4"] = "PR"
    ws["A5"] = date(2008, 9, 15)
    ws["B5"] = "Withdrawal"
    ws["C5"] = 0
    ws["D5"] = "SN"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 20
    ws.freeze_panes = "A4"


def build_projection(ws) -> None:
    title(ws, "Forward Projection", 8)
    headers = [
        "Effective Date", "Event Type", "Premiums", "Withdrawals", "Cash Value",
        "Portfolio Value", "Total New Money Value", "Calced New Money Buckets",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(3, col, header)
    style_header_row(ws, 3, 1, len(headers))

    for row in range(4, 4 + MAX_EVENTS):
        idx = row - 3
        if row == 4:
            ws.cell(row, 1, "=Inputs!$B$3")
        else:
            ws.cell(row, 1, next_event_date_formula(row))
        ws.cell(row, 2, f'=IF(A{row}="","",Calc!B{idx + 1})')
        ws.cell(row, 3, f'=IF(A{row}="","",Calc!C{idx + 1})')
        ws.cell(row, 4, f'=IF(A{row}="","",Calc!D{idx + 1})')
        ws.cell(row, 5, f'=IF(A{row}="","",Calc!I{idx + 1})')
        ws.cell(row, 6, f'=IF(A{row}="","",Calc!G{idx + 1})')
        ws.cell(row, 7, f'=IF(A{row}="","",Calc!H{idx + 1})')
        ws.cell(row, 8, f'=IF(A{row}="","",Calc!K{idx + 1})')
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).border = BORDER
        ws.cell(row, 1).number_format = "m/d/yyyy"
        for col in range(3, 8):
            ws.cell(row, col).number_format = "$#,##0.00"

    widths = [18, 28, 15, 15, 16, 18, 22, 60]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"


def build_calc(ws) -> None:
    bucket_count = MAX_START_BUCKETS + MAX_TRANSACTIONS
    before_start_col = 12
    after_start_col = before_start_col + bucket_count
    roll_start_col = after_start_col + bucket_count
    label_start_col = roll_start_col + bucket_count
    last_col = label_start_col + bucket_count - 1

    headers = [
        "Event Date", "Event Type", "Premiums", "Withdrawals", "Days", "Portfolio Before Tx",
        "Portfolio After Tx", "Total New Money", "Cash Value", "Withdrawal After Portfolio",
        "Bucket Display",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)

    for idx in range(bucket_count):
        ws.cell(1, before_start_col + idx, f"Before Bucket {idx + 1}")
        ws.cell(1, after_start_col + idx, f"After Bucket {idx + 1}")
        ws.cell(1, roll_start_col + idx, f"Roll Date {idx + 1}")
        ws.cell(1, label_start_col + idx, f"Label {idx + 1}")

    style_header_row(ws, 1, 1, last_col)

    for row in range(2, 2 + MAX_EVENTS):
        prev_row = row - 1
        projection_row = row + 2
        ws.cell(row, 1, f'=Projection!A{projection_row}')
        ws.cell(row, 2, event_type_formula(row))
        ws.cell(row, 3, f'=IF(A{row}="","",SUMIFS(Transactions!$C$4:$C$253,Transactions!$A$4:$A$253,A{row},Transactions!$B$4:$B$253,"Premium"))')
        ws.cell(row, 4, f'=IF(A{row}="","",SUMIFS(Transactions!$C$4:$C$253,Transactions!$A$4:$A$253,A{row},Transactions!$B$4:$B$253,"Withdrawal"))')
        ws.cell(row, 5, f'=IF(A{row}="","",IF(ROW()=2,0,A{row}-A{prev_row}))')
        ws.cell(row, 6, portfolio_before_formula(row, prev_row, before_start_col, after_start_col, roll_start_col, bucket_count))
        ws.cell(row, 10, f'=IF(A{row}="","",MAX(0,D{row}-MAX(0,F{row})))')
        ws.cell(row, 7, f'=IF(A{row}="","",F{row}-D{row})')
        ws.cell(row, 8, f'=IF(A{row}="","",SUM({col_range(after_start_col, row, bucket_count)}))')
        ws.cell(row, 9, f'=IF(A{row}="","",G{row}+H{row})')
        ws.cell(row, 11, f'=IF(A{row}="","",TEXTJOIN(", ",TRUE,{bucket_display_range(row, after_start_col, bucket_count)}))')

        for bucket_idx in range(bucket_count):
            before_col = before_start_col + bucket_idx
            after_col = after_start_col + bucket_idx
            roll_col = roll_start_col + bucket_idx
            label_col = label_start_col + bucket_idx
            ws.cell(row, before_col, bucket_before_formula(row, prev_row, bucket_idx, before_col, after_col, roll_col))
            ws.cell(row, after_col, bucket_after_formula(row, before_start_col, before_col))
            ws.cell(row, roll_col, roll_date_formula(row, bucket_idx))
            ws.cell(row, label_col, bucket_label_formula(row, bucket_idx))

        for col in range(1, last_col + 1):
            ws.cell(row, col).number_format = "$#,##0.00" if col not in (1, 2, 5, 11) and col < roll_start_col else "General"
        ws.cell(row, 1).number_format = "m/d/yyyy"

    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["K"].width = 80
    ws.freeze_panes = "A2"


def next_event_date_formula(row: int) -> str:
    prev_row = row - 1
    return (
        f'=IF(A{prev_row}="","",IF(A{prev_row}>=Inputs!$B$7,"",'
        f'MIN(Inputs!$B$7,'
        f'IF(EOMONTH(A{prev_row},0)>A{prev_row},EOMONTH(A{prev_row},0),EOMONTH(A{prev_row},1)),'
        f'IFERROR(AGGREGATE(15,6,Transactions!$A$4:$A$253/'
        f'(Transactions!$A$4:$A$253>A{prev_row})/'
        f'(Transactions!$A$4:$A$253<=Inputs!$B$7),1),Inputs!$B$7))))'
    )


def event_type_formula(row: int) -> str:
    return (
        f'=IF(A{row}="","",IF(A{row}=Inputs!$B$3,"Start",'
        f'IF(AND(C{row}>0,D{row}>0),"Premium, Withdrawal",'
        f'IF(C{row}>0,"Premium",IF(D{row}>0,"Withdrawal",'
        f'IF(A{row}=Inputs!$B$7,"Projection End","Month End"))))))'
    )


def portfolio_before_formula(row: int, prev_row: int, before_start_col: int, after_start_col: int, roll_start_col: int, bucket_count: int) -> str:
    if row == 2:
        return f'=IF(A{row}="","",Inputs!$B$4-SUM(\'{"Starting Buckets"}\'!$C$4:$C$33))'
    after_range = col_range(after_start_col, prev_row, bucket_count)
    roll_range = col_range(roll_start_col, row, bucket_count)
    rolled_sum = (
        f'SUMPRODUCT(({after_range}>0)*({roll_range}>A{prev_row})*({roll_range}<=A{row})*'
        f'{after_range}*POWER(1+Inputs!$B$5,({roll_range}-A{prev_row})/365)*'
        f'POWER(1+Inputs!$B$6,(A{row}-{roll_range})/365))'
    )
    return f'=IF(A{row}="","",G{prev_row}*POWER(1+Inputs!$B$6,E{row}/365)+{rolled_sum})'


def bucket_before_formula(row: int, prev_row: int, bucket_idx: int, before_col: int, after_col: int, roll_col: int) -> str:
    before = f"{get_column_letter(before_col)}{row}"
    after_prev = f"{get_column_letter(after_col)}{prev_row}"
    roll = f"{get_column_letter(roll_col)}{row}"
    if bucket_idx < MAX_START_BUCKETS:
        bucket_row = bucket_idx + 4
        if row == 2:
            return f'=IF(A{row}="","",IF(AND(\'Starting Buckets\'!$C${bucket_row}>0,{roll}>A{row}),\'Starting Buckets\'!$C${bucket_row},0))'
        return f'=IF(A{row}="","",IF(OR({after_prev}=0,{roll}<=A{row}),0,{after_prev}*POWER(1+Inputs!$B$5,E{row}/365)))'

    trans_row = bucket_idx - MAX_START_BUCKETS + 4
    if row == 2:
        return f'=IF(A{row}="","",IF(AND(Transactions!$B${trans_row}="Premium",Transactions!$A${trans_row}=A{row}),Transactions!$C${trans_row},0))'
    return (
        f'=IF(A{row}="","",IF(AND(Transactions!$B${trans_row}="Premium",Transactions!$A${trans_row}=A{row}),Transactions!$C${trans_row},'
        f'IF(OR({after_prev}=0,{roll}<=A{row}),0,{after_prev}*POWER(1+Inputs!$B$5,E{row}/365))))'
    )


def bucket_after_formula(row: int, before_start_col: int, before_col: int) -> str:
    before = f"{get_column_letter(before_col)}{row}"
    return f'=IF(A{row}="","",{before})'


def roll_date_formula(row: int, bucket_idx: int) -> str:
    if bucket_idx < MAX_START_BUCKETS:
        bucket_row = bucket_idx + 4
        return f'=IF(\'Starting Buckets\'!$D${bucket_row}<>"",\'Starting Buckets\'!$D${bucket_row},0)'
    trans_row = bucket_idx - MAX_START_BUCKETS + 4
    return f'=IF(Transactions!$F${trans_row}<>"",Transactions!$F${trans_row},0)'


def bucket_label_formula(row: int, bucket_idx: int) -> str:
    if bucket_idx < MAX_START_BUCKETS:
        bucket_row = bucket_idx + 4
        return f'=IF(\'Starting Buckets\'!$E${bucket_row}<>"",\'Starting Buckets\'!$E${bucket_row},"")'
    trans_row = bucket_idx - MAX_START_BUCKETS + 4
    return f'=IF(Transactions!$B${trans_row}="Premium","Premium "&TEXT(Transactions!$A${trans_row},"m/d/yyyy"),"")'


def col_range(start_col: int, row: int, width: int) -> str:
    return f"{get_column_letter(start_col)}{row}:{get_column_letter(start_col + width - 1)}{row}"


def bucket_display_range(row: int, after_start_col: int, bucket_count: int) -> str:
    start = get_column_letter(after_start_col)
    end = get_column_letter(after_start_col + bucket_count - 1)
    return f'IF({start}{row}:{end}{row}>0,TEXT({start}{row}:{end}{row},"#,##0.00"),"")'


if __name__ == "__main__":
    main()