"""
Generate an Excel workbook with the field mapping review data.
"""
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb_src = xlrd.open_workbook("docs/COBOLDB2translation.xls")
ws_src = wb_src.sheet_by_name("Translation")

# Build field indexes from workbook
field_to_tables = {}
table_to_fields = {}
for r in range(1, ws_src.nrows):
    table = str(ws_src.cell_value(r, 3)).strip()
    field = str(ws_src.cell_value(r, 4)).strip()
    desc = str(ws_src.cell_value(r, 5)).strip() if ws_src.ncols > 5 else ""
    if table and field:
        if field not in field_to_tables:
            field_to_tables[field] = []
        field_to_tables[field].append((table, desc))
        if table not in table_to_fields:
            table_to_fields[table] = {}
        table_to_fields[table][field] = desc

# ── Data ────────────────────────────────────────────────────────────────
# (confidence, code_table, code_field, suggested_table, suggested_field, reasoning)
rows = [
    # HIGH CONFIDENCE
    ("HIGH", "LH_BAS_POL", "BL_DAY_NBR", "LH_BAS_POL", "BIL_DAY_NBR", "Workbook has BIL_DAY_NBR = 'Billing Day'"),
    ("HIGH", "LH_BAS_POL", "NXT_MVRY_DT", "LH_BAS_POL", "NXT_MVRY_PRC_DT", "Workbook: 'Next Monthliversary Date'"),
    ("HIGH", "LH_BAS_POL", "TMN_DT", "LH_BAS_POL", "PLN_TMN_DT", "Workbook: 'Next Potential Termination Event Date'"),
    ("HIGH", "LH_COV_PHA", "ANN_PRM_AMT", "LH_COV_PHA", "ANN_PRM_UNT_AMT", "Workbook: 'Annual Premium Per Unit'"),
    ("HIGH", "LH_COV_PHA", "ELM_PER_CD", "LH_COV_PHA", "AH_ACC_ELM_PER_CD", "DI elimination period (or AH_SIC_ELM_PER_CD)"),
    ("HIGH", "LH_COV_PHA", "BNF_PER_CD", "LH_COV_PHA", "AH_ACC_BNF_PER_CD", "DI benefit period (or AH_SIC_BNF_PER_CD)"),
    ("HIGH", "LH_COV_INS_RNL_RT", "ISS_AGE", "LH_COV_PHA", "INS_ISS_AGE", "Issue Age lives on LH_COV_PHA"),
    ("HIGH", "LH_BNF_INS_RNL_RT", "ISS_AGE", "LH_COV_PHA", "INS_ISS_AGE", "Same pattern"),
    ("HIGH", "LH_COV_TARGET", "TAR_VAL_AMT", "LH_COV_TARGET", "TAR_PRM_AMT", "Likely meant 'Target Premium Amount'"),
    ("HIGH", "LH_CSH_VAL_LOAN", "LN_ITS_RT", "LH_CSH_VAL_LOAN", "LN_CRG_ITS_RT", "Workbook: 'Interest Rate charged on this loan'"),
    ("HIGH", "LH_FND_VAL_LOAN", "LN_ITS_RT", "LH_FND_VAL_LOAN", "LN_CRG_ITS_RT", "Same pattern as CSH_VAL_LOAN"),
    ("HIGH", "LH_FND_VAL_LOAN", "LN_ITS_AMT", "LH_FND_VAL_LOAN", "POL_LN_ITS_AMT", "Workbook has this on same table"),
    ("HIGH", "LH_POL_FND_VAL_TOT", "BKT_STR_DT", "LH_POL_FND_VAL_TOT", "ITS_PER_STR_DT", "Workbook: 'Period Start Date'"),
    ("HIGH", "LH_APPLIED_PTP", "PTP_APL_TYP_CD", "LH_APPLIED_PTP", "CK_PTP_TYP_CD", "Workbook: 'C=Coupon, D=Dividend'"),
    ("HIGH", "LH_UNAPPLIED_PTP", "PTP_TYP_CD", "LH_UNAPPLIED_PTP", "CK_PTP_TYP_CD", "Same pattern"),
    ("HIGH", "LH_PTP_ON_DEP", "PTP_TYP_CD", "LH_PTP_ON_DEP", "CK_PTP_TYP_CD", "Same pattern"),
    ("HIGH", "LH_ONE_YR_TRM_ADD", "OYT_FCE_AMT", "LH_ONE_YR_TRM_ADD", "OYT_ADD_AMT", "Only amount field on this table"),
    ("HIGH", "LH_PAID_UP_ADD", "PUA_FCE_AMT", "LH_PAID_UP_ADD", "PUA_AMT", "Only amount field on this table"),

    # MEDIUM CONFIDENCE
    ("MEDIUM", "LH_BAS_POL", "ISSUE_DT", "LH_COV_PHA", "ISSUE_DT", "Field exists on LH_COV_PHA, not LH_BAS_POL; may be a view alias"),
    ("MEDIUM", "LH_BAS_POL", "NXT_ANV_DT", "", "", "Workbook has LST_ANV_DT but no NXT_ANV_DT; might be computed"),
    ("MEDIUM", "LH_BAS_POL", "REG_PRM_AMT", "LH_BAS_POL", "POL_PRM_AMT", "Workbook: 'Mode Premium' — may not be same semantics"),
    ("MEDIUM", "TH_BAS_POL", "TAR_PRM_AMT", "LH_POL_TARGET", "TAR_PRM_AMT", "Exists on target tables; TH_BAS_POL only has 12 fields in workbook"),
    ("MEDIUM", "TH_BAS_POL", "DTH_BNF_PLN_OPT_CD", "LH_NON_TRD_POL", "DTH_BNF_PLN_OPT_CD", "Exists on LH_NON_TRD_POL; TH_BAS_POL underspec'd in workbook"),
    ("MEDIUM", "TH_BAS_POL", "TFDF_CD", "LH_NON_TRD_POL", "TFDF_CD", "Exists on LH_NON_TRD_POL"),
    ("MEDIUM", "TH_BAS_POL", "MIN_PRM_AMT", "", "", "No clear match; TH_BAS_POL only 12 fields in workbook"),
    ("MEDIUM", "LH_COV_PHA", "PRM_PAY_STS_CD", "", "", "No clear match; may be a view alias for premium pay status"),
    ("MEDIUM", "LH_COV_PHA", "TMN_DT", "LH_BAS_POL", "PLN_TMN_DT", "On LH_BAS_POL, not LH_COV_PHA"),
    ("MEDIUM", "LH_COV_PHA", "VLU_PER_UNT_AMT", "LH_COV_PHA", "COV_VPU_AMT", "Value Per Unit — code already has fallback to COV_VPU_AMT"),
    ("MEDIUM", "LH_CSH_VAL_LOAN", "LN_ITS_STS_CD", "LH_CSH_VAL_LOAN", "LN_ITS_AMT_TYP_CD", "Possibly 'Interest Amount Type Code'"),
    ("MEDIUM", "LH_FND_VAL_LOAN", "LN_ITS_STS_CD", "LH_FND_VAL_LOAN", "LN_ITS_AMT_TYP_CD", "Same pattern"),
    ("MEDIUM", "LH_POL_FND_VAL_TOT", "CRE_ITS_RT", "", "", "No clear match on this table"),
    ("MEDIUM", "LH_POL_MVRY_VAL", "CSH_SUR_VAL_AMT", "", "", "Probably a view column name"),
    ("MEDIUM", "LH_POL_MVRY_VAL", "DTH_BNF_AMT", "", "", "Probably a view column name"),
    ("MEDIUM", "LH_LN_RPY_TRM", "LN_ITS_AMT", "LH_LN_RPY_TRM", "LN_RPY_AMT", "Workbook: 'Repayment Amount'"),
    ("MEDIUM", "FH_FIXED", "TRN_TYP_CD", "FH_FIXED", "TRANS", "Workbook uses TRANS for transaction type"),
    ("MEDIUM", "FH_FIXED", "TRN_SBY_CD", "FH_FIXED", "TRANS", "May be embedded in TRANS field"),
    ("MEDIUM", "FH_FIXED", "TOT_TRS_AMT", "FH_FIXED", "GROSS_AMT", "Workbook names gross amount differently"),
    ("MEDIUM", "FH_FIXED", "ACC_VAL_GRS_AMT", "FH_FIXED", "NET_AMT", "Workbook: NET_AMT for accumulation value"),
    ("MEDIUM", "FH_FIXED", "FND_ID_CD", "FH_FIXED", "FUND_ID", "Different naming convention"),
    ("MEDIUM", "FH_FIXED", "COV_PHA_NBR", "FH_FIXED", "PHASE", "Different naming convention"),

    # LOW CONFIDENCE — Workbook likely incomplete for these
    ("LOW", "LH_BAS_POL", "STS_CD", "", "", "No match at all — status code; likely a view alias"),
    ("LOW", "LH_BAS_POL", "GRC_IND", "", "", "No match — grace indicator"),
    ("LOW", "LH_BAS_POL", "DEFRA_IND", "", "", "No match at all"),
    ("LOW", "LH_BAS_POL", "GSP_AMT", "", "", "Guideline Single Premium — not in workbook"),
    ("LOW", "LH_BAS_POL", "GLP_AMT", "", "", "Guideline Level Premium — not in workbook"),
    ("LOW", "TH_BAS_POL", "VUL_IND", "", "", "TH_BAS_POL only 12 fields in workbook; likely more on actual DB2"),
    ("LOW", "TH_BAS_POL", "IUL_IND", "", "", "Same — workbook TH_BAS_POL underspecified"),
    ("LOW", "LH_COV_PHA", "ISSUE_AGE_YR_NBR", "LH_COV_PHA", "INS_ISS_AGE", "May be alternate field for issue age"),
    ("LOW", "TH_COV_PHA", "CV_AMT", "", "", "TH_COV_PHA only 6 fields in workbook"),
    ("LOW", "TH_COV_PHA", "NSP_AMT", "", "", "TH_COV_PHA only 6 fields in workbook"),
    ("LOW", "TH_COV_PHA", "OPT_EXER_IND", "", "", "GIO option exercise indicator — not in workbook"),
    ("LOW", "LH_BNF_INS_RNL_RT", "JT_INS_IND", "LH_COV_INS_RNL_RT", "JT_INS_IND", "Table not in workbook; exists on LH_COV_INS_RNL_RT"),
    ("LOW", "LH_COV_SKIPPED_PER", "SKP_FRM_DT", "", "", "Table exists in workbook but these fields not listed"),
    ("LOW", "LH_COV_SKIPPED_PER", "SKP_TO_DT", "", "", "Table exists in workbook but these fields not listed"),
    ("LOW", "LH_COV_SKIPPED_PER", "SKP_TYP_CD", "", "", "Table exists in workbook but these fields not listed"),
    ("LOW", "LH_AGT_COM_AMT", "COM_PCT", "LH_AGT_COM_RT", "COM_RT_PCT", "LH_AGT_COM_AMT only 10 fields; COM_RT_PCT is on LH_AGT_COM_RT"),
    ("LOW", "LH_AGT_COM_AMT", "MKT_ORG_CD", "", "", "Not found on LH_AGT_COM_AMT"),
    ("LOW", "LH_AGT_COM_AMT", "SVC_AGT_IND", "", "", "Not found on LH_AGT_COM_AMT"),
    ("LOW", "LH_TAMRA_MEC_PRM", "MEC_IND", "", "", "Table not well documented in workbook"),
    ("LOW", "LH_LN_RPY_TRM", "PMT_NBR", "", "", "Table exists but field names differ"),
    ("LOW", "LH_LN_RPY_TRM", "PMT_DT", "", "", "No match found"),
    ("LOW", "LH_LN_RPY_TRM", "PMT_AMT", "", "", "No match found"),
    ("LOW", "LH_LN_RPY_TRM", "LN_PRI_AMT", "LH_CSH_VAL_LOAN", "LN_PRI_AMT", "Exists on CSH_VAL_LOAN / FND_VAL_LOAN, not RPY_TRM"),
    ("LOW", "LH_APPLIED_PTP", "PTP_APL_DT", "", "", "No match found"),
    ("LOW", "LH_APPLIED_PTP", "PTP_GRS_AMT", "", "", "No match found"),
    ("LOW", "LH_APPLIED_PTP", "PTP_NET_AMT", "", "", "No match found"),
    ("LOW", "LH_APPLIED_PTP", "POL_DUR_NBR", "LH_POL_MVRY_VAL", "POL_DUR_NBR", "Exists on other tables, not LH_APPLIED_PTP"),
    ("LOW", "LH_UNAPPLIED_PTP", "PTP_PRO_DT", "", "", "No match found"),
    ("LOW", "LH_UNAPPLIED_PTP", "PTP_GRS_AMT", "", "", "No match found"),
    ("LOW", "LH_UNAPPLIED_PTP", "PTP_NET_AMT", "", "", "No match found"),
    ("LOW", "LH_UNAPPLIED_PTP", "POL_DUR_NBR", "LH_POL_MVRY_VAL", "POL_DUR_NBR", "Exists on other tables"),
    ("LOW", "LH_ONE_YR_TRM_ADD", "COV_PHA_NBR", "LH_COV_PHA", "COV_PHA_NBR", "Exists on LH_COV_PHA, not this table"),
    ("LOW", "LH_ONE_YR_TRM_ADD", "OYT_ISS_DT", "", "", "No match found"),
    ("LOW", "LH_ONE_YR_TRM_ADD", "OYT_CSV_AMT", "LH_ONE_YR_TRM_ADD", "OYT_ADD_AMT", "Only amount field on this table"),
    ("LOW", "LH_PAID_UP_ADD", "PUA_ISS_DT", "", "", "No match found"),
    ("LOW", "LH_PAID_UP_ADD", "PUA_CSV_AMT", "LH_PAID_UP_ADD", "PUA_AMT", "Only amount field on this table"),
    ("LOW", "LH_PTP_ON_DEP", "DEP_DT", "", "", "No match found"),
    ("LOW", "LH_PTP_ON_DEP", "CUM_DEP_AMT", "LH_PTP_ON_DEP", "PTP_DEP_AMT", "Partial — workbook: deposit amount"),
    ("LOW", "LH_PTP_ON_DEP", "ITS_AMT", "LH_PTP_ON_DEP", "DEP_ITS_AMT", "Partial — workbook: 'Interest Amount'"),
]

# ── Build Excel ─────────────────────────────────────────────────────────
wb_out = Workbook()
ws = wb_out.active
ws.title = "Field Mapping Review"

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = [
    "#", "Confidence", "Code Table", "Code Field",
    "Suggested Table", "Suggested Field", "Reasoning",
    "Action (fill in)"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 26
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 26
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 55
ws.column_dimensions["H"].width = 30

# Data rows
for i, (conf, code_tbl, code_fld, sug_tbl, sug_fld, reason) in enumerate(rows, 1):
    row_num = i + 1
    ws.cell(row=row_num, column=1, value=i).border = thin_border

    conf_cell = ws.cell(row=row_num, column=2, value=conf)
    conf_cell.border = thin_border
    if conf == "HIGH":
        conf_cell.fill = high_fill
    elif conf == "MEDIUM":
        conf_cell.fill = med_fill
    else:
        conf_cell.fill = low_fill

    for col, val in [(3, code_tbl), (4, code_fld), (5, sug_tbl), (6, sug_fld), (7, reason)]:
        c = ws.cell(row=row_num, column=col, value=val)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True)

    # Action column — blank for user
    action_cell = ws.cell(row=row_num, column=8, value="")
    action_cell.border = thin_border

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

# Save
output_path = "docs/DB2_Field_Mapping_Review.xlsx"
wb_out.save(output_path)
print(f"Saved {len(rows)} entries to {output_path}")
