"""Temporary script to extract shadow account values from RERUN workbook."""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('RERUN (v19.1).xlsm', data_only=True, read_only=True)
ws = wb['CalcEngine']

wp_col = column_index_from_string('WP')
xw_col = column_index_from_string('XW')

# Headers
headers = []
for col in range(wp_col, xw_col + 1):
    h = ws.cell(row=5, column=col).value or get_column_letter(col)
    headers.append(f"{get_column_letter(col)}:{h}")

sep = " | "
print(sep.join(headers))
print("-" * 200)

# Data rows 6-18 (month 0 through ~12)
for row in range(6, 19):
    vals = []
    for col in range(wp_col, xw_col + 1):
        v = ws.cell(row=row, column=col).value
        if v is None:
            vals.append("")
        elif isinstance(v, (int, float)):
            vals.append(f"{v:.6f}" if abs(v) < 1 else f"{v:.2f}")
        else:
            vals.append(str(v))
    print(f"Row {row}: {sep.join(vals)}")

# Also get some reference columns: duration, attained age, AV
print("\n=== Reference columns (month, att_age, AV, DB) ===")
for row in range(6, 19):
    dur = ws.cell(row=row, column=column_index_from_string('F')).value
    att = ws.cell(row=row, column=column_index_from_string('L')).value
    av = ws.cell(row=row, column=column_index_from_string('UL')).value
    db = ws.cell(row=row, column=column_index_from_string('UP')).value
    print(f"Row {row}: dur={dur}, att_age={att}, AV={av}, DB={db}")

wb.close()
