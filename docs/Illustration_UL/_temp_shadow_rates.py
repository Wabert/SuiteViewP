"""Extract PolicyRates shadow columns and check CalcEngine for active shadow."""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('RERUN (v19.1).xlsm', data_only=True, read_only=True)

# --- PolicyRates: Shadow rate columns GM-GU ---
ws_pr = wb['PolicyRates']
print("=== PolicyRates: Shadow Columns GM-GU, Rows 1-8 ===")
gm = column_index_from_string('GM')
gu = column_index_from_string('GU')
for row in range(1, 9):
    for col in range(gm, gu + 1):
        v = ws_pr.cell(row=row, column=col).value
        if v is not None:
            print(f"  {get_column_letter(col)}{row} = {v}")
    print()

# Headers at row 3 and 4
print("=== PolicyRates: Shadow Headers ===")
for col in range(gm, gu + 1):
    h3 = ws_pr.cell(row=3, column=col).value
    h4 = ws_pr.cell(row=4, column=col).value
    print(f"  {get_column_letter(col)}: row3={h3}, row4={h4}")

# --- Check CalcEngine for vCCV_Active and sShadowInherent ---
ws_ce = wb['CalcEngine']
print("\n=== CalcEngine: checking vCCV_Active ===")

# Search for vCCV_Active in named ranges - look at a few key columns
# Let's check what INPUT has for CCV
ws_in = wb['INPUT']
# Search a few key cells
for row in range(1, 100):
    for col in range(1, 30):
        v = ws_in.cell(row=row, column=col).value
        if v and isinstance(v, str) and 'CCV' in str(v).upper():
            print(f"  INPUT {get_column_letter(col)}{row} = {v}")
        if v and isinstance(v, str) and 'Shadow' in str(v):
            print(f"  INPUT {get_column_letter(col)}{row} = {v}")

wb.close()
